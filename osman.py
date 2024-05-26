import pickle
import streamlit as st
import streamlit_authenticator as stauth
from pathlib import Path
from PIL import Image
import pandas as pd
import numpy as np
from streamlit_option_menu import option_menu
import openpyxl
from openpyxl.styles import Font, PatternFill
import tempfile

# User Authentication
names = ["TI Polda NF 1", "TI Polda NF 2"]
usernames = ["admin1", "admin2"]

# load hashed kd_akses
file_path = Path(__file__).parent / "hashed_pw.pkl"
with file_path.open("rb") as file:
    hashed_kd_akses = pickle.load(file)

authenticator = stauth.Authenticate(
    names, usernames, hashed_kd_akses, "lookup", "abcdef")
name, authentication_status, username = authenticator.login("Login", "main")

if authentication_status == False:
    st.error("Username/kode akses salah")

if authentication_status == None:
    st.warning("Silahkan masukan username dan kode akses")

url = "https://osmanlokpts.streamlit.app/"
url_panduan = "https://docs.google.com/document/d/1zc9W_Tt51J9POZaybez1KBVbNWhpEiRsGsfo5VsRQJI/edit?usp=sharing"

if authentication_status:
    authenticator.logout("Logout", "sidebar")
    with st.sidebar:
        # Tombol untuk URL utama dengan warna GreenYellow
        st.markdown(f'''
<a href="{url}"><button style="background-color:GreenYellow; border:none; color:white; padding:10px 24px; text-align:center; display:inline-block; font-size:16px; margin:4px 2px; cursor:pointer;">Report Lokasi</button></a>
''', unsafe_allow_html=True)

        # Tombol untuk URL panduan dengan warna Tomato
        st.markdown(f'''
<a href="{url_panduan}"><button style="background-color:Tomato; border:none; color:white; padding:10px 24px; text-align:center; display:inline-block; font-size:16px; margin:4px 2px; cursor:pointer;">Panduan v. 1.0</button></a>
''', unsafe_allow_html=True)

        # Pilihan file
        selected_file = option_menu(
            menu_title="Pilih file:",
            options=["Pivot",
                     "Nilai Std. SD (K13), SMP (K13-KM), 10 SMA (KM)",
                     "Nilai Std. 8 SMP (KM-MTK SB)",
                     "Nilai Std. SD (KM)",
                     "Nilai Std. 10, 11 IPS (K13)",
                     "Nilai Std. 11 SMA (KM)",
                     "Nilai Std. 10, 11, PPLS IPA",
                     "Nilai Std. PPLS IPS"],
        )
   
    # k13
    k13_4sd_mat = ''
    k13_4sd_ind = ''
    k13_4sd_eng = ''
    k13_4sd_ipa = ''
    k13_4sd_ips = ''
    k13_5sd_mat = ''
    k13_5sd_ind = ''
    k13_5sd_eng = ''
    k13_5sd_ipa = ''
    k13_5sd_ips = ''
    k13_6sd_mat = ''
    k13_6sd_ind = ''
    k13_6sd_eng = ''
    k13_6sd_ipa = ''
    k13_6sd_ips = ''
    k13_7smp_mat = ''
    k13_7smp_ind = ''
    k13_7smp_eng = ''
    k13_7smp_ipa = ''
    k13_7smp_ips = ''
    k13_8smp_mat = ''
    k13_8smp_ind = ''
    k13_8smp_eng = ''
    k13_8smp_ipa = ''
    k13_8smp_ips = ''
    k13_9smp_mat = ''
    k13_9smp_ind = ''
    k13_9smp_eng = ''
    k13_9smp_ipa = ''
    k13_9smp_ips = ''
    k13_10ipa_mat = ''
    k13_10ipa_bio = ''
    k13_10ipa_fis = ''
    k13_10ipa_kim = ''
    k13_10ips_mat = ''
    k13_10ips_ind = ''
    k13_10ips_eng = ''
    k13_10ips_sej = ''
    k13_10ips_eko = ''
    k13_10ips_sos = ''
    k13_10ips_geo = ''
    k13_11ipa_mat = ''
    k13_11ipa_bio = ''
    k13_11ipa_fis = ''
    k13_11ipa_kim = ''
    k13_11ips_mat = ''
    k13_11ips_ind = ''
    k13_11ips_eng = ''
    k13_11ips_sej = ''
    k13_11ips_eko = ''
    k13_11ips_sos = ''
    k13_11ips_geo = ''
    
    # km
    km_4sd_mat = ''
    km_4sd_ind = ''
    km_4sd_eng = ''
    km_4sd_ipas = ''
    km_5sd_mat = ''
    km_5sd_ind = ''
    km_5sd_eng = ''
    km_5sd_ipas = ''
    km_6sd_mat = ''
    km_6sd_ind = ''
    km_6sd_eng = ''
    km_6sd_ipa = ''
    km_6sd_ips = ''
    km_7smp_mat = ''
    km_7smp_ind = ''
    km_7smp_eng = ''
    km_7smp_ipa = ''
    km_7smp_ips = ''
    km_8smp_mat = ''
    km_8smp_mat_sb = ''
    km_8smp_ind = ''
    km_8smp_eng = ''
    km_8smp_ipa = ''
    km_8smp_ips = ''
    km_9smp_mat = ''
    km_9smp_ind = ''
    km_9smp_eng = ''
    km_9smp_ipa = ''
    km_9smp_ips = ''
    km_10sma_mat = ''
    km_10sma_ind = ''
    km_10sma_eng = ''
    km_10sma_ipa = ''
    km_10sma_ips = ''
    km_11sma_mat_1 = ''
    km_11sma_mat_2 = ''
    km_11sma_ind = ''
    km_11sma_eng = ''
    km_11sma_sej = ''
    km_11sma_eko = ''
    km_11sma_sos = ''
    km_11sma_geo = ''
    km_11sma_ant = ''
    km_11sma_bio = ''
    km_11sma_fis = ''
    km_11sma_kim_1 = ''
    km_11sma_kim_2 = ''
    
    # ppls
    ppls_ipa_mat = ''
    ppls_ipa_bio = ''
    ppls_ipa_fis = ''
    ppls_ipa_kim = ''
    ppls_ips_geo = ''
    ppls_ips_eko = ''
    ppls_ips_sej = ''
    ppls_ips_sos = ''


    if selected_file == "Pivot":
    
        image = Image.open('logo resmi nf resize.png')
        st.image(image)

        st.title("PIVOT")

        col1 = st.container()
        with col1:
            KURIKULUM = st.selectbox(
                "KURIKULUM",
                ("--Pilih Kurikulum--", "K13", "KM", "PPLS"))

        col2 = st.container()
        with col2:
            KELAS = st.selectbox(
                "KELAS",
                ("--Pilih Kelas--", "4 SD", "5 SD", "6 SD", "7 SMP", "8 SMP", "8 SMP SB", "9 SMP", "10 IPA", "10 IPS", "10 SMA", "11 IPA", "11 IPS", "11 SMA", "PPLS IPA", "PPLS IPS"))

        # Kode Paket 4 SD K13
        if KURIKULUM == 'K13' and KELAS == '4 SD':
            st.subheader("Input Kode Paket Kelas 4 SD K13")
            col3, col4, col5, col6, col7 = st.columns(5)
            with col3:
                k13_4sd_mat = st.text_input("Kode Paket MTK",
                                placeholder="M4d2O0123-24K13")
            with col4:
                k13_4sd_ind = st.text_input("Kode Paket B.IND",
                                placeholder="I4d2O0123-24K13")
            with col5:
                k13_4sd_eng = st.text_input("Kode Paket B.ING",
                                placeholder="E4d2O0123-24K13")
            with col6:
                k13_4sd_ipa = st.text_input("Kode Paket IPA",
                                placeholder="A4d2O0123-24K13")
            with col7:
                k13_4sd_ips = st.text_input("Kode Paket IPS",
                                placeholder="Z4d2O0123-24K13")
            k13_4sd = [k13_4sd_mat, k13_4sd_ind,
                   k13_4sd_eng, k13_4sd_ipa, k13_4sd_ips]
            column_order_k13_4sd = ['IDTAHUN', 'NAMA', 'NONF', 'KELAS', 'NAMA_SKLH', 'KD_LOK', 'MAT_4SD', 'IND_4SD',
                                    'ENG_4SD', 'IPA_4SD', 'IPS_4SD']

        # Kode Paket 4 SD KM
        elif KURIKULUM == 'KM' and KELAS == '4 SD':
            st.subheader("Input Kode Paket Kelas 4 SD KM")
            col3, col4, col5, col6 = st.columns(4)
            with col3:
                km_4sd_mat = st.text_input("Kode Paket MTK",
                                placeholder="M4d2O0123-24KM")
            with col4:
                km_4sd_ind = st.text_input("Kode Paket B.IND",
                                placeholder="I4d2O0123-24KM")
            with col5:
                km_4sd_eng = st.text_input("Kode Paket B.ING",
                                placeholder="E4d2O0123-24KM")
            with col6:
                km_4sd_ipas = st.text_input("Kode Paket IPAS",
                                placeholder="Z4d2O0123-24KM")
            km_4sd = [km_4sd_mat, km_4sd_ind,
                  km_4sd_eng, km_4sd_ipas]
            column_order_km_4sd = ['IDTAHUN', 'NAMA', 'NONF', 'KELAS', 'NAMA_SKLH', 'KD_LOK', 'MAT_4SD', 'IND_4SD',
                                'ENG_4SD', 'IPAS_4SD']
        
        # Kode Paket 5 SD K13
        elif KURIKULUM == 'K13' and KELAS == '5 SD':
            st.subheader("Input Kode Paket Kelas 5 SD K13")
            col3, col4, col5, col6, col7 = st.columns(5)
            with col3:
                k13_5sd_mat = st.text_input("Kode Paket MTK",
                                placeholder="M5d2O0123-24K13")
            with col4:
                k13_5sd_ind = st.text_input("Kode Paket B.IND",
                                placeholder="I5d2O0123-24K13")
            with col5:
                k13_5sd_eng = st.text_input("Kode Paket B.ING",
                                placeholder="E5d2O0123-24K13")
            with col6:
                k13_5sd_ipa = st.text_input("Kode Paket IPA",
                                placeholder="A5d2O0123-24K13")
            with col7:
                k13_5sd_ips = st.text_input("Kode Paket IPS",
                                placeholder="Z5d2O0123-24K13")
            k13_5sd = [k13_5sd_mat, k13_5sd_ind,
                   k13_5sd_eng, k13_5sd_ipa, k13_5sd_ips]
            column_order_k13_5sd = ['IDTAHUN', 'NAMA', 'NONF', 'KELAS', 'NAMA_SKLH', 'KD_LOK', 'MAT_5SD', 'IND_5SD',
                                    'ENG_5SD', 'IPA_5SD', 'IPS_5SD']

        # Kode Paket 5 SD KM
        elif KURIKULUM == 'KM' and KELAS == '5 SD':
            st.subheader("Input Kode Paket Kelas 5 SD KM")
            col3, col4, col5, col6 = st.columns(4)
            with col3:
                km_5sd_mat = st.text_input("Kode Paket MTK",
                                placeholder="M5d2O0123-24KM")
            with col4:
                km_5sd_ind = st.text_input("Kode Paket B.IND",
                                placeholder="I5d2O0123-24KM")
            with col5:
                km_5sd_eng = st.text_input("Kode Paket B.ING",
                                placeholder="E5d2O0123-24KM")
            with col6:
                km_5sd_ipas = st.text_input("Kode Paket IPAS",
                                placeholder="2241D223-24")
            km_5sd = [km_5sd_mat, km_5sd_ind,
                  km_5sd_eng, km_5sd_ipas]
            column_order_km_5sd = ['IDTAHUN', 'NAMA', 'NONF', 'KELAS', 'NAMA_SKLH', 'KD_LOK', 'MAT_5SD', 'IND_5SD',
                                'ENG_5SD', 'IPAS_5SD']

        # Kode Paket 6 SD K13
        elif KURIKULUM == 'K13' and KELAS == '6 SD':
            st.subheader("Input Kode Paket Kelas 6 SD K13")
            col3, col4, col5, col6, col7 = st.columns(5)
            with col3:
                k13_6sd_mat = st.text_input("Kode Paket MTK",
                                placeholder="M6d2O0123-24K13")
            with col4:
                k13_6sd_ind = st.text_input("Kode Paket B.IND",
                                placeholder="I6d2O0123-24K13")
            with col5:
                k13_6sd_eng = st.text_input("Kode Paket B.ING",
                                placeholder="E6d2O0123-24K13")
            with col6:
                k13_6sd_ipa = st.text_input("Kode Paket IPA",
                                placeholder="A6d2O0123-24K13")
            with col7:
                k13_6sd_ips = st.text_input("Kode Paket IPS",
                                placeholder="Z6d2O0123-24K13")
            k13_6sd = [k13_6sd_mat, k13_6sd_ind,
                   k13_6sd_eng, k13_6sd_ipa, k13_6sd_ips]
            column_order_k13_6sd = ['IDTAHUN', 'NAMA', 'NONF', 'KELAS', 'NAMA_SKLH', 'KD_LOK', 'MAT_6SD', 'IND_6SD',
                                    'ENG_6SD', 'IPA_6SD', 'IPS_6SD']

        # Kode Paket 7 SMP K13
        elif KURIKULUM == 'K13' and KELAS == '7 SMP':
            st.subheader("Input Kode Paket Kelas 7 SMP K13")
            col3, col4, col5, col6, col7 = st.columns(5)
            with col3:
                k13_7smp_mat = st.text_input("Kode Paket MTK",
                                placeholder="M1p2O0123-24K13")
            with col4:
                k13_7smp_ind = st.text_input("Kode Paket B.IND",
                                placeholder="I1p2O0123-24K13")
            with col5:
                k13_7smp_eng = st.text_input("Kode Paket B.ING",
                                placeholder="E1p2O0123-24K13")
            with col6:
                k13_7smp_ipa = st.text_input("Kode Paket IPA",
                                placeholder="A1p2O0123-24K13")
            with col7:
                k13_7smp_ips = st.text_input("Kode Paket IPS",
                                placeholder="Z1p2O0123-24K13")
            k13_7smp = [k13_7smp_mat, k13_7smp_ind,
                   k13_7smp_eng, k13_7smp_ipa, k13_7smp_ips]
            column_order_k13_7smp = ['IDTAHUN', 'NAMA', 'NONF', 'KELAS', 'NAMA_SKLH', 'KD_LOK', 'MAT_7SMP', 'IND_7SMP',
                                    'ENG_7SMP', 'IPA_7SMP', 'IPS_7SMP']
        
        # Kode Paket 7 SMP KM
        elif KURIKULUM == 'KM' and KELAS == '7 SMP':
            st.subheader("Input Kode Paket Kelas 7 SMP KM")
            col3, col4, col5, col6, col7 = st.columns(5)
            with col3:
                km_7smp_mat = st.text_input("Kode Paket MTK",
                                placeholder="M1p2O0123-24K13")
            with col4:
                km_7smp_ind = st.text_input("Kode Paket B.IND",
                                placeholder="I1p2O0123-24K13")
            with col5:
                km_7smp_eng = st.text_input("Kode Paket B.ING",
                                placeholder="E1p2O0123-24K13")
            with col6:
                km_7smp_ipa = st.text_input("Kode Paket IPA",
                                placeholder="A1p2O0123-24K13")
            with col7:
                km_7smp_ips = st.text_input("Kode Paket IPS",
                                placeholder="Z1p2O0123-24K13")
            km_7smp = [km_7smp_mat, km_7smp_ind,
                   km_7smp_eng, km_7smp_ipa, km_7smp_ips]
            column_order_km_7smp = ['IDTAHUN', 'NAMA', 'NONF', 'KELAS', 'NAMA_SKLH', 'KD_LOK', 'MAT_7SMP', 'IND_7SMP',
                                    'ENG_7SMP', 'IPA_7SMP', 'IPS_7SMP']

        # Kode Paket 8 SMP K13
        elif KURIKULUM == 'K13' and KELAS == '8 SMP':
            st.subheader("Input Kode Paket Kelas 8 SMP K13")
            col3, col4, col5, col6, col7 = st.columns(5)
            with col3:
                k13_8smp_mat = st.text_input("Kode Paket MTK",
                                placeholder="M2p2O0123-24K13")
            with col4:
                k13_8smp_ind = st.text_input("Kode Paket B.IND",
                                placeholder="I2p2O0123-24K13")
            with col5:
                k13_8smp_eng = st.text_input("Kode Paket B.ING",
                                placeholder="E2p2O0123-24K13")
            with col6:
                k13_8smp_ipa = st.text_input("Kode Paket IPA",
                                placeholder="A2p2O0123-24K13")
            with col7:
                k13_8smp_ips = st.text_input("Kode Paket IPS",
                                placeholder="Z2p2O0123-24K13")
            k13_8smp = [k13_8smp_mat, k13_8smp_ind,
                   k13_8smp_eng, k13_8smp_ipa, k13_8smp_ips]
            column_order_k13_8smp = ['IDTAHUN', 'NAMA', 'NONF', 'KELAS', 'NAMA_SKLH', 'KD_LOK', 'MAT_8SMP', 'IND_8SMP',
                                    'ENG_8SMP', 'IPA_8SMP', 'IPS_8SMP']
        
        # Kode Paket 8 SMP KM
        elif KURIKULUM == 'KM' and KELAS == '8 SMP':
            st.subheader("Input Kode Paket Kelas 8 SMP KM")
            col3, col4, col5, col6, col7 = st.columns(5)
            with col3:
                km_8smp_mat = st.text_input("Kode Paket MTK",
                                placeholder="M2p2O0123-24KM")
            with col4:
                km_8smp_ind = st.text_input("Kode Paket B.IND",
                                placeholder="I2p2O0123-24KM")
            with col5:
                km_8smp_eng = st.text_input("Kode Paket B.ING",
                                placeholder="E2p2O0123-24KM")
            with col6:
                km_8smp_ipa = st.text_input("Kode Paket IPA",
                                placeholder="A2p2O0123-24KM")
            with col7:
                km_8smp_ips = st.text_input("Kode Paket IPS",
                                placeholder="Z2p2O0123-24KM")
            km_8smp = [km_8smp_mat, km_8smp_ind,
                   km_8smp_eng, km_8smp_ipa, km_8smp_ips]
            column_order_km_8smp = ['IDTAHUN', 'NAMA', 'NONF', 'KELAS', 'NAMA_SKLH', 'KD_LOK', 'MAT_8SMP', 'IND_8SMP',
                                    'ENG_8SMP', 'IPA_8SMP', 'IPS_8SMP']

        # Kode Paket 8 SMP KM SB
        elif KURIKULUM == 'KM' and KELAS == '8 SMP SB':
            st.subheader("Input Kode Paket Kelas 8 SMP KM SB")
            col3, col4, col5, col6, col7 = st.columns(5)
            with col3:
                km_8smp_mat_sb = st.text_input("Kode Paket MTK SB",
                                placeholder="M2p2O0123-24KM")
            with col4:
                km_8smp_ind = st.text_input("Kode Paket B.IND",
                                placeholder="I2p2O0123-24KM")
            with col5:
                km_8smp_eng = st.text_input("Kode Paket B.ING",
                                placeholder="E2p2O0123-24KM")
            with col6:
                km_8smp_ipa = st.text_input("Kode Paket IPA",
                                placeholder="A2p2O0123-24KM")
            with col7:
                km_8smp_ips = st.text_input("Kode Paket IPS",
                                placeholder="Z2p2O0123-24KM")
            km_8smp_sb = [km_8smp_mat_sb, km_8smp_ind,
                    km_8smp_eng, km_8smp_ipa, km_8smp_ips]
            column_order_km_8smp_sb = ['IDTAHUN', 'NAMA', 'NONF', 'KELAS', 'NAMA_SKLH', 'KD_LOK', 'MAT_SB_8SMP', 'IND_8SMP',
                                    'ENG_8SMP', 'IPA_8SMP', 'IPS_8SMP']

        # Kode Paket 9 SMP K13
        elif KURIKULUM == 'K13' and KELAS == '9 SMP':
            st.subheader("Input Kode Paket Kelas 9 SMP K13")
            col3, col4, col5, col6, col7 = st.columns(5)
            with col3:
                k13_9smp_mat = st.text_input("Kode Paket MTK",
                                placeholder="M3p2O0123-24K13")
            with col4:
                k13_9smp_ind = st.text_input("Kode Paket B.IND",
                                placeholder="I3p2O0123-24K13")
            with col5:
                k13_9smp_eng = st.text_input("Kode Paket B.ING",
                                placeholder="E3p2O0123-24K13")
            with col6:
                k13_9smp_ipa = st.text_input("Kode Paket IPA",
                                placeholder="6141A223-24")
            with col7:
                k13_9smp_ips = st.text_input("Kode Paket IPS",
                                placeholder="O3p2O0123-24K13")
            k13_9smp = [k13_9smp_mat, k13_9smp_ind,
                        k13_9smp_eng, k13_9smp_ipa, k13_9smp_ips]
            column_order_k13_9smp = ['IDTAHUN', 'NAMA', 'NONF', 'KELAS', 'NAMA_SKLH', 'KD_LOK', 'MAT_9SMP', 'IND_9SMP',
                                    'ENG_9SMP', 'IPA_9SMP', 'IPS_9SMP']

        # Kode Paket 9 SMP KM
        elif KURIKULUM == 'KM' and KELAS == '9 SMP':
            st.subheader("Input Kode Paket Kelas 9 SMP KM")
            col3, col4, col5, col6, col7 = st.columns(5)
            with col3:
                km_9smp_mat = st.text_input("Kode Paket MTK",
                                placeholder="M3p2O0123-24KM")
            with col4:
                km_9smp_ind = st.text_input("Kode Paket B.IND",
                                placeholder="I3p2O0123-24KM")
            with col5:
                km_9smp_eng = st.text_input("Kode Paket B.ING",
                                placeholder="E3p2O0123-24KM")
            with col6:
                km_9smp_ipa = st.text_input("Kode Paket IPA",
                                placeholder="6141A223-24")
            with col7:
                km_9smp_ips = st.text_input("Kode Paket IPS",
                                placeholder="O3p2O0123-24KM")
            km_9smp = [km_9smp_mat, km_9smp_ind,
                        km_9smp_eng, km_9smp_ipa, km_9smp_ips]
            column_order_km_9smp = ['IDTAHUN', 'NAMA', 'NONF', 'KELAS', 'NAMA_SKLH', 'KD_LOK', 'MAT_9SMP', 'IND_9SMP',
                                    'ENG_9SMP', 'IPA_9SMP', 'IPS_9SMP']
        
        # Kode Paket 10 SMA KM
        elif KURIKULUM == 'KM' and KELAS == '10 SMA':
            st.subheader("Input Kode Paket Kelas 10 SMA KM")
            col3, col4, col5, col6, col7 = st.columns(5)
            with col3:
                km_10sma_mat = st.text_input("Kode Paket MTK",
                                placeholder="M4x2O0023-24KM")
            with col4:
                km_10sma_ind = st.text_input("Kode Paket B.IND",
                                placeholder="I4x2O0023-24KM")
            with col5:
                km_10sma_eng = st.text_input("Kode Paket B.ING",
                                placeholder="E4x2O0023-24KM")
            with col6:
                km_10sma_ipa = st.text_input("Kode Paket IPA",
                                placeholder="9240A223-24")
            with col7:
                km_10sma_ips = st.text_input("Kode Paket IPS",
                                placeholder="9240S223-24")
            km_10sma = [km_10sma_mat, km_10sma_ind,
                        km_10sma_eng, km_10sma_ipa, km_10sma_ips]
            column_order_km_10sma = ['IDTAHUN', 'NAMA', 'NONF', 'KELAS', 'NAMA_SKLH', 'KD_LOK', 'MAT_10SMA', 'IND_10SMA',
                                    'ENG_10SMA', 'IPA_10SMA', 'IPS_10SMA']
        
        # Kode Paket 10 IPA K13
        elif KURIKULUM == 'K13' and KELAS == '10 IPA':
            st.subheader("Input Kode Paket Kelas 10 IPA K13")
            col3, col4, col5, col6 = st.columns(4)
            with col3:
                k13_10ipa_mat = st.text_input("Kode Paket MTK",
                                placeholder="M4a2O0023-24K13")
            with col4:
                k13_10ipa_bio = st.text_input("Kode Paket BIO",
                                placeholder="B4a2O0023-24K13")
            with col5:
                k13_10ipa_fis = st.text_input("Kode Paket FIS",
                                placeholder="F4a2O0023-24K13")
            with col6:
                k13_10ipa_kim = st.text_input("Kode Paket KIM",
                                placeholder="K4a2O0023-24K13")
            k13_10ipa = [k13_10ipa_mat, k13_10ipa_bio,
                  k13_10ipa_fis, k13_10ipa_kim]
            column_order_k13_10ipa = ['IDTAHUN', 'NAMA', 'NONF', 'KELAS', 'NAMA_SKLH', 'KD_LOK', 'MAT_10IPA',
                                'FIS_10IPA', 'KIM_10IPA', 'BIO_10IPA']
        
        # Kode Paket 10 IPS K13
        elif KURIKULUM == 'K13' and KELAS == '10 IPS':
            st.subheader("Input Kode Paket Kelas 10 IPS K13")
            col3, col4, col5, col6, col7, col8, col9 = st.columns(7)
            with col3:
                k13_10ips_mat = st.text_input("Kode Paket MTK",
                                placeholder="M4a2O0023-24K13")
            with col4:
                k13_10ips_ind = st.text_input("Kode Paket B.IND",
                                placeholder="B4a2O0023-24K13")
            with col5:
                k13_10ips_eng = st.text_input("Kode Paket B.ING",
                                placeholder="F4a2O0023-24K13")
            with col6:
                k13_10ips_sej = st.text_input("Kode Paket SEJ",
                                placeholder="K4a2O0023-24K13")
            with col7:
                k13_10ips_eko = st.text_input("Kode Paket EKO",
                                placeholder="K4a2O0023-24K13")
            with col8:
                k13_10ips_sos = st.text_input("Kode Paket SOS",
                                placeholder="K4a2O0023-24K13")
            with col9:
                k13_10ips_geo = st.text_input("Kode Paket GEO",
                                placeholder="K4a2O0023-24K13")
            k13_10ips = [k13_10ips_mat, k13_10ips_ind,
                  k13_10ips_eng, k13_10ips_sej, k13_10ips_eko, k13_10ips_sos, k13_10ips_geo]
            column_order_k13_10ips = ['IDTAHUN', 'NAMA', 'NONF', 'KELAS', 'NAMA_SKLH', 'KD_LOK', 'MAT_10IPS', 'IND_10IPS',
                                'ENG_10IPS', 'SEJ_10IPS', 'GEO_10IPS', 'EKO_10IPS', 'SOS_10IPS']
        
        # Kode Paket 11 SMA KM
        elif KURIKULUM == 'KM' and KELAS == '11 SMA':
            st.subheader("Input Kode Paket Kelas 11 SMA KM")
            col3, col4, col5, col6 = st.columns(4)
            with col3:
                km_11sma_mat_1 = st.text_input("Kode Paket MTK 1",
                                placeholder="Q5x2O0023-24KM2")
            with col4:
                km_11sma_mat_2 = st.text_input("Kode Paket MTK 2",
                                placeholder="R5x2O0023-24KM2")
            with col5:
                km_11sma_ind = st.text_input("Kode Paket B.IND",
                                placeholder="I5x2O0023-24KM2")
            with col6:
                km_11sma_eng = st.text_input("Kode Paket B.ING",
                                placeholder="E5x2O0023-24KM2")
            col7, col8, col9, col10, col11 = st.columns(5)
            with col7:
                km_11sma_sej = st.text_input("Kode Paket SEJ",
                                placeholder="S5x2O0023-24KM2")
            with col8:
                km_11sma_eko = st.text_input("Kode Paket EKO",
                                placeholder="O5x2O0023-24KM2")
            with col9:
                km_11sma_sos = st.text_input("Kode Paket SOS",
                                placeholder="L5x2O0023-24KM2")
            with col10:
                km_11sma_geo = st.text_input("Kode Paket GEO",
                                placeholder="G5x2O0023-24KM2")
            with col11:
                km_11sma_ant = st.text_input("Kode Paket ANT",
                                placeholder="N5x2O0023-24KM2")
            col12, col13, col14, col15 = st.columns(4)
            with col12:
                km_11sma_bio = st.text_input("Kode Paket BIO",
                                placeholder="B5x2O0023-24KM2")
            with col13:
                km_11sma_fis = st.text_input("Kode Paket FIS",
                                placeholder="F5x2O0023-24KM2")
            with col14:
                km_11sma_kim_1 = st.text_input("Kode Paket KIM 1",
                                placeholder="K5x2O0123-24KM")
            with col15:
                km_11sma_kim_2 = st.text_input("Kode Paket KIM 2",
                                placeholder="K5x2O0023-24KM2")
            km_11sma = [km_11sma_mat_1, km_11sma_mat_2, km_11sma_ind,
                        km_11sma_eng, km_11sma_sej, km_11sma_eko, km_11sma_sos, km_11sma_geo, km_11sma_ant,
                        km_11sma_bio, km_11sma_fis, km_11sma_kim_1, km_11sma_kim_2]
            column_order_km_11sma = ['IDTAHUN', 'NAMA', 'NONF', 'KELAS', 'NAMA_SKLH', 'KD_LOK', 'MAT_1_11SMA', 'MAT_2_11SMA', 'IND_11SMA',
                                    'ENG_11SMA', 'SEJ_11SMA', 'EKO_11SMA', 'SOS_11SMA', 'GEO_11SMA', 'ANT_11SMA',
                                    'BIO_11SMA', 'FIS_11SMA', 'KIM_1_11SMA', 'KIM_2_11SMA']
        
        # Kode Paket 11 IPA K13
        elif KURIKULUM == 'K13' and KELAS == '11 IPA':
            st.subheader("Input Kode Paket Kelas 11 IPA K13")
            col3, col4, col5, col6 = st.columns(4)
            with col3:
                k13_11ipa_mat = st.text_input("Kode Paket MTK",
                                placeholder="M5a2O0023-24K13")
            with col4:
                k13_11ipa_bio = st.text_input("Kode Paket BIO",
                                placeholder="B5a2O0023-24K13")
            with col5:
                k13_11ipa_fis = st.text_input("Kode Paket FIS",
                                placeholder="F5a2O0023-24K13")
            with col6:
                k13_11ipa_kim = st.text_input("Kode Paket KIM",
                                placeholder="K5a2O0023-24K13")
            k13_11ipa = [k13_11ipa_mat, k13_11ipa_bio,
                  k13_11ipa_fis, k13_11ipa_kim]
            column_order_k13_11ipa = ['IDTAHUN', 'NAMA', 'NONF', 'KELAS', 'NAMA_SKLH', 'KD_LOK', 'MAT_11IPA',
                                'FIS_11IPA', 'KIM_11IPA', 'BIO_11IPA']

        # Kode Paket 11 IPS K13
        elif KURIKULUM == 'K13' and KELAS == '11 IPS':
            st.subheader("Input Kode Paket Kelas 11 IPS K13")
            col3, col4, col5, col6, col7, col8, col9 = st.columns(7)
            with col3:
                k13_11ips_mat = st.text_input("Kode Paket MTK",
                                placeholder="M4a2O0023-24K13")
            with col4:
                k13_11ips_ind = st.text_input("Kode Paket B.IND",
                                placeholder="B4a2O0023-24K13")
            with col5:
                k13_11ips_eng = st.text_input("Kode Paket B.ING",
                                placeholder="F4a2O0023-24K13")
            with col6:
                k13_11ips_sej = st.text_input("Kode Paket SEJ",
                                placeholder="K4a2O0023-24K13")
            with col7:
                k13_11ips_eko = st.text_input("Kode Paket EKO",
                                placeholder="K4a2O0023-24K13")
            with col8:
                k13_11ips_sos = st.text_input("Kode Paket SOS",
                                placeholder="K4a2O0023-24K13")
            with col9:
                k13_11ips_geo = st.text_input("Kode Paket GEO",
                                placeholder="K4a2O0023-24K13")
            k13_11ips = [k13_11ips_mat, k13_11ips_ind,
                  k13_11ips_eng, k13_11ips_sej, k13_11ips_eko, k13_11ips_sos, k13_11ips_geo]
            column_order_k13_11ips = ['IDTAHUN', 'NAMA', 'NONF', 'KELAS', 'NAMA_SKLH', 'KD_LOK', 'MAT_11IPS', 'IND_11IPS',
                                'ENG_11IPS', 'SEJ_11IPS', 'GEO_11IPS', 'EKO_11IPS', 'SOS_11IPS']

        # Kode Paket PPLS IPA
        elif KURIKULUM == 'PPLS' and KELAS == 'PPLS IPA':
            st.subheader("Input Kode Paket Kelas PPLS IPA")
            col3, col4, col5, col6 = st.columns(4)
            with col3:
                ppls_ipa_mat = st.text_input("Kode Paket MTK",
                                placeholder="M9a2O0123-24PPLS")
            with col4:
                ppls_ipa_bio = st.text_input("Kode Paket BIO",
                                placeholder="B9a2O0123-24PPLS")
            with col5:
                ppls_ipa_fis = st.text_input("Kode Paket FIS",
                                placeholder="F9a2O0123-24PPLS")
            with col6:
                ppls_ipa_kim = st.text_input("Kode Paket KIM",
                                placeholder="K9a2O0123-24PPLS")
            ppls_ipa = [ppls_ipa_mat, ppls_ipa_bio,
                    ppls_ipa_fis, ppls_ipa_kim]
            column_order_ppls_ipa = ['IDTAHUN', 'NAMA', 'NONF', 'KELAS', 'NAMA_SKLH', 'KD_LOK', 'MAT_PPLS_IPA',
                                    'FIS_PPLS_IPA', 'KIM_PPLS_IPA', 'BIO_PPLS_IPA',]
                            
        # Kode Paket PPLS IPS
        elif KURIKULUM == 'PPLS' and KELAS == 'PPLS IPS':
            st.subheader("Input Kode Paket Kelas PPLS IPS")
            col3, col4, col5, col6 = st.columns(4)
            with col3:
                ppls_ips_geo = st.text_input("Kode Paket GEO",
                                placeholder="G9s2O0123-24PPLS")
            with col4:
                ppls_ips_eko = st.text_input("Kode Paket EKO",
                                placeholder="O9s2O0123-24PPLS")
            with col5:
                ppls_ips_sej = st.text_input("Kode Paket SEJ",
                                placeholder="S9s2O0123-24PPLS")
            with col6:
                ppls_ips_sos = st.text_input("Kode Paket SOS",
                                placeholder="L9s2O0123-24PPLS")
            ppls_ips = [ppls_ips_geo, ppls_ips_eko,
                        ppls_ips_sej, ppls_ips_sos]
            column_order_ppls_ips = ['IDTAHUN', 'NAMA', 'NONF', 'KELAS', 'NAMA_SKLH', 'KD_LOK', 'GEO_PPLS_IPS',
                                    'EKO_PPLS_IPS', 'SEJ_PPLS_IPS', 'SOS_PPLS_IPS',]

        col8 = st.container()
        with col8:
            SEMESTER = st.selectbox(
                "SEMESTER",
                ("--Pilih Semester--", "SEMESTER 1", "SEMESTER 2"))

        col9 = st.container()
        with col9:
            PENILAIAN = st.selectbox(
                "PENILAIAN",
                ("--Pilih Penilaian--", "PENILAIAN TENGAH SEMESTER", "SUMATIF TENGAH SEMESTER", "PENILAIAN AKHIR TAHUN", "SUMATIF AKHIR TAHUN", "TO UJIAN SEKOLAH"))

        TAHUN = st.text_input("Masukkan Tahun Ajaran",
                              placeholder="contoh: 2022-2023")

        url_contoh_detail ='https://docs.google.com/spreadsheets/d/1hq9P44M9P51GrJn07DyWsHA1xKFj--dG/edit?usp=sharing&ouid=100219691055398475739&rtpof=true&sd=true'
        uploaded_detail = st.file_uploader(
            'Letakkan file excel Detail Siswa', type='xlsx')
        url_contoh_to_pts = 'https://docs.google.com/spreadsheets/d/1K-bWwSvkd0fgJ-xVtn-CWQcZH_-MoM8F/edit?usp=sharing&ouid=100219691055398475739&rtpof=true&sd=true'
        uploaded_to_pts = st.file_uploader(
            'Letakkan file excel TO', type='xlsx')

        detail = None
        to_pts = None

        if uploaded_detail is not None:
            detail = pd.read_excel(uploaded_detail)

        if uploaded_to_pts is not None:
            to_pts = pd.read_excel(uploaded_to_pts)

        if detail is not None and to_pts is not None:
            detail = detail.drop(['user_id', 'is_test_access', 'no_hp', 'lokasi_id', 'jenjang_id',
                                  'riwayat_jenjang', 'jenjang_dipilih_id', 'kode_level', 'kode_kelas',
                                  'tempat_lahir', 'tanggal_lahir', 'semester', 'tahun_ajar',
                                  'program', 'pin', 'join_skolla', 'created_at', 'updated_at'], axis=1)  # Menghilangkan kolom sebelum dilakukan merge
            if KELAS == "6 SD" and KURIKULUM == "K13":
                to_pts['kelas_id'] = "'" + to_pts['kelas_id'].astype(str)
            else:
                to_pts['kelas_id'] = to_pts['kelas_id']
            result = pd.merge(detail, to_pts[['no_nf', 'kode_paket', 'tahun_ajaran', 'kelas_id',
                                              'lokasi_id', 'jumlah_benar']], on='no_nf', how='left')
            # Mengganti nilai 0 pada kolom 'jumlah_benar' menjadi NaN (kosong)
            result['jumlah_benar'] = result['jumlah_benar'].replace(0, np.nan)
            # Menghapus nilai NaN dari kolom 'kode_paket'
            result = result.dropna(subset=['kode_paket'])

            # k13
            if KELAS == "4 SD" and KURIKULUM == "K13":
                kode_kls_kur = k13_4sd
                column_order = column_order_k13_4sd
            elif KELAS == "5 SD" and KURIKULUM == "K13":
                kode_kls_kur = k13_5sd
                column_order = column_order_k13_5sd
            elif KELAS == "6 SD" and KURIKULUM == "K13":
                kode_kls_kur = k13_6sd
                column_order = column_order_k13_6sd
            elif KELAS == "7 SMP" and KURIKULUM == "K13":
                kode_kls_kur = k13_7smp
                column_order = column_order_k13_7smp
            elif KELAS == "8 SMP" and KURIKULUM == "K13":
                kode_kls_kur = k13_8smp
                column_order = column_order_k13_8smp
            elif KELAS == "9 SMP" and KURIKULUM == "K13":
                kode_kls_kur = k13_9smp
                column_order = column_order_k13_9smp
            elif KELAS == "10 IPA" and KURIKULUM == "K13":
                kode_kls_kur = k13_10ipa
                column_order = column_order_k13_10ipa
            elif KELAS == "10 IPS" and KURIKULUM == "K13":
                kode_kls_kur = k13_10ips
                column_order = column_order_k13_10ips
            elif KELAS == "11 IPA" and KURIKULUM == "K13":
                kode_kls_kur = k13_11ipa
                column_order = column_order_k13_11ipa
            elif KELAS == "11 IPS" and KURIKULUM == "K13":
                kode_kls_kur = k13_11ips
                column_order = column_order_k13_11ips
            # km
            elif KELAS == "4 SD" and KURIKULUM == "KM":
                kode_kls_kur = km_4sd
                column_order = column_order_km_4sd
            elif KELAS == "5 SD" and KURIKULUM == "KM":
                kode_kls_kur = km_5sd
                column_order = column_order_km_5sd
            elif KELAS == "7 SMP" and KURIKULUM == "KM":
                kode_kls_kur = km_7smp
                column_order = column_order_km_7smp
            elif KELAS == "8 SMP" and KURIKULUM == "KM":
                kode_kls_kur = km_8smp
                column_order = column_order_km_8smp
            elif KELAS == "8 SMP SB" and KURIKULUM == "KM":
                kode_kls_kur = km_8smp_sb
                column_order = column_order_km_8smp_sb
            elif KELAS == "9 SMP" and KURIKULUM == "KM":
                kode_kls_kur = km_9smp
                column_order = column_order_km_9smp
            elif KELAS == "10 SMA" and KURIKULUM == "KM":
                kode_kls_kur = km_10sma
                column_order = column_order_km_10sma
            elif KELAS == "11 SMA" and KURIKULUM == "KM":
                kode_kls_kur = km_11sma
                column_order = column_order_km_11sma
            # ppls
            elif KELAS == "PPLS IPA" and KURIKULUM == "PPLS":
                kode_kls_kur = ppls_ipa
                column_order = column_order_ppls_ipa
            elif KELAS == "PPLS IPS" and KURIKULUM == "PPLS":
                kode_kls_kur = ppls_ips
                column_order = column_order_ppls_ips

            result_filtered = result[result['kode_paket'].isin(kode_kls_kur)]
            result_filtered.drop_duplicates(
                subset=['no_nf', 'kode_paket'], keep='first', inplace=True)

            # Menggunakan pivot_table untuk menjadikan konten kolom 'kode_paket' sebagai header dan menghilangkan duplikat
            result_pivot = pd.pivot_table(result_filtered, index=[
                'name', 'no_nf', 'lokasi_id', 'sekolah', 'kelas_id', 'tahun_ajaran'], columns='kode_paket', values='jumlah_benar', aggfunc='first')
            result_pivot.reset_index(inplace=True)  # Mengatur ulang indeks

            # Ubah nama kolom
            result_pivot = result_pivot.rename(
                columns={'name': 'NAMA', 'no_nf': 'NONF', 'lokasi_id': 'KD_LOK', 'sekolah': 'NAMA_SKLH', 'kelas_id': 'KELAS', 'tahun_ajaran': 'IDTAHUN',
                         k13_4sd_mat: 'MAT_4SD', k13_4sd_ind: 'IND_4SD', k13_4sd_eng: 'ENG_4SD', k13_4sd_ipa: 'IPA_4SD', k13_4sd_ips: 'IPS_4SD',
                         k13_5sd_mat: 'MAT_5SD', k13_5sd_ind: 'IND_5SD', k13_5sd_eng: 'ENG_5SD', k13_5sd_ipa: 'IPA_5SD', k13_5sd_ips: 'IPS_5SD',
                         k13_6sd_mat: 'MAT_6SD', k13_6sd_ind: 'IND_6SD', k13_6sd_eng: 'ENG_6SD', k13_6sd_ipa: 'IPA_6SD', k13_6sd_ips: 'IPS_6SD',
                         k13_7smp_mat: 'MAT_7SMP', k13_7smp_ind: 'IND_7SMP', k13_7smp_eng: 'ENG_7SMP', k13_7smp_ipa: 'IPA_7SMP', k13_7smp_ips: 'IPS_7SMP',
                         k13_8smp_mat: 'MAT_8SMP', k13_8smp_ind: 'IND_8SMP', k13_8smp_eng: 'ENG_8SMP', k13_8smp_ipa: 'IPA_8SMP', k13_8smp_ips: 'IPS_8SMP',
                         k13_9smp_mat: 'MAT_9SMP', k13_9smp_ind: 'IND_9SMP', k13_9smp_eng: 'ENG_9SMP', k13_9smp_ipa: 'IPA_9SMP', k13_9smp_ips: 'IPS_9SMP',
                         k13_10ipa_mat: 'MAT_10IPA', k13_10ipa_bio: 'BIO_10IPA', k13_10ipa_fis: 'FIS_10IPA', k13_10ipa_kim: 'KIM_10IPA',
                         k13_10ips_mat: 'MAT_10IPS', k13_10ips_ind: 'IND_10IPS', k13_10ips_eng: 'ENG_10IPS', k13_10ips_sej: 'SEJ_10IPS', k13_10ips_eko: 'EKO_10IPS',k13_10ips_sos: 'SOS_10IPS',k13_10ips_geo: 'GEO_10IPS',
                         k13_11ipa_mat: 'MAT_11IPA', k13_11ipa_bio: 'BIO_11IPA', k13_11ipa_fis: 'FIS_11IPA', k13_11ipa_kim: 'KIM_11IPA',
                         k13_11ips_mat: 'MAT_11IPS', k13_11ips_ind: 'IND_11IPS', k13_11ips_eng: 'ENG_11IPS', k13_11ips_sej: 'SEJ_11IPS', k13_11ips_eko: 'EKO_11IPS',k13_11ips_sos: 'SOS_11IPS',k13_11ips_geo: 'GEO_11IPS',
                         km_4sd_mat: 'MAT_4SD', km_4sd_ind: 'IND_4SD', km_4sd_eng: 'ENG_4SD', km_4sd_ipas: 'IPAS_4SD',
                         km_5sd_mat: 'MAT_5SD', km_5sd_ind: 'IND_5SD', km_5sd_eng: 'ENG_5SD', km_5sd_ipas: 'IPAS_5SD',
                         km_7smp_mat: 'MAT_7SMP', km_7smp_ind: 'IND_7SMP', km_7smp_eng: 'ENG_7SMP', km_7smp_ipa: 'IPA_7SMP', km_7smp_ips: 'IPS_7SMP',
                         km_8smp_mat: 'MAT_8SMP', km_8smp_ind: 'IND_8SMP', km_8smp_eng: 'ENG_8SMP', km_8smp_ipa: 'IPA_8SMP', km_8smp_ips: 'IPS_8SMP', km_8smp_mat_sb: 'MAT_SB_8SMP',
                         km_9smp_mat: 'MAT_9SMP', km_9smp_ind: 'IND_9SMP', km_9smp_eng: 'ENG_9SMP', km_9smp_ipa: 'IPA_9SMP', km_9smp_ips: 'IPS_9SMP',
                         km_10sma_mat: 'MAT_10SMA', km_10sma_ind: 'IND_10SMA', km_10sma_eng: 'ENG_10SMA', km_10sma_ipa: 'IPA_10SMA', km_10sma_ips: 'IPS_10SMA',
                         km_11sma_mat_1: 'MAT_1_11SMA', km_11sma_mat_2: 'MAT_2_11SMA', km_11sma_ind: 'IND_11SMA', km_11sma_eng: 'ENG_11SMA', km_11sma_sej: 'SEJ_11SMA', km_11sma_eko: 'EKO_11SMA', km_11sma_sos: 'SOS_11SMA', km_11sma_geo: 'GEO_11SMA', km_11sma_ant: 'ANT_11SMA', km_11sma_bio: 'BIO_11SMA', km_11sma_fis: 'FIS_11SMA', km_11sma_kim_1: 'KIM_1_11SMA', km_11sma_kim_2: 'KIM_2_11SMA',
                         ppls_ipa_mat: 'MAT_PPLS_IPA', ppls_ipa_fis: 'FIS_PPLS_IPA', ppls_ipa_kim: 'KIM_PPLS_IPA', ppls_ipa_bio: 'BIO_PPLS_IPA',
                         ppls_ips_geo: 'GEO_PPLS_IPS', ppls_ips_eko: 'EKO_PPLS_IPS', ppls_ips_sej: 'SEJ_PPLS_IPS', ppls_ips_sos: 'SOS_PPLS_IPS'})

            result_pivot = result_pivot.reindex(columns=column_order)

            kelas = KELAS.lower().replace(" ", "")
            kurikulum = KURIKULUM.lower()
            tahun = TAHUN.replace("-", "")
            semester = SEMESTER.lower()
            penilaian = PENILAIAN.lower()

            path_file = f"{kelas}_{penilaian}_{semester}_{kurikulum}_{tahun}_pivot.xlsx"

            # Simpan file ke direktori temporer
            temp_dir = tempfile.gettempdir()
            file_path = temp_dir + '/' + path_file
            # wb.save(file_path)

            # Menyimpan DataFrame ke file Excel
            result_pivot.to_excel(file_path, index=False)
            st.success("File siap diunduh!")

            # Tombol unduh file
            with open(file_path, "rb") as f:
                bytes_data = f.read()
            st.download_button(label="Unduh File", data=bytes_data,
                               file_name=path_file)

            st.warning(
                "Buka file unduhan, klik 'Enable Editing' dan 'Save'")
            
    if selected_file == "Nilai Std. SD (K13), SMP (K13-KM), 10 SMA (KM)" :  
    # menghilangkan hamburger
        st.markdown("""
        <style>
        .css-1rs6os.edgvbvh3
        {
            visibility:hidden;
        }
        .css-1lsmgbg.egzxvld0
        {
            visibility:hidden;
        }
        </style>
        """, unsafe_allow_html=True)

        image = Image.open('logo resmi nf resize.png')
        st.image(image)

        st.title("Olah Nilai Standar K13-KM")

        st.header("SD-SMP")

        col6 = st.container()

        with col6:
            KELAS = st.selectbox(
                "KELAS",
                ("--Pilih Kelas--", "4 SD", "5 SD", "6 SD", "7 SMP", "8 SMP", "9 SMP", "10 SMA"))

        col7 = st.container()

        with col7:
            SEMESTER = st.selectbox(
                "SEMESTER",
                ("--Pilih Semester--", "SEMESTER 1", "SEMESTER 2"))

        col8 = st.container()

        with col8:
            PENILAIAN = st.selectbox(
                "PENILAIAN",
                ("--Pilih Penilaian--", "PENILAIAN TENGAH SEMESTER", "SUMATIF TENGAH SEMESTER", "PENILAIAN AKHIR TAHUN", "SUMATIF AKHIR TAHUN", "TO UJIAN SEKOLAH"))

        col9 = st.container()

        with col9:
            KURIKULUM = st.selectbox(
                "KURIKULUM",
                ("--Pilih Kurikulum--", "K13", "KM"))

        TAHUN = st.text_input("Masukkan Tahun Ajaran",
                              placeholder="contoh: 2022-2023")

        col1, col2, col3, col4, col5 = st.columns(5)

        with col1:
            MTK = st.selectbox(
                "JML. SOAL MAT.",
                ("--Pilih--", 25, 30, 35, 40, 45))

        with col2:
            IND = st.selectbox(
                "JML. SOAL IND.",
                ("--Pilih--", 25, 30, 35, 40, 45))

        with col3:
            ENG = st.selectbox(
                "JML. SOAL ENG.",
                ("--Pilih--", 25, 30, 35, 40, 45))

        with col4:
            IPA = st.selectbox(
                "JML. SOAL IPA.",
                ("--Pilih--", 25, 30, 35, 40, 45))

        with col5:
            IPS = st.selectbox(
                "JML. SOAL IPS.",
                ("--Pilih--", 25, 30, 35, 40, 45))

        JML_SOAL_MAT = MTK
        JML_SOAL_IND = IND
        JML_SOAL_ENG = ENG
        JML_SOAL_IPA = IPA
        JML_SOAL_IPS = IPS

        uploaded_file = st.file_uploader(
            'Letakkan file excel', type='xlsx')

        if uploaded_file is not None:

            wb = openpyxl.load_workbook(uploaded_file)
            ws = wb['Sheet1']

            q = len(ws['K'])
            r = len(ws['K'])+2
            s = len(ws['K'])+3
            t = len(ws['K'])+4
            u = len(ws['K'])+5
            v = len(ws['K'])+6
            w = len(ws['K'])+7
            x = len(ws['K'])+8

            ws['G{}'.format(r)] = "=ROUND(AVERAGE(G2:G{}),2)".format(q)  # mat
            ws['H{}'.format(r)] = "=ROUND(AVERAGE(H2:H{}),2)".format(q)  # ind
            ws['I{}'.format(r)] = "=ROUND(AVERAGE(I2:I{}),2)".format(q)  # eng
            ws['J{}'.format(r)] = "=ROUND(AVERAGE(J2:J{}),2)".format(q)  # ipa
            ws['K{}'.format(r)] = "=ROUND(AVERAGE(K2:K{}),2)".format(q)  # ips
            ws['L{}'.format(r)] = "=ROUND(AVERAGE(L2:L{}),2)".format(q)  # jml
            ws['G{}'.format(s)] = "=STDEV(G2:G{})".format(q)
            ws['H{}'.format(s)] = "=STDEV(H2:H{})".format(q)
            ws['I{}'.format(s)] = "=STDEV(I2:I{})".format(q)
            ws['J{}'.format(s)] = "=STDEV(J2:J{})".format(q)
            ws['K{}'.format(s)] = "=STDEV(K2:K{})".format(q)
            ws['G{}'.format(t)] = "=MAX(G2:G{})".format(q)
            ws['H{}'.format(t)] = "=MAX(H2:H{})".format(q)
            ws['I{}'.format(t)] = "=MAX(I2:I{})".format(q)
            ws['J{}'.format(t)] = "=MAX(J2:J{})".format(q)
            ws['K{}'.format(t)] = "=MAX(K2:K{})".format(q)
            ws['L{}'.format(t)] = "=MAX(L2:L{})".format(q)
            ws['M{}'.format(r)] = "=MAX(M2:M{})".format(q)
            ws['N{}'.format(r)] = "=MAX(N2:N{})".format(q)
            ws['O{}'.format(r)] = "=MAX(O2:O{})".format(q)
            ws['P{}'.format(r)] = "=MAX(P2:P{})".format(q)
            ws['Q{}'.format(r)] = "=MAX(Q2:Q{})".format(q)
            ws['R{}'.format(r)] = "=MAX(R2:R{})".format(q)
            ws['S{}'.format(r)] = "=MAX(S2:S{})".format(q)
            ws['T{}'.format(r)] = "=MAX(T2:T{})".format(q)
            ws['U{}'.format(r)] = "=MAX(U2:U{})".format(q)
            ws['V{}'.format(r)] = "=MAX(V2:V{})".format(q)
            ws['W{}'.format(r)] = "=ROUND(MAX(W2:W{}),2)".format(q)
            ws['G{}'.format(u)] = "=MIN(G2:G{})".format(q)
            ws['H{}'.format(u)] = "=MIN(H2:H{})".format(q)
            ws['I{}'.format(u)] = "=MIN(I2:I{})".format(q)
            ws['J{}'.format(u)] = "=MIN(J2:J{})".format(q)
            ws['K{}'.format(u)] = "=MIN(K2:K{})".format(q)
            ws['L{}'.format(u)] = "=MIN(L2:L{})".format(q)
            ws['R{}'.format(s)] = "=MIN(R2:R{})".format(q)
            ws['S{}'.format(s)] = "=MIN(S2:S{})".format(q)
            ws['T{}'.format(s)] = "=MIN(T2:T{})".format(q)
            ws['U{}'.format(s)] = "=MIN(U2:U{})".format(q)
            ws['V{}'.format(s)] = "=MIN(V2:V{})".format(q)
            ws['W{}'.format(s)] = "=MIN(W2:W{})".format(q)
            ws['R{}'.format(t)] = "=ROUND(AVERAGE(R2:R{}),2)".format(q)
            ws['S{}'.format(t)] = "=ROUND(AVERAGE(S2:S{}),2)".format(q)
            ws['T{}'.format(t)] = "=ROUND(AVERAGE(T2:T{}),2)".format(q)
            ws['U{}'.format(t)] = "=ROUND(AVERAGE(U2:U{}),2)".format(q)
            ws['V{}'.format(t)] = "=ROUND(AVERAGE(V2:V{}),2)".format(q)
            ws['W{}'.format(t)] = "=ROUND(AVERAGE(W2:W{}),2)".format(q)
            ws['X{}'.format(r)] = "=MAX(X2:X{})".format(q)
            ws['Z{}'.format(r)] = "=SUM(Z2:Z{})".format(q)
            ws['AA{}'.format(r)] = "=SUM(AA2:AA{})".format(q)
            ws['AB{}'.format(r)] = "=SUM(AB2:AB{})".format(q)
            ws['AC{}'.format(r)] = "=SUM(AC2:AC{})".format(q)
            ws['AD{}'.format(r)] = "=SUM(AD2:AD{})".format(q)
            # new
            # iterasi 1 rata-rata - 1
            ws['F{}'.format(v)] = 'JUMLAH SOAL'
            ws['G{}'.format(v)] = JML_SOAL_MAT
            ws['H{}'.format(v)] = JML_SOAL_IND
            ws['I{}'.format(v)] = JML_SOAL_ENG
            ws['J{}'.format(v)] = JML_SOAL_IPA
            ws['K{}'.format(v)] = JML_SOAL_IPS
            ws['AK{}'.format(r)] = "=IF($Z${}=0,$G${},$G${}-1)".format(r, r, r)
            ws['AK{}'.format(s)] = "=STDEV(AK2:AK{})".format(q)
            ws['AK{}'.format(t)] = "=MAX(AK2:AK{})".format(q)
            ws['AK{}'.format(u)] = "=MIN(AK2:AK{})".format(q)
            ws['AL{}'.format(
                r)] = "=IF($AA${}=0,$H${},$H${}-1)".format(r, r, r)
            ws['AL{}'.format(s)] = "=STDEV(AL2:AL{})".format(q)
            ws['AL{}'.format(t)] = "=MAX(AL2:AL{})".format(q)
            ws['AL{}'.format(u)] = "=MIN(AL2:AL{})".format(q)
            ws['AM{}'.format(
                r)] = "=IF($AB${}=0,$I${},$I${}-1)".format(r, r, r)
            ws['AM{}'.format(s)] = "=STDEV(AM2:AM{})".format(q)
            ws['AM{}'.format(t)] = "=MAX(AM2:AM{})".format(q)
            ws['AM{}'.format(u)] = "=MIN(AM2:AM{})".format(q)
            ws['AN{}'.format(
                r)] = "=IF($AC${}=0,$J${},$J${}-1)".format(r, r, r)
            ws['AN{}'.format(s)] = "=STDEV(AN2:AN{})".format(q)
            ws['AN{}'.format(t)] = "=MAX(AN2:AN{})".format(q)
            ws['AN{}'.format(u)] = "=MIN(AN2:AN{})".format(q)
            ws['AO{}'.format(
                r)] = "=IF($AD${}=0,$K${},$K${}-1)".format(r, r, r)
            ws['AO{}'.format(s)] = "=STDEV(AO2:AO{})".format(q)
            ws['AO{}'.format(t)] = "=MAX(AO2:AO{})".format(q)
            ws['AO{}'.format(u)] = "=MIN(AO2:AO{})".format(q)
            ws['AP{}'.format(r)] = "=ROUND(AVERAGE(AP2:AP{}),2)".format(q)
            ws['AP{}'.format(t)] = "=MAX(AP2:AP{})".format(q)
            ws['AP{}'.format(u)] = "=MIN(AP2:AP{})".format(q)
            ws['AQ{}'.format(r)] = "=MAX(AQ2:AQ{})".format(q)
            ws['AR{}'.format(r)] = "=MAX(AR2:AR{})".format(q)
            ws['AS{}'.format(r)] = "=MAX(AS2:AS{})".format(q)
            ws['AT{}'.format(r)] = "=MAX(AT2:AT{})".format(q)
            ws['AU{}'.format(r)] = "=MAX(AU2:AU{})".format(q)
            ws['AV{}'.format(r)] = "=MAX(AV2:AV{})".format(q)
            ws['AV{}'.format(s)] = "=MIN(AV2:AV{})".format(q)
            ws['AV{}'.format(t)] = "=ROUND(AVERAGE(AV2:AV{}),2)".format(q)
            ws['AW{}'.format(r)] = "=MAX(AW2:AW{})".format(q)
            ws['AW{}'.format(s)] = "=MIN(AW2:AW{})".format(q)
            ws['AW{}'.format(t)] = "=ROUND(AVERAGE(AW2:AW{}),2)".format(q)
            ws['AX{}'.format(r)] = "=MAX(AX2:AX{})".format(q)
            ws['AX{}'.format(s)] = "=MIN(AX2:AX{})".format(q)
            ws['AX{}'.format(t)] = "=ROUND(AVERAGE(AX2:AX{}),2)".format(q)
            ws['AY{}'.format(r)] = "=MAX(AY2:AY{})".format(q)
            ws['AY{}'.format(s)] = "=MIN(AY2:AY{})".format(q)
            ws['AY{}'.format(t)] = "=ROUND(AVERAGE(AY2:AY{}),2)".format(q)
            ws['AZ{}'.format(r)] = "=MAX(AZ2:AZ{})".format(q)
            ws['AZ{}'.format(s)] = "=MIN(AZ2:AZ{})".format(q)
            ws['AZ{}'.format(t)] = "=ROUND(AVERAGE(AZ2:AZ{}),2)".format(q)
            ws['BA{}'.format(r)] = "=MAX(BA2:BA{})".format(q)
            ws['BA{}'.format(s)] = "=MIN(BA2:BA{})".format(q)
            ws['BA{}'.format(t)] = "=ROUND(AVERAGE(BA2:BA{}),2)".format(q)
            ws['BD{}'.format(r)] = "=SUM(BD2:BD{})".format(q)
            ws['BE{}'.format(r)] = "=SUM(BE2:BE{})".format(q)
            ws['BF{}'.format(r)] = "=SUM(BF2:BF{})".format(q)
            ws['BG{}'.format(r)] = "=SUM(BG2:BG{})".format(q)
            ws['BH{}'.format(r)] = "=SUM(BH2:BH{})".format(q)

            # iterasi 2 rata-rata - 1
            ws['BO{}'.format(
                r)] = "=IF($BD${}=0,$AK${},$AK${}-1)".format(r, r, r)
            ws['BO{}'.format(s)] = "=STDEV(BO2:BO{})".format(q)
            ws['BO{}'.format(t)] = "=MAX(BO2:BO{})".format(q)
            ws['BO{}'.format(u)] = "=MIN(BO2:BO{})".format(q)
            ws['BP{}'.format(
                r)] = "=IF($BE${}=0,$AL${},$AL${}-1)".format(r, r, r)
            ws['BP{}'.format(s)] = "=STDEV(BP2:BP{})".format(q)
            ws['BP{}'.format(t)] = "=MAX(BP2:BP{})".format(q)
            ws['BP{}'.format(u)] = "=MIN(BP2:BP{})".format(q)
            ws['BQ{}'.format(
                r)] = "=IF($BF${}=0,$AM${},$AM${}-1)".format(r, r, r)
            ws['BQ{}'.format(s)] = "=STDEV(BQ2:BQ{})".format(q)
            ws['BQ{}'.format(t)] = "=MAX(BQ2:BQ{})".format(q)
            ws['BQ{}'.format(u)] = "=MIN(BQ2:BQ{})".format(q)
            ws['BR{}'.format(
                r)] = "=IF($BG${}=0,$AN${},$AN${}-1)".format(r, r, r)
            ws['BR{}'.format(s)] = "=STDEV(BR2:BR{})".format(q)
            ws['BR{}'.format(t)] = "=MAX(BR2:BR{})".format(q)
            ws['BR{}'.format(u)] = "=MIN(BR2:BR{})".format(q)
            ws['BS{}'.format(
                r)] = "=IF($BH${}=0,$AO${},$AO${}-1)".format(r, r, r)
            ws['BS{}'.format(s)] = "=STDEV(BS2:BS{})".format(q)
            ws['BS{}'.format(t)] = "=MAX(BS2:BS{})".format(q)
            ws['BS{}'.format(u)] = "=MIN(BS2:BS{})".format(q)
            ws['BT{}'.format(r)] = "=ROUND(AVERAGE(BT2:BT{}),2)".format(q)
            ws['BT{}'.format(t)] = "=MAX(BT2:BT{})".format(q)
            ws['BT{}'.format(u)] = "=MIN(BT2:BT{})".format(q)
            ws['BU{}'.format(r)] = "=MAX(BU2:BU{})".format(q)
            ws['BV{}'.format(r)] = "=MAX(BV2:BV{})".format(q)
            ws['BW{}'.format(r)] = "=MAX(BW2:BW{})".format(q)
            ws['BX{}'.format(r)] = "=MAX(BX2:BX{})".format(q)
            ws['BY{}'.format(r)] = "=MAX(BY2:BY{})".format(q)
            ws['BZ{}'.format(r)] = "=MAX(BZ2:BZ{})".format(q)
            ws['BZ{}'.format(s)] = "=MIN(BZ2:BZ{})".format(q)
            ws['BZ{}'.format(t)] = "=ROUND(AVERAGE(BZ2:BZ{}),2)".format(q)
            ws['CA{}'.format(r)] = "=MAX(CA2:CA{})".format(q)
            ws['CA{}'.format(s)] = "=MIN(CA2:CA{})".format(q)
            ws['CA{}'.format(t)] = "=ROUND(AVERAGE(CA2:CA{}),2)".format(q)
            ws['CB{}'.format(r)] = "=MAX(CB2:CB{})".format(q)
            ws['CB{}'.format(s)] = "=MIN(CB2:CB{})".format(q)
            ws['CB{}'.format(t)] = "=ROUND(AVERAGE(CB2:CB{}),2)".format(q)
            ws['CC{}'.format(r)] = "=MAX(CC2:CC{})".format(q)
            ws['CC{}'.format(s)] = "=MIN(CC2:CC{})".format(q)
            ws['CC{}'.format(t)] = "=ROUND(AVERAGE(CC2:CC{}),2)".format(q)
            ws['CD{}'.format(r)] = "=MAX(CD2:CD{})".format(q)
            ws['CD{}'.format(s)] = "=MIN(CD2:CD{})".format(q)
            ws['CD{}'.format(t)] = "=ROUND(AVERAGE(CD2:CD{}),2)".format(q)
            ws['CE{}'.format(r)] = "=MAX(CE2:CE{})".format(q)
            ws['CE{}'.format(s)] = "=MIN(CE2:CE{})".format(q)
            ws['CE{}'.format(t)] = "=ROUND(AVERAGE(CE2:CE{}),2)".format(q)
            ws['CH{}'.format(r)] = "=SUM(CH2:CH{})".format(q)
            ws['CI{}'.format(r)] = "=SUM(CI2:CI{})".format(q)
            ws['CJ{}'.format(r)] = "=SUM(CJ2:CJ{})".format(q)
            ws['CK{}'.format(r)] = "=SUM(CK2:CK{})".format(q)
            ws['CL{}'.format(r)] = "=SUM(CL2:CL{})".format(q)

            # iterasi 3 rata-rata - 1
            ws['CS{}'.format(
                r)] = "=IF($CH${}=0,$BO${},$BO${}-1)".format(r, r, r)
            ws['CS{}'.format(s)] = "=STDEV(CS2:CS{})".format(q)
            ws['CS{}'.format(t)] = "=MAX(CS2:CS{})".format(q)
            ws['CS{}'.format(u)] = "=MIN(CS2:CS{})".format(q)
            ws['CT{}'.format(
                r)] = "=IF($CI${}=0,$BP${},$BP${}-1)".format(r, r, r)
            ws['CT{}'.format(s)] = "=STDEV(CT2:CT{})".format(q)
            ws['CT{}'.format(t)] = "=MAX(CT2:CT{})".format(q)
            ws['CT{}'.format(u)] = "=MIN(CT2:CT{})".format(q)
            ws['CU{}'.format(
                r)] = "=IF($CJ${}=0,$BQ${},$BQ${}-1)".format(r, r, r)
            ws['CU{}'.format(s)] = "=STDEV(CU2:CU{})".format(q)
            ws['CU{}'.format(t)] = "=MAX(CU2:CU{})".format(q)
            ws['CU{}'.format(u)] = "=MIN(CU2:CU{})".format(q)
            ws['CV{}'.format(
                r)] = "=IF($CK${}=0,$BR${},$BR${}-1)".format(r, r, r)
            ws['CV{}'.format(s)] = "=STDEV(CV2:CV{})".format(q)
            ws['CV{}'.format(t)] = "=MAX(CV2:CV{})".format(q)
            ws['CV{}'.format(u)] = "=MIN(CV2:CV{})".format(q)
            ws['CW{}'.format(
                r)] = "=IF($CL${}=0,$BS${},$BS${}-1)".format(r, r, r)
            ws['CW{}'.format(s)] = "=STDEV(CW2:CW{})".format(q)
            ws['CW{}'.format(t)] = "=MAX(CW2:CW{})".format(q)
            ws['CW{}'.format(u)] = "=MIN(CW2:CW{})".format(q)
            ws['CX{}'.format(r)] = "=ROUND(AVERAGE(CX2:CX{}),2)".format(q)
            ws['CX{}'.format(t)] = "=MAX(CX2:CX{})".format(q)
            ws['CX{}'.format(u)] = "=MIN(CX2:CX{})".format(q)
            ws['CY{}'.format(r)] = "=MAX(CY2:CY{})".format(q)
            ws['CZ{}'.format(r)] = "=MAX(CZ2:CZ{})".format(q)
            ws['DA{}'.format(r)] = "=MAX(DA2:DA{})".format(q)
            ws['DB{}'.format(r)] = "=MAX(DB2:DB{})".format(q)
            ws['DC{}'.format(r)] = "=MAX(DC2:DC{})".format(q)
            ws['DD{}'.format(r)] = "=MAX(DD2:DD{})".format(q)
            ws['DD{}'.format(s)] = "=MIN(DD2:DD{})".format(q)
            ws['DD{}'.format(t)] = "=ROUND(AVERAGE(DD2:DD{}),2)".format(q)
            ws['DE{}'.format(r)] = "=MAX(DE2:DE{})".format(q)
            ws['DE{}'.format(s)] = "=MIN(DE2:DE{})".format(q)
            ws['DE{}'.format(t)] = "=ROUND(AVERAGE(DE2:DE{}),2)".format(q)
            ws['DF{}'.format(r)] = "=MAX(DF2:DF{})".format(q)
            ws['DF{}'.format(s)] = "=MIN(DF2:DF{})".format(q)
            ws['DF{}'.format(t)] = "=ROUND(AVERAGE(DF2:DF{}),2)".format(q)
            ws['DG{}'.format(r)] = "=MAX(DG2:DG{})".format(q)
            ws['DG{}'.format(s)] = "=MIN(DG2:DG{})".format(q)
            ws['DG{}'.format(t)] = "=ROUND(AVERAGE(DG2:DG{}),2)".format(q)
            ws['DH{}'.format(r)] = "=MAX(DH2:DH{})".format(q)
            ws['DH{}'.format(s)] = "=MIN(DH2:DH{})".format(q)
            ws['DH{}'.format(t)] = "=ROUND(AVERAGE(DH2:DH{}),2)".format(q)
            ws['DI{}'.format(r)] = "=MAX(DI2:DI{})".format(q)
            ws['DI{}'.format(s)] = "=MIN(DI2:DI{})".format(q)
            ws['DI{}'.format(t)] = "=ROUND(AVERAGE(DI2:DI{}),2)".format(q)
            ws['DL{}'.format(r)] = "=SUM(DL2:DL{})".format(q)
            ws['DM{}'.format(r)] = "=SUM(DM2:DM{})".format(q)
            ws['DN{}'.format(r)] = "=SUM(DN2:DN{})".format(q)
            ws['DO{}'.format(r)] = "=SUM(DO2:DO{})".format(q)
            ws['DP{}'.format(r)] = "=SUM(DP2:DP{})".format(q)

            # iterasi 4 rata-rata - 1
            ws['DW{}'.format(
                r)] = "=IF($DL${}=0,$CS${},$CS${}-1)".format(r, r, r)
            ws['DW{}'.format(s)] = "=STDEV(DW2:DW{})".format(q)
            ws['DW{}'.format(t)] = "=MAX(DW2:DW{})".format(q)
            ws['DW{}'.format(u)] = "=MIN(DW2:DW{})".format(q)
            ws['DX{}'.format(
                r)] = "=IF($DM${}=0,$CT${},$CT${}-1)".format(r, r, r)
            ws['DX{}'.format(s)] = "=STDEV(DX2:DX{})".format(q)
            ws['DX{}'.format(t)] = "=MAX(DX2:DX{})".format(q)
            ws['DX{}'.format(u)] = "=MIN(DX2:DX{})".format(q)
            ws['DY{}'.format(
                r)] = "=IF($DN${}=0,$CU${},$CU${}-1)".format(r, r, r)
            ws['DY{}'.format(s)] = "=STDEV(DY2:DY{})".format(q)
            ws['DY{}'.format(t)] = "=MAX(DY2:DY{})".format(q)
            ws['DY{}'.format(u)] = "=MIN(DY2:DY{})".format(q)
            ws['DZ{}'.format(
                r)] = "=IF($DO${}=0,$CV${},$CV${}-1)".format(r, r, r)
            ws['DZ{}'.format(s)] = "=STDEV(DZ2:DZ{})".format(q)
            ws['DZ{}'.format(t)] = "=MAX(DZ2:DZ{})".format(q)
            ws['DZ{}'.format(u)] = "=MIN(DZ2:DZ{})".format(q)
            ws['EA{}'.format(
                r)] = "=IF($DP${}=0,$CW${},$CW${}-1)".format(r, r, r)
            ws['EA{}'.format(s)] = "=STDEV(EA2:EA{})".format(q)
            ws['EA{}'.format(t)] = "=MAX(EA2:EA{})".format(q)
            ws['EA{}'.format(u)] = "=MIN(EA2:EA{})".format(q)
            ws['EB{}'.format(r)] = "=ROUND(AVERAGE(EB2:EB{}),2)".format(q)
            ws['EB{}'.format(t)] = "=MAX(EB2:EB{})".format(q)
            ws['EB{}'.format(u)] = "=MIN(EB2:EB{})".format(q)
            ws['EC{}'.format(r)] = "=MAX(EC2:EC{})".format(q)
            ws['ED{}'.format(r)] = "=MAX(ED2:ED{})".format(q)
            ws['EE{}'.format(r)] = "=MAX(EE2:EE{})".format(q)
            ws['EF{}'.format(r)] = "=MAX(EF2:EF{})".format(q)
            ws['EG{}'.format(r)] = "=MAX(EG2:EG{})".format(q)
            ws['EH{}'.format(r)] = "=MAX(EH2:EH{})".format(q)
            ws['EH{}'.format(s)] = "=MIN(EH2:EH{})".format(q)
            ws['EH{}'.format(t)] = "=ROUND(AVERAGE(EH2:EH{}),2)".format(q)
            ws['EI{}'.format(r)] = "=MAX(EI2:EI{})".format(q)
            ws['EI{}'.format(s)] = "=MIN(EI2:EI{})".format(q)
            ws['EI{}'.format(t)] = "=ROUND(AVERAGE(EI2:EI{}),2)".format(q)
            ws['EJ{}'.format(r)] = "=MAX(EJ2:EJ{})".format(q)
            ws['EJ{}'.format(s)] = "=MIN(EJ2:EJ{})".format(q)
            ws['EJ{}'.format(t)] = "=ROUND(AVERAGE(EJ2:EJ{}),2)".format(q)
            ws['EK{}'.format(r)] = "=MAX(EK2:EK{})".format(q)
            ws['EK{}'.format(s)] = "=MIN(EK2:EK{})".format(q)
            ws['EK{}'.format(t)] = "=ROUND(AVERAGE(EK2:EK{}),2)".format(q)
            ws['EL{}'.format(r)] = "=MAX(EL2:EL{})".format(q)
            ws['EL{}'.format(s)] = "=MIN(EL2:EL{})".format(q)
            ws['EL{}'.format(t)] = "=ROUND(AVERAGE(EL2:EL{}),2)".format(q)
            ws['EM{}'.format(r)] = "=MAX(EM2:EM{})".format(q)
            ws['EM{}'.format(s)] = "=MIN(EM2:EM{})".format(q)
            ws['EM{}'.format(t)] = "=ROUND(AVERAGE(EM2:EM{}),2)".format(q)
            ws['EP{}'.format(r)] = "=SUM(EP2:EP{})".format(q)
            ws['EQ{}'.format(r)] = "=SUM(EQ2:EQ{})".format(q)
            ws['ER{}'.format(r)] = "=SUM(ER2:ER{})".format(q)
            ws['ES{}'.format(r)] = "=SUM(ES2:ES{})".format(q)
            ws['ET{}'.format(r)] = "=SUM(ET2:ET{})".format(q)

            # iterasi 5 rata-rata - 1
            ws['FA{}'.format(
                r)] = "=IF($EP${}=0,$DW${},$DW${}-1)".format(r, r, r)
            ws['FA{}'.format(s)] = "=STDEV(FA2:FA{})".format(q)
            ws['FA{}'.format(t)] = "=MAX(FA2:FA{})".format(q)
            ws['FA{}'.format(u)] = "=MIN(FA2:FA{})".format(q)
            ws['FB{}'.format(
                r)] = "=IF($EQ${}=0,$DX${},$DX${}-1)".format(r, r, r)
            ws['FB{}'.format(s)] = "=STDEV(FB2:FB{})".format(q)
            ws['FB{}'.format(t)] = "=MAX(FB2:FB{})".format(q)
            ws['FB{}'.format(u)] = "=MIN(FB2:FB{})".format(q)
            ws['FC{}'.format(
                r)] = "=IF($ER${}=0,$DY${},$DY${}-1)".format(r, r, r)
            ws['FC{}'.format(s)] = "=STDEV(FC2:FC{})".format(q)
            ws['FC{}'.format(t)] = "=MAX(FC2:FC{})".format(q)
            ws['FC{}'.format(u)] = "=MIN(FC2:FC{})".format(q)
            ws['FD{}'.format(
                r)] = "=IF($ES${}=0,$DZ${},$DZ${}-1)".format(r, r, r)
            ws['FD{}'.format(s)] = "=STDEV(FD2:FD{})".format(q)
            ws['FD{}'.format(t)] = "=MAX(FD2:FD{})".format(q)
            ws['FD{}'.format(u)] = "=MIN(FD2:FD{})".format(q)
            ws['FE{}'.format(
                r)] = "=IF($ET${}=0,$EA${},$EA${}-1)".format(r, r, r)
            ws['FE{}'.format(s)] = "=STDEV(FE2:FE{})".format(q)
            ws['FE{}'.format(t)] = "=MAX(FE2:FE{})".format(q)
            ws['FE{}'.format(u)] = "=MIN(FE2:FE{})".format(q)
            ws['FF{}'.format(r)] = "=ROUND(AVERAGE(FF2:FF{}),2)".format(q)
            ws['FF{}'.format(t)] = "=MAX(FF2:FF{})".format(q)
            ws['FF{}'.format(u)] = "=MIN(FF2:FF{})".format(q)
            ws['FG{}'.format(r)] = "=MAX(FG2:FG{})".format(q)
            ws['FH{}'.format(r)] = "=MAX(FH2:FH{})".format(q)
            ws['FI{}'.format(r)] = "=MAX(FI2:FI{})".format(q)
            ws['FJ{}'.format(r)] = "=MAX(FJ2:FJ{})".format(q)
            ws['FK{}'.format(r)] = "=MAX(FK2:FK{})".format(q)
            ws['FL{}'.format(r)] = "=MAX(FL2:FL{})".format(q)
            ws['FL{}'.format(s)] = "=MIN(FL2:FL{})".format(q)
            ws['FL{}'.format(t)] = "=ROUND(AVERAGE(FL2:FL{}),2)".format(q)
            ws['FM{}'.format(r)] = "=MAX(FM2:FM{})".format(q)
            ws['FM{}'.format(s)] = "=MIN(FM2:FM{})".format(q)
            ws['FM{}'.format(t)] = "=ROUND(AVERAGE(FM2:FM{}),2)".format(q)
            ws['FN{}'.format(r)] = "=MAX(FN2:FN{})".format(q)
            ws['FN{}'.format(s)] = "=MIN(FN2:FN{})".format(q)
            ws['FN{}'.format(t)] = "=ROUND(AVERAGE(FN2:FN{}),2)".format(q)
            ws['FO{}'.format(r)] = "=MAX(FO2:FO{})".format(q)
            ws['FO{}'.format(s)] = "=MIN(FO2:FO{})".format(q)
            ws['FO{}'.format(t)] = "=ROUND(AVERAGE(FO2:FO{}),2)".format(q)
            ws['FP{}'.format(r)] = "=MAX(FP2:FP{})".format(q)
            ws['FP{}'.format(s)] = "=MIN(FP2:FP{})".format(q)
            ws['FP{}'.format(t)] = "=ROUND(AVERAGE(FP2:FP{}),2)".format(q)
            ws['FQ{}'.format(r)] = "=MAX(FQ2:FQ{})".format(q)
            ws['FQ{}'.format(s)] = "=MIN(FQ2:FQ{})".format(q)
            ws['FQ{}'.format(t)] = "=ROUND(AVERAGE(FQ2:FQ{}),2)".format(q)
            ws['FT{}'.format(r)] = "=SUM(FT2:FT{})".format(q)
            ws['FU{}'.format(r)] = "=SUM(FU2:FU{})".format(q)
            ws['FV{}'.format(r)] = "=SUM(FV2:FV{})".format(q)
            ws['FW{}'.format(r)] = "=SUM(FW2:FW{})".format(q)
            ws['FX{}'.format(r)] = "=SUM(FX2:FX{})".format(q)

            # Z Score
            ws['B1'] = 'NAMA_SISWA_1'
            ws['C1'] = 'NOMOR_NF_1'
            ws['D1'] = 'KELAS_1'
            ws['E1'] = 'NAMA_SEKOLAH_1'
            ws['F1'] = 'LOKASI_1'
            ws['G1'] = 'MAT_1'
            ws['H1'] = 'IND_1'
            ws['I1'] = 'ENG_1'
            ws['J1'] = 'IPA_1'
            ws['K1'] = 'IPS_1'
            ws['L1'] = 'JML_1'
            ws['M1'] = 'Z_MAT_1'
            ws['N1'] = 'Z_IND_1'
            ws['O1'] = 'Z_ENG_1'
            ws['P1'] = 'Z_IPA_1'
            ws['Q1'] = 'Z_IPS_1'
            ws['R1'] = 'S_MAT_1'
            ws['S1'] = 'S_IND_1'
            ws['T1'] = 'S_ENG_1'
            ws['U1'] = 'S_IPA_1'
            ws['V1'] = 'S_IPS_1'
            ws['W1'] = 'S_JML_1'
            ws['X1'] = 'RANK_NAS._1'
            ws['Y1'] = 'RANK_LOK._1'
            ws['M1'].font = Font(bold=False, name='Calibri', size=11)
            ws['N1'].font = Font(bold=False, name='Calibri', size=11)
            ws['O1'].font = Font(bold=False, name='Calibri', size=11)
            ws['P1'].font = Font(bold=False, name='Calibri', size=11)
            ws['Q1'].font = Font(bold=False, name='Calibri', size=11)
            ws['R1'].font = Font(bold=False, name='Calibri', size=11)
            ws['S1'].font = Font(bold=False, name='Calibri', size=11)
            ws['T1'].font = Font(bold=False, name='Calibri', size=11)
            ws['U1'].font = Font(bold=False, name='Calibri', size=11)
            ws['V1'].font = Font(bold=False, name='Calibri', size=11)
            ws['W1'].font = Font(bold=False, name='Calibri', size=11)
            ws['X1'].font = Font(bold=False, name='Calibri', size=11)
            ws['Y1'].font = Font(bold=False, name='Calibri', size=11)
        # FILL
            ws['B1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['C1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['D1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['E1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['F1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['G1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['H1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['I1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['J1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['K1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['L1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['M1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['N1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['O1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['P1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['Q1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['R1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['S1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['T1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['U1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['V1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['W1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['X1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['Y1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            # tambahan
            ws['Z1'] = 'MAT_20_1'
            ws['AA1'] = 'IND_20_1'
            ws['AB1'] = 'ENG_20_1'
            ws['AC1'] = 'IPA_20_1'
            ws['AD1'] = 'IPS_20_1'
            ws['Z1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AA1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AB1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AC1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AD1'].font = Font(bold=False, name='Calibri', size=11)
            ws['Z1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AA1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AB1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AC1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AD1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')

            for row in range(2, q+1):
                ws['L{}'.format(
                    row)] = '=SUM(G{}:K{})'.format(row, row, row)
                ws['M{}'.format(
                    row)] = '=IFERROR(ROUND(IF(G{}="","",(G{}-G${})/G${}),2),"")'.format(row, row, r, s)
                ws['N{}'.format(
                    row)] = '=IFERROR(ROUND(IF(H{}="","",(H{}-H${})/H${}),2),"")'.format(row, row, r, s)
                ws['O{}'.format(
                    row)] = '=IFERROR(ROUND(IF(I{}="","",(I{}-I${})/I${}),2),"")'.format(row, row, r, s)
                ws['P{}'.format(
                    row)] = '=IFERROR(ROUND(IF(J{}="","",(J{}-J${})/J${}),2),"")'.format(row, row, r, s)
                ws['Q{}'.format(
                    row)] = '=IFERROR(ROUND(IF(K{}="","",(K{}-K${})/K${}),2),"")'.format(row, row, r, s)
                ws['R{}'.format(
                    row)] = '=IFERROR(ROUND(IF(G{}="","",IF(70+30*M{}/$M${}<20,20,70+30*M{}/$M${})),2),"")'.format(row, row, r, row, r)
                ws['S{}'.format(
                    row)] = '=IFERROR(ROUND(IF(H{}="","",IF(70+30*N{}/$N${}<20,20,70+30*N{}/$N${})),2),"")'.format(row, row, r, row, r)
                ws['T{}'.format(
                    row)] = '=IFERROR(ROUND(IF(I{}="","",IF(70+30*O{}/$O${}<20,20,70+30*O{}/$O${})),2),"")'.format(row, row, r, row, r)
                ws['U{}'.format(
                    row)] = '=IFERROR(ROUND(IF(J{}="","",IF(70+30*P{}/$P${}<20,20,70+30*P{}/$P${})),2),"")'.format(row, row, r, row, r)
                ws['V{}'.format(
                    row)] = '=IFERROR(ROUND(IF(K{}="","",IF(70+30*Q{}/$Q${}<20,20,70+30*Q{}/$Q${})),2),"")'.format(row, row, r, row, r)

                ws['W{}'.format(row)] = '=IF(SUM(R{}:V{})=0,"",SUM(R{}:V{}))'.format(
                    row, row, row, row)
                ws['X{}'.format(row)] = '=IF(W{}="","",RANK(W{},$W$2:$W${}))'.format(
                    row, row, q)
                ws['Y{}'.format(
                    row)] = '=IF(X{}="","",COUNTIFS($F$2:$F${},F{},$X$2:$X${},"<"&X{})+1)'.format(row, q, row, q, row)
                # TAMBAHAN
                ws['Z{}'.format(row)] = '=IF($G${}=25,IF(AND(G{}>4,R{}=20),1,""),IF($G${}=30,IF(AND(G{}>5,R{}=20),1,""),IF($G${}=35,IF(AND(G{}>6,R{}=20),1,""),IF($G${}=40,IF(AND(G{}>7,R{}=20),1,""),IF($G${}=45,IF(AND(G{}>8,R{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['AA{}'.format(row)] = '=IF($H${}=25,IF(AND(H{}>4,S{}=20),1,""),IF($H${}=30,IF(AND(H{}>5,S{}=20),1,""),IF($H${}=35,IF(AND(H{}>6,S{}=20),1,""),IF($H${}=40,IF(AND(H{}>7,S{}=20),1,""),IF($H${}=45,IF(AND(H{}>8,S{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['AB{}'.format(row)] = '=IF($I${}=25,IF(AND(I{}>4,T{}=20),1,""),IF($I${}=30,IF(AND(I{}>5,T{}=20),1,""),IF($I${}=35,IF(AND(I{}>6,T{}=20),1,""),IF($I${}=40,IF(AND(I{}>7,T{}=20),1,""),IF($I${}=45,IF(AND(I{}>8,T{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['AC{}'.format(row)] = '=IF($J${}=25,IF(AND(J{}>4,U{}=20),1,""),IF($J${}=30,IF(AND(J{}>5,U{}=20),1,""),IF($J${}=35,IF(AND(J{}>6,U{}=20),1,""),IF($J${}=40,IF(AND(J{}>7,U{}=20),1,""),IF($J${}=45,IF(AND(J{}>8,U{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['AD{}'.format(row)] = '=IF($K${}=25,IF(AND(K{}>4,V{}=20),1,""),IF($K${}=30,IF(AND(K{}>5,V{}=20),1,""),IF($K${}=35,IF(AND(K{}>6,V{}=20),1,""),IF($K${}=40,IF(AND(K{}>7,V{}=20),1,""),IF($K${}=45,IF(AND(K{}>8,V{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

        # new Z Score
            ws['AF1'] = 'NAMA_SISWA_2'
            ws['AG1'] = 'NOMOR_NF_2'
            ws['AH1'] = 'KELAS_2'
            ws['AI1'] = 'NAMA_SEKOLAH_2'
            ws['AJ1'] = 'LOKASI_2'
            ws['AK1'] = 'MAT_2'
            ws['AL1'] = 'IND_2'
            ws['AM1'] = 'ENG_2'
            ws['AN1'] = 'IPA_2'
            ws['AO1'] = 'IPS_2'
            ws['AP1'] = 'JML_2'
            ws['AQ1'] = 'Z_MAT_2'
            ws['AR1'] = 'Z_IND_2'
            ws['AS1'] = 'Z_ENG_2'
            ws['AT1'] = 'Z_IPA_2'
            ws['AU1'] = 'Z_IPS_2'
            ws['AV1'] = 'S_MAT_2'
            ws['AW1'] = 'S_IND_2'
            ws['AX1'] = 'S_ENG_2'
            ws['AY1'] = 'S_IPA_2'
            ws['AZ1'] = 'S_IPS_2'
            ws['BA1'] = 'S_JML_2'
            ws['BB1'] = 'RANK_NAS._2'
            ws['BC1'] = 'RANK_LOK._2'
            ws['AQ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AR1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AS1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AT1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AU1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AV1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AW1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AX1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AY1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AZ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BA1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BB1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BC1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['AF1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AG1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AH1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AI1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AJ1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AK1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AL1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AM1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AN1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AO1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AP1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AQ1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AR1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AS1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AT1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AU1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AV1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AW1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AX1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AY1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AZ1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BA1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BB1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BC1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')

            # tambahan
            ws['BD1'] = 'MAT_20_2'
            ws['BE1'] = 'IND_20_2'
            ws['BF1'] = 'ENG_20_2'
            ws['BG1'] = 'IPA_20_2'
            ws['BH1'] = 'IPS_20_2'
            ws['BD1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BE1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BF1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BG1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BH1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BD1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BE1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BF1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BG1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BH1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')

            for row in range(2, q+1):
                ws['AF{}'.format(row)] = '=B{}'.format(row)
                ws['AG{}'.format(row)] = '=C{}'.format(row, row)
                ws['AH{}'.format(row)] = '=D{}'.format(row, row)
                ws['AI{}'.format(row)] = '=E{}'.format(row, row)
                ws['AJ{}'.format(row)] = '=F{}'.format(row, row)
                ws['AK{}'.format(row)] = '=IF(G{}="","",G{})'.format(row, row)
                ws['AL{}'.format(row)] = '=IF(H{}="","",H{})'.format(row, row)
                ws['AM{}'.format(row)] = '=IF(I{}="","",I{})'.format(row, row)
                ws['AN{}'.format(row)] = '=IF(J{}="","",J{})'.format(row, row)
                ws['AO{}'.format(row)] = '=IF(K{}="","",K{})'.format(row, row)
                ws['AP{}'.format(row)] = '=IF(L{}="","",L{})'.format(row, row)
                ws['AQ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AK{}="","",(AK{}-AK${})/AK${}),2),"")'.format(row, row, r, s)
                ws['AR{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AL{}="","",(AL{}-AL${})/AL${}),2),"")'.format(row, row, r, s)
                ws['AS{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AM{}="","",(AM{}-AM${})/AM${}),2),"")'.format(row, row, r, s)
                ws['AT{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AN{}="","",(AN{}-AN${})/AN${}),2),"")'.format(row, row, r, s)
                ws['AU{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AO{}="","",(AO{}-AO${})/AO${}),2),"")'.format(row, row, r, s)
                ws['AV{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AK{}="","",IF(70+30*AQ{}/$AQ${}<20,20,70+30*AQ{}/$AQ${})),2),"")'.format(row, row, r, row, r)
                ws['AW{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AL{}="","",IF(70+30*AR{}/$AR${}<20,20,70+30*AR{}/$AR${})),2),"")'.format(row, row, r, row, r)
                ws['AX{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AM{}="","",IF(70+30*AS{}/$AS${}<20,20,70+30*AS{}/$AS${})),2),"")'.format(row, row, r, row, r)
                ws['AY{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AN{}="","",IF(70+30*AT{}/$AT${}<20,20,70+30*AT{}/$AT${})),2),"")'.format(row, row, r, row, r)
                ws['AZ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AO{}="","",IF(70+30*AU{}/$AU${}<20,20,70+30*AU{}/$AU${})),2),"")'.format(row, row, r, row, r)

                ws['BA{}'.format(row)] = '=IF(SUM(AV{}:AZ{})=0,"",SUM(AV{}:AZ{}))'.format(
                    row, row, row, row)
                ws['BB{}'.format(row)] = '=IF(BA{}="","",RANK(BA{},$BA$2:$BA${}))'.format(
                    row, row, q)
                ws['BC{}'.format(
                    row)] = '=IF(BB{}="","",COUNTIFS($AJ$2:$AJ${},F{},$BB$2:$BB${},"<"&BB{})+1)'.format(row, q, row, q, row)
            #     TAMBAHAN
                ws['BD{}'.format(row)] = '=IF($G${}=25,IF(AND(AK{}>4,AV{}=20),1,""),IF($G${}=30,IF(AND(AK{}>5,AV{}=20),1,""),IF($G${}=35,IF(AND(AK{}>6,AV{}=20),1,""),IF($G${}=40,IF(AND(AK{}>7,AV{}=20),1,""),IF($G${}=45,IF(AND(AK{}>8,AV{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['BE{}'.format(row)] = '=IF($H${}=25,IF(AND(AL{}>4,AW{}=20),1,""),IF($H${}=30,IF(AND(AL{}>5,AW{}=20),1,""),IF($H${}=35,IF(AND(AL{}>6,AW{}=20),1,""),IF($H${}=40,IF(AND(AL{}>7,AW{}=20),1,""),IF($H${}=45,IF(AND(AL{}>8,AW{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['BF{}'.format(row)] = '=IF($I${}=25,IF(AND(AM{}>4,AX{}=20),1,""),IF($I${}=30,IF(AND(AM{}>5,AX{}=20),1,""),IF($I${}=35,IF(AND(AM{}>6,AX{}=20),1,""),IF($I${}=40,IF(AND(AM{}>7,AX{}=20),1,""),IF($I${}=45,IF(AND(AM{}>8,AX{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['BG{}'.format(row)] = '=IF($J${}=25,IF(AND(AN{}>4,AY{}=20),1,""),IF($J${}=30,IF(AND(AN{}>5,AY{}=20),1,""),IF($J${}=35,IF(AND(AN{}>6,AY{}=20),1,""),IF($J${}=40,IF(AND(AN{}>7,AY{}=20),1,""),IF($J${}=45,IF(AND(AN{}>8,AY{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['BH{}'.format(row)] = '=IF($K${}=25,IF(AND(AO{}>4,AZ{}=20),1,""),IF($K${}=30,IF(AND(AO{}>5,AZ{}=20),1,""),IF($K${}=35,IF(AND(AO{}>6,AZ{}=20),1,""),IF($K${}=40,IF(AND(AO{}>7,AZ{}=20),1,""),IF($K${}=45,IF(AND(AO{}>8,AZ{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

                # new Z Score [2]
            ws['BJ1'] = 'NAMA_SISWA_3'
            ws['BK1'] = 'NOMOR_NF_3'
            ws['BL1'] = 'KELAS_3'
            ws['BM1'] = 'NAMA_SEKOLAH_3'
            ws['BN1'] = 'LOKASI_3'
            ws['BO1'] = 'MAT_3'
            ws['BP1'] = 'IND_3'
            ws['BQ1'] = 'ENG_3'
            ws['BR1'] = 'IPA_3'
            ws['BS1'] = 'IPS_3'
            ws['BT1'] = 'JML_3'
            ws['BU1'] = 'Z_MAT_3'
            ws['BV1'] = 'Z_IND_3'
            ws['BW1'] = 'Z_ENG_3'
            ws['BX1'] = 'Z_IPA_3'
            ws['BY1'] = 'Z_IPS_3'
            ws['BZ1'] = 'S_MAT_3'
            ws['CA1'] = 'S_IND_3'
            ws['CB1'] = 'S_ENG_3'
            ws['CC1'] = 'S_IPA_3'
            ws['CD1'] = 'S_IPS_3'
            ws['CE1'] = 'S_JML_3'
            ws['CF1'] = 'RANK_NAS._3'
            ws['CG1'] = 'RANK_LOK._3'
            ws['BU1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BV1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BW1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BX1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BY1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BZ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CA1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CB1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CC1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CD1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CE1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CF1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CG1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['BJ1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BK1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BL1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BM1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BN1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BO1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BP1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BQ1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BR1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BS1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BT1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BU1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BV1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BW1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BX1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BY1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BZ1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CA1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CB1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CC1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CD1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CE1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CF1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CG1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')

            # tambahan
            ws['CH1'] = 'MAT_20_3'
            ws['CI1'] = 'IND_20_3'
            ws['CJ1'] = 'ENG_20_3'
            ws['CK1'] = 'IPA_20_3'
            ws['CL1'] = 'IPS_20_3'
            ws['CH1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CI1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CJ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CK1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CL1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CH1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CI1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CJ1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CK1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CL1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')

            for row in range(2, q+1):
                ws['BJ{}'.format(row)] = '=B{}'.format(row)
                ws['BK{}'.format(row)] = '=C{}'.format(row, row)
                ws['BL{}'.format(row)] = '=D{}'.format(row, row)
                ws['BM{}'.format(row)] = '=E{}'.format(row, row)
                ws['BN{}'.format(row)] = '=F{}'.format(row, row)
                ws['BO{}'.format(row)] = '=IF(G{}="","",G{})'.format(row, row)
                ws['BP{}'.format(row)] = '=IF(H{}="","",H{})'.format(row, row)
                ws['BQ{}'.format(row)] = '=IF(I{}="","",I{})'.format(row, row)
                ws['BR{}'.format(row)] = '=IF(J{}="","",J{})'.format(row, row)
                ws['BS{}'.format(row)] = '=IF(K{}="","",K{})'.format(row, row)
                ws['BT{}'.format(row)] = '=IF(L{}="","",L{})'.format(row, row)
                ws['BU{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BO{}="","",(BO{}-BO${})/BO${}),2),"")'.format(row, row, r, s)
                ws['BV{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BP{}="","",(BP{}-BP${})/BP${}),2),"")'.format(row, row, r, s)
                ws['BW{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BQ{}="","",(BQ{}-BQ${})/BQ${}),2),"")'.format(row, row, r, s)
                ws['BX{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BR{}="","",(BR{}-BR${})/BR${}),2),"")'.format(row, row, r, s)
                ws['BY{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BS{}="","",(BS{}-BS${})/BS${}),2),"")'.format(row, row, r, s)
                ws['BZ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BO{}="","",IF(70+30*BU{}/$BU${}<20,20,70+30*BU{}/$BU${})),2),"")'.format(row, row, r, row, r)
                ws['CA{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BP{}="","",IF(70+30*BV{}/$BV${}<20,20,70+30*BV{}/$BV${})),2),"")'.format(row, row, r, row, r)
                ws['CB{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BQ{}="","",IF(70+30*BW{}/$BW${}<20,20,70+30*BW{}/$BW${})),2),"")'.format(row, row, r, row, r)
                ws['CC{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BR{}="","",IF(70+30*BX{}/$BX${}<20,20,70+30*BX{}/$BX${})),2),"")'.format(row, row, r, row, r)
                ws['CD{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BS{}="","",IF(70+30*BY{}/$BY${}<20,20,70+30*BY{}/$BY${})),2),"")'.format(row, row, r, row, r)

                ws['CE{}'.format(row)] = '=IF(SUM(BZ{}:CD{})=0,"",SUM(BZ{}:CD{}))'.format(
                    row, row, row, row)
                ws['CF{}'.format(row)] = '=IF(CE{}="","",RANK(CE{},$CE$2:$CE${}))'.format(
                    row, row, q)
                ws['CG{}'.format(
                    row)] = '=IF(CF{}="","",COUNTIFS($BN$2:$BN${},F{},$CF$2:$CF${},"<"&CF{})+1)'.format(row, q, row, q, row)
                #     TAMBAHAN
                ws['CH{}'.format(row)] = '=IF($G${}=25,IF(AND(BO{}>4,BZ{}=20),1,""),IF($G${}=30,IF(AND(BO{}>5,BZ{}=20),1,""),IF($G${}=35,IF(AND(BO{}>6,BZ{}=20),1,""),IF($G${}=40,IF(AND(BO{}>7,BZ{}=20),1,""),IF($G${}=45,IF(AND(BO{}>8,BZ{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['CI{}'.format(row)] = '=IF($H${}=25,IF(AND(BP{}>4,CA{}=20),1,""),IF($H${}=30,IF(AND(BP{}>5,CA{}=20),1,""),IF($H${}=35,IF(AND(BP{}>6,CA{}=20),1,""),IF($H${}=40,IF(AND(BP{}>7,CA{}=20),1,""),IF($H${}=45,IF(AND(BP{}>8,CA{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['CJ{}'.format(row)] = '=IF($I${}=25,IF(AND(BQ{}>4,CB{}=20),1,""),IF($I${}=30,IF(AND(BQ{}>5,CB{}=20),1,""),IF($I${}=35,IF(AND(BQ{}>6,CB{}=20),1,""),IF($I${}=40,IF(AND(BQ{}>7,CB{}=20),1,""),IF($I${}=45,IF(AND(BQ{}>8,CB{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['CK{}'.format(row)] = '=IF($J${}=25,IF(AND(BR{}>4,CC{}=20),1,""),IF($J${}=30,IF(AND(BR{}>5,CC{}=20),1,""),IF($J${}=35,IF(AND(BR{}>6,CC{}=20),1,""),IF($J${}=40,IF(AND(BR{}>7,CC{}=20),1,""),IF($J${}=45,IF(AND(BR{}>8,CC{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['CL{}'.format(row)] = '=IF($K${}=25,IF(AND(BS{}>4,CD{}=20),1,""),IF($K${}=30,IF(AND(BS{}>5,CD{}=20),1,""),IF($K${}=35,IF(AND(BS{}>6,CD{}=20),1,""),IF($K${}=40,IF(AND(BS{}>7,CD{}=20),1,""),IF($K${}=45,IF(AND(BS{}>8,CD{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

                # new Z Score [3]
            ws['CN1'] = 'NAMA_SISWA_4'
            ws['CO1'] = 'NOMOR_NF_4'
            ws['CP1'] = 'KELAS_4'
            ws['CQ1'] = 'NAMA_SEKOLAH_4'
            ws['CR1'] = 'LOKASI_4'
            ws['CS1'] = 'MAT_4'
            ws['CT1'] = 'IND_4'
            ws['CU1'] = 'ENG_4'
            ws['CV1'] = 'IPA_4'
            ws['CW1'] = 'IPS_4'
            ws['CX1'] = 'JML_4'
            ws['CY1'] = 'Z_MAT_4'
            ws['CZ1'] = 'Z_IND_4'
            ws['DA1'] = 'Z_ENG_4'
            ws['DB1'] = 'Z_IPA_4'
            ws['DC1'] = 'Z_IPS_4'
            ws['DD1'] = 'S_MAT_4'
            ws['DE1'] = 'S_IND_4'
            ws['DF1'] = 'S_ENG_4'
            ws['DG1'] = 'S_IPA_4'
            ws['DH1'] = 'S_IPS_4'
            ws['DI1'] = 'S_JML_4'
            ws['DJ1'] = 'RANK_NAS._4'
            ws['DK1'] = 'RANK_LOK._4'
            ws['CY1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CZ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DA1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DB1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DC1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DD1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DE1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DF1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DG1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DH1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DI1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DJ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DK1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['CN1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CO1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CP1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CQ1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CR1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CS1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CT1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CU1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CV1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CW1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CX1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CY1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CZ1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DA1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DB1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DC1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DD1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DE1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DF1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DG1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DH1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DI1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DJ1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DK1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')

            # tambahan
            ws['DL1'] = 'MAT_20_4'
            ws['DM1'] = 'IND_20_4'
            ws['DN1'] = 'ENG_20_4'
            ws['DO1'] = 'IPA_20_4'
            ws['DP1'] = 'IPS_20_4'
            ws['DL1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DM1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DN1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DO1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DP1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DL1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DM1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DN1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DO1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DP1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')

            for row in range(2, q+1):
                ws['CN{}'.format(row)] = '=B{}'.format(row)
                ws['CO{}'.format(row)] = '=C{}'.format(row, row)
                ws['CP{}'.format(row)] = '=D{}'.format(row, row)
                ws['CQ{}'.format(row)] = '=E{}'.format(row, row)
                ws['CR{}'.format(row)] = '=F{}'.format(row, row)
                ws['CS{}'.format(row)] = '=IF(G{}="","",G{})'.format(row, row)
                ws['CT{}'.format(row)] = '=IF(H{}="","",H{})'.format(row, row)
                ws['CU{}'.format(row)] = '=IF(I{}="","",I{})'.format(row, row)
                ws['CV{}'.format(row)] = '=IF(J{}="","",J{})'.format(row, row)
                ws['CW{}'.format(row)] = '=IF(K{}="","",K{})'.format(row, row)
                ws['CX{}'.format(row)] = '=IF(L{}="","",L{})'.format(row, row)
                ws['CY{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CS{}="","",(CS{}-CS${})/CS${}),2),"")'.format(row, row, r, s)
                ws['CZ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CT{}="","",(CT{}-CT${})/CT${}),2),"")'.format(row, row, r, s)
                ws['DA{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CU{}="","",(CU{}-CU${})/CU${}),2),"")'.format(row, row, r, s)
                ws['DB{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CV{}="","",(CV{}-CV${})/CV${}),2),"")'.format(row, row, r, s)
                ws['DC{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CW{}="","",(CW{}-CW${})/CW${}),2),"")'.format(row, row, r, s)
                ws['DD{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CY{}="","",IF(70+30*CY{}/$CY${}<20,20,70+30*CY{}/$CY${})),2),"")'.format(row, row, r, row, r)
                ws['DE{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CZ{}="","",IF(70+30*CZ{}/$CZ${}<20,20,70+30*CZ{}/$CZ${})),2),"")'.format(row, row, r, row, r)
                ws['DF{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DA{}="","",IF(70+30*DA{}/$DA${}<20,20,70+30*DA{}/$DA${})),2),"")'.format(row, row, r, row, r)
                ws['DG{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DB{}="","",IF(70+30*DB{}/$DB${}<20,20,70+30*DB{}/$DB${})),2),"")'.format(row, row, r, row, r)
                ws['DH{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DC{}="","",IF(70+30*DC{}/$DC${}<20,20,70+30*DC{}/$DC${})),2),"")'.format(row, row, r, row, r)

                ws['DI{}'.format(row)] = '=IF(SUM(DD{}:DH{})=0,"",SUM(DD{}:DH{}))'.format(
                    row, row, row, row)
                ws['DJ{}'.format(row)] = '=IF(DI{}="","",RANK(DI{},$DI$2:$DI${}))'.format(
                    row, row, q)
                ws['DK{}'.format(
                    row)] = '=IF(DJ{}="","",COUNTIFS($CR$2:$CR${},F{},$DJ$2:$DJ${},"<"&DJ{})+1)'.format(row, q, row, q, row)
                #     TAMBAHAN
                ws['DL{}'.format(row)] = '=IF($G${}=25,IF(AND(CS{}>4,DD{}=20),1,""),IF($G${}=30,IF(AND(CS{}>5,DD{}=20),1,""),IF($G${}=35,IF(AND(CS{}>6,DD{}=20),1,""),IF($G${}=40,IF(AND(CS{}>7,DD{}=20),1,""),IF($G${}=45,IF(AND(CS{}>8,DD{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['DM{}'.format(row)] = '=IF($H${}=25,IF(AND(CT{}>4,DE{}=20),1,""),IF($H${}=30,IF(AND(CT{}>5,DE{}=20),1,""),IF($H${}=35,IF(AND(CT{}>6,DE{}=20),1,""),IF($H${}=40,IF(AND(CT{}>7,DE{}=20),1,""),IF($H${}=45,IF(AND(CT{}>8,DE{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['DN{}'.format(row)] = '=IF($I${}=25,IF(AND(CU{}>4,DF{}=20),1,""),IF($I${}=30,IF(AND(CU{}>5,DF{}=20),1,""),IF($I${}=35,IF(AND(CU{}>6,DF{}=20),1,""),IF($I${}=40,IF(AND(CU{}>7,DF{}=20),1,""),IF($I${}=45,IF(AND(CU{}>8,DF{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['DO{}'.format(row)] = '=IF($J${}=25,IF(AND(CV{}>4,DG{}=20),1,""),IF($J${}=30,IF(AND(CV{}>5,DG{}=20),1,""),IF($J${}=35,IF(AND(CV{}>6,DG{}=20),1,""),IF($J${}=40,IF(AND(CV{}>7,DG{}=20),1,""),IF($J${}=45,IF(AND(CV{}>8,DG{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['DP{}'.format(row)] = '=IF($K${}=25,IF(AND(CW{}>4,DH{}=20),1,""),IF($K${}=30,IF(AND(CW{}>5,DH{}=20),1,""),IF($K${}=35,IF(AND(CW{}>6,DH{}=20),1,""),IF($K${}=40,IF(AND(CW{}>7,DH{}=20),1,""),IF($K${}=45,IF(AND(CW{}>8,DH{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

            # new Z Score [4]
            ws['DR1'] = 'NAMA_SISWA_5'
            ws['DS1'] = 'NOMOR_NF_5'
            ws['DT1'] = 'KELAS_5'
            ws['DU1'] = 'NAMA_SEKOLAH_5'
            ws['DV1'] = 'LOKASI_5'
            ws['DW1'] = 'MAT_5'
            ws['DX1'] = 'IND_5'
            ws['DY1'] = 'ENG_5'
            ws['DZ1'] = 'IPA_5'
            ws['EA1'] = 'IPS_5'
            ws['EB1'] = 'JML_5'
            ws['EC1'] = 'Z_MAT_5'
            ws['ED1'] = 'Z_IND_5'
            ws['EE1'] = 'Z_ENG_5'
            ws['EF1'] = 'Z_IPA_5'
            ws['EG1'] = 'Z_IPS_5'
            ws['EH1'] = 'S_MAT_5'
            ws['EI1'] = 'S_IND_5'
            ws['EJ1'] = 'S_ENG_5'
            ws['EK1'] = 'S_IPA_5'
            ws['EL1'] = 'S_IPS_5'
            ws['EM1'] = 'S_JML_5'
            ws['EN1'] = 'RANK_NAS._5'
            ws['EO1'] = 'RANK_LOK._5'
            ws['EC1'].font = Font(bold=False, name='Calibri', size=11)
            ws['ED1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EE1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EF1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EG1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EH1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EI1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EJ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EK1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EL1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EM1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EN1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EO1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['DR1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DS1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DT1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DU1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DV1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DW1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DX1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DY1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DZ1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EA1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EB1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EC1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['ED1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EE1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EF1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EG1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EH1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EI1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EJ1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EK1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EL1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EM1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EN1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EO1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')

            # tambahan
            ws['EP1'] = 'MAT_20_5'
            ws['EQ1'] = 'IND_20_5'
            ws['ER1'] = 'ENG_20_5'
            ws['ES1'] = 'IPA_20_5'
            ws['ET1'] = 'IPS_20_5'
            ws['EP1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EQ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['ER1'].font = Font(bold=False, name='Calibri', size=11)
            ws['ES1'].font = Font(bold=False, name='Calibri', size=11)
            ws['ET1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EP1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EQ1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['ER1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['ES1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['ET1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')

            for row in range(2, q+1):
                ws['DR{}'.format(row)] = '=B{}'.format(row)
                ws['DS{}'.format(row)] = '=C{}'.format(row, row)
                ws['DT{}'.format(row)] = '=D{}'.format(row, row)
                ws['DU{}'.format(row)] = '=E{}'.format(row, row)
                ws['DV{}'.format(row)] = '=F{}'.format(row, row)
                ws['DW{}'.format(row)] = '=IF(G{}="","",G{})'.format(row, row)
                ws['DX{}'.format(row)] = '=IF(H{}="","",H{})'.format(row, row)
                ws['DY{}'.format(row)] = '=IF(I{}="","",I{})'.format(row, row)
                ws['DZ{}'.format(row)] = '=IF(J{}="","",J{})'.format(row, row)
                ws['EA{}'.format(row)] = '=IF(K{}="","",K{})'.format(row, row)
                ws['EB{}'.format(row)] = '=IF(L{}="","",L{})'.format(row, row)
                ws['EC{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DW{}="","",(DW{}-DW${})/DW${}),2),"")'.format(row, row, r, s)
                ws['ED{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DX{}="","",(DX{}-DX${})/DX${}),2),"")'.format(row, row, r, s)
                ws['EE{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DY{}="","",(DY{}-DY${})/DY${}),2),"")'.format(row, row, r, s)
                ws['EF{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DZ{}="","",(DZ{}-DZ${})/DZ${}),2),"")'.format(row, row, r, s)
                ws['EG{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EA{}="","",(EA{}-EA${})/EA${}),2),"")'.format(row, row, r, s)
                ws['EH{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EC{}="","",IF(70+30*EC{}/$EC${}<20,20,70+30*EC{}/$EC${})),2),"")'.format(row, row, r, row, r)
                ws['EI{}'.format(
                    row)] = '=IFERROR(ROUND(IF(ED{}="","",IF(70+30*ED{}/$ED${}<20,20,70+30*ED{}/$ED${})),2),"")'.format(row, row, r, row, r)
                ws['EJ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EE{}="","",IF(70+30*EE{}/$EE${}<20,20,70+30*EE{}/$EE${})),2),"")'.format(row, row, r, row, r)
                ws['EK{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EF{}="","",IF(70+30*EF{}/$EF${}<20,20,70+30*EF{}/$EF${})),2),"")'.format(row, row, r, row, r)
                ws['EL{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EG{}="","",IF(70+30*EG{}/$EG${}<20,20,70+30*EG{}/$EG${})),2),"")'.format(row, row, r, row, r)

                ws['EM{}'.format(row)] = '=IF(SUM(EH{}:EL{})=0,"",SUM(EH{}:EL{}))'.format(
                    row, row, row, row)
                ws['EN{}'.format(row)] = '=IF(EM{}="","",RANK(EM{},$EM$2:$EM${}))'.format(
                    row, row, q)
                ws['EO{}'.format(
                    row)] = '=IF(EN{}="","",COUNTIFS($DV$2:$DV${},F{},$EN$2:$EN${},"<"&EN{})+1)'.format(row, q, row, q, row)
                #     TAMBAHAN
                ws['EP{}'.format(row)] = '=IF($G${}=25,IF(AND(DW{}>4,EH{}=20),1,""),IF($G${}=30,IF(AND(DW{}>5,EH{}=20),1,""),IF($G${}=35,IF(AND(DW{}>6,EH{}=20),1,""),IF($G${}=40,IF(AND(DW{}>7,EH{}=20),1,""),IF($G${}=45,IF(AND(DW{}>8,EH{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['EQ{}'.format(row)] = '=IF($H${}=25,IF(AND(DX{}>4,EI{}=20),1,""),IF($H${}=30,IF(AND(DX{}>5,EI{}=20),1,""),IF($H${}=35,IF(AND(DX{}>6,EI{}=20),1,""),IF($H${}=40,IF(AND(DX{}>7,EI{}=20),1,""),IF($H${}=45,IF(AND(DX{}>8,EI{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['ER{}'.format(row)] = '=IF($I${}=25,IF(AND(DY{}>4,EJ{}=20),1,""),IF($I${}=30,IF(AND(DY{}>5,EJ{}=20),1,""),IF($I${}=35,IF(AND(DY{}>6,EJ{}=20),1,""),IF($I${}=40,IF(AND(DY{}>7,EJ{}=20),1,""),IF($I${}=45,IF(AND(DY{}>8,EJ{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['ES{}'.format(row)] = '=IF($J${}=25,IF(AND(DZ{}>4,EK{}=20),1,""),IF($J${}=30,IF(AND(DZ{}>5,EK{}=20),1,""),IF($J${}=35,IF(AND(DZ{}>6,EK{}=20),1,""),IF($J${}=40,IF(AND(DZ{}>7,EK{}=20),1,""),IF($J${}=45,IF(AND(DZ{}>8,EK{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['ET{}'.format(row)] = '=IF($K${}=25,IF(AND(EA{}>4,EL{}=20),1,""),IF($K${}=30,IF(AND(EA{}>5,EL{}=20),1,""),IF($K${}=35,IF(AND(EA{}>6,EL{}=20),1,""),IF($K${}=40,IF(AND(EA{}>7,EL{}=20),1,""),IF($K${}=45,IF(AND(EA{}>8,EL{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

                # new Z Score [5]
            ws['EV1'] = 'NAMA SISWA'
            ws['EW1'] = 'NOMOR NF'
            ws['EX1'] = 'KELAS'
            ws['EY1'] = 'NAMA SEKOLAH'
            ws['EZ1'] = 'LOKASI'
            ws['FA1'] = 'MAT'
            ws['FB1'] = 'IND'
            ws['FC1'] = 'ENG'
            ws['FD1'] = 'IPA'
            ws['FE1'] = 'IPS'
            ws['FF1'] = 'JML'
            ws['FG1'] = 'Z_MAT'
            ws['FH1'] = 'Z_IND'
            ws['FI1'] = 'Z_ENG'
            ws['FJ1'] = 'Z_IPA'
            ws['FK1'] = 'Z_IPS'
            ws['FL1'] = 'S_MAT'
            ws['FM1'] = 'S_IND'
            ws['FN1'] = 'S_ENG'
            ws['FO1'] = 'S_IPA'
            ws['FP1'] = 'S_IPS'
            ws['FQ1'] = 'S_JML'
            ws['FR1'] = 'RANK NAS.'
            ws['FS1'] = 'RANK LOK.'
            ws['FG1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FH1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FI1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FJ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FK1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FL1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FM1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FN1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FO1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FP1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FQ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FR1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FS1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['EV1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EW1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EX1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EY1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EZ1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FA1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FB1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FC1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FD1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FE1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FF1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FG1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FH1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FI1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FJ1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FK1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FL1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FM1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FN1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FO1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FP1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FQ1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FR1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FS1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')

            # tambahan
            ws['FT1'] = 'MAT_20'
            ws['FU1'] = 'IND_20'
            ws['FV1'] = 'ENG_20'
            ws['FW1'] = 'IPA_20'
            ws['FX1'] = 'IPS_20'
            ws['FT1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FU1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FV1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FW1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FX1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FT1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FU1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FV1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FW1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FX1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')

            for row in range(2, q+1):
                ws['EV{}'.format(row)] = '=B{}'.format(row)
                ws['EW{}'.format(row)] = '=C{}'.format(row, row)
                ws['EX{}'.format(row)] = '=D{}'.format(row, row)
                ws['EY{}'.format(row)] = '=E{}'.format(row, row)
                ws['EZ{}'.format(row)] = '=F{}'.format(row, row)
                ws['FA{}'.format(row)] = '=IF(G{}="","",G{})'.format(row, row)
                ws['FB{}'.format(row)] = '=IF(H{}="","",H{})'.format(row, row)
                ws['FC{}'.format(row)] = '=IF(I{}="","",I{})'.format(row, row)
                ws['FD{}'.format(row)] = '=IF(J{}="","",J{})'.format(row, row)
                ws['FE{}'.format(row)] = '=IF(K{}="","",K{})'.format(row, row)
                ws['FF{}'.format(row)] = '=IF(L{}="","",L{})'.format(row, row)
                ws['FG{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FA{}="","",(FA{}-FA${})/FA${}),2),"")'.format(row, row, r, s)
                ws['FH{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FB{}="","",(FB{}-FB${})/FB${}),2),"")'.format(row, row, r, s)
                ws['FI{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FC{}="","",(FC{}-FC${})/FC${}),2),"")'.format(row, row, r, s)
                ws['FJ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FD{}="","",(FD{}-FD${})/FD${}),2),"")'.format(row, row, r, s)
                ws['FK{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FE{}="","",(FE{}-FE${})/FE${}),2),"")'.format(row, row, r, s)
                ws['FL{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FG{}="","",IF(70+30*FG{}/$FG${}<20,20,70+30*FG{}/$FG${})),2),"")'.format(row, row, r, row, r)
                ws['FM{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FH{}="","",IF(70+30*FH{}/$FH${}<20,20,70+30*FH{}/$FH${})),2),"")'.format(row, row, r, row, r)
                ws['FN{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FI{}="","",IF(70+30*FI{}/$FI${}<20,20,70+30*FI{}/$FI${})),2),"")'.format(row, row, r, row, r)
                ws['FO{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FJ{}="","",IF(70+30*FJ{}/$FJ${}<20,20,70+30*FJ{}/$FJ${})),2),"")'.format(row, row, r, row, r)
                ws['FP{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FK{}="","",IF(70+30*FK{}/$FK${}<20,20,70+30*FK{}/$FK${})),2),"")'.format(row, row, r, row, r)

                ws['FQ{}'.format(row)] = '=IF(SUM(FL{}:FP{})=0,"",SUM(FL{}:FP{}))'.format(
                    row, row, row, row)
                ws['FR{}'.format(row)] = '=IF(FQ{}="","",RANK(FQ{},$FQ$2:$FQ${}))'.format(
                    row, row, q)
                ws['FS{}'.format(
                    row)] = '=IF(FR{}="","",COUNTIFS($EZ$2:$EZ${},F{},$FR$2:$FR${},"<"&FR{})+1)'.format(row, q, row, q, row)
                #     TAMBAHAN
                ws['FT{}'.format(row)] = '=IF($G${}=25,IF(AND(FA{}>4,FL{}=20),1,""),IF($G${}=30,IF(AND(FA{}>5,FL{}=20),1,""),IF($G${}=35,IF(AND(FA{}>6,FL{}=20),1,""),IF($G${}=40,IF(AND(FA{}>7,FL{}=20),1,""),IF($G${}=45,IF(AND(FA{}>8,FL{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['FU{}'.format(row)] = '=IF($H${}=25,IF(AND(FB{}>4,FM{}=20),1,""),IF($H${}=30,IF(AND(FB{}>5,FM{}=20),1,""),IF($H${}=35,IF(AND(FB{}>6,FM{}=20),1,""),IF($H${}=40,IF(AND(FB{}>7,FM{}=20),1,""),IF($H${}=45,IF(AND(FB{}>8,FM{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['FV{}'.format(row)] = '=IF($I${}=25,IF(AND(FC{}>4,FN{}=20),1,""),IF($I${}=30,IF(AND(FC{}>5,FN{}=20),1,""),IF($I${}=35,IF(AND(FC{}>6,FN{}=20),1,""),IF($I${}=40,IF(AND(FC{}>7,FN{}=20),1,""),IF($I${}=45,IF(AND(FC{}>8,FN{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['FW{}'.format(row)] = '=IF($J${}=25,IF(AND(FD{}>4,FO{}=20),1,""),IF($J${}=30,IF(AND(FD{}>5,FO{}=20),1,""),IF($J${}=35,IF(AND(FD{}>6,FO{}=20),1,""),IF($J${}=40,IF(AND(FD{}>7,FO{}=20),1,""),IF($J${}=45,IF(AND(FD{}>8,FO{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['FX{}'.format(row)] = '=IF($K${}=25,IF(AND(FE{}>4,FP{}=20),1,""),IF($K${}=30,IF(AND(FE{}>5,FP{}=20),1,""),IF($K${}=35,IF(AND(FE{}>6,FP{}=20),1,""),IF($K${}=40,IF(AND(FE{}>7,FP{}=20),1,""),IF($K${}=45,IF(AND(FE{}>8,FP{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

            # Mengubah 'KELAS' sesuai dengan nilai yang dipilih dari selectbox 'KELAS'
            kelas = KELAS.lower().replace(" ", "")
            semester = SEMESTER.lower()
            tahun = TAHUN.replace("-", "")
            penilaian = PENILAIAN.lower()
            kurikulum = KURIKULUM.lower()

            path_file = f"{kelas}_{penilaian}_{semester}_{kurikulum}_{tahun}_nilai_std.xlsx"

            # Simpan file ke direktori temporer
            temp_dir = tempfile.gettempdir()
            file_path = temp_dir + '/' + path_file
            wb.save(file_path)

            st.success(
                "File siap diunduh!")

            # Tombol unduh file
            with open(file_path, "rb") as f:
                bytes_data = f.read()
            st.download_button(label="Unduh File", data=bytes_data,
                               file_name=path_file)

            st.warning(
                "Buka file unduhan, klik 'Enable Editing' dan 'Save'")
    if selected_file == "Nilai Std. 8 SMP (KM-MTK SB)":
        # menghilangkan hamburger
        st.markdown("""
        <style>
        .css-1rs6os.edgvbvh3
        {
            visibility:hidden;
        }
        .css-1lsmgbg.egzxvld0
        {
            visibility:hidden;
        }
        </style>
        """, unsafe_allow_html=True)

        image = Image.open('logo resmi nf resize.png')
        st.image(image)

        st.title("Olah Nilai Standar KM")

        st.header("8 SMP-MTK SB")

        col6 = st.container()

        with col6:
            KELAS = st.selectbox(
                "KELAS",
                ("--Pilih Kelas--", "8 SMP SB"))

        col7 = st.container()

        with col7:
            SEMESTER = st.selectbox(
                "SEMESTER",
                ("--Pilih Semester--", "SEMESTER 1", "SEMESTER 2"))

        col8 = st.container()

        with col8:
            PENILAIAN = st.selectbox(
                "PENILAIAN",
                ("--Pilih Penilaian--", "SUMATIF TENGAH SEMESTER", "SUMATIF AKHIR TAHUN"))

        col9 = st.container()

        with col9:
            KURIKULUM = st.selectbox(
                "KURIKULUM",
                ("--Pilih Kurikulum--", "KM"))

        TAHUN = st.text_input("Masukkan Tahun Ajaran",
                              placeholder="contoh: 2022-2023")

        col1, col2, col3, col4, col5 = st.columns(5)

        with col1:
            MTK = st.selectbox(
                "JML. SOAL MAT. SB.",
                ("--Pilih--", 25, 30, 35, 40, 45))

        with col2:
            IND = st.selectbox(
                "JML. SOAL IND.",
                ("--Pilih--", 25, 30, 35, 40, 45))

        with col3:
            ENG = st.selectbox(
                "JML. SOAL ENG.",
                ("--Pilih--", 25, 30, 35, 40, 45))

        with col4:
            IPA = st.selectbox(
                "JML. SOAL IPA.",
                ("--Pilih--", 25, 30, 35, 40, 45))

        with col5:
            IPS = st.selectbox(
                "JML. SOAL IPS.",
                ("--Pilih--", 25, 30, 35, 40, 45))

        JML_SOAL_MAT = MTK
        JML_SOAL_IND = IND
        JML_SOAL_ENG = ENG
        JML_SOAL_IPA = IPA
        JML_SOAL_IPS = IPS

        uploaded_file = st.file_uploader(
            'Letakkan file excel', type='xlsx')

        if uploaded_file is not None:

            wb = openpyxl.load_workbook(uploaded_file)
            ws = wb['Sheet1']

            q = len(ws['K'])
            r = len(ws['K'])+2
            s = len(ws['K'])+3
            t = len(ws['K'])+4
            u = len(ws['K'])+5
            v = len(ws['K'])+6
            w = len(ws['K'])+7
            x = len(ws['K'])+8

            ws['G{}'.format(r)] = "=ROUND(AVERAGE(G2:G{}),2)".format(q)  # mat
            ws['H{}'.format(r)] = "=ROUND(AVERAGE(H2:H{}),2)".format(q)  # ind
            ws['I{}'.format(r)] = "=ROUND(AVERAGE(I2:I{}),2)".format(q)  # eng
            ws['J{}'.format(r)] = "=ROUND(AVERAGE(J2:J{}),2)".format(q)  # ipa
            ws['K{}'.format(r)] = "=ROUND(AVERAGE(K2:K{}),2)".format(q)  # ips
            ws['L{}'.format(r)] = "=ROUND(AVERAGE(L2:L{}),2)".format(q)  # jml
            ws['G{}'.format(s)] = "=STDEV(G2:G{})".format(q)
            ws['H{}'.format(s)] = "=STDEV(H2:H{})".format(q)
            ws['I{}'.format(s)] = "=STDEV(I2:I{})".format(q)
            ws['J{}'.format(s)] = "=STDEV(J2:J{})".format(q)
            ws['K{}'.format(s)] = "=STDEV(K2:K{})".format(q)
            ws['G{}'.format(t)] = "=MAX(G2:G{})".format(q)
            ws['H{}'.format(t)] = "=MAX(H2:H{})".format(q)
            ws['I{}'.format(t)] = "=MAX(I2:I{})".format(q)
            ws['J{}'.format(t)] = "=MAX(J2:J{})".format(q)
            ws['K{}'.format(t)] = "=MAX(K2:K{})".format(q)
            ws['L{}'.format(t)] = "=MAX(L2:L{})".format(q)
            ws['M{}'.format(r)] = "=MAX(M2:M{})".format(q)
            ws['N{}'.format(r)] = "=MAX(N2:N{})".format(q)
            ws['O{}'.format(r)] = "=MAX(O2:O{})".format(q)
            ws['P{}'.format(r)] = "=MAX(P2:P{})".format(q)
            ws['Q{}'.format(r)] = "=MAX(Q2:Q{})".format(q)
            ws['R{}'.format(r)] = "=MAX(R2:R{})".format(q)
            ws['S{}'.format(r)] = "=MAX(S2:S{})".format(q)
            ws['T{}'.format(r)] = "=MAX(T2:T{})".format(q)
            ws['U{}'.format(r)] = "=MAX(U2:U{})".format(q)
            ws['V{}'.format(r)] = "=MAX(V2:V{})".format(q)
            ws['W{}'.format(r)] = "=ROUND(MAX(W2:W{}),2)".format(q)
            ws['G{}'.format(u)] = "=MIN(G2:G{})".format(q)
            ws['H{}'.format(u)] = "=MIN(H2:H{})".format(q)
            ws['I{}'.format(u)] = "=MIN(I2:I{})".format(q)
            ws['J{}'.format(u)] = "=MIN(J2:J{})".format(q)
            ws['K{}'.format(u)] = "=MIN(K2:K{})".format(q)
            ws['L{}'.format(u)] = "=MIN(L2:L{})".format(q)
            ws['R{}'.format(s)] = "=MIN(R2:R{})".format(q)
            ws['S{}'.format(s)] = "=MIN(S2:S{})".format(q)
            ws['T{}'.format(s)] = "=MIN(T2:T{})".format(q)
            ws['U{}'.format(s)] = "=MIN(U2:U{})".format(q)
            ws['V{}'.format(s)] = "=MIN(V2:V{})".format(q)
            ws['W{}'.format(s)] = "=MIN(W2:W{})".format(q)
            ws['R{}'.format(t)] = "=ROUND(AVERAGE(R2:R{}),2)".format(q)
            ws['S{}'.format(t)] = "=ROUND(AVERAGE(S2:S{}),2)".format(q)
            ws['T{}'.format(t)] = "=ROUND(AVERAGE(T2:T{}),2)".format(q)
            ws['U{}'.format(t)] = "=ROUND(AVERAGE(U2:U{}),2)".format(q)
            ws['V{}'.format(t)] = "=ROUND(AVERAGE(V2:V{}),2)".format(q)
            ws['W{}'.format(t)] = "=ROUND(AVERAGE(W2:W{}),2)".format(q)
            ws['X{}'.format(r)] = "=MAX(X2:X{})".format(q)
            ws['Z{}'.format(r)] = "=SUM(Z2:Z{})".format(q)
            ws['AA{}'.format(r)] = "=SUM(AA2:AA{})".format(q)
            ws['AB{}'.format(r)] = "=SUM(AB2:AB{})".format(q)
            ws['AC{}'.format(r)] = "=SUM(AC2:AC{})".format(q)
            ws['AD{}'.format(r)] = "=SUM(AD2:AD{})".format(q)
            # new
            # iterasi 1 rata-rata - 1
            ws['F{}'.format(v)] = 'JUMLAH SOAL'
            ws['G{}'.format(v)] = JML_SOAL_MAT
            ws['H{}'.format(v)] = JML_SOAL_IND
            ws['I{}'.format(v)] = JML_SOAL_ENG
            ws['J{}'.format(v)] = JML_SOAL_IPA
            ws['K{}'.format(v)] = JML_SOAL_IPS
            ws['AK{}'.format(r)] = "=IF($Z${}=0,$G${},$G${}-1)".format(r, r, r)
            ws['AK{}'.format(s)] = "=STDEV(AK2:AK{})".format(q)
            ws['AK{}'.format(t)] = "=MAX(AK2:AK{})".format(q)
            ws['AK{}'.format(u)] = "=MIN(AK2:AK{})".format(q)
            ws['AL{}'.format(
                r)] = "=IF($AA${}=0,$H${},$H${}-1)".format(r, r, r)
            ws['AL{}'.format(s)] = "=STDEV(AL2:AL{})".format(q)
            ws['AL{}'.format(t)] = "=MAX(AL2:AL{})".format(q)
            ws['AL{}'.format(u)] = "=MIN(AL2:AL{})".format(q)
            ws['AM{}'.format(
                r)] = "=IF($AB${}=0,$I${},$I${}-1)".format(r, r, r)
            ws['AM{}'.format(s)] = "=STDEV(AM2:AM{})".format(q)
            ws['AM{}'.format(t)] = "=MAX(AM2:AM{})".format(q)
            ws['AM{}'.format(u)] = "=MIN(AM2:AM{})".format(q)
            ws['AN{}'.format(
                r)] = "=IF($AC${}=0,$J${},$J${}-1)".format(r, r, r)
            ws['AN{}'.format(s)] = "=STDEV(AN2:AN{})".format(q)
            ws['AN{}'.format(t)] = "=MAX(AN2:AN{})".format(q)
            ws['AN{}'.format(u)] = "=MIN(AN2:AN{})".format(q)
            ws['AO{}'.format(
                r)] = "=IF($AD${}=0,$K${},$K${}-1)".format(r, r, r)
            ws['AO{}'.format(s)] = "=STDEV(AO2:AO{})".format(q)
            ws['AO{}'.format(t)] = "=MAX(AO2:AO{})".format(q)
            ws['AO{}'.format(u)] = "=MIN(AO2:AO{})".format(q)
            ws['AP{}'.format(r)] = "=ROUND(AVERAGE(AP2:AP{}),2)".format(q)
            ws['AP{}'.format(t)] = "=MAX(AP2:AP{})".format(q)
            ws['AP{}'.format(u)] = "=MIN(AP2:AP{})".format(q)
            ws['AQ{}'.format(r)] = "=MAX(AQ2:AQ{})".format(q)
            ws['AR{}'.format(r)] = "=MAX(AR2:AR{})".format(q)
            ws['AS{}'.format(r)] = "=MAX(AS2:AS{})".format(q)
            ws['AT{}'.format(r)] = "=MAX(AT2:AT{})".format(q)
            ws['AU{}'.format(r)] = "=MAX(AU2:AU{})".format(q)
            ws['AV{}'.format(r)] = "=MAX(AV2:AV{})".format(q)
            ws['AV{}'.format(s)] = "=MIN(AV2:AV{})".format(q)
            ws['AV{}'.format(t)] = "=ROUND(AVERAGE(AV2:AV{}),2)".format(q)
            ws['AW{}'.format(r)] = "=MAX(AW2:AW{})".format(q)
            ws['AW{}'.format(s)] = "=MIN(AW2:AW{})".format(q)
            ws['AW{}'.format(t)] = "=ROUND(AVERAGE(AW2:AW{}),2)".format(q)
            ws['AX{}'.format(r)] = "=MAX(AX2:AX{})".format(q)
            ws['AX{}'.format(s)] = "=MIN(AX2:AX{})".format(q)
            ws['AX{}'.format(t)] = "=ROUND(AVERAGE(AX2:AX{}),2)".format(q)
            ws['AY{}'.format(r)] = "=MAX(AY2:AY{})".format(q)
            ws['AY{}'.format(s)] = "=MIN(AY2:AY{})".format(q)
            ws['AY{}'.format(t)] = "=ROUND(AVERAGE(AY2:AY{}),2)".format(q)
            ws['AZ{}'.format(r)] = "=MAX(AZ2:AZ{})".format(q)
            ws['AZ{}'.format(s)] = "=MIN(AZ2:AZ{})".format(q)
            ws['AZ{}'.format(t)] = "=ROUND(AVERAGE(AZ2:AZ{}),2)".format(q)
            ws['BA{}'.format(r)] = "=MAX(BA2:BA{})".format(q)
            ws['BA{}'.format(s)] = "=MIN(BA2:BA{})".format(q)
            ws['BA{}'.format(t)] = "=ROUND(AVERAGE(BA2:BA{}),2)".format(q)
            ws['BD{}'.format(r)] = "=SUM(BD2:BD{})".format(q)
            ws['BE{}'.format(r)] = "=SUM(BE2:BE{})".format(q)
            ws['BF{}'.format(r)] = "=SUM(BF2:BF{})".format(q)
            ws['BG{}'.format(r)] = "=SUM(BG2:BG{})".format(q)
            ws['BH{}'.format(r)] = "=SUM(BH2:BH{})".format(q)

            # iterasi 2 rata-rata - 1
            ws['BO{}'.format(
                r)] = "=IF($BD${}=0,$AK${},$AK${}-1)".format(r, r, r)
            ws['BO{}'.format(s)] = "=STDEV(BO2:BO{})".format(q)
            ws['BO{}'.format(t)] = "=MAX(BO2:BO{})".format(q)
            ws['BO{}'.format(u)] = "=MIN(BO2:BO{})".format(q)
            ws['BP{}'.format(
                r)] = "=IF($BE${}=0,$AL${},$AL${}-1)".format(r, r, r)
            ws['BP{}'.format(s)] = "=STDEV(BP2:BP{})".format(q)
            ws['BP{}'.format(t)] = "=MAX(BP2:BP{})".format(q)
            ws['BP{}'.format(u)] = "=MIN(BP2:BP{})".format(q)
            ws['BQ{}'.format(
                r)] = "=IF($BF${}=0,$AM${},$AM${}-1)".format(r, r, r)
            ws['BQ{}'.format(s)] = "=STDEV(BQ2:BQ{})".format(q)
            ws['BQ{}'.format(t)] = "=MAX(BQ2:BQ{})".format(q)
            ws['BQ{}'.format(u)] = "=MIN(BQ2:BQ{})".format(q)
            ws['BR{}'.format(
                r)] = "=IF($BG${}=0,$AN${},$AN${}-1)".format(r, r, r)
            ws['BR{}'.format(s)] = "=STDEV(BR2:BR{})".format(q)
            ws['BR{}'.format(t)] = "=MAX(BR2:BR{})".format(q)
            ws['BR{}'.format(u)] = "=MIN(BR2:BR{})".format(q)
            ws['BS{}'.format(
                r)] = "=IF($BH${}=0,$AO${},$AO${}-1)".format(r, r, r)
            ws['BS{}'.format(s)] = "=STDEV(BS2:BS{})".format(q)
            ws['BS{}'.format(t)] = "=MAX(BS2:BS{})".format(q)
            ws['BS{}'.format(u)] = "=MIN(BS2:BS{})".format(q)
            ws['BT{}'.format(r)] = "=ROUND(AVERAGE(BT2:BT{}),2)".format(q)
            ws['BT{}'.format(t)] = "=MAX(BT2:BT{})".format(q)
            ws['BT{}'.format(u)] = "=MIN(BT2:BT{})".format(q)
            ws['BU{}'.format(r)] = "=MAX(BU2:BU{})".format(q)
            ws['BV{}'.format(r)] = "=MAX(BV2:BV{})".format(q)
            ws['BW{}'.format(r)] = "=MAX(BW2:BW{})".format(q)
            ws['BX{}'.format(r)] = "=MAX(BX2:BX{})".format(q)
            ws['BY{}'.format(r)] = "=MAX(BY2:BY{})".format(q)
            ws['BZ{}'.format(r)] = "=MAX(BZ2:BZ{})".format(q)
            ws['BZ{}'.format(s)] = "=MIN(BZ2:BZ{})".format(q)
            ws['BZ{}'.format(t)] = "=ROUND(AVERAGE(BZ2:BZ{}),2)".format(q)
            ws['CA{}'.format(r)] = "=MAX(CA2:CA{})".format(q)
            ws['CA{}'.format(s)] = "=MIN(CA2:CA{})".format(q)
            ws['CA{}'.format(t)] = "=ROUND(AVERAGE(CA2:CA{}),2)".format(q)
            ws['CB{}'.format(r)] = "=MAX(CB2:CB{})".format(q)
            ws['CB{}'.format(s)] = "=MIN(CB2:CB{})".format(q)
            ws['CB{}'.format(t)] = "=ROUND(AVERAGE(CB2:CB{}),2)".format(q)
            ws['CC{}'.format(r)] = "=MAX(CC2:CC{})".format(q)
            ws['CC{}'.format(s)] = "=MIN(CC2:CC{})".format(q)
            ws['CC{}'.format(t)] = "=ROUND(AVERAGE(CC2:CC{}),2)".format(q)
            ws['CD{}'.format(r)] = "=MAX(CD2:CD{})".format(q)
            ws['CD{}'.format(s)] = "=MIN(CD2:CD{})".format(q)
            ws['CD{}'.format(t)] = "=ROUND(AVERAGE(CD2:CD{}),2)".format(q)
            ws['CE{}'.format(r)] = "=MAX(CE2:CE{})".format(q)
            ws['CE{}'.format(s)] = "=MIN(CE2:CE{})".format(q)
            ws['CE{}'.format(t)] = "=ROUND(AVERAGE(CE2:CE{}),2)".format(q)
            ws['CH{}'.format(r)] = "=SUM(CH2:CH{})".format(q)
            ws['CI{}'.format(r)] = "=SUM(CI2:CI{})".format(q)
            ws['CJ{}'.format(r)] = "=SUM(CJ2:CJ{})".format(q)
            ws['CK{}'.format(r)] = "=SUM(CK2:CK{})".format(q)
            ws['CL{}'.format(r)] = "=SUM(CL2:CL{})".format(q)

            # iterasi 3 rata-rata - 1
            ws['CS{}'.format(
                r)] = "=IF($CH${}=0,$BO${},$BO${}-1)".format(r, r, r)
            ws['CS{}'.format(s)] = "=STDEV(CS2:CS{})".format(q)
            ws['CS{}'.format(t)] = "=MAX(CS2:CS{})".format(q)
            ws['CS{}'.format(u)] = "=MIN(CS2:CS{})".format(q)
            ws['CT{}'.format(
                r)] = "=IF($CI${}=0,$BP${},$BP${}-1)".format(r, r, r)
            ws['CT{}'.format(s)] = "=STDEV(CT2:CT{})".format(q)
            ws['CT{}'.format(t)] = "=MAX(CT2:CT{})".format(q)
            ws['CT{}'.format(u)] = "=MIN(CT2:CT{})".format(q)
            ws['CU{}'.format(
                r)] = "=IF($CJ${}=0,$BQ${},$BQ${}-1)".format(r, r, r)
            ws['CU{}'.format(s)] = "=STDEV(CU2:CU{})".format(q)
            ws['CU{}'.format(t)] = "=MAX(CU2:CU{})".format(q)
            ws['CU{}'.format(u)] = "=MIN(CU2:CU{})".format(q)
            ws['CV{}'.format(
                r)] = "=IF($CK${}=0,$BR${},$BR${}-1)".format(r, r, r)
            ws['CV{}'.format(s)] = "=STDEV(CV2:CV{})".format(q)
            ws['CV{}'.format(t)] = "=MAX(CV2:CV{})".format(q)
            ws['CV{}'.format(u)] = "=MIN(CV2:CV{})".format(q)
            ws['CW{}'.format(
                r)] = "=IF($CL${}=0,$BS${},$BS${}-1)".format(r, r, r)
            ws['CW{}'.format(s)] = "=STDEV(CW2:CW{})".format(q)
            ws['CW{}'.format(t)] = "=MAX(CW2:CW{})".format(q)
            ws['CW{}'.format(u)] = "=MIN(CW2:CW{})".format(q)
            ws['CX{}'.format(r)] = "=ROUND(AVERAGE(CX2:CX{}),2)".format(q)
            ws['CX{}'.format(t)] = "=MAX(CX2:CX{})".format(q)
            ws['CX{}'.format(u)] = "=MIN(CX2:CX{})".format(q)
            ws['CY{}'.format(r)] = "=MAX(CY2:CY{})".format(q)
            ws['CZ{}'.format(r)] = "=MAX(CZ2:CZ{})".format(q)
            ws['DA{}'.format(r)] = "=MAX(DA2:DA{})".format(q)
            ws['DB{}'.format(r)] = "=MAX(DB2:DB{})".format(q)
            ws['DC{}'.format(r)] = "=MAX(DC2:DC{})".format(q)
            ws['DD{}'.format(r)] = "=MAX(DD2:DD{})".format(q)
            ws['DD{}'.format(s)] = "=MIN(DD2:DD{})".format(q)
            ws['DD{}'.format(t)] = "=ROUND(AVERAGE(DD2:DD{}),2)".format(q)
            ws['DE{}'.format(r)] = "=MAX(DE2:DE{})".format(q)
            ws['DE{}'.format(s)] = "=MIN(DE2:DE{})".format(q)
            ws['DE{}'.format(t)] = "=ROUND(AVERAGE(DE2:DE{}),2)".format(q)
            ws['DF{}'.format(r)] = "=MAX(DF2:DF{})".format(q)
            ws['DF{}'.format(s)] = "=MIN(DF2:DF{})".format(q)
            ws['DF{}'.format(t)] = "=ROUND(AVERAGE(DF2:DF{}),2)".format(q)
            ws['DG{}'.format(r)] = "=MAX(DG2:DG{})".format(q)
            ws['DG{}'.format(s)] = "=MIN(DG2:DG{})".format(q)
            ws['DG{}'.format(t)] = "=ROUND(AVERAGE(DG2:DG{}),2)".format(q)
            ws['DH{}'.format(r)] = "=MAX(DH2:DH{})".format(q)
            ws['DH{}'.format(s)] = "=MIN(DH2:DH{})".format(q)
            ws['DH{}'.format(t)] = "=ROUND(AVERAGE(DH2:DH{}),2)".format(q)
            ws['DI{}'.format(r)] = "=MAX(DI2:DI{})".format(q)
            ws['DI{}'.format(s)] = "=MIN(DI2:DI{})".format(q)
            ws['DI{}'.format(t)] = "=ROUND(AVERAGE(DI2:DI{}),2)".format(q)
            ws['DL{}'.format(r)] = "=SUM(DL2:DL{})".format(q)
            ws['DM{}'.format(r)] = "=SUM(DM2:DM{})".format(q)
            ws['DN{}'.format(r)] = "=SUM(DN2:DN{})".format(q)
            ws['DO{}'.format(r)] = "=SUM(DO2:DO{})".format(q)
            ws['DP{}'.format(r)] = "=SUM(DP2:DP{})".format(q)

            # iterasi 4 rata-rata - 1
            ws['DW{}'.format(
                r)] = "=IF($DL${}=0,$CS${},$CS${}-1)".format(r, r, r)
            ws['DW{}'.format(s)] = "=STDEV(DW2:DW{})".format(q)
            ws['DW{}'.format(t)] = "=MAX(DW2:DW{})".format(q)
            ws['DW{}'.format(u)] = "=MIN(DW2:DW{})".format(q)
            ws['DX{}'.format(
                r)] = "=IF($DM${}=0,$CT${},$CT${}-1)".format(r, r, r)
            ws['DX{}'.format(s)] = "=STDEV(DX2:DX{})".format(q)
            ws['DX{}'.format(t)] = "=MAX(DX2:DX{})".format(q)
            ws['DX{}'.format(u)] = "=MIN(DX2:DX{})".format(q)
            ws['DY{}'.format(
                r)] = "=IF($DN${}=0,$CU${},$CU${}-1)".format(r, r, r)
            ws['DY{}'.format(s)] = "=STDEV(DY2:DY{})".format(q)
            ws['DY{}'.format(t)] = "=MAX(DY2:DY{})".format(q)
            ws['DY{}'.format(u)] = "=MIN(DY2:DY{})".format(q)
            ws['DZ{}'.format(
                r)] = "=IF($DO${}=0,$CV${},$CV${}-1)".format(r, r, r)
            ws['DZ{}'.format(s)] = "=STDEV(DZ2:DZ{})".format(q)
            ws['DZ{}'.format(t)] = "=MAX(DZ2:DZ{})".format(q)
            ws['DZ{}'.format(u)] = "=MIN(DZ2:DZ{})".format(q)
            ws['EA{}'.format(
                r)] = "=IF($DP${}=0,$CW${},$CW${}-1)".format(r, r, r)
            ws['EA{}'.format(s)] = "=STDEV(EA2:EA{})".format(q)
            ws['EA{}'.format(t)] = "=MAX(EA2:EA{})".format(q)
            ws['EA{}'.format(u)] = "=MIN(EA2:EA{})".format(q)
            ws['EB{}'.format(r)] = "=ROUND(AVERAGE(EB2:EB{}),2)".format(q)
            ws['EB{}'.format(t)] = "=MAX(EB2:EB{})".format(q)
            ws['EB{}'.format(u)] = "=MIN(EB2:EB{})".format(q)
            ws['EC{}'.format(r)] = "=MAX(EC2:EC{})".format(q)
            ws['ED{}'.format(r)] = "=MAX(ED2:ED{})".format(q)
            ws['EE{}'.format(r)] = "=MAX(EE2:EE{})".format(q)
            ws['EF{}'.format(r)] = "=MAX(EF2:EF{})".format(q)
            ws['EG{}'.format(r)] = "=MAX(EG2:EG{})".format(q)
            ws['EH{}'.format(r)] = "=MAX(EH2:EH{})".format(q)
            ws['EH{}'.format(s)] = "=MIN(EH2:EH{})".format(q)
            ws['EH{}'.format(t)] = "=ROUND(AVERAGE(EH2:EH{}),2)".format(q)
            ws['EI{}'.format(r)] = "=MAX(EI2:EI{})".format(q)
            ws['EI{}'.format(s)] = "=MIN(EI2:EI{})".format(q)
            ws['EI{}'.format(t)] = "=ROUND(AVERAGE(EI2:EI{}),2)".format(q)
            ws['EJ{}'.format(r)] = "=MAX(EJ2:EJ{})".format(q)
            ws['EJ{}'.format(s)] = "=MIN(EJ2:EJ{})".format(q)
            ws['EJ{}'.format(t)] = "=ROUND(AVERAGE(EJ2:EJ{}),2)".format(q)
            ws['EK{}'.format(r)] = "=MAX(EK2:EK{})".format(q)
            ws['EK{}'.format(s)] = "=MIN(EK2:EK{})".format(q)
            ws['EK{}'.format(t)] = "=ROUND(AVERAGE(EK2:EK{}),2)".format(q)
            ws['EL{}'.format(r)] = "=MAX(EL2:EL{})".format(q)
            ws['EL{}'.format(s)] = "=MIN(EL2:EL{})".format(q)
            ws['EL{}'.format(t)] = "=ROUND(AVERAGE(EL2:EL{}),2)".format(q)
            ws['EM{}'.format(r)] = "=MAX(EM2:EM{})".format(q)
            ws['EM{}'.format(s)] = "=MIN(EM2:EM{})".format(q)
            ws['EM{}'.format(t)] = "=ROUND(AVERAGE(EM2:EM{}),2)".format(q)
            ws['EP{}'.format(r)] = "=SUM(EP2:EP{})".format(q)
            ws['EQ{}'.format(r)] = "=SUM(EQ2:EQ{})".format(q)
            ws['ER{}'.format(r)] = "=SUM(ER2:ER{})".format(q)
            ws['ES{}'.format(r)] = "=SUM(ES2:ES{})".format(q)
            ws['ET{}'.format(r)] = "=SUM(ET2:ET{})".format(q)

            # iterasi 5 rata-rata - 1
            ws['FA{}'.format(
                r)] = "=IF($EP${}=0,$DW${},$DW${}-1)".format(r, r, r)
            ws['FA{}'.format(s)] = "=STDEV(FA2:FA{})".format(q)
            ws['FA{}'.format(t)] = "=MAX(FA2:FA{})".format(q)
            ws['FA{}'.format(u)] = "=MIN(FA2:FA{})".format(q)
            ws['FB{}'.format(
                r)] = "=IF($EQ${}=0,$DX${},$DX${}-1)".format(r, r, r)
            ws['FB{}'.format(s)] = "=STDEV(FB2:FB{})".format(q)
            ws['FB{}'.format(t)] = "=MAX(FB2:FB{})".format(q)
            ws['FB{}'.format(u)] = "=MIN(FB2:FB{})".format(q)
            ws['FC{}'.format(
                r)] = "=IF($ER${}=0,$DY${},$DY${}-1)".format(r, r, r)
            ws['FC{}'.format(s)] = "=STDEV(FC2:FC{})".format(q)
            ws['FC{}'.format(t)] = "=MAX(FC2:FC{})".format(q)
            ws['FC{}'.format(u)] = "=MIN(FC2:FC{})".format(q)
            ws['FD{}'.format(
                r)] = "=IF($ES${}=0,$DZ${},$DZ${}-1)".format(r, r, r)
            ws['FD{}'.format(s)] = "=STDEV(FD2:FD{})".format(q)
            ws['FD{}'.format(t)] = "=MAX(FD2:FD{})".format(q)
            ws['FD{}'.format(u)] = "=MIN(FD2:FD{})".format(q)
            ws['FE{}'.format(
                r)] = "=IF($ET${}=0,$EA${},$EA${}-1)".format(r, r, r)
            ws['FE{}'.format(s)] = "=STDEV(FE2:FE{})".format(q)
            ws['FE{}'.format(t)] = "=MAX(FE2:FE{})".format(q)
            ws['FE{}'.format(u)] = "=MIN(FE2:FE{})".format(q)
            ws['FF{}'.format(r)] = "=ROUND(AVERAGE(FF2:FF{}),2)".format(q)
            ws['FF{}'.format(t)] = "=MAX(FF2:FF{})".format(q)
            ws['FF{}'.format(u)] = "=MIN(FF2:FF{})".format(q)
            ws['FG{}'.format(r)] = "=MAX(FG2:FG{})".format(q)
            ws['FH{}'.format(r)] = "=MAX(FH2:FH{})".format(q)
            ws['FI{}'.format(r)] = "=MAX(FI2:FI{})".format(q)
            ws['FJ{}'.format(r)] = "=MAX(FJ2:FJ{})".format(q)
            ws['FK{}'.format(r)] = "=MAX(FK2:FK{})".format(q)
            ws['FL{}'.format(r)] = "=MAX(FL2:FL{})".format(q)
            ws['FL{}'.format(s)] = "=MIN(FL2:FL{})".format(q)
            ws['FL{}'.format(t)] = "=ROUND(AVERAGE(FL2:FL{}),2)".format(q)
            ws['FM{}'.format(r)] = "=MAX(FM2:FM{})".format(q)
            ws['FM{}'.format(s)] = "=MIN(FM2:FM{})".format(q)
            ws['FM{}'.format(t)] = "=ROUND(AVERAGE(FM2:FM{}),2)".format(q)
            ws['FN{}'.format(r)] = "=MAX(FN2:FN{})".format(q)
            ws['FN{}'.format(s)] = "=MIN(FN2:FN{})".format(q)
            ws['FN{}'.format(t)] = "=ROUND(AVERAGE(FN2:FN{}),2)".format(q)
            ws['FO{}'.format(r)] = "=MAX(FO2:FO{})".format(q)
            ws['FO{}'.format(s)] = "=MIN(FO2:FO{})".format(q)
            ws['FO{}'.format(t)] = "=ROUND(AVERAGE(FO2:FO{}),2)".format(q)
            ws['FP{}'.format(r)] = "=MAX(FP2:FP{})".format(q)
            ws['FP{}'.format(s)] = "=MIN(FP2:FP{})".format(q)
            ws['FP{}'.format(t)] = "=ROUND(AVERAGE(FP2:FP{}),2)".format(q)
            ws['FQ{}'.format(r)] = "=MAX(FQ2:FQ{})".format(q)
            ws['FQ{}'.format(s)] = "=MIN(FQ2:FQ{})".format(q)
            ws['FQ{}'.format(t)] = "=ROUND(AVERAGE(FQ2:FQ{}),2)".format(q)
            ws['FT{}'.format(r)] = "=SUM(FT2:FT{})".format(q)
            ws['FU{}'.format(r)] = "=SUM(FU2:FU{})".format(q)
            ws['FV{}'.format(r)] = "=SUM(FV2:FV{})".format(q)
            ws['FW{}'.format(r)] = "=SUM(FW2:FW{})".format(q)
            ws['FX{}'.format(r)] = "=SUM(FX2:FX{})".format(q)

            # Z Score
            ws['B1'] = 'NAMA_SISWA_1'
            ws['C1'] = 'NOMOR_NF_1'
            ws['D1'] = 'KELAS_1'
            ws['E1'] = 'NAMA_SEKOLAH_1'
            ws['F1'] = 'LOKASI_1'
            ws['G1'] = 'MAT_1'
            ws['H1'] = 'IND_1'
            ws['I1'] = 'ENG_1'
            ws['J1'] = 'IPA_1'
            ws['K1'] = 'IPS_1'
            ws['L1'] = 'JML_1'
            ws['M1'] = 'Z_MAT_1'
            ws['N1'] = 'Z_IND_1'
            ws['O1'] = 'Z_ENG_1'
            ws['P1'] = 'Z_IPA_1'
            ws['Q1'] = 'Z_IPS_1'
            ws['R1'] = 'S_MAT_1'
            ws['S1'] = 'S_IND_1'
            ws['T1'] = 'S_ENG_1'
            ws['U1'] = 'S_IPA_1'
            ws['V1'] = 'S_IPS_1'
            ws['W1'] = 'S_JML_1'
            ws['X1'] = 'RANK_NAS._1'
            ws['Y1'] = 'RANK_LOK._1'
            ws['M1'].font = Font(bold=False, name='Calibri', size=11)
            ws['N1'].font = Font(bold=False, name='Calibri', size=11)
            ws['O1'].font = Font(bold=False, name='Calibri', size=11)
            ws['P1'].font = Font(bold=False, name='Calibri', size=11)
            ws['Q1'].font = Font(bold=False, name='Calibri', size=11)
            ws['R1'].font = Font(bold=False, name='Calibri', size=11)
            ws['S1'].font = Font(bold=False, name='Calibri', size=11)
            ws['T1'].font = Font(bold=False, name='Calibri', size=11)
            ws['U1'].font = Font(bold=False, name='Calibri', size=11)
            ws['V1'].font = Font(bold=False, name='Calibri', size=11)
            ws['W1'].font = Font(bold=False, name='Calibri', size=11)
            ws['X1'].font = Font(bold=False, name='Calibri', size=11)
            ws['Y1'].font = Font(bold=False, name='Calibri', size=11)
        # FILL
            ws['B1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['C1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['D1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['E1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['F1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['G1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['H1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['I1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['J1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['K1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['L1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['M1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['N1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['O1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['P1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['Q1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['R1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['S1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['T1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['U1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['V1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['W1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['X1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['Y1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            # tambahan
            ws['Z1'] = 'MAT_20_1'
            ws['AA1'] = 'IND_20_1'
            ws['AB1'] = 'ENG_20_1'
            ws['AC1'] = 'IPA_20_1'
            ws['AD1'] = 'IPS_20_1'
            ws['Z1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AA1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AB1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AC1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AD1'].font = Font(bold=False, name='Calibri', size=11)
            ws['Z1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AA1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AB1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AC1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AD1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')

            for row in range(2, q+1):
                ws['L{}'.format(
                    row)] = '=SUM(G{}:K{})'.format(row, row, row)
                ws['M{}'.format(
                    row)] = '=IFERROR(ROUND(IF(G{}="","",(G{}-G${})/G${}),2),"")'.format(row, row, r, s)
                ws['N{}'.format(
                    row)] = '=IFERROR(ROUND(IF(H{}="","",(H{}-H${})/H${}),2),"")'.format(row, row, r, s)
                ws['O{}'.format(
                    row)] = '=IFERROR(ROUND(IF(I{}="","",(I{}-I${})/I${}),2),"")'.format(row, row, r, s)
                ws['P{}'.format(
                    row)] = '=IFERROR(ROUND(IF(J{}="","",(J{}-J${})/J${}),2),"")'.format(row, row, r, s)
                ws['Q{}'.format(
                    row)] = '=IFERROR(ROUND(IF(K{}="","",(K{}-K${})/K${}),2),"")'.format(row, row, r, s)
                ws['R{}'.format(
                    row)] = '=IFERROR(ROUND(IF(G{}="","",IF(70+30*M{}/$M${}<20,20,70+30*M{}/$M${})),2),"")'.format(row, row, r, row, r)
                ws['S{}'.format(
                    row)] = '=IFERROR(ROUND(IF(H{}="","",IF(70+30*N{}/$N${}<20,20,70+30*N{}/$N${})),2),"")'.format(row, row, r, row, r)
                ws['T{}'.format(
                    row)] = '=IFERROR(ROUND(IF(I{}="","",IF(70+30*O{}/$O${}<20,20,70+30*O{}/$O${})),2),"")'.format(row, row, r, row, r)
                ws['U{}'.format(
                    row)] = '=IFERROR(ROUND(IF(J{}="","",IF(70+30*P{}/$P${}<20,20,70+30*P{}/$P${})),2),"")'.format(row, row, r, row, r)
                ws['V{}'.format(
                    row)] = '=IFERROR(ROUND(IF(K{}="","",IF(70+30*Q{}/$Q${}<20,20,70+30*Q{}/$Q${})),2),"")'.format(row, row, r, row, r)

                ws['W{}'.format(row)] = '=IF(SUM(R{}:V{})=0,"",SUM(R{}:V{}))'.format(
                    row, row, row, row)
                ws['X{}'.format(row)] = '=IF(W{}="","",RANK(W{},$W$2:$W${}))'.format(
                    row, row, q)
                ws['Y{}'.format(
                    row)] = '=IF(X{}="","",COUNTIFS($F$2:$F${},F{},$X$2:$X${},"<"&X{})+1)'.format(row, q, row, q, row)
                # TAMBAHAN
                ws['Z{}'.format(row)] = '=IF($G${}=25,IF(AND(G{}>4,R{}=20),1,""),IF($G${}=30,IF(AND(G{}>5,R{}=20),1,""),IF($G${}=35,IF(AND(G{}>6,R{}=20),1,""),IF($G${}=40,IF(AND(G{}>7,R{}=20),1,""),IF($G${}=45,IF(AND(G{}>8,R{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['AA{}'.format(row)] = '=IF($H${}=25,IF(AND(H{}>4,S{}=20),1,""),IF($H${}=30,IF(AND(H{}>5,S{}=20),1,""),IF($H${}=35,IF(AND(H{}>6,S{}=20),1,""),IF($H${}=40,IF(AND(H{}>7,S{}=20),1,""),IF($H${}=45,IF(AND(H{}>8,S{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['AB{}'.format(row)] = '=IF($I${}=25,IF(AND(I{}>4,T{}=20),1,""),IF($I${}=30,IF(AND(I{}>5,T{}=20),1,""),IF($I${}=35,IF(AND(I{}>6,T{}=20),1,""),IF($I${}=40,IF(AND(I{}>7,T{}=20),1,""),IF($I${}=45,IF(AND(I{}>8,T{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['AC{}'.format(row)] = '=IF($J${}=25,IF(AND(J{}>4,U{}=20),1,""),IF($J${}=30,IF(AND(J{}>5,U{}=20),1,""),IF($J${}=35,IF(AND(J{}>6,U{}=20),1,""),IF($J${}=40,IF(AND(J{}>7,U{}=20),1,""),IF($J${}=45,IF(AND(J{}>8,U{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['AD{}'.format(row)] = '=IF($K${}=25,IF(AND(K{}>4,V{}=20),1,""),IF($K${}=30,IF(AND(K{}>5,V{}=20),1,""),IF($K${}=35,IF(AND(K{}>6,V{}=20),1,""),IF($K${}=40,IF(AND(K{}>7,V{}=20),1,""),IF($K${}=45,IF(AND(K{}>8,V{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

        # new Z Score
            ws['AF1'] = 'NAMA_SISWA_2'
            ws['AG1'] = 'NOMOR_NF_2'
            ws['AH1'] = 'KELAS_2'
            ws['AI1'] = 'NAMA_SEKOLAH_2'
            ws['AJ1'] = 'LOKASI_2'
            ws['AK1'] = 'MAT_2'
            ws['AL1'] = 'IND_2'
            ws['AM1'] = 'ENG_2'
            ws['AN1'] = 'IPA_2'
            ws['AO1'] = 'IPS_2'
            ws['AP1'] = 'JML_2'
            ws['AQ1'] = 'Z_MAT_2'
            ws['AR1'] = 'Z_IND_2'
            ws['AS1'] = 'Z_ENG_2'
            ws['AT1'] = 'Z_IPA_2'
            ws['AU1'] = 'Z_IPS_2'
            ws['AV1'] = 'S_MAT_2'
            ws['AW1'] = 'S_IND_2'
            ws['AX1'] = 'S_ENG_2'
            ws['AY1'] = 'S_IPA_2'
            ws['AZ1'] = 'S_IPS_2'
            ws['BA1'] = 'S_JML_2'
            ws['BB1'] = 'RANK_NAS._2'
            ws['BC1'] = 'RANK_LOK._2'
            ws['AQ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AR1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AS1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AT1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AU1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AV1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AW1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AX1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AY1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AZ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BA1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BB1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BC1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['AF1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AG1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AH1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AI1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AJ1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AK1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AL1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AM1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AN1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AO1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AP1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AQ1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AR1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AS1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AT1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AU1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AV1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AW1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AX1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AY1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AZ1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BA1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BB1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BC1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')

            # tambahan
            ws['BD1'] = 'MAT_20_2'
            ws['BE1'] = 'IND_20_2'
            ws['BF1'] = 'ENG_20_2'
            ws['BG1'] = 'IPA_20_2'
            ws['BH1'] = 'IPS_20_2'
            ws['BD1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BE1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BF1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BG1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BH1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BD1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BE1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BF1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BG1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BH1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')

            for row in range(2, q+1):
                ws['AF{}'.format(row)] = '=B{}'.format(row)
                ws['AG{}'.format(row)] = '=C{}'.format(row, row)
                ws['AH{}'.format(row)] = '=D{}'.format(row, row)
                ws['AI{}'.format(row)] = '=E{}'.format(row, row)
                ws['AJ{}'.format(row)] = '=F{}'.format(row, row)
                ws['AK{}'.format(row)] = '=IF(G{}="","",G{})'.format(row, row)
                ws['AL{}'.format(row)] = '=IF(H{}="","",H{})'.format(row, row)
                ws['AM{}'.format(row)] = '=IF(I{}="","",I{})'.format(row, row)
                ws['AN{}'.format(row)] = '=IF(J{}="","",J{})'.format(row, row)
                ws['AO{}'.format(row)] = '=IF(K{}="","",K{})'.format(row, row)
                ws['AP{}'.format(row)] = '=IF(L{}="","",L{})'.format(row, row)
                ws['AQ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AK{}="","",(AK{}-AK${})/AK${}),2),"")'.format(row, row, r, s)
                ws['AR{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AL{}="","",(AL{}-AL${})/AL${}),2),"")'.format(row, row, r, s)
                ws['AS{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AM{}="","",(AM{}-AM${})/AM${}),2),"")'.format(row, row, r, s)
                ws['AT{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AN{}="","",(AN{}-AN${})/AN${}),2),"")'.format(row, row, r, s)
                ws['AU{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AO{}="","",(AO{}-AO${})/AO${}),2),"")'.format(row, row, r, s)
                ws['AV{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AK{}="","",IF(70+30*AQ{}/$AQ${}<20,20,70+30*AQ{}/$AQ${})),2),"")'.format(row, row, r, row, r)
                ws['AW{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AL{}="","",IF(70+30*AR{}/$AR${}<20,20,70+30*AR{}/$AR${})),2),"")'.format(row, row, r, row, r)
                ws['AX{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AM{}="","",IF(70+30*AS{}/$AS${}<20,20,70+30*AS{}/$AS${})),2),"")'.format(row, row, r, row, r)
                ws['AY{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AN{}="","",IF(70+30*AT{}/$AT${}<20,20,70+30*AT{}/$AT${})),2),"")'.format(row, row, r, row, r)
                ws['AZ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AO{}="","",IF(70+30*AU{}/$AU${}<20,20,70+30*AU{}/$AU${})),2),"")'.format(row, row, r, row, r)

                ws['BA{}'.format(row)] = '=IF(SUM(AV{}:AZ{})=0,"",SUM(AV{}:AZ{}))'.format(
                    row, row, row, row)
                ws['BB{}'.format(row)] = '=IF(BA{}="","",RANK(BA{},$BA$2:$BA${}))'.format(
                    row, row, q)
                ws['BC{}'.format(
                    row)] = '=IF(BB{}="","",COUNTIFS($AJ$2:$AJ${},F{},$BB$2:$BB${},"<"&BB{})+1)'.format(row, q, row, q, row)
            #     TAMBAHAN
                ws['BD{}'.format(row)] = '=IF($G${}=25,IF(AND(AK{}>4,AV{}=20),1,""),IF($G${}=30,IF(AND(AK{}>5,AV{}=20),1,""),IF($G${}=35,IF(AND(AK{}>6,AV{}=20),1,""),IF($G${}=40,IF(AND(AK{}>7,AV{}=20),1,""),IF($G${}=45,IF(AND(AK{}>8,AV{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['BE{}'.format(row)] = '=IF($H${}=25,IF(AND(AL{}>4,AW{}=20),1,""),IF($H${}=30,IF(AND(AL{}>5,AW{}=20),1,""),IF($H${}=35,IF(AND(AL{}>6,AW{}=20),1,""),IF($H${}=40,IF(AND(AL{}>7,AW{}=20),1,""),IF($H${}=45,IF(AND(AL{}>8,AW{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['BF{}'.format(row)] = '=IF($I${}=25,IF(AND(AM{}>4,AX{}=20),1,""),IF($I${}=30,IF(AND(AM{}>5,AX{}=20),1,""),IF($I${}=35,IF(AND(AM{}>6,AX{}=20),1,""),IF($I${}=40,IF(AND(AM{}>7,AX{}=20),1,""),IF($I${}=45,IF(AND(AM{}>8,AX{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['BG{}'.format(row)] = '=IF($J${}=25,IF(AND(AN{}>4,AY{}=20),1,""),IF($J${}=30,IF(AND(AN{}>5,AY{}=20),1,""),IF($J${}=35,IF(AND(AN{}>6,AY{}=20),1,""),IF($J${}=40,IF(AND(AN{}>7,AY{}=20),1,""),IF($J${}=45,IF(AND(AN{}>8,AY{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['BH{}'.format(row)] = '=IF($K${}=25,IF(AND(AO{}>4,AZ{}=20),1,""),IF($K${}=30,IF(AND(AO{}>5,AZ{}=20),1,""),IF($K${}=35,IF(AND(AO{}>6,AZ{}=20),1,""),IF($K${}=40,IF(AND(AO{}>7,AZ{}=20),1,""),IF($K${}=45,IF(AND(AO{}>8,AZ{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

                # new Z Score [2]
            ws['BJ1'] = 'NAMA_SISWA_3'
            ws['BK1'] = 'NOMOR_NF_3'
            ws['BL1'] = 'KELAS_3'
            ws['BM1'] = 'NAMA_SEKOLAH_3'
            ws['BN1'] = 'LOKASI_3'
            ws['BO1'] = 'MAT_3'
            ws['BP1'] = 'IND_3'
            ws['BQ1'] = 'ENG_3'
            ws['BR1'] = 'IPA_3'
            ws['BS1'] = 'IPS_3'
            ws['BT1'] = 'JML_3'
            ws['BU1'] = 'Z_MAT_3'
            ws['BV1'] = 'Z_IND_3'
            ws['BW1'] = 'Z_ENG_3'
            ws['BX1'] = 'Z_IPA_3'
            ws['BY1'] = 'Z_IPS_3'
            ws['BZ1'] = 'S_MAT_3'
            ws['CA1'] = 'S_IND_3'
            ws['CB1'] = 'S_ENG_3'
            ws['CC1'] = 'S_IPA_3'
            ws['CD1'] = 'S_IPS_3'
            ws['CE1'] = 'S_JML_3'
            ws['CF1'] = 'RANK_NAS._3'
            ws['CG1'] = 'RANK_LOK._3'
            ws['BU1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BV1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BW1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BX1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BY1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BZ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CA1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CB1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CC1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CD1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CE1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CF1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CG1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['BJ1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BK1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BL1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BM1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BN1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BO1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BP1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BQ1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BR1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BS1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BT1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BU1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BV1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BW1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BX1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BY1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BZ1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CA1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CB1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CC1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CD1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CE1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CF1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CG1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')

            # tambahan
            ws['CH1'] = 'MAT_20_3'
            ws['CI1'] = 'IND_20_3'
            ws['CJ1'] = 'ENG_20_3'
            ws['CK1'] = 'IPA_20_3'
            ws['CL1'] = 'IPS_20_3'
            ws['CH1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CI1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CJ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CK1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CL1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CH1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CI1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CJ1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CK1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CL1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')

            for row in range(2, q+1):
                ws['BJ{}'.format(row)] = '=B{}'.format(row)
                ws['BK{}'.format(row)] = '=C{}'.format(row, row)
                ws['BL{}'.format(row)] = '=D{}'.format(row, row)
                ws['BM{}'.format(row)] = '=E{}'.format(row, row)
                ws['BN{}'.format(row)] = '=F{}'.format(row, row)
                ws['BO{}'.format(row)] = '=IF(G{}="","",G{})'.format(row, row)
                ws['BP{}'.format(row)] = '=IF(H{}="","",H{})'.format(row, row)
                ws['BQ{}'.format(row)] = '=IF(I{}="","",I{})'.format(row, row)
                ws['BR{}'.format(row)] = '=IF(J{}="","",J{})'.format(row, row)
                ws['BS{}'.format(row)] = '=IF(K{}="","",K{})'.format(row, row)
                ws['BT{}'.format(row)] = '=IF(L{}="","",L{})'.format(row, row)
                ws['BU{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BO{}="","",(BO{}-BO${})/BO${}),2),"")'.format(row, row, r, s)
                ws['BV{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BP{}="","",(BP{}-BP${})/BP${}),2),"")'.format(row, row, r, s)
                ws['BW{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BQ{}="","",(BQ{}-BQ${})/BQ${}),2),"")'.format(row, row, r, s)
                ws['BX{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BR{}="","",(BR{}-BR${})/BR${}),2),"")'.format(row, row, r, s)
                ws['BY{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BS{}="","",(BS{}-BS${})/BS${}),2),"")'.format(row, row, r, s)
                ws['BZ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BO{}="","",IF(70+30*BU{}/$BU${}<20,20,70+30*BU{}/$BU${})),2),"")'.format(row, row, r, row, r)
                ws['CA{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BP{}="","",IF(70+30*BV{}/$BV${}<20,20,70+30*BV{}/$BV${})),2),"")'.format(row, row, r, row, r)
                ws['CB{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BQ{}="","",IF(70+30*BW{}/$BW${}<20,20,70+30*BW{}/$BW${})),2),"")'.format(row, row, r, row, r)
                ws['CC{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BR{}="","",IF(70+30*BX{}/$BX${}<20,20,70+30*BX{}/$BX${})),2),"")'.format(row, row, r, row, r)
                ws['CD{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BS{}="","",IF(70+30*BY{}/$BY${}<20,20,70+30*BY{}/$BY${})),2),"")'.format(row, row, r, row, r)

                ws['CE{}'.format(row)] = '=IF(SUM(BZ{}:CD{})=0,"",SUM(BZ{}:CD{}))'.format(
                    row, row, row, row)
                ws['CF{}'.format(row)] = '=IF(CE{}="","",RANK(CE{},$CE$2:$CE${}))'.format(
                    row, row, q)
                ws['CG{}'.format(
                    row)] = '=IF(CF{}="","",COUNTIFS($BN$2:$BN${},F{},$CF$2:$CF${},"<"&CF{})+1)'.format(row, q, row, q, row)
                #     TAMBAHAN
                ws['CH{}'.format(row)] = '=IF($G${}=25,IF(AND(BO{}>4,BZ{}=20),1,""),IF($G${}=30,IF(AND(BO{}>5,BZ{}=20),1,""),IF($G${}=35,IF(AND(BO{}>6,BZ{}=20),1,""),IF($G${}=40,IF(AND(BO{}>7,BZ{}=20),1,""),IF($G${}=45,IF(AND(BO{}>8,BZ{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['CI{}'.format(row)] = '=IF($H${}=25,IF(AND(BP{}>4,CA{}=20),1,""),IF($H${}=30,IF(AND(BP{}>5,CA{}=20),1,""),IF($H${}=35,IF(AND(BP{}>6,CA{}=20),1,""),IF($H${}=40,IF(AND(BP{}>7,CA{}=20),1,""),IF($H${}=45,IF(AND(BP{}>8,CA{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['CJ{}'.format(row)] = '=IF($I${}=25,IF(AND(BQ{}>4,CB{}=20),1,""),IF($I${}=30,IF(AND(BQ{}>5,CB{}=20),1,""),IF($I${}=35,IF(AND(BQ{}>6,CB{}=20),1,""),IF($I${}=40,IF(AND(BQ{}>7,CB{}=20),1,""),IF($I${}=45,IF(AND(BQ{}>8,CB{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['CK{}'.format(row)] = '=IF($J${}=25,IF(AND(BR{}>4,CC{}=20),1,""),IF($J${}=30,IF(AND(BR{}>5,CC{}=20),1,""),IF($J${}=35,IF(AND(BR{}>6,CC{}=20),1,""),IF($J${}=40,IF(AND(BR{}>7,CC{}=20),1,""),IF($J${}=45,IF(AND(BR{}>8,CC{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['CL{}'.format(row)] = '=IF($K${}=25,IF(AND(BS{}>4,CD{}=20),1,""),IF($K${}=30,IF(AND(BS{}>5,CD{}=20),1,""),IF($K${}=35,IF(AND(BS{}>6,CD{}=20),1,""),IF($K${}=40,IF(AND(BS{}>7,CD{}=20),1,""),IF($K${}=45,IF(AND(BS{}>8,CD{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

                # new Z Score [3]
            ws['CN1'] = 'NAMA_SISWA_4'
            ws['CO1'] = 'NOMOR_NF_4'
            ws['CP1'] = 'KELAS_4'
            ws['CQ1'] = 'NAMA_SEKOLAH_4'
            ws['CR1'] = 'LOKASI_4'
            ws['CS1'] = 'MAT_4'
            ws['CT1'] = 'IND_4'
            ws['CU1'] = 'ENG_4'
            ws['CV1'] = 'IPA_4'
            ws['CW1'] = 'IPS_4'
            ws['CX1'] = 'JML_4'
            ws['CY1'] = 'Z_MAT_4'
            ws['CZ1'] = 'Z_IND_4'
            ws['DA1'] = 'Z_ENG_4'
            ws['DB1'] = 'Z_IPA_4'
            ws['DC1'] = 'Z_IPS_4'
            ws['DD1'] = 'S_MAT_4'
            ws['DE1'] = 'S_IND_4'
            ws['DF1'] = 'S_ENG_4'
            ws['DG1'] = 'S_IPA_4'
            ws['DH1'] = 'S_IPS_4'
            ws['DI1'] = 'S_JML_4'
            ws['DJ1'] = 'RANK_NAS._4'
            ws['DK1'] = 'RANK_LOK._4'
            ws['CY1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CZ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DA1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DB1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DC1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DD1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DE1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DF1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DG1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DH1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DI1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DJ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DK1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['CN1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CO1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CP1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CQ1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CR1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CS1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CT1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CU1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CV1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CW1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CX1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CY1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CZ1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DA1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DB1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DC1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DD1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DE1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DF1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DG1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DH1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DI1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DJ1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DK1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')

            # tambahan
            ws['DL1'] = 'MAT_20_4'
            ws['DM1'] = 'IND_20_4'
            ws['DN1'] = 'ENG_20_4'
            ws['DO1'] = 'IPA_20_4'
            ws['DP1'] = 'IPS_20_4'
            ws['DL1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DM1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DN1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DO1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DP1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DL1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DM1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DN1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DO1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DP1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')

            for row in range(2, q+1):
                ws['CN{}'.format(row)] = '=B{}'.format(row)
                ws['CO{}'.format(row)] = '=C{}'.format(row, row)
                ws['CP{}'.format(row)] = '=D{}'.format(row, row)
                ws['CQ{}'.format(row)] = '=E{}'.format(row, row)
                ws['CR{}'.format(row)] = '=F{}'.format(row, row)
                ws['CS{}'.format(row)] = '=IF(G{}="","",G{})'.format(row, row)
                ws['CT{}'.format(row)] = '=IF(H{}="","",H{})'.format(row, row)
                ws['CU{}'.format(row)] = '=IF(I{}="","",I{})'.format(row, row)
                ws['CV{}'.format(row)] = '=IF(J{}="","",J{})'.format(row, row)
                ws['CW{}'.format(row)] = '=IF(K{}="","",K{})'.format(row, row)
                ws['CX{}'.format(row)] = '=IF(L{}="","",L{})'.format(row, row)
                ws['CY{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CS{}="","",(CS{}-CS${})/CS${}),2),"")'.format(row, row, r, s)
                ws['CZ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CT{}="","",(CT{}-CT${})/CT${}),2),"")'.format(row, row, r, s)
                ws['DA{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CU{}="","",(CU{}-CU${})/CU${}),2),"")'.format(row, row, r, s)
                ws['DB{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CV{}="","",(CV{}-CV${})/CV${}),2),"")'.format(row, row, r, s)
                ws['DC{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CW{}="","",(CW{}-CW${})/CW${}),2),"")'.format(row, row, r, s)
                ws['DD{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CY{}="","",IF(70+30*CY{}/$CY${}<20,20,70+30*CY{}/$CY${})),2),"")'.format(row, row, r, row, r)
                ws['DE{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CZ{}="","",IF(70+30*CZ{}/$CZ${}<20,20,70+30*CZ{}/$CZ${})),2),"")'.format(row, row, r, row, r)
                ws['DF{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DA{}="","",IF(70+30*DA{}/$DA${}<20,20,70+30*DA{}/$DA${})),2),"")'.format(row, row, r, row, r)
                ws['DG{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DB{}="","",IF(70+30*DB{}/$DB${}<20,20,70+30*DB{}/$DB${})),2),"")'.format(row, row, r, row, r)
                ws['DH{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DC{}="","",IF(70+30*DC{}/$DC${}<20,20,70+30*DC{}/$DC${})),2),"")'.format(row, row, r, row, r)

                ws['DI{}'.format(row)] = '=IF(SUM(DD{}:DH{})=0,"",SUM(DD{}:DH{}))'.format(
                    row, row, row, row)
                ws['DJ{}'.format(row)] = '=IF(DI{}="","",RANK(DI{},$DI$2:$DI${}))'.format(
                    row, row, q)
                ws['DK{}'.format(
                    row)] = '=IF(DJ{}="","",COUNTIFS($CR$2:$CR${},F{},$DJ$2:$DJ${},"<"&DJ{})+1)'.format(row, q, row, q, row)
                #     TAMBAHAN
                ws['DL{}'.format(row)] = '=IF($G${}=25,IF(AND(CS{}>4,DD{}=20),1,""),IF($G${}=30,IF(AND(CS{}>5,DD{}=20),1,""),IF($G${}=35,IF(AND(CS{}>6,DD{}=20),1,""),IF($G${}=40,IF(AND(CS{}>7,DD{}=20),1,""),IF($G${}=45,IF(AND(CS{}>8,DD{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['DM{}'.format(row)] = '=IF($H${}=25,IF(AND(CT{}>4,DE{}=20),1,""),IF($H${}=30,IF(AND(CT{}>5,DE{}=20),1,""),IF($H${}=35,IF(AND(CT{}>6,DE{}=20),1,""),IF($H${}=40,IF(AND(CT{}>7,DE{}=20),1,""),IF($H${}=45,IF(AND(CT{}>8,DE{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['DN{}'.format(row)] = '=IF($I${}=25,IF(AND(CU{}>4,DF{}=20),1,""),IF($I${}=30,IF(AND(CU{}>5,DF{}=20),1,""),IF($I${}=35,IF(AND(CU{}>6,DF{}=20),1,""),IF($I${}=40,IF(AND(CU{}>7,DF{}=20),1,""),IF($I${}=45,IF(AND(CU{}>8,DF{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['DO{}'.format(row)] = '=IF($J${}=25,IF(AND(CV{}>4,DG{}=20),1,""),IF($J${}=30,IF(AND(CV{}>5,DG{}=20),1,""),IF($J${}=35,IF(AND(CV{}>6,DG{}=20),1,""),IF($J${}=40,IF(AND(CV{}>7,DG{}=20),1,""),IF($J${}=45,IF(AND(CV{}>8,DG{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['DP{}'.format(row)] = '=IF($K${}=25,IF(AND(CW{}>4,DH{}=20),1,""),IF($K${}=30,IF(AND(CW{}>5,DH{}=20),1,""),IF($K${}=35,IF(AND(CW{}>6,DH{}=20),1,""),IF($K${}=40,IF(AND(CW{}>7,DH{}=20),1,""),IF($K${}=45,IF(AND(CW{}>8,DH{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

            # new Z Score [4]
            ws['DR1'] = 'NAMA_SISWA_5'
            ws['DS1'] = 'NOMOR_NF_5'
            ws['DT1'] = 'KELAS_5'
            ws['DU1'] = 'NAMA_SEKOLAH_5'
            ws['DV1'] = 'LOKASI_5'
            ws['DW1'] = 'MAT_5'
            ws['DX1'] = 'IND_5'
            ws['DY1'] = 'ENG_5'
            ws['DZ1'] = 'IPA_5'
            ws['EA1'] = 'IPS_5'
            ws['EB1'] = 'JML_5'
            ws['EC1'] = 'Z_MAT_5'
            ws['ED1'] = 'Z_IND_5'
            ws['EE1'] = 'Z_ENG_5'
            ws['EF1'] = 'Z_IPA_5'
            ws['EG1'] = 'Z_IPS_5'
            ws['EH1'] = 'S_MAT_5'
            ws['EI1'] = 'S_IND_5'
            ws['EJ1'] = 'S_ENG_5'
            ws['EK1'] = 'S_IPA_5'
            ws['EL1'] = 'S_IPS_5'
            ws['EM1'] = 'S_JML_5'
            ws['EN1'] = 'RANK_NAS._5'
            ws['EO1'] = 'RANK_LOK._5'
            ws['EC1'].font = Font(bold=False, name='Calibri', size=11)
            ws['ED1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EE1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EF1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EG1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EH1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EI1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EJ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EK1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EL1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EM1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EN1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EO1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['DR1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DS1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DT1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DU1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DV1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DW1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DX1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DY1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DZ1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EA1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EB1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EC1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['ED1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EE1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EF1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EG1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EH1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EI1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EJ1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EK1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EL1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EM1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EN1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EO1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')

            # tambahan
            ws['EP1'] = 'MAT_20_5'
            ws['EQ1'] = 'IND_20_5'
            ws['ER1'] = 'ENG_20_5'
            ws['ES1'] = 'IPA_20_5'
            ws['ET1'] = 'IPS_20_5'
            ws['EP1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EQ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['ER1'].font = Font(bold=False, name='Calibri', size=11)
            ws['ES1'].font = Font(bold=False, name='Calibri', size=11)
            ws['ET1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EP1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EQ1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['ER1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['ES1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['ET1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')

            for row in range(2, q+1):
                ws['DR{}'.format(row)] = '=B{}'.format(row)
                ws['DS{}'.format(row)] = '=C{}'.format(row, row)
                ws['DT{}'.format(row)] = '=D{}'.format(row, row)
                ws['DU{}'.format(row)] = '=E{}'.format(row, row)
                ws['DV{}'.format(row)] = '=F{}'.format(row, row)
                ws['DW{}'.format(row)] = '=IF(G{}="","",G{})'.format(row, row)
                ws['DX{}'.format(row)] = '=IF(H{}="","",H{})'.format(row, row)
                ws['DY{}'.format(row)] = '=IF(I{}="","",I{})'.format(row, row)
                ws['DZ{}'.format(row)] = '=IF(J{}="","",J{})'.format(row, row)
                ws['EA{}'.format(row)] = '=IF(K{}="","",K{})'.format(row, row)
                ws['EB{}'.format(row)] = '=IF(L{}="","",L{})'.format(row, row)
                ws['EC{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DW{}="","",(DW{}-DW${})/DW${}),2),"")'.format(row, row, r, s)
                ws['ED{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DX{}="","",(DX{}-DX${})/DX${}),2),"")'.format(row, row, r, s)
                ws['EE{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DY{}="","",(DY{}-DY${})/DY${}),2),"")'.format(row, row, r, s)
                ws['EF{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DZ{}="","",(DZ{}-DZ${})/DZ${}),2),"")'.format(row, row, r, s)
                ws['EG{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EA{}="","",(EA{}-EA${})/EA${}),2),"")'.format(row, row, r, s)
                ws['EH{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EC{}="","",IF(70+30*EC{}/$EC${}<20,20,70+30*EC{}/$EC${})),2),"")'.format(row, row, r, row, r)
                ws['EI{}'.format(
                    row)] = '=IFERROR(ROUND(IF(ED{}="","",IF(70+30*ED{}/$ED${}<20,20,70+30*ED{}/$ED${})),2),"")'.format(row, row, r, row, r)
                ws['EJ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EE{}="","",IF(70+30*EE{}/$EE${}<20,20,70+30*EE{}/$EE${})),2),"")'.format(row, row, r, row, r)
                ws['EK{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EF{}="","",IF(70+30*EF{}/$EF${}<20,20,70+30*EF{}/$EF${})),2),"")'.format(row, row, r, row, r)
                ws['EL{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EG{}="","",IF(70+30*EG{}/$EG${}<20,20,70+30*EG{}/$EG${})),2),"")'.format(row, row, r, row, r)

                ws['EM{}'.format(row)] = '=IF(SUM(EH{}:EL{})=0,"",SUM(EH{}:EL{}))'.format(
                    row, row, row, row)
                ws['EN{}'.format(row)] = '=IF(EM{}="","",RANK(EM{},$EM$2:$EM${}))'.format(
                    row, row, q)
                ws['EO{}'.format(
                    row)] = '=IF(EN{}="","",COUNTIFS($DV$2:$DV${},F{},$EN$2:$EN${},"<"&EN{})+1)'.format(row, q, row, q, row)
                #     TAMBAHAN
                ws['EP{}'.format(row)] = '=IF($G${}=25,IF(AND(DW{}>4,EH{}=20),1,""),IF($G${}=30,IF(AND(DW{}>5,EH{}=20),1,""),IF($G${}=35,IF(AND(DW{}>6,EH{}=20),1,""),IF($G${}=40,IF(AND(DW{}>7,EH{}=20),1,""),IF($G${}=45,IF(AND(DW{}>8,EH{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['EQ{}'.format(row)] = '=IF($H${}=25,IF(AND(DX{}>4,EI{}=20),1,""),IF($H${}=30,IF(AND(DX{}>5,EI{}=20),1,""),IF($H${}=35,IF(AND(DX{}>6,EI{}=20),1,""),IF($H${}=40,IF(AND(DX{}>7,EI{}=20),1,""),IF($H${}=45,IF(AND(DX{}>8,EI{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['ER{}'.format(row)] = '=IF($I${}=25,IF(AND(DY{}>4,EJ{}=20),1,""),IF($I${}=30,IF(AND(DY{}>5,EJ{}=20),1,""),IF($I${}=35,IF(AND(DY{}>6,EJ{}=20),1,""),IF($I${}=40,IF(AND(DY{}>7,EJ{}=20),1,""),IF($I${}=45,IF(AND(DY{}>8,EJ{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['ES{}'.format(row)] = '=IF($J${}=25,IF(AND(DZ{}>4,EK{}=20),1,""),IF($J${}=30,IF(AND(DZ{}>5,EK{}=20),1,""),IF($J${}=35,IF(AND(DZ{}>6,EK{}=20),1,""),IF($J${}=40,IF(AND(DZ{}>7,EK{}=20),1,""),IF($J${}=45,IF(AND(DZ{}>8,EK{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['ET{}'.format(row)] = '=IF($K${}=25,IF(AND(EA{}>4,EL{}=20),1,""),IF($K${}=30,IF(AND(EA{}>5,EL{}=20),1,""),IF($K${}=35,IF(AND(EA{}>6,EL{}=20),1,""),IF($K${}=40,IF(AND(EA{}>7,EL{}=20),1,""),IF($K${}=45,IF(AND(EA{}>8,EL{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

                # new Z Score [5]
            ws['EV1'] = 'NAMA SISWA'
            ws['EW1'] = 'NOMOR NF'
            ws['EX1'] = 'KELAS'
            ws['EY1'] = 'NAMA SEKOLAH'
            ws['EZ1'] = 'LOKASI'
            ws['FA1'] = 'MAT'
            ws['FB1'] = 'IND'
            ws['FC1'] = 'ENG'
            ws['FD1'] = 'IPA'
            ws['FE1'] = 'IPS'
            ws['FF1'] = 'JML'
            ws['FG1'] = 'Z_MAT'
            ws['FH1'] = 'Z_IND'
            ws['FI1'] = 'Z_ENG'
            ws['FJ1'] = 'Z_IPA'
            ws['FK1'] = 'Z_IPS'
            ws['FL1'] = 'S_MAT'
            ws['FM1'] = 'S_IND'
            ws['FN1'] = 'S_ENG'
            ws['FO1'] = 'S_IPA'
            ws['FP1'] = 'S_IPS'
            ws['FQ1'] = 'S_JML'
            ws['FR1'] = 'RANK NAS.'
            ws['FS1'] = 'RANK LOK.'
            ws['FG1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FH1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FI1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FJ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FK1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FL1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FM1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FN1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FO1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FP1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FQ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FR1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FS1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['EV1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EW1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EX1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EY1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EZ1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FA1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FB1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FC1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FD1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FE1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FF1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FG1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FH1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FI1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FJ1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FK1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FL1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FM1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FN1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FO1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FP1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FQ1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FR1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FS1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')

            # tambahan
            ws['FT1'] = 'MAT_20'
            ws['FU1'] = 'IND_20'
            ws['FV1'] = 'ENG_20'
            ws['FW1'] = 'IPA_20'
            ws['FX1'] = 'IPS_20'
            ws['FT1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FU1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FV1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FW1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FX1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FT1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FU1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FV1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FW1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FX1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')

            for row in range(2, q+1):
                ws['EV{}'.format(row)] = '=B{}'.format(row)
                ws['EW{}'.format(row)] = '=C{}'.format(row, row)
                ws['EX{}'.format(row)] = '=D{}'.format(row, row)
                ws['EY{}'.format(row)] = '=E{}'.format(row, row)
                ws['EZ{}'.format(row)] = '=F{}'.format(row, row)
                ws['FA{}'.format(row)] = '=IF(G{}="","",G{})'.format(row, row)
                ws['FB{}'.format(row)] = '=IF(H{}="","",H{})'.format(row, row)
                ws['FC{}'.format(row)] = '=IF(I{}="","",I{})'.format(row, row)
                ws['FD{}'.format(row)] = '=IF(J{}="","",J{})'.format(row, row)
                ws['FE{}'.format(row)] = '=IF(K{}="","",K{})'.format(row, row)
                ws['FF{}'.format(row)] = '=IF(L{}="","",L{})'.format(row, row)
                ws['FG{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FA{}="","",(FA{}-FA${})/FA${}),2),"")'.format(row, row, r, s)
                ws['FH{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FB{}="","",(FB{}-FB${})/FB${}),2),"")'.format(row, row, r, s)
                ws['FI{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FC{}="","",(FC{}-FC${})/FC${}),2),"")'.format(row, row, r, s)
                ws['FJ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FD{}="","",(FD{}-FD${})/FD${}),2),"")'.format(row, row, r, s)
                ws['FK{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FE{}="","",(FE{}-FE${})/FE${}),2),"")'.format(row, row, r, s)
                ws['FL{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FG{}="","",IF(70+30*FG{}/$FG${}<20,20,70+30*FG{}/$FG${})),2),"")'.format(row, row, r, row, r)
                ws['FM{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FH{}="","",IF(70+30*FH{}/$FH${}<20,20,70+30*FH{}/$FH${})),2),"")'.format(row, row, r, row, r)
                ws['FN{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FI{}="","",IF(70+30*FI{}/$FI${}<20,20,70+30*FI{}/$FI${})),2),"")'.format(row, row, r, row, r)
                ws['FO{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FJ{}="","",IF(70+30*FJ{}/$FJ${}<20,20,70+30*FJ{}/$FJ${})),2),"")'.format(row, row, r, row, r)
                ws['FP{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FK{}="","",IF(70+30*FK{}/$FK${}<20,20,70+30*FK{}/$FK${})),2),"")'.format(row, row, r, row, r)

                ws['FQ{}'.format(row)] = '=IF(SUM(FL{}:FP{})=0,"",SUM(FL{}:FP{}))'.format(
                    row, row, row, row)
                ws['FR{}'.format(row)] = '=IF(FQ{}="","",RANK(FQ{},$FQ$2:$FQ${}))'.format(
                    row, row, q)
                ws['FS{}'.format(
                    row)] = '=IF(FR{}="","",COUNTIFS($EZ$2:$EZ${},F{},$FR$2:$FR${},"<"&FR{})+1)'.format(row, q, row, q, row)
                #     TAMBAHAN
                ws['FT{}'.format(row)] = '=IF($G${}=25,IF(AND(FA{}>4,FL{}=20),1,""),IF($G${}=30,IF(AND(FA{}>5,FL{}=20),1,""),IF($G${}=35,IF(AND(FA{}>6,FL{}=20),1,""),IF($G${}=40,IF(AND(FA{}>7,FL{}=20),1,""),IF($G${}=45,IF(AND(FA{}>8,FL{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['FU{}'.format(row)] = '=IF($H${}=25,IF(AND(FB{}>4,FM{}=20),1,""),IF($H${}=30,IF(AND(FB{}>5,FM{}=20),1,""),IF($H${}=35,IF(AND(FB{}>6,FM{}=20),1,""),IF($H${}=40,IF(AND(FB{}>7,FM{}=20),1,""),IF($H${}=45,IF(AND(FB{}>8,FM{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['FV{}'.format(row)] = '=IF($I${}=25,IF(AND(FC{}>4,FN{}=20),1,""),IF($I${}=30,IF(AND(FC{}>5,FN{}=20),1,""),IF($I${}=35,IF(AND(FC{}>6,FN{}=20),1,""),IF($I${}=40,IF(AND(FC{}>7,FN{}=20),1,""),IF($I${}=45,IF(AND(FC{}>8,FN{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['FW{}'.format(row)] = '=IF($J${}=25,IF(AND(FD{}>4,FO{}=20),1,""),IF($J${}=30,IF(AND(FD{}>5,FO{}=20),1,""),IF($J${}=35,IF(AND(FD{}>6,FO{}=20),1,""),IF($J${}=40,IF(AND(FD{}>7,FO{}=20),1,""),IF($J${}=45,IF(AND(FD{}>8,FO{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['FX{}'.format(row)] = '=IF($K${}=25,IF(AND(FE{}>4,FP{}=20),1,""),IF($K${}=30,IF(AND(FE{}>5,FP{}=20),1,""),IF($K${}=35,IF(AND(FE{}>6,FP{}=20),1,""),IF($K${}=40,IF(AND(FE{}>7,FP{}=20),1,""),IF($K${}=45,IF(AND(FE{}>8,FP{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

            # Mengubah 'KELAS' sesuai dengan nilai yang dipilih dari selectbox 'KELAS'
            kelas = KELAS.lower().replace(" ", "")
            semester = SEMESTER.lower()
            tahun = TAHUN.replace("-", "")
            penilaian = PENILAIAN.lower()
            kurikulum = KURIKULUM.lower()

            path_file = f"{kelas}_{penilaian}_{semester}_{kurikulum}_{tahun}_nilai_std.xlsx"

            # Simpan file ke direktori temporer
            temp_dir = tempfile.gettempdir()
            file_path = temp_dir + '/' + path_file
            wb.save(file_path)

            st.success(
                "File siap diunduh!")

            # Tombol unduh file
            with open(file_path, "rb") as f:
                bytes_data = f.read()
            st.download_button(label="Unduh File", data=bytes_data,
                               file_name=path_file)

            st.warning(
                "Buka file unduhan, klik 'Enable Editing' dan 'Save'")  
    if selected_file == "Nilai Std. SD (KM)":
        # menghilangkan hamburger
        st.markdown("""
        <style>
        .css-1rs6os.edgvbvh3
        {
            visibility:hidden;
        }
        .css-1lsmgbg.egzxvld0
        {
            visibility:hidden;
        }
        </style>
        """, unsafe_allow_html=True)

        image = Image.open('logo resmi nf resize.png')
        st.image(image)

        st.title("Olah Nilai Standar 4, 5 (KM)")
        st.header("SD KM")

        col6 = st.container()

        with col6:
            KELAS = st.selectbox(
                "KELAS",
                ("--Pilih Kelas--", "4 SD", "5 SD"))

        col7 = st.container()

        with col7:
            SEMESTER = st.selectbox(
                "SEMESTER",
                ("--Pilih Semester--", "SEMESTER 1", "SEMESTER 2"))

        col8 = st.container()

        with col8:
            PENILAIAN = st.selectbox(
                "PENILAIAN",
                ("--Pilih Penilaian--", "SUMATIF TENGAH SEMESTER", "SUMATIF AKHIR TAHUN"))

        col9 = st.container()

        with col9:
            KURIKULUM = st.selectbox(
                "KURIKULUM",
                ("--Pilih Kurikulum--", "KM", "K13"))

        TAHUN = st.text_input("Masukkan Tahun Ajaran",
                              placeholder="contoh: 2022-2023")

        
        col1, col2, col3, col4 = st.columns(4)

        with col1:
            MTK = st.selectbox(
                "JML. SOAL MAT.",
                (15, 20, 25, 30, 35, 40, 45, 50))

        with col2:
            IND = st.selectbox(
                "JML. SOAL IND.",
                (15, 20, 25, 30, 35, 40, 45, 50))

        with col3:
            ENG = st.selectbox(
                "JML. SOAL ENG.",
                (15, 20, 25, 30, 35, 40, 45, 50))

        with col4:
            IPAS = st.selectbox(
                "JML. SOAL IPAS.",
                (15, 20, 25, 30, 35, 40, 45, 50))

        JML_SOAL_MAT = MTK
        JML_SOAL_IND = IND
        JML_SOAL_ENG = ENG
        JML_SOAL_IPAS = IPAS

        uploaded_file = st.file_uploader(
            'Letakkan file excel', type='xlsx')

        if uploaded_file is not None:
            wb = openpyxl.load_workbook(uploaded_file)
            ws = wb['Sheet1']

            q = len(ws['K'])
            r = len(ws['K'])+2
            s = len(ws['K'])+3
            t = len(ws['K'])+4
            u = len(ws['K'])+5
            v = len(ws['K'])+6
            w = len(ws['K'])+7
            x = len(ws['K'])+8

            ws['G{}'.format(r)] = "=ROUND(AVERAGE(G2:G{}),2)".format(q)
            ws['H{}'.format(r)] = "=ROUND(AVERAGE(H2:H{}),2)".format(q)
            ws['I{}'.format(r)] = "=ROUND(AVERAGE(I2:I{}),2)".format(q)
            ws['J{}'.format(r)] = "=ROUND(AVERAGE(J2:J{}),2)".format(q)
            ws['K{}'.format(r)] = "=ROUND(AVERAGE(K2:K{}),2)".format(q)
            ws['G{}'.format(s)] = "=STDEV(G2:G{})".format(q)
            ws['H{}'.format(s)] = "=STDEV(H2:H{})".format(q)
            ws['I{}'.format(s)] = "=STDEV(I2:I{})".format(q)
            ws['J{}'.format(s)] = "=STDEV(J2:J{})".format(q)
            ws['G{}'.format(t)] = "=MAX(G2:G{})".format(q)
            ws['H{}'.format(t)] = "=MAX(H2:H{})".format(q)
            ws['I{}'.format(t)] = "=MAX(I2:I{})".format(q)
            ws['J{}'.format(t)] = "=MAX(J2:J{})".format(q)
            ws['K{}'.format(t)] = "=MAX(K2:K{})".format(q)
            ws['L{}'.format(r)] = "=MAX(L2:L{})".format(q)
            ws['M{}'.format(r)] = "=MAX(M2:M{})".format(q)
            ws['N{}'.format(r)] = "=MAX(N2:N{})".format(q)
            ws['O{}'.format(r)] = "=MAX(O2:O{})".format(q)
            ws['P{}'.format(r)] = "=MAX(P2:P{})".format(q)
            ws['Q{}'.format(r)] = "=MAX(Q2:Q{})".format(q)
            ws['R{}'.format(r)] = "=MAX(R2:R{})".format(q)
            ws['S{}'.format(r)] = "=MAX(S2:S{})".format(q)
            ws['T{}'.format(r)] = "=ROUND(MAX(T2:T{}),2)".format(q)
            ws['U{}'.format(r)] = "=MAX(U2:U{})".format(q)
            ws['G{}'.format(u)] = "=MIN(G2:G{})".format(q)
            ws['H{}'.format(u)] = "=MIN(H2:H{})".format(q)
            ws['I{}'.format(u)] = "=MIN(I2:I{})".format(q)
            ws['J{}'.format(u)] = "=MIN(J2:J{})".format(q)
            ws['K{}'.format(u)] = "=MIN(K2:K{})".format(q)
            ws['P{}'.format(s)] = "=MIN(P2:P{})".format(q)
            ws['Q{}'.format(s)] = "=MIN(Q2:R{})".format(q)
            ws['R{}'.format(s)] = "=MIN(R2:S{})".format(q)
            ws['S{}'.format(s)] = "=MIN(S2:T{})".format(q)
            ws['T{}'.format(s)] = "=MIN(T2:T{})".format(q)
            ws['P{}'.format(t)] = "=ROUND(AVERAGE(P2:P{}),2)".format(q)
            ws['Q{}'.format(t)] = "=ROUND(AVERAGE(Q2:Q{}),2)".format(q)
            ws['R{}'.format(t)] = "=ROUND(AVERAGE(R2:R{}),2)".format(q)
            ws['S{}'.format(t)] = "=ROUND(AVERAGE(S2:S{}),2)".format(q)
            ws['T{}'.format(t)] = "=ROUND(AVERAGE(T2:T{}),2)".format(q)
            ws['W{}'.format(r)] = "=SUM(W2:W{})".format(q)
            ws['X{}'.format(r)] = "=SUM(X2:X{})".format(q)
            ws['Y{}'.format(r)] = "=SUM(Y2:Y{})".format(q)
            ws['Z{}'.format(r)] = "=SUM(Z2:Z{})".format(q)

            # new
            # iterasi 1 rata-rata - 1

            # MAPEL NORMAL
            ws['AG{}'.format(r)] = "=IF($W${}=0,$G${},$G${}-1)".format(r, r, r)
            ws['AG{}'.format(s)] = "=STDEV(AG2:AG{})".format(q)
            ws['AG{}'.format(t)] = "=MAX(AG2:AG{})".format(q)
            ws['AG{}'.format(u)] = "=MIN(AG2:AG{})".format(q)
            ws['AH{}'.format(r)] = "=IF($X${}=0,$H${},$H${}-1)".format(r, r, r)
            ws['AH{}'.format(s)] = "=STDEV(AH2:AH{})".format(q)
            ws['AH{}'.format(t)] = "=MAX(AH2:AH{})".format(q)
            ws['AH{}'.format(u)] = "=MIN(AH2:AH{})".format(q)
            ws['AI{}'.format(r)] = "=IF($Y${}=0,$I${},$I${}-1)".format(r, r, r)
            ws['AI{}'.format(s)] = "=STDEV(AI2:AI{})".format(q)
            ws['AI{}'.format(t)] = "=MAX(AI2:AI{})".format(q)
            ws['AI{}'.format(u)] = "=MIN(AI2:AI{})".format(q)
            ws['AJ{}'.format(r)] = "=IF($Z${}=0,$J${},$J${}-1)".format(r, r, r)
            ws['AJ{}'.format(s)] = "=STDEV(AJ2:AJ{})".format(q)
            ws['AJ{}'.format(t)] = "=MAX(AJ2:AJ{})".format(q)
            ws['AJ{}'.format(u)] = "=MIN(AJ2:AJ{})".format(q)

            # JUMLAH MAPEL NORMAL
            ws['AK{}'.format(r)] = "=ROUND(AVERAGE(AK2:AK{}),2)".format(q)
            ws['AK{}'.format(t)] = "=MAX(AK2:AK{})".format(q)
            ws['AK{}'.format(u)] = "=MIN(AK2:AK{})".format(q)

            # Z SCORE
            ws['AL{}'.format(r)] = "=MAX(AL2:AL{})".format(q)
            ws['AM{}'.format(r)] = "=MAX(AM2:AM{})".format(q)
            ws['AN{}'.format(r)] = "=MAX(AN2:AN{})".format(q)
            ws['AO{}'.format(r)] = "=MAX(AO2:AO{})".format(q)

            # NILAI STANDAR
            ws['AP{}'.format(r)] = "=MAX(AP2:AP{})".format(q)
            ws['AP{}'.format(s)] = "=MIN(AP2:AP{})".format(q)
            ws['AP{}'.format(t)] = "=ROUND(AVERAGE(AP2:AP{}),2)".format(q)
            ws['AQ{}'.format(r)] = "=MAX(AQ2:AQ{})".format(q)
            ws['AQ{}'.format(s)] = "=MIN(AQ2:AQ{})".format(q)
            ws['AQ{}'.format(t)] = "=ROUND(AVERAGE(AQ2:AQ{}),2)".format(q)
            ws['AR{}'.format(r)] = "=MAX(AR2:AR{})".format(q)
            ws['AR{}'.format(s)] = "=MIN(AR2:AR{})".format(q)
            ws['AR{}'.format(t)] = "=ROUND(AVERAGE(AR2:AR{}),2)".format(q)
            ws['AS{}'.format(r)] = "=MAX(AS2:AS{})".format(q)
            ws['AS{}'.format(s)] = "=MIN(AS2:AS{})".format(q)
            ws['AS{}'.format(t)] = "=ROUND(AVERAGE(AS2:AS{}),2)".format(q)
            ws['AT{}'.format(r)] = "=MAX(AT2:AT{})".format(q)
            ws['AT{}'.format(s)] = "=MIN(AT2:AT{})".format(q)
            ws['AT{}'.format(t)] = "=ROUND(AVERAGE(AT2:AT{}),2)".format(q)

            # INISIASI MAPEL
            ws['AW{}'.format(r)] = "=SUM(AW2:AW{})".format(q)
            ws['AX{}'.format(r)] = "=SUM(AX2:AX{})".format(q)
            ws['AY{}'.format(r)] = "=SUM(AY2:AY{})".format(q)
            ws['AZ{}'.format(r)] = "=SUM(AZ2:AZ{})".format(q)

            # iterasi 2 rata-rata - 1
            # MAPEL NORMAL
            ws['BG{}'.format(
                r)] = "=IF($AW${}=0,$AG${},$AG${}-1)".format(r, r, r)
            ws['BG{}'.format(s)] = "=STDEV(BG2:BG{})".format(q)
            ws['BG{}'.format(t)] = "=MAX(BG2:BG{})".format(q)
            ws['BG{}'.format(u)] = "=MIN(BG2:BG{})".format(q)
            ws['BH{}'.format(
                r)] = "=IF($AX${}=0,$AH${},$AH${}-1)".format(r, r, r)
            ws['BH{}'.format(s)] = "=STDEV(BH2:BH{})".format(q)
            ws['BH{}'.format(t)] = "=MAX(BH2:BH{})".format(q)
            ws['BH{}'.format(u)] = "=MIN(BH2:BH{})".format(q)
            ws['BI{}'.format(
                r)] = "=IF($AY${}=0,$AI${},$AI${}-1)".format(r, r, r)
            ws['BI{}'.format(s)] = "=STDEV(BI2:BI{})".format(q)
            ws['BI{}'.format(t)] = "=MAX(BI2:BI{})".format(q)
            ws['BI{}'.format(u)] = "=MIN(BI2:BI{})".format(q)
            ws['BJ{}'.format(
                r)] = "=IF($AZ${}=0,$AJ${},$AJ${}-1)".format(r, r, r)
            ws['BJ{}'.format(s)] = "=STDEV(BJ2:BJ{})".format(q)
            ws['BJ{}'.format(t)] = "=MAX(BJ2:BJ{})".format(q)
            ws['BJ{}'.format(u)] = "=MIN(BJ2:BJ{})".format(q)

            # JUMLAH MAPEL NORMAL
            ws['BK{}'.format(r)] = "=ROUND(AVERAGE(BK2:BK{}),2)".format(q)
            ws['BK{}'.format(t)] = "=MAX(BK2:BK{})".format(q)
            ws['BK{}'.format(u)] = "=MIN(BK2:BK{})".format(q)

            # Z SCORE
            ws['BL{}'.format(r)] = "=MAX(BL2:BL{})".format(q)
            ws['BM{}'.format(r)] = "=MAX(BM2:BM{})".format(q)
            ws['BN{}'.format(r)] = "=MAX(BN2:BN{})".format(q)
            ws['BO{}'.format(r)] = "=MAX(BO2:BO{})".format(q)

            # NILAI STANDAR
            ws['BP{}'.format(r)] = "=MAX(BP2:BP{})".format(q)
            ws['BP{}'.format(s)] = "=MIN(BP2:BP{})".format(q)
            ws['BP{}'.format(t)] = "=ROUND(AVERAGE(BP2:BP{}),2)".format(q)
            ws['BQ{}'.format(r)] = "=MAX(BQ2:BQ{})".format(q)
            ws['BQ{}'.format(s)] = "=MIN(BQ2:BQ{})".format(q)
            ws['BQ{}'.format(t)] = "=ROUND(AVERAGE(BQ2:BQ{}),2)".format(q)
            ws['BR{}'.format(r)] = "=MAX(BR2:BR{})".format(q)
            ws['BR{}'.format(s)] = "=MIN(BR2:BR{})".format(q)
            ws['BR{}'.format(t)] = "=ROUND(AVERAGE(BR2:BR{}),2)".format(q)
            ws['BS{}'.format(r)] = "=MAX(BS2:BS{})".format(q)
            ws['BS{}'.format(s)] = "=MIN(BS2:BS{})".format(q)
            ws['BS{}'.format(t)] = "=ROUND(AVERAGE(BS2:BS{}),2)".format(q)
            ws['BT{}'.format(r)] = "=MAX(BT2:BT{})".format(q)
            ws['BT{}'.format(s)] = "=MIN(BT2:BT{})".format(q)
            ws['BT{}'.format(t)] = "=ROUND(AVERAGE(BT2:BT{}),2)".format(q)

            # INISIASI MAPEL
            ws['BW{}'.format(r)] = "=SUM(BW2:BW{})".format(q)
            ws['BX{}'.format(r)] = "=SUM(BX2:BX{})".format(q)
            ws['BY{}'.format(r)] = "=SUM(BY2:BY{})".format(q)
            ws['BZ{}'.format(r)] = "=SUM(BZ2:BZ{})".format(q)

            # iterasi 3 rata-rata - 1
            # MAPEL NORMAL
            ws['CG{}'.format(
                r)] = "=IF($BW${}=0,$BG${},$BG${}-1)".format(r, r, r)
            ws['CG{}'.format(s)] = "=STDEV(CG2:CG{})".format(q)
            ws['CG{}'.format(t)] = "=MAX(CG2:CG{})".format(q)
            ws['CG{}'.format(u)] = "=MIN(CG2:CG{})".format(q)
            ws['CH{}'.format(
                r)] = "=IF($BX${}=0,$BH${},$BH${}-1)".format(r, r, r)
            ws['CH{}'.format(s)] = "=STDEV(CH2:CH{})".format(q)
            ws['CH{}'.format(t)] = "=MAX(CH2:CH{})".format(q)
            ws['CH{}'.format(u)] = "=MIN(CH2:CH{})".format(q)
            ws['CI{}'.format(
                r)] = "=IF($BY${}=0,$BI${},$BI${}-1)".format(r, r, r)
            ws['CI{}'.format(s)] = "=STDEV(CI2:CI{})".format(q)
            ws['CI{}'.format(t)] = "=MAX(CI2:CI{})".format(q)
            ws['CI{}'.format(u)] = "=MIN(CI2:CI{})".format(q)
            ws['CJ{}'.format(
                r)] = "=IF($BZ${}=0,$BJ${},$BJ${}-1)".format(r, r, r)
            ws['CJ{}'.format(s)] = "=STDEV(CJ2:CJ{})".format(q)
            ws['CJ{}'.format(t)] = "=MAX(CJ2:CJ{})".format(q)
            ws['CJ{}'.format(u)] = "=MIN(CJ2:CJ{})".format(q)

            # JUMLAH MAPEL NORMAL
            ws['CK{}'.format(r)] = "=ROUND(AVERAGE(CK2:CK{}),2)".format(q)
            ws['CK{}'.format(t)] = "=MAX(CK2:CK{})".format(q)
            ws['CK{}'.format(u)] = "=MIN(CK2:CK{})".format(q)

            # Z SCORE
            ws['CL{}'.format(r)] = "=MAX(CL2:CL{})".format(q)
            ws['CM{}'.format(r)] = "=MAX(CM2:CM{})".format(q)
            ws['CN{}'.format(r)] = "=MAX(CN2:CN{})".format(q)
            ws['CO{}'.format(r)] = "=MAX(CO2:CO{})".format(q)

            # NILAI STANDAR
            ws['CP{}'.format(r)] = "=MAX(CP2:CP{})".format(q)
            ws['CP{}'.format(s)] = "=MIN(CP2:CP{})".format(q)
            ws['CP{}'.format(t)] = "=ROUND(AVERAGE(CP2:CP{}),2)".format(q)
            ws['CQ{}'.format(r)] = "=MAX(CQ2:CQ{})".format(q)
            ws['CQ{}'.format(s)] = "=MIN(CQ2:CQ{})".format(q)
            ws['CQ{}'.format(t)] = "=ROUND(AVERAGE(CQ2:CQ{}),2)".format(q)
            ws['CR{}'.format(r)] = "=MAX(CR2:CR{})".format(q)
            ws['CR{}'.format(s)] = "=MIN(CR2:CR{})".format(q)
            ws['CR{}'.format(t)] = "=ROUND(AVERAGE(CR2:CR{}),2)".format(q)
            ws['CS{}'.format(r)] = "=MAX(CS2:CS{})".format(q)
            ws['CS{}'.format(s)] = "=MIN(CS2:CS{})".format(q)
            ws['CS{}'.format(t)] = "=ROUND(AVERAGE(CS2:CS{}),2)".format(q)
            ws['CT{}'.format(r)] = "=MAX(CT2:CT{})".format(q)
            ws['CT{}'.format(s)] = "=MIN(CT2:CT{})".format(q)
            ws['CT{}'.format(t)] = "=ROUND(AVERAGE(CT2:CT{}),2)".format(q)

            # INISIASI MAPEL
            ws['CW{}'.format(r)] = "=SUM(CW2:CW{})".format(q)
            ws['CX{}'.format(r)] = "=SUM(CX2:CX{})".format(q)
            ws['CY{}'.format(r)] = "=SUM(CY2:CY{})".format(q)
            ws['CZ{}'.format(r)] = "=SUM(CZ2:CZ{})".format(q)

            # iterasi 4 rata-rata - 1
            # MAPEL NORMAL
            ws['DG{}'.format(
                r)] = "=IF($CW${}=0,$CG${},$CG${}-1)".format(r, r, r)
            ws['DG{}'.format(s)] = "=STDEV(DG2:DG{})".format(q)
            ws['DG{}'.format(t)] = "=MAX(DG2:DG{})".format(q)
            ws['DG{}'.format(u)] = "=MIN(DG2:DG{})".format(q)
            ws['DH{}'.format(
                r)] = "=IF($CX${}=0,$CH${},$CH${}-1)".format(r, r, r)
            ws['DH{}'.format(s)] = "=STDEV(DH2:DH{})".format(q)
            ws['DH{}'.format(t)] = "=MAX(DH2:DH{})".format(q)
            ws['DH{}'.format(u)] = "=MIN(DH2:DH{})".format(q)
            ws['DI{}'.format(
                r)] = "=IF($CY${}=0,$CI${},$CI${}-1)".format(r, r, r)
            ws['DI{}'.format(s)] = "=STDEV(DI2:DI{})".format(q)
            ws['DI{}'.format(t)] = "=MAX(DI2:DI{})".format(q)
            ws['DI{}'.format(u)] = "=MIN(DI2:DI{})".format(q)
            ws['DJ{}'.format(
                r)] = "=IF($CZ${}=0,$CJ${},$CJ${}-1)".format(r, r, r)
            ws['DJ{}'.format(s)] = "=STDEV(DJ2:DJ{})".format(q)
            ws['DJ{}'.format(t)] = "=MAX(DJ2:DJ{})".format(q)
            ws['DJ{}'.format(u)] = "=MIN(DJ2:DJ{})".format(q)

            # JUMLAH MAPEL NORMAL
            ws['DK{}'.format(r)] = "=ROUND(AVERAGE(DK2:DK{}),2)".format(q)
            ws['DK{}'.format(t)] = "=MAX(DK2:DK{})".format(q)
            ws['DK{}'.format(u)] = "=MIN(DK2:DK{})".format(q)

            # Z SCORE
            ws['DL{}'.format(r)] = "=MAX(DL2:DL{})".format(q)
            ws['DM{}'.format(r)] = "=MAX(DM2:DM{})".format(q)
            ws['DN{}'.format(r)] = "=MAX(DN2:DN{})".format(q)
            ws['DO{}'.format(r)] = "=MAX(DO2:DO{})".format(q)

            # NILAI STANDAR
            ws['DP{}'.format(r)] = "=MAX(DP2:DP{})".format(q)
            ws['DP{}'.format(s)] = "=MIN(DP2:DP{})".format(q)
            ws['DP{}'.format(t)] = "=ROUND(AVERAGE(DP2:DP{}),2)".format(q)
            ws['DQ{}'.format(r)] = "=MAX(DQ2:DQ{})".format(q)
            ws['DQ{}'.format(s)] = "=MIN(DQ2:DQ{})".format(q)
            ws['DQ{}'.format(t)] = "=ROUND(AVERAGE(DQ2:DQ{}),2)".format(q)
            ws['DR{}'.format(r)] = "=MAX(DR2:DR{})".format(q)
            ws['DR{}'.format(s)] = "=MIN(DR2:DR{})".format(q)
            ws['DR{}'.format(t)] = "=ROUND(AVERAGE(DR2:DR{}),2)".format(q)
            ws['DS{}'.format(r)] = "=MAX(DS2:DS{})".format(q)
            ws['DS{}'.format(s)] = "=MIN(DS2:DS{})".format(q)
            ws['DS{}'.format(t)] = "=ROUND(AVERAGE(DS2:DS{}),2)".format(q)
            ws['DT{}'.format(r)] = "=MAX(DT2:DT{})".format(q)
            ws['DT{}'.format(s)] = "=MIN(DT2:DT{})".format(q)
            ws['DT{}'.format(t)] = "=ROUND(AVERAGE(DT2:DT{}),2)".format(q)

            # INISIASI MAPEL
            ws['DW{}'.format(r)] = "=SUM(DW2:DW{})".format(q)
            ws['DX{}'.format(r)] = "=SUM(DX2:DX{})".format(q)
            ws['DY{}'.format(r)] = "=SUM(DY2:DY{})".format(q)
            ws['DZ{}'.format(r)] = "=SUM(DZ2:DZ{})".format(q)

            # iterasi 5 rata-rata - 1
            # MAPEL NORMAL
            ws['EG{}'.format(
                r)] = "=IF($DW${}=0,$DG${},$DG${}-1)".format(r, r, r)
            ws['EG{}'.format(s)] = "=STDEV(EG2:EG{})".format(q)
            ws['EG{}'.format(t)] = "=MAX(EG2:EG{})".format(q)
            ws['EG{}'.format(u)] = "=MIN(EG2:EG{})".format(q)
            ws['EH{}'.format(
                r)] = "=IF($DX${}=0,$DH${},$DH${}-1)".format(r, r, r)
            ws['EH{}'.format(s)] = "=STDEV(EH2:EH{})".format(q)
            ws['EH{}'.format(t)] = "=MAX(EH2:EH{})".format(q)
            ws['EH{}'.format(u)] = "=MIN(EH2:EH{})".format(q)
            ws['EI{}'.format(
                r)] = "=IF($DY${}=0,$DI${},$DI${}-1)".format(r, r, r)
            ws['EI{}'.format(s)] = "=STDEV(EI2:EI{})".format(q)
            ws['EI{}'.format(t)] = "=MAX(EI2:EI{})".format(q)
            ws['EI{}'.format(u)] = "=MIN(EI2:EI{})".format(q)
            ws['EJ{}'.format(
                r)] = "=IF($DZ${}=0,$DJ${},$DJ${}-1)".format(r, r, r)
            ws['EJ{}'.format(s)] = "=STDEV(EJ2:EJ{})".format(q)
            ws['EJ{}'.format(t)] = "=MAX(EJ2:EJ{})".format(q)
            ws['EJ{}'.format(u)] = "=MIN(EJ2:EJ{})".format(q)

            # JUMLAH MAPEL NORMAL
            ws['EK{}'.format(r)] = "=ROUND(AVERAGE(EK2:EK{}),2)".format(q)
            ws['EK{}'.format(t)] = "=MAX(EK2:EK{})".format(q)
            ws['EK{}'.format(u)] = "=MIN(EK2:EK{})".format(q)

            # Z SCORE
            ws['EL{}'.format(r)] = "=MAX(EL2:EL{})".format(q)
            ws['EM{}'.format(r)] = "=MAX(EM2:EM{})".format(q)
            ws['EN{}'.format(r)] = "=MAX(EN2:EN{})".format(q)
            ws['EO{}'.format(r)] = "=MAX(EO2:EO{})".format(q)

            # NILAI STANDAR
            ws['EP{}'.format(r)] = "=MAX(EP2:EP{})".format(q)
            ws['EP{}'.format(s)] = "=MIN(EP2:EP{})".format(q)
            ws['EP{}'.format(t)] = "=ROUND(AVERAGE(EP2:EP{}),2)".format(q)
            ws['EQ{}'.format(r)] = "=MAX(EQ2:EQ{})".format(q)
            ws['EQ{}'.format(s)] = "=MIN(EQ2:EQ{})".format(q)
            ws['EQ{}'.format(t)] = "=ROUND(AVERAGE(EQ2:EQ{}),2)".format(q)
            ws['ER{}'.format(r)] = "=MAX(ER2:ER{})".format(q)
            ws['ER{}'.format(s)] = "=MIN(ER2:ER{})".format(q)
            ws['ER{}'.format(t)] = "=ROUND(AVERAGE(ER2:ER{}),2)".format(q)
            ws['ES{}'.format(r)] = "=MAX(ES2:ES{})".format(q)
            ws['ES{}'.format(s)] = "=MIN(ES2:ES{})".format(q)
            ws['ES{}'.format(t)] = "=ROUND(AVERAGE(ES2:ES{}),2)".format(q)
            ws['ET{}'.format(r)] = "=MAX(ET2:ET{})".format(q)
            ws['ET{}'.format(s)] = "=MIN(ET2:ET{})".format(q)
            ws['ET{}'.format(t)] = "=ROUND(AVERAGE(ET2:ET{}),2)".format(q)

            # INISIASI MAPEL
            ws['EW{}'.format(r)] = "=SUM(EW2:EW{})".format(q)
            ws['EX{}'.format(r)] = "=SUM(EX2:EX{})".format(q)
            ws['EY{}'.format(r)] = "=SUM(EY2:EY{})".format(q)
            ws['EZ{}'.format(r)] = "=SUM(EZ2:EZ{})".format(q)

            # Jumlah Soal
            ws['F{}'.format(v)] = 'JUMLAH SOAL'
            ws['G{}'.format(v)] = JML_SOAL_MAT
            ws['H{}'.format(v)] = JML_SOAL_IND
            ws['I{}'.format(v)] = JML_SOAL_ENG
            ws['J{}'.format(v)] = JML_SOAL_IPAS

            # Z Score
            ws['B1'] = 'NAMA SISWA_A'
            ws['C1'] = 'NOMOR NF_A'
            ws['D1'] = 'KELAS_A'
            ws['E1'] = 'NAMA SEKOLAH_A'
            ws['F1'] = 'LOKASI_A'
            ws['G1'] = 'MAT_A'
            ws['H1'] = 'IND_A'
            ws['I1'] = 'ENG_A'
            ws['J1'] = 'IPAS_A'
            ws['K1'] = 'JML_A'
            ws['L1'] = 'Z_MAT_A'
            ws['M1'] = 'Z_IND_A'
            ws['N1'] = 'Z_ENG_A'
            ws['O1'] = 'Z_IPAS_A'
            ws['P1'] = 'S_MAT_A'
            ws['Q1'] = 'S_IND_A'
            ws['R1'] = 'S_ENG_A'
            ws['S1'] = 'S_IPAS_A'
            ws['T1'] = 'S_JML_A'
            ws['U1'] = 'RANK NAS._A'
            ws['V1'] = 'RANK LOK._A'

            ws['L1'].font = Font(bold=False, name='Calibri', size=11)
            ws['M1'].font = Font(bold=False, name='Calibri', size=11)
            ws['N1'].font = Font(bold=False, name='Calibri', size=11)
            ws['O1'].font = Font(bold=False, name='Calibri', size=11)
            ws['P1'].font = Font(bold=False, name='Calibri', size=11)
            ws['Q1'].font = Font(bold=False, name='Calibri', size=11)
            ws['R1'].font = Font(bold=False, name='Calibri', size=11)
            ws['S1'].font = Font(bold=False, name='Calibri', size=11)
            ws['T1'].font = Font(bold=False, name='Calibri', size=11)
            ws['U1'].font = Font(bold=False, name='Calibri', size=11)
            ws['V1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['B1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['C1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['D1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['E1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['F1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['G1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['H1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['I1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['J1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['K1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['L1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['M1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['N1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['O1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['P1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['Q1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['R1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['S1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['T1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['U1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['V1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')

            # tambahan
            ws['W1'] = 'MAT_20_A'
            ws['X1'] = 'IND_20_A'
            ws['Y1'] = 'ENG_20_A'
            ws['Z1'] = 'IPAS_20_A'
            ws['W1'].font = Font(bold=False, name='Calibri', size=11)
            ws['X1'].font = Font(bold=False, name='Calibri', size=11)
            ws['Y1'].font = Font(bold=False, name='Calibri', size=11)
            ws['Z1'].font = Font(bold=False, name='Calibri', size=11)
            ws['W1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['X1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['Y1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['Z1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')

            for row in range(2, q+1):
                ws['K{}'.format(
                    row)] = '=SUM(G{}:J{})'.format(row, row, row)
                ws['L{}'.format(
                    row)] = '=IFERROR(ROUND(IF(G{}="","",(G{}-G${})/G${}),2),"")'.format(row, row, r, s)
                ws['M{}'.format(
                    row)] = '=IFERROR(ROUND(IF(H{}="","",(H{}-H${})/H${}),2),"")'.format(row, row, r, s)
                ws['N{}'.format(
                    row)] = '=IFERROR(ROUND(IF(I{}="","",(I{}-I${})/I${}),2),"")'.format(row, row, r, s)
                ws['O{}'.format(
                    row)] = '=IFERROR(ROUND(IF(J{}="","",(J{}-J${})/J${}),2),"")'.format(row, row, r, s)
                ws['P{}'.format(
                    row)] = '=IFERROR(ROUND(IF(G{}="","",IF(70+30*L{}/$L${}<20,20,70+30*L{}/$L${})),2),"")'.format(row, row, r, row, r)
                ws['Q{}'.format(
                    row)] = '=IFERROR(ROUND(IF(H{}="","",IF(70+30*M{}/$M${}<20,20,70+30*M{}/$M${})),2),"")'.format(row, row, r, row, r)
                ws['R{}'.format(
                    row)] = '=IFERROR(ROUND(IF(I{}="","",IF(70+30*N{}/$N${}<20,20,70+30*N{}/$N${})),2),"")'.format(row, row, r, row, r)
                ws['S{}'.format(
                    row)] = '=IFERROR(ROUND(IF(J{}="","",IF(70+30*O{}/$O${}<20,20,70+30*O{}/$P${})),2),"")'.format(row, row, r, row, r)

                ws['T{}'.format(row)] = '=IF(SUM(P{}:S{})=0,"",SUM(P{}:S{}))'.format(
                    row, row, row, row)
                ws['U{}'.format(row)] = '=IF(T{}="","",RANK(T{},$T$2:$T${}))'.format(
                    row, row, q)
                ws['V{}'.format(
                    row)] = '=IF(U{}="","",COUNTIFS($F$2:$F${},F{},$U$2:$U${},"<"&U{})+1)'.format(row, q, row, q, row)
                # TAMBAHAN
                ws['W{}'.format(row)] = '=IF($G${}=25,IF(AND(G{}>4,P{}=20),1,""),IF($G${}=30,IF(AND(G{}>5,P{}=20),1,""),IF($G${}=35,IF(AND(G{}>6,P{}=20),1,""),IF($G${}=40,IF(AND(G{}>7,P{}=20),1,""),IF($G${}=45,IF(AND(G{}>8,P{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['X{}'.format(row)] = '=IF($H${}=25,IF(AND(H{}>4,Q{}=20),1,""),IF($H${}=30,IF(AND(H{}>5,Q{}=20),1,""),IF($H${}=35,IF(AND(H{}>6,Q{}=20),1,""),IF($H${}=40,IF(AND(H{}>7,Q{}=20),1,""),IF($H${}=45,IF(AND(H{}>8,Q{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['Y{}'.format(row)] = '=IF($I${}=25,IF(AND(I{}>4,R{}=20),1,""),IF($I${}=30,IF(AND(I{}>5,R{}=20),1,""),IF($I${}=35,IF(AND(I{}>6,R{}=20),1,""),IF($I${}=40,IF(AND(I{}>7,R{}=20),1,""),IF($I${}=45,IF(AND(I{}>8,R{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['Z{}'.format(row)] = '=IF($J${}=25,IF(AND(J{}>4,S{}=20),1,""),IF($J${}=30,IF(AND(J{}>5,S{}=20),1,""),IF($J${}=35,IF(AND(J{}>6,S{}=20),1,""),IF($J${}=40,IF(AND(J{}>7,S{}=20),1,""),IF($J${}=45,IF(AND(J{}>8,S{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

            # Z Score iterasi 1
            ws['AB1'] = 'NAMA SISWA_B'
            ws['AC1'] = 'NOMOR NF_B'
            ws['AD1'] = 'KELAS_B'
            ws['AE1'] = 'NAMA SEKOLAH_B'
            ws['AF1'] = 'LOKASI_B'
            ws['AG1'] = 'MAT_B'
            ws['AH1'] = 'IND_B'
            ws['AI1'] = 'ENG_B'
            ws['AJ1'] = 'IPAS_B'
            ws['AK1'] = 'JML_B'
            ws['AL1'] = 'Z_MAT_B'
            ws['AM1'] = 'Z_IND_B'
            ws['AN1'] = 'Z_ENG_B'
            ws['AO1'] = 'Z_IPAS_B'
            ws['AP1'] = 'S_MAT_B'
            ws['AQ1'] = 'S_IND_B'
            ws['AR1'] = 'S_ENG_B'
            ws['AS1'] = 'S_IPAS_B'
            ws['AT1'] = 'S_JML_B'
            ws['AU1'] = 'RANK NAS._B'
            ws['AV1'] = 'RANK LOK._B'

            ws['AL1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AM1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AN1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AO1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AP1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AQ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AR1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AS1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AT1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AU1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AV1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['AB1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AC1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AD1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AE1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AF1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AG1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AH1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AI1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AJ1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AK1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AL1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AM1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AN1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AO1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AP1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AQ1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AR1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AS1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AT1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AU1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AV1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')

            # tambahan
            ws['AW1'] = 'MAT_20'
            ws['AX1'] = 'IND_20'
            ws['AY1'] = 'ENG_20'
            ws['AZ1'] = 'IPAS_20'
            ws['AW1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AX1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AY1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AZ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AW1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AX1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AY1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AZ1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')

            for row in range(2, q+1):
                # Tambahan
                ws['AB{}'.format(row)] = '=B{}'.format(row)
                ws['AC{}'.format(row)] = '=C{}'.format(row, row)
                ws['AD{}'.format(row)] = '=D{}'.format(row, row)
                ws['AE{}'.format(row)] = '=E{}'.format(row, row)
                ws['AF{}'.format(row)] = '=F{}'.format(row, row)
                ws['AG{}'.format(row)] = '=IF(G{}="","",G{})'.format(row, row)
                ws['AH{}'.format(row)] = '=IF(H{}="","",H{})'.format(row, row)
                ws['AI{}'.format(row)] = '=IF(I{}="","",I{})'.format(row, row)
                ws['AJ{}'.format(row)] = '=IF(J{}="","",J{})'.format(row, row)
                ws['AK{}'.format(row)] = '=IF(K{}="","",K{})'.format(row, row)

                ws['AL{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AG{}="","",(AG{}-AG${})/AG${}),2),"")'.format(row, row, r, s)
                ws['AM{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AH{}="","",(AH{}-AH${})/AH${}),2),"")'.format(row, row, r, s)
                ws['AN{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AI{}="","",(AI{}-AI${})/AI${}),2),"")'.format(row, row, r, s)
                ws['AO{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AJ{}="","",(AJ{}-AJ${})/AJ${}),2),"")'.format(row, row, r, s)

                ws['AP{}'.format(
                    row)] = '=IFERROR(ROUND(IF(G{}="","",IF(70+30*AL{}/$AL${}<20,20,70+30*AL{}/$AL${})),2),"")'.format(row, row, r, row, r)
                ws['AQ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(H{}="","",IF(70+30*AM{}/$AM{}<20,20,70+30*AM{}/$AM${})),2),"")'.format(row, row, r, row, r)
                ws['AR{}'.format(
                    row)] = '=IFERROR(ROUND(IF(I{}="","",IF(70+30*AN{}/$AN${}<20,20,70+30*AN{}/$AN${})),2),"")'.format(row, row, r, row, r)
                ws['AS{}'.format(
                    row)] = '=IFERROR(ROUND(IF(J{}="","",IF(70+30*AO{}/$AO${}<20,20,70+30*AO{}/$AO${})),2),"")'.format(row, row, r, row, r)

                ws['AT{}'.format(row)] = '=IF(SUM(AP{}:AS{})=0,"",SUM(AP{}:AS{}))'.format(
                    row, row, row, row)
                ws['AU{}'.format(row)] = '=IF(AT{}="","",RANK(AT{},$AT$2:$AT${}))'.format(
                    row, row, q)
                ws['AV{}'.format(
                    row)] = '=IF(AU{}="","",COUNTIFS($AF$2:$AF${},AF{},$AU$2:$AU${},"<"&AU{})+1)'.format(row, q, row, q, row)
                # TAMBAHAN
                ws['AW{}'.format(row)] = '=IF($G${}=25,IF(AND(AG{}>4,AP{}=20),1,""),IF($G${}=30,IF(AND(AG{}>5,AP{}=20),1,""),IF($G${}=35,IF(AND(AG{}>6,AP{}=20),1,""),IF($G${}=40,IF(AND(AG{}>7,AP{}=20),1,""),IF($G${}=45,IF(AND(AG{}>8,AP{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['AX{}'.format(row)] = '=IF($H${}=25,IF(AND(AH{}>4,AQ{}=20),1,""),IF($H${}=30,IF(AND(AH{}>5,AQ{}=20),1,""),IF($H${}=35,IF(AND(AH{}>6,AQ{}=20),1,""),IF($H${}=40,IF(AND(AH{}>7,AQ{}=20),1,""),IF($H${}=45,IF(AND(AH{}>8,AQ{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['AY{}'.format(row)] = '=IF($I${}=25,IF(AND(AI{}>4,AR{}=20),1,""),IF($I${}=30,IF(AND(AI{}>5,AR{}=20),1,""),IF($I${}=35,IF(AND(AI{}>6,AR{}=20),1,""),IF($I${}=40,IF(AND(AI{}>7,AR{}=20),1,""),IF($I${}=45,IF(AND(AI{}>8,AR{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['AZ{}'.format(row)] = '=IF($J${}=25,IF(AND(AJ{}>4,AS{}=20),1,""),IF($J${}=30,IF(AND(AJ{}>5,AS{}=20),1,""),IF($J${}=35,IF(AND(AJ{}>6,AS{}=20),1,""),IF($J${}=40,IF(AND(AJ{}>7,AS{}=20),1,""),IF($J${}=45,IF(AND(AJ{}>8,AS{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

            # Z Score iterasi 2
            ws['BB1'] = 'NAMA SISWA_C'
            ws['BC1'] = 'NOMOR NF_c'
            ws['BD1'] = 'KELAS_C'
            ws['BE1'] = 'NAMA SEKOLAH_C'
            ws['BF1'] = 'LOKASI_C'
            ws['BG1'] = 'MAT_C'
            ws['BH1'] = 'IND_C'
            ws['BI1'] = 'ENG_C'
            ws['BJ1'] = 'IPAS_C'
            ws['BK1'] = 'JML_C'
            ws['BL1'] = 'Z_MAT_C'
            ws['BM1'] = 'Z_IND_C'
            ws['BN1'] = 'Z_ENG_C'
            ws['BO1'] = 'Z_IPAS_C'
            ws['BP1'] = 'S_MAT_C'
            ws['BQ1'] = 'S_IND_C'
            ws['BR1'] = 'S_ENG_C'
            ws['BS1'] = 'S_IPAS_C'
            ws['BT1'] = 'S_JML_C'
            ws['BU1'] = 'RANK NAS._C'
            ws['BV1'] = 'RANK LOK._C'

            ws['BL1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BM1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BN1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BO1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BP1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BQ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BR1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BS1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BT1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BU1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BV1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['BB1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BC1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BD1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BE1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BF1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BG1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BH1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BI1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BJ1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BK1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BL1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BM1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BN1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BO1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BP1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BQ1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BR1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BS1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BT1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BU1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BV1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')

            # tambahan
            ws['BW1'] = 'MAT_20_C'
            ws['BX1'] = 'IND_20_C'
            ws['BY1'] = 'ENG_20_C'
            ws['BZ1'] = 'IPAS_20_C'
            ws['BW1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BX1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BY1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BZ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BW1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BX1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BY1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BZ1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')

            for row in range(2, q+1):
                # Tambahan
                ws['BB{}'.format(row)] = '=AB{}'.format(row)
                ws['BC{}'.format(row)] = '=AC{}'.format(row, row)
                ws['BD{}'.format(row)] = '=AD{}'.format(row, row)
                ws['BE{}'.format(row)] = '=AE{}'.format(row, row)
                ws['BF{}'.format(row)] = '=AF{}'.format(row, row)
                ws['BG{}'.format(row)] = '=IF(AG{}="","",AG{})'.format(
                    row, row)
                ws['BH{}'.format(row)] = '=IF(AH{}="","",AH{})'.format(
                    row, row)
                ws['BI{}'.format(row)] = '=IF(AI{}="","",AI{})'.format(
                    row, row)
                ws['BJ{}'.format(row)] = '=IF(AJ{}="","",AJ{})'.format(
                    row, row)
                ws['BK{}'.format(row)] = '=IF(AK{}="","",AK{})'.format(
                    row, row)

                ws['BL{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BG{}="","",(BG{}-BG${})/BG${}),2),"")'.format(row, row, r, s)
                ws['BM{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BH{}="","",(BH{}-BH${})/BH${}),2),"")'.format(row, row, r, s)
                ws['BN{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BI{}="","",(BI{}-BI${})/BI${}),2),"")'.format(row, row, r, s)
                ws['BO{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BJ{}="","",(BJ{}-BJ${})/BJ${}),2),"")'.format(row, row, r, s)

                ws['BP{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BG{}="","",IF(70+30*BL{}/$BL${}<20,20,70+30*BL{}/$BL${})),2),"")'.format(row, row, r, row, r)
                ws['BQ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BH{}="","",IF(70+30*BM{}/$BM{}<20,20,70+30*BM{}/$BM${})),2),"")'.format(row, row, r, row, r)
                ws['BR{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BI{}="","",IF(70+30*BN{}/$BN${}<20,20,70+30*BN{}/$BN${})),2),"")'.format(row, row, r, row, r)
                ws['BS{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BJ{}="","",IF(70+30*BO{}/$BO${}<20,20,70+30*BO{}/$BO${})),2),"")'.format(row, row, r, row, r)

                ws['BT{}'.format(row)] = '=IF(SUM(BP{}:BS{})=0,"",SUM(BP{}:BS{}))'.format(
                    row, row, row, row)
                ws['BU{}'.format(row)] = '=IF(BT{}="","",RANK(BT{},$BT$2:$BT${}))'.format(
                    row, row, q)
                ws['BV{}'.format(
                    row)] = '=IF(BU{}="","",COUNTIFS($BF$2:$BF${},BF{},$BU$2:$BU${},"<"&BU{})+1)'.format(row, q, row, q, row)
                # TAMBAHAN
                ws['BW{}'.format(row)] = '=IF($G${}=25,IF(AND(BG{}>4,BP{}=20),1,""),IF($G${}=30,IF(AND(BG{}>5,BP{}=20),1,""),IF($G${}=35,IF(AND(BG{}>6,BP{}=20),1,""),IF($G${}=40,IF(AND(BG{}>7,BP{}=20),1,""),IF($G${}=45,IF(AND(BG{}>8,BP{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['BX{}'.format(row)] = '=IF($H${}=25,IF(AND(BH{}>4,BQ{}=20),1,""),IF($H${}=30,IF(AND(BH{}>5,BQ{}=20),1,""),IF($H${}=35,IF(AND(BH{}>6,BQ{}=20),1,""),IF($H${}=40,IF(AND(BH{}>7,BQ{}=20),1,""),IF($H${}=45,IF(AND(BH{}>8,BQ{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['BY{}'.format(row)] = '=IF($I${}=25,IF(AND(BI{}>4,BR{}=20),1,""),IF($I${}=30,IF(AND(BI{}>5,BR{}=20),1,""),IF($I${}=35,IF(AND(BI{}>6,BR{}=20),1,""),IF($I${}=40,IF(AND(BI{}>7,BR{}=20),1,""),IF($I${}=45,IF(AND(BI{}>8,BR{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['BZ{}'.format(row)] = '=IF($J${}=25,IF(AND(BJ{}>4,BS{}=20),1,""),IF($J${}=30,IF(AND(BJ{}>5,BS{}=20),1,""),IF($J${}=35,IF(AND(BJ{}>6,BS{}=20),1,""),IF($J${}=40,IF(AND(BJ{}>7,BS{}=20),1,""),IF($J${}=45,IF(AND(BJ{}>8,BS{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

            # Z Score iterasi 3
            ws['CB1'] = 'NAMA SISWA_D'
            ws['CC1'] = 'NOMOR NF_D'
            ws['CD1'] = 'KELAS_D'
            ws['CE1'] = 'NAMA SEKOLAH_D'
            ws['CF1'] = 'LOKASI_D'
            ws['CG1'] = 'MAT_D'
            ws['CH1'] = 'IND_D'
            ws['CI1'] = 'ENG_D'
            ws['CJ1'] = 'IPAS_D'
            ws['CK1'] = 'JML_D'
            ws['CL1'] = 'Z_MAT_D'
            ws['CM1'] = 'Z_IND_D'
            ws['CN1'] = 'Z_ENG_D'
            ws['CO1'] = 'Z_IPAS_D'
            ws['CP1'] = 'S_MAT_D'
            ws['CQ1'] = 'S_IND_D'
            ws['CR1'] = 'S_ENG_D'
            ws['CS1'] = 'S_IPAS_D'
            ws['CT1'] = 'S_JML_D'
            ws['CU1'] = 'RANK NAS._D'
            ws['CV1'] = 'RANK LOK._D'

            ws['CL1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CM1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CN1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CO1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CP1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CQ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CR1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CS1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CT1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CU1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CV1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['CB1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CC1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CD1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CE1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CF1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CG1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CH1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CI1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CJ1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CK1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CL1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CM1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CN1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CO1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CP1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CQ1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CR1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CS1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CT1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CU1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CV1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')

            # tambahan
            ws['CW1'] = 'MAT_20_D'
            ws['CX1'] = 'IND_20_D'
            ws['CY1'] = 'ENG_20_D'
            ws['CZ1'] = 'IPAS_20_D'
            ws['CW1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CX1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CY1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CZ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CW1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CX1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CY1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CZ1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')

            for row in range(2, q+1):
                ws['CB{}'.format(row)] = '=BB{}'.format(row)
                ws['CC{}'.format(row)] = '=BC{}'.format(row, row)
                ws['CD{}'.format(row)] = '=BD{}'.format(row, row)
                ws['CE{}'.format(row)] = '=BE{}'.format(row, row)
                ws['CF{}'.format(row)] = '=BF{}'.format(row, row)
                ws['CG{}'.format(row)] = '=IF(BG{}="","",BG{})'.format(
                    row, row)
                ws['CH{}'.format(row)] = '=IF(BH{}="","",BH{})'.format(
                    row, row)
                ws['CI{}'.format(row)] = '=IF(BI{}="","",BI{})'.format(
                    row, row)
                ws['CJ{}'.format(row)] = '=IF(BJ{}="","",BJ{})'.format(
                    row, row)
                ws['CK{}'.format(row)] = '=IF(BK{}="","",BK{})'.format(
                    row, row)

                ws['CL{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CG{}="","",(CG{}-CG${})/CG${}),2),"")'.format(row, row, r, s)
                ws['CM{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CH{}="","",(CH{}-CH${})/CH${}),2),"")'.format(row, row, r, s)
                ws['CN{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CI{}="","",(CI{}-CI${})/CI${}),2),"")'.format(row, row, r, s)
                ws['CO{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CJ{}="","",(CJ{}-CJ${})/CJ${}),2),"")'.format(row, row, r, s)

                ws['CP{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CG{}="","",IF(70+30*CL{}/$CL${}<20,20,70+30*CL{}/$CL${})),2),"")'.format(row, row, r, row, r)
                ws['CQ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CH{}="","",IF(70+30*CM{}/$CM{}<20,20,70+30*CM{}/$CM${})),2),"")'.format(row, row, r, row, r)
                ws['CR{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CI{}="","",IF(70+30*CN{}/$CN${}<20,20,70+30*CN{}/$CN${})),2),"")'.format(row, row, r, row, r)
                ws['CS{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CJ{}="","",IF(70+30*CO{}/$CO${}<20,20,70+30*CO{}/$CO${})),2),"")'.format(row, row, r, row, r)

                ws['CT{}'.format(row)] = '=IF(SUM(CP{}:CS{})=0,"",SUM(CP{}:CS{}))'.format(
                    row, row, row, row)
                ws['CU{}'.format(row)] = '=IF(CT{}="","",RANK(CT{},$CT$2:$CT${}))'.format(
                    row, row, q)
                ws['CV{}'.format(
                    row)] = '=IF(CU{}="","",COUNTIFS($CF$2:$CF${},CF{},$CU$2:$CU${},"<"&CU{})+1)'.format(row, q, row, q, row)
                # TAMBAHAN
                ws['CW{}'.format(row)] = '=IF($G${}=25,IF(AND(CG{}>4,CP{}=20),1,""),IF($G${}=30,IF(AND(CG{}>5,CP{}=20),1,""),IF($G${}=35,IF(AND(CG{}>6,CP{}=20),1,""),IF($G${}=40,IF(AND(CG{}>7,CP{}=20),1,""),IF($G${}=45,IF(AND(CG{}>8,CP{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['CX{}'.format(row)] = '=IF($H${}=25,IF(AND(CH{}>4,CQ{}=20),1,""),IF($H${}=30,IF(AND(CH{}>5,CQ{}=20),1,""),IF($H${}=35,IF(AND(CH{}>6,CQ{}=20),1,""),IF($H${}=40,IF(AND(CH{}>7,CQ{}=20),1,""),IF($H${}=45,IF(AND(CH{}>8,CQ{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['CY{}'.format(row)] = '=IF($I${}=25,IF(AND(CI{}>4,CR{}=20),1,""),IF($I${}=30,IF(AND(CI{}>5,CR{}=20),1,""),IF($I${}=35,IF(AND(CI{}>6,CR{}=20),1,""),IF($I${}=40,IF(AND(CI{}>7,CR{}=20),1,""),IF($I${}=45,IF(AND(CI{}>8,CR{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['CZ{}'.format(row)] = '=IF($J${}=25,IF(AND(CJ{}>4,CS{}=20),1,""),IF($J${}=30,IF(AND(CJ{}>5,CS{}=20),1,""),IF($J${}=35,IF(AND(CJ{}>6,CS{}=20),1,""),IF($J${}=40,IF(AND(CJ{}>7,CS{}=20),1,""),IF($J${}=45,IF(AND(CJ{}>8,CS{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

            # Z Score iterasi 4
            ws['DB1'] = 'NAMA SISWA_E'
            ws['DC1'] = 'NOMOR NF_E'
            ws['DD1'] = 'KELAS_E'
            ws['DE1'] = 'NAMA SEKOLAH_E'
            ws['DF1'] = 'LOKASI_E'
            ws['DG1'] = 'MAT_E'
            ws['DH1'] = 'IND_E'
            ws['DI1'] = 'ENG_E'
            ws['DJ1'] = 'IPAS_E'
            ws['DK1'] = 'JML_E'
            ws['DL1'] = 'Z_MAT_E'
            ws['DM1'] = 'Z_IND_E'
            ws['DN1'] = 'Z_ENG_E'
            ws['DO1'] = 'Z_IPAS_E'
            ws['DP1'] = 'S_MAT_E'
            ws['DQ1'] = 'S_IND_E'
            ws['DR1'] = 'S_ENG_E'
            ws['DS1'] = 'S_IPAS_E'
            ws['DT1'] = 'S_JML_E'
            ws['DU1'] = 'RANK NAS._E'
            ws['DV1'] = 'RANK LOK._E'

            ws['DL1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DM1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DN1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DO1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DP1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DQ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DR1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DS1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DT1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DU1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DV1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['DB1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DC1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DD1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DE1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DF1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DG1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DH1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DI1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DJ1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DK1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DL1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DM1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DN1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DO1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DP1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DQ1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DR1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DS1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DT1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DU1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DV1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')

            # tambahan
            ws['DW1'] = 'MAT_20'
            ws['DX1'] = 'IND_20'
            ws['DY1'] = 'ENG_20'
            ws['DZ1'] = 'IPAS_20'
            ws['DW1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DX1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DY1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DZ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DW1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DX1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DY1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DZ1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')

            for row in range(2, q+1):
                # Tambahan
                ws['DB{}'.format(row)] = '=CB{}'.format(row)
                ws['DC{}'.format(row)] = '=CC{}'.format(row, row)
                ws['DD{}'.format(row)] = '=CD{}'.format(row, row)
                ws['DE{}'.format(row)] = '=CE{}'.format(row, row)
                ws['DF{}'.format(row)] = '=CF{}'.format(row, row)
                ws['DG{}'.format(row)] = '=IF(CG{}="","",CG{})'.format(
                    row, row)
                ws['DH{}'.format(row)] = '=IF(CH{}="","",CH{})'.format(
                    row, row)
                ws['DI{}'.format(row)] = '=IF(CI{}="","",CI{})'.format(
                    row, row)
                ws['DJ{}'.format(row)] = '=IF(CJ{}="","",CJ{})'.format(
                    row, row)
                ws['DK{}'.format(row)] = '=IF(CK{}="","",CK{})'.format(
                    row, row)

                ws['DL{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DG{}="","",(DG{}-DG${})/DG${}),2),"")'.format(row, row, r, s)
                ws['DM{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DH{}="","",(DH{}-DH${})/DH${}),2),"")'.format(row, row, r, s)
                ws['DN{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DI{}="","",(DI{}-DI${})/DI${}),2),"")'.format(row, row, r, s)
                ws['DO{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DJ{}="","",(DJ{}-DJ${})/DJ${}),2),"")'.format(row, row, r, s)

                ws['DP{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DG{}="","",IF(70+30*DL{}/$DL${}<20,20,70+30*DL{}/$DL${})),2),"")'.format(row, row, r, row, r)
                ws['DQ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DH{}="","",IF(70+30*DM{}/$DM{}<20,20,70+30*DM{}/$DM${})),2),"")'.format(row, row, r, row, r)
                ws['DR{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DI{}="","",IF(70+30*DN{}/$DN${}<20,20,70+30*DN{}/$DN${})),2),"")'.format(row, row, r, row, r)
                ws['DS{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DJ{}="","",IF(70+30*DO{}/$DO${}<20,20,70+30*DO{}/$DO${})),2),"")'.format(row, row, r, row, r)

                ws['DT{}'.format(row)] = '=IF(SUM(DP{}:DS{})=0,"",SUM(DP{}:DS{}))'.format(
                    row, row, row, row)
                ws['DU{}'.format(row)] = '=IF(DT{}="","",RANK(DT{},$DT$2:$DT${}))'.format(
                    row, row, q)
                ws['DV{}'.format(
                    row)] = '=IF(DU{}="","",COUNTIFS($DF$2:$DF${},DF{},$DU$2:$DU${},"<"&DU{})+1)'.format(row, q, row, q, row)
                # TAMBAHAN
                ws['DW{}'.format(row)] = '=IF($G${}=25,IF(AND(DG{}>4,DP{}=20),1,""),IF($G${}=30,IF(AND(DG{}>5,DP{}=20),1,""),IF($G${}=35,IF(AND(DG{}>6,DP{}=20),1,""),IF($G${}=40,IF(AND(DG{}>7,DP{}=20),1,""),IF($G${}=45,IF(AND(DG{}>8,DP{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['DX{}'.format(row)] = '=IF($H${}=25,IF(AND(DH{}>4,DQ{}=20),1,""),IF($H${}=30,IF(AND(DH{}>5,DQ{}=20),1,""),IF($H${}=35,IF(AND(DH{}>6,DQ{}=20),1,""),IF($H${}=40,IF(AND(DH{}>7,DQ{}=20),1,""),IF($H${}=45,IF(AND(DH{}>8,DQ{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['DY{}'.format(row)] = '=IF($I${}=25,IF(AND(DI{}>4,DR{}=20),1,""),IF($I${}=30,IF(AND(DI{}>5,DR{}=20),1,""),IF($I${}=35,IF(AND(DI{}>6,DR{}=20),1,""),IF($I${}=40,IF(AND(DI{}>7,DR{}=20),1,""),IF($I${}=45,IF(AND(DI{}>8,DR{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['DZ{}'.format(row)] = '=IF($J${}=25,IF(AND(DJ{}>4,DS{}=20),1,""),IF($J${}=30,IF(AND(DJ{}>5,DS{}=20),1,""),IF($J${}=35,IF(AND(DJ{}>6,DS{}=20),1,""),IF($J${}=40,IF(AND(DJ{}>7,DS{}=20),1,""),IF($J${}=45,IF(AND(DJ{}>8,DS{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

            # Z Score iterasi 5
            ws['EB1'] = 'NAMA SISWA'
            ws['EC1'] = 'NOMOR NF'
            ws['ED1'] = 'KELAS'
            ws['EE1'] = 'NAMA SEKOLAH'
            ws['EF1'] = 'LOKASI'
            ws['EG1'] = 'MAT'
            ws['EH1'] = 'IND'
            ws['EI1'] = 'ENG'
            ws['EJ1'] = 'IPAS'
            ws['EK1'] = 'JML'
            ws['EL1'] = 'Z_MAT'
            ws['EM1'] = 'Z_IND'
            ws['EN1'] = 'Z_ENG'
            ws['EO1'] = 'Z_IPAS'
            ws['EP1'] = 'S_MAT'
            ws['EQ1'] = 'S_IND'
            ws['ER1'] = 'S_ENG'
            ws['ES1'] = 'S_IPAS'
            ws['ET1'] = 'S_JML'
            ws['EU1'] = 'RANK NAS.'
            ws['EV1'] = 'RANK LOK.'

            ws['EL1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EM1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EN1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EO1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EP1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EQ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['ER1'].font = Font(bold=False, name='Calibri', size=11)
            ws['ES1'].font = Font(bold=False, name='Calibri', size=11)
            ws['ET1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EU1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EV1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['EB1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EC1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['ED1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EE1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EF1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EG1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EH1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EI1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EJ1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EK1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EL1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EM1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EN1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EO1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EP1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EQ1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['ER1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['ES1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['ET1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EU1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EV1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')

            # tambahan
            ws['EW1'] = 'MAT_20'
            ws['EX1'] = 'IND_20'
            ws['EY1'] = 'ENG_20'
            ws['EZ1'] = 'IPAS_20'
            ws['EW1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EX1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EY1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EZ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EW1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EX1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EY1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EZ1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')

            for row in range(2, q+1):
                # Tambahan
                ws['EB{}'.format(row)] = '=DB{}'.format(row)
                ws['EC{}'.format(row)] = '=DC{}'.format(row, row)
                ws['ED{}'.format(row)] = '=DD{}'.format(row, row)
                ws['EE{}'.format(row)] = '=DE{}'.format(row, row)
                ws['EF{}'.format(row)] = '=DF{}'.format(row, row)
                ws['EG{}'.format(row)] = '=IF(DG{}="","",DG{})'.format(
                    row, row)
                ws['EH{}'.format(row)] = '=IF(DH{}="","",DH{})'.format(
                    row, row)
                ws['EI{}'.format(row)] = '=IF(DI{}="","",DI{})'.format(
                    row, row)
                ws['EJ{}'.format(row)] = '=IF(DJ{}="","",DJ{})'.format(
                    row, row)
                ws['EK{}'.format(row)] = '=IF(DK{}="","",DK{})'.format(
                    row, row)

                ws['EL{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EG{}="","",(EG{}-EG${})/EG${}),2),"")'.format(row, row, r, s)
                ws['EM{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EH{}="","",(EH{}-EH${})/EH${}),2),"")'.format(row, row, r, s)
                ws['EN{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EI{}="","",(EI{}-EI${})/EI${}),2),"")'.format(row, row, r, s)
                ws['EO{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EJ{}="","",(EJ{}-EJ${})/EJ${}),2),"")'.format(row, row, r, s)

                ws['EP{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EG{}="","",IF(70+30*EL{}/$EL${}<20,20,70+30*EL{}/$EL${})),2),"")'.format(row, row, r, row, r)
                ws['EQ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EH{}="","",IF(70+30*EM{}/$EM{}<20,20,70+30*EM{}/$EM${})),2),"")'.format(row, row, r, row, r)
                ws['ER{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EI{}="","",IF(70+30*EN{}/$EN${}<20,20,70+30*EN{}/$EN${})),2),"")'.format(row, row, r, row, r)
                ws['ES{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EJ{}="","",IF(70+30*EO{}/$EO${}<20,20,70+30*EO{}/$EO${})),2),"")'.format(row, row, r, row, r)

                ws['ET{}'.format(row)] = '=IF(SUM(EP{}:ES{})=0,"",SUM(EP{}:ES{}))'.format(
                    row, row, row, row)
                ws['EU{}'.format(row)] = '=IF(ET{}="","",RANK(ET{},$ET$2:$ET${}))'.format(
                    row, row, q)
                ws['EV{}'.format(
                    row)] = '=IF(EU{}="","",COUNTIFS($EF$2:$EF${},EF{},$EU$2:$EU${},"<"&EU{})+1)'.format(row, q, row, q, row)
                # TAMBAHAN
                ws['EW{}'.format(row)] = '=IF($G${}=25,IF(AND(EG{}>4,EP{}=20),1,""),IF($G${}=30,IF(AND(EG{}>5,EP{}=20),1,""),IF($G${}=35,IF(AND(EG{}>6,EP{}=20),1,""),IF($G${}=40,IF(AND(EG{}>7,EP{}=20),1,""),IF($G${}=45,IF(AND(EG{}>8,EP{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['EX{}'.format(row)] = '=IF($H${}=25,IF(AND(EH{}>4,EQ{}=20),1,""),IF($H${}=30,IF(AND(EH{}>5,EQ{}=20),1,""),IF($H${}=35,IF(AND(EH{}>6,EQ{}=20),1,""),IF($H${}=40,IF(AND(EH{}>7,EQ{}=20),1,""),IF($H${}=45,IF(AND(EH{}>8,EQ{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['EY{}'.format(row)] = '=IF($I${}=25,IF(AND(EI{}>4,ER{}=20),1,""),IF($I${}=30,IF(AND(EI{}>5,ER{}=20),1,""),IF($I${}=35,IF(AND(EI{}>6,ER{}=20),1,""),IF($I${}=40,IF(AND(EI{}>7,ER{}=20),1,""),IF($I${}=45,IF(AND(EI{}>8,ER{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['EZ{}'.format(row)] = '=IF($J${}=25,IF(AND(EJ{}>4,ES{}=20),1,""),IF($J${}=30,IF(AND(EJ{}>5,ES{}=20),1,""),IF($J${}=35,IF(AND(EJ{}>6,ES{}=20),1,""),IF($J${}=40,IF(AND(EJ{}>7,ES{}=20),1,""),IF($J${}=45,IF(AND(EJ{}>8,ES{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

            # Mengubah 'KELAS' sesuai dengan nilai yang dipilih dari selectbox 'KELAS'
            kelas = KELAS.lower().replace(" ", "")
            semester = SEMESTER.lower()
            tahun = TAHUN.replace("-", "")
            penilaian = PENILAIAN.lower()
            kurikulum = KURIKULUM.lower()

            path_file = f"{kelas}_{penilaian}_{semester}_{kurikulum}_{tahun}_nilai_std.xlsx"

            # Simpan file ke direktori temporer
            temp_dir = tempfile.gettempdir()
            file_path = temp_dir + '/' + path_file
            wb.save(file_path)

            st.success(
                "File siap diunduh!")

            # Tombol unduh file
            with open(file_path, "rb") as f:
                bytes_data = f.read()
            st.download_button(label="Unduh File", data=bytes_data,
                               file_name=path_file)

            st.warning(
                "Buka file unduhan, klik 'Enable Editing' dan 'Save'")
            
    if selected_file == "Nilai Std. 10, 11 IPS (K13)":
        # menghilangkan hamburger
        st.markdown("""
        <style>
        .css-1rs6os.edgvbvh3
        {
            visibility:hidden;
        }
        .css-1lsmgbg.egzxvld0
        {
            visibility:hidden;
        }
        </style>
        """, unsafe_allow_html=True)

        image = Image.open('logo resmi nf resize.png')
        st.image(image)

        st.title("Olah Nilai Standar K13")
        st.header("10-11 SMA IPS")

        col6 = st.container()

        with col6:
            KELAS = st.selectbox(
                "KELAS",
                ("--Pilih Kelas--", "10 IPS", "11 IPS"))

        col7 = st.container()

        with col7:
            SEMESTER = st.selectbox(
                "SEMESTER",
                ("--Pilih Semester--", "SEMESTER 1", "SEMESTER 2"))

        col8 = st.container()

        with col8:
            PENILAIAN = st.selectbox(
                "PENILAIAN",
                ("--Pilih Penilaian--", "PENILAIAN TENGAH SEMESTER", "PENILAIAN AKHIR SEMESTER", "PENILAIAN AKHIR TAHUN"))

        col9 = st.container()

        with col9:
            KURIKULUM = st.selectbox(
                "KURIKULUM",
                ("--Pilih Kurikulum--", "K13"))

        TAHUN = st.text_input("Masukkan Tahun Ajaran",
                              placeholder="contoh: 2022-2023")

        col1, col2, col3, col4, col5, col6, col7 = st.columns(7)

        with col1:
            MTK = st.selectbox(
                "JML. SOAL MAT.",
                (15, 20, 25, 30, 35, 40, 45, 50))

        with col2:
            IND = st.selectbox(
                "JML. SOAL IND.",
                (15, 20, 25, 30, 35, 40, 45, 50))

        with col3:
            ENG = st.selectbox(
                "JML. SOAL ENG.",
                (15, 20, 25, 30, 35, 40, 45, 50))

        with col4:
            SEJ = st.selectbox(
                "JML. SOAL SEJ.",
                (15, 20, 25, 30, 35, 40, 45, 50))

        with col5:
            GEO = st.selectbox(
                "JML. SOAL GEO.",
                (15, 20, 25, 30, 35, 40, 45, 50))

        with col6:
            EKO = st.selectbox(
                "JML. SOAL EKO.",
                (15, 20, 25, 30, 35, 40, 45, 50))

        with col7:
            SOS = st.selectbox(
                "JML. SOAL SOS.",
                (15, 20, 25, 30, 35, 40, 45, 50))

        JML_SOAL_MAT = MTK
        JML_SOAL_IND = IND
        JML_SOAL_ENG = ENG
        JML_SOAL_SEJ = SEJ
        JML_SOAL_GEO = GEO
        JML_SOAL_EKO = EKO
        JML_SOAL_SOS = SOS

        uploaded_file = st.file_uploader(
            'Letakkan file excel IPS', type='xlsx')

        if uploaded_file is not None:
            wb = openpyxl.load_workbook(uploaded_file)
            ws = wb['Sheet1']

            q = len(ws['K'])
            r = len(ws['K'])+2
            s = len(ws['K'])+3
            t = len(ws['K'])+4
            u = len(ws['K'])+5
            v = len(ws['K'])+6
            w = len(ws['K'])+7
            x = len(ws['K'])+8

            ws['G{}'.format(r)] = "=ROUND(AVERAGE(G2:G{}),2)".format(q)
            ws['H{}'.format(r)] = "=ROUND(AVERAGE(H2:H{}),2)".format(q)
            ws['I{}'.format(r)] = "=ROUND(AVERAGE(I2:I{}),2)".format(q)
            ws['J{}'.format(r)] = "=ROUND(AVERAGE(J2:J{}),2)".format(q)
            ws['K{}'.format(r)] = "=ROUND(AVERAGE(K2:K{}),2)".format(q)
            ws['L{}'.format(r)] = "=ROUND(AVERAGE(L2:L{}),2)".format(q)
            ws['M{}'.format(r)] = "=ROUND(AVERAGE(M2:M{}),2)".format(q)
            ws['N{}'.format(r)] = "=ROUND(AVERAGE(N2:N{}),2)".format(q)

            ws['G{}'.format(s)] = "=STDEV(G2:G{})".format(q)
            ws['H{}'.format(s)] = "=STDEV(H2:H{})".format(q)
            ws['I{}'.format(s)] = "=STDEV(I2:I{})".format(q)
            ws['J{}'.format(s)] = "=STDEV(J2:J{})".format(q)
            ws['K{}'.format(s)] = "=STDEV(K2:K{})".format(q)
            ws['L{}'.format(s)] = "=STDEV(L2:L{})".format(q)
            ws['M{}'.format(s)] = "=STDEV(M2:M{})".format(q)

            ws['G{}'.format(t)] = "=MAX(G2:G{})".format(q)
            ws['H{}'.format(t)] = "=MAX(H2:H{})".format(q)
            ws['I{}'.format(t)] = "=MAX(I2:I{})".format(q)
            ws['J{}'.format(t)] = "=MAX(J2:J{})".format(q)
            ws['K{}'.format(t)] = "=MAX(K2:K{})".format(q)
            ws['L{}'.format(t)] = "=MAX(L2:L{})".format(q)
            ws['M{}'.format(t)] = "=MAX(M2:M{})".format(q)

            ws['N{}'.format(r)] = "=MAX(N2:N{})".format(q)
            ws['O{}'.format(r)] = "=MAX(O2:O{})".format(q)
            ws['P{}'.format(r)] = "=MAX(P2:P{})".format(q)
            ws['Q{}'.format(r)] = "=MAX(Q2:Q{})".format(q)
            ws['R{}'.format(r)] = "=MAX(R2:R{})".format(q)
            ws['S{}'.format(r)] = "=MAX(S2:S{})".format(q)
            ws['T{}'.format(r)] = "=MAX(T2:T{})".format(q)
            ws['U{}'.format(r)] = "=MAX(U2:U{})".format(q)
            ws['V{}'.format(r)] = "=MAX(V2:V{})".format(q)
            ws['W{}'.format(r)] = "=MAX(W2:W{})".format(q)
            ws['X{}'.format(r)] = "=MAX(X2:X{})".format(q)
            ws['Y{}'.format(r)] = "=MAX(Y2:Y{})".format(q)
            ws['Z{}'.format(r)] = "=MAX(Z2:Z{})".format(q)
            ws['AA{}'.format(r)] = "=MAX(AA2:AA{})".format(q)

            ws['AB{}'.format(r)] = "=ROUND(MAX(AB2:AB{}),2)".format(q)
            ws['AC{}'.format(r)] = "=MAX(AC2:AC{})".format(q)
            ws['AD{}'.format(r)] = "=MAX(AD2:AD{})".format(q)

            ws['G{}'.format(u)] = "=MIN(G2:G{})".format(q)
            ws['H{}'.format(u)] = "=MIN(H2:H{})".format(q)
            ws['I{}'.format(u)] = "=MIN(I2:I{})".format(q)
            ws['J{}'.format(u)] = "=MIN(J2:J{})".format(q)
            ws['K{}'.format(u)] = "=MIN(K2:K{})".format(q)
            ws['L{}'.format(u)] = "=MIN(L2:L{})".format(q)
            ws['M{}'.format(u)] = "=MIN(M2:M{})".format(q)
            ws['N{}'.format(u)] = "=MIN(N2:N{})".format(q)

            ws['V{}'.format(s)] = "=MIN(V2:V{})".format(q)
            ws['W{}'.format(s)] = "=MIN(W2:W{})".format(q)
            ws['X{}'.format(s)] = "=MIN(X2:X{})".format(q)
            ws['Y{}'.format(s)] = "=MIN(Y2:Y{})".format(q)
            ws['Z{}'.format(s)] = "=MIN(Z2:Z{})".format(q)
            ws['AA{}'.format(s)] = "=MIN(AA2:AA{})".format(q)
            ws['AB{}'.format(s)] = "=MIN(AB2:AB{})".format(q)
            ws['AC{}'.format(s)] = "=MIN(AC2:AC{})".format(q)

            ws['V{}'.format(t)] = "=ROUND(AVERAGE(V2:V{}),2)".format(q)
            ws['W{}'.format(t)] = "=ROUND(AVERAGE(W2:W{}),2)".format(q)
            ws['X{}'.format(t)] = "=ROUND(AVERAGE(X2:X{}),2)".format(q)
            ws['Y{}'.format(t)] = "=ROUND(AVERAGE(Y2:Y{}),2)".format(q)
            ws['Z{}'.format(t)] = "=ROUND(AVERAGE(Z2:Z{}),2)".format(q)
            ws['AA{}'.format(t)] = "=ROUND(AVERAGE(AA2:AA{}),2)".format(q)
            ws['AB{}'.format(t)] = "=ROUND(AVERAGE(AB2:AB{}),2)".format(q)
            ws['AC{}'.format(t)] = "=ROUND(AVERAGE(AC2:AC{}),2)".format(q)

            ws['AF{}'.format(r)] = "=SUM(AF2:AF{})".format(q)
            ws['AG{}'.format(r)] = "=SUM(AG2:AG{})".format(q)
            ws['AH{}'.format(r)] = "=SUM(AH2:AH{})".format(q)
            ws['AI{}'.format(r)] = "=SUM(AI2:AI{})".format(q)
            ws['AJ{}'.format(r)] = "=SUM(AJ2:AJ{})".format(q)
            ws['AK{}'.format(r)] = "=SUM(AK2:AK{})".format(q)
            ws['AL{}'.format(r)] = "=SUM(AL2:AL{})".format(q)

            # Jumlah Soal
            ws['F{}'.format(v)] = 'JUMLAH SOAL'
            ws['G{}'.format(v)] = JML_SOAL_MAT
            ws['H{}'.format(v)] = JML_SOAL_IND
            ws['I{}'.format(v)] = JML_SOAL_ENG
            ws['J{}'.format(v)] = JML_SOAL_SEJ
            ws['K{}'.format(v)] = JML_SOAL_GEO
            ws['L{}'.format(v)] = JML_SOAL_EKO
            ws['M{}'.format(v)] = JML_SOAL_SOS

            # new
            # iterasi 1 rata-rata - 1
            # rata" MTK ke MTK tambahan dan mapel MTK awal
            ws['AS{}'.format(
                r)] = "=IF($AF${}=0,$G${},$G${}-1)".format(r, r, r)
            ws['AS{}'.format(s)] = "=STDEV(AS2:AS{})".format(q)
            ws['AS{}'.format(t)] = "=MAX(AS2:AS{})".format(q)
            ws['AS{}'.format(u)] = "=MIN(AS2:AS{})".format(q)
            # rata" IND ke IND tambahan dan mapel IND awal
            ws['AT{}'.format(
                r)] = "=IF($AG${}=0,$H${},$H${}-1)".format(r, r, r)
            ws['AT{}'.format(s)] = "=STDEV(AT2:AT{})".format(q)
            ws['AT{}'.format(t)] = "=MAX(AT2:AT{})".format(q)
            ws['AT{}'.format(u)] = "=MIN(AT2:AT{})".format(q)
            # rata" ENG ke ENG tambahan dan mapel ENG awal
            ws['AU{}'.format(
                r)] = "=IF($AH${}=0,$I${},$I${}-1)".format(r, r, r)
            ws['AU{}'.format(s)] = "=STDEV(AU2:AU{})".format(q)
            ws['AU{}'.format(t)] = "=MAX(AU2:AU{})".format(q)
            ws['AU{}'.format(u)] = "=MIN(AU2:AU{})".format(q)
            # rata" SEJ ke SEJ tambahan dan mapel SEJ awal
            ws['AV{}'.format(
                r)] = "=IF($AI${}=0,$J${},$J${}-1)".format(r, r, r)
            ws['AV{}'.format(s)] = "=STDEV(AV2:AV{})".format(q)
            ws['AV{}'.format(t)] = "=MAX(AV2:AV{})".format(q)
            ws['AV{}'.format(u)] = "=MIN(AV2:AV{})".format(q)
            # rata" GEO ke GEO tambahan dan mapel GEO awal
            ws['AW{}'.format(
                r)] = "=IF($AJ${}=0,$K${},$K${}-1)".format(r, r, r)
            ws['AW{}'.format(s)] = "=STDEV(AW2:AW{})".format(q)
            ws['AW{}'.format(t)] = "=MAX(AW2:AW{})".format(q)
            ws['AW{}'.format(u)] = "=MIN(AW2:AW{})".format(q)
            # rata" EKO ke EKO tambahan dan mapel EKO awal
            ws['AX{}'.format(
                r)] = "=IF($AK${}=0,$L${},$L${}-1)".format(r, r, r)
            ws['AX{}'.format(s)] = "=STDEV(AX2:AX{})".format(q)
            ws['AX{}'.format(t)] = "=MAX(AX2:AX{})".format(q)
            ws['AX{}'.format(u)] = "=MIN(AX2:AX{})".format(q)
            # rata" SOS ke SOS tambahan dan mapel SOS awal
            ws['AY{}'.format(
                r)] = "=IF($AL${}=0,$M${},$M${}-1)".format(r, r, r)
            ws['AY{}'.format(s)] = "=STDEV(AY2:AY{})".format(q)
            ws['AY{}'.format(t)] = "=MAX(AY2:AY{})".format(q)
            ws['AY{}'.format(u)] = "=MIN(AY2:AY{})".format(q)
            # jml MAPEL
            ws['AZ{}'.format(r)] = "=ROUND(AVERAGE(AZ2:AZ{}),2)".format(q)
            ws['AZ{}'.format(t)] = "=MAX(AZ2:AZ{})".format(q)
            ws['AZ{}'.format(u)] = "=MIN(AZ2:AZ{})".format(q)
            # MAX Z SCORE
            ws['BA{}'.format(r)] = "=MAX(BA2:BA{})".format(q)
            ws['BB{}'.format(r)] = "=MAX(BB2:BB{})".format(q)
            ws['BC{}'.format(r)] = "=MAX(BC2:BC{})".format(q)
            ws['BD{}'.format(r)] = "=MAX(BD2:BD{})".format(q)
            ws['BE{}'.format(r)] = "=MAX(BE2:BE{})".format(q)
            ws['BF{}'.format(r)] = "=MAX(BF2:BF{})".format(q)
            ws['BG{}'.format(r)] = "=MAX(BG2:BG{})".format(q)
            # NILAI STANDAR MTK
            ws['BH{}'.format(r)] = "=MAX(BH2:BH{})".format(q)
            ws['BH{}'.format(s)] = "=MIN(BH2:BH{})".format(q)
            ws['BH{}'.format(t)] = "=ROUND(AVERAGE(BH2:BH{}),2)".format(q)
            # NILAI STANDAR IND
            ws['BI{}'.format(r)] = "=MAX(BI2:BI{})".format(q)
            ws['BI{}'.format(s)] = "=MIN(BI2:BI{})".format(q)
            ws['BI{}'.format(t)] = "=ROUND(AVERAGE(BI2:BI{}),2)".format(q)
            # NILAI STANDAR ENG
            ws['BJ{}'.format(r)] = "=MAX(BJ2:BJ{})".format(q)
            ws['BJ{}'.format(s)] = "=MIN(BJ2:BJ{})".format(q)
            ws['BJ{}'.format(t)] = "=ROUND(AVERAGE(BJ2:BJ{}),2)".format(q)
            # NILAI STANDAR SEJ
            ws['BK{}'.format(r)] = "=MAX(BK2:BK{})".format(q)
            ws['BK{}'.format(s)] = "=MIN(BK2:BK{})".format(q)
            ws['BK{}'.format(t)] = "=ROUND(AVERAGE(BK2:BK{}),2)".format(q)
            # NILAI STANDAR GEO
            ws['BL{}'.format(r)] = "=MAX(BL2:BL{})".format(q)
            ws['BL{}'.format(s)] = "=MIN(BL2:BL{})".format(q)
            ws['BL{}'.format(t)] = "=ROUND(AVERAGE(BL2:BL{}),2)".format(q)
            # NILAI STANDAR EKO
            ws['BM{}'.format(r)] = "=MAX(BM2:BM{})".format(q)
            ws['BM{}'.format(s)] = "=MIN(BM2:BM{})".format(q)
            ws['BM{}'.format(t)] = "=ROUND(AVERAGE(BM2:BM{}),2)".format(q)
            # NILAI STANDAR SOS
            ws['BN{}'.format(r)] = "=MAX(BN2:BN{})".format(q)
            ws['BN{}'.format(s)] = "=MIN(BN2:BN{})".format(q)
            ws['BN{}'.format(t)] = "=ROUND(AVERAGE(BN2:BN{}),2)".format(q)
            # NILAI STANDAR S.JML
            ws['BO{}'.format(r)] = "=MAX(BO2:BO{})".format(q)
            ws['BO{}'.format(s)] = "=MIN(BO2:BO{})".format(q)
            ws['BO{}'.format(t)] = "=ROUND(AVERAGE(BO2:BO{}),2)".format(q)

            # TAMBAHAN
            ws['BR{}'.format(r)] = "=SUM(BR2:BR{})".format(q)
            ws['BS{}'.format(r)] = "=SUM(BS2:BS{})".format(q)
            ws['BT{}'.format(r)] = "=SUM(BT2:BT{})".format(q)
            ws['BU{}'.format(r)] = "=SUM(BU2:BU{})".format(q)
            ws['BV{}'.format(r)] = "=SUM(BV2:BV{})".format(q)
            ws['BW{}'.format(r)] = "=SUM(BW2:BW{})".format(q)
            ws['BX{}'.format(r)] = "=SUM(BX2:BX{})".format(q)

            # iterasi 2 rata-rata - 2
            # rata" MTK ke MTK tambahan dan mapel MTK awal
            ws['CE{}'.format(
                r)] = "=IF($BR${}=0,$AS${},$AS${}-1)".format(r, r, r)
            ws['CE{}'.format(s)] = "=STDEV(CE2:CE{})".format(q)
            ws['CE{}'.format(t)] = "=MAX(CE2:CE{})".format(q)
            ws['CE{}'.format(u)] = "=MIN(CE2:CE{})".format(q)
            # rata" IND ke IND tambahan dan mapel IND awal
            ws['CF{}'.format(
                r)] = "=IF($BS${}=0,$AT${},$AT${}-1)".format(r, r, r)
            ws['CF{}'.format(s)] = "=STDEV(CF2:CF{})".format(q)
            ws['CF{}'.format(t)] = "=MAX(CF2:CF{})".format(q)
            ws['CF{}'.format(u)] = "=MIN(CF2:CF{})".format(q)
            # rata" ENG ke ENG tambahan dan mapel ENG awal
            ws['CG{}'.format(
                r)] = "=IF($BT${}=0,$AU${},$AU${}-1)".format(r, r, r)
            ws['CG{}'.format(s)] = "=STDEV(CG2:CG{})".format(q)
            ws['CG{}'.format(t)] = "=MAX(CG2:CG{})".format(q)
            ws['CG{}'.format(u)] = "=MIN(CG2:CG{})".format(q)
            # rata" SEJ ke SEJ tambahan dan mapel SEJ awal
            ws['CH{}'.format(
                r)] = "=IF($BU${}=0,$AV${},$AV${}-1)".format(r, r, r)
            ws['CH{}'.format(s)] = "=STDEV(CH2:CH{})".format(q)
            ws['CH{}'.format(t)] = "=MAX(CH2:CH{})".format(q)
            ws['CH{}'.format(u)] = "=MIN(CH2:CH{})".format(q)
            # rata" GEO ke GEO tambahan dan mapel GEO awal
            ws['CI{}'.format(
                r)] = "=IF($BV${}=0,$AW${},$AW${}-1)".format(r, r, r)
            ws['CI{}'.format(s)] = "=STDEV(CI2:CI{})".format(q)
            ws['CI{}'.format(t)] = "=MAX(CI2:CI{})".format(q)
            ws['CI{}'.format(u)] = "=MIN(CI2:CI{})".format(q)
            # rata" EKO ke EKO tambahan dan mapel EKO awal
            ws['CJ{}'.format(
                r)] = "=IF($BW${}=0,$AX${},$AX${}-1)".format(r, r, r)
            ws['CJ{}'.format(s)] = "=STDEV(CJ2:CJ{})".format(q)
            ws['CJ{}'.format(t)] = "=MAX(CJ2:CJ{})".format(q)
            ws['CJ{}'.format(u)] = "=MIN(CJ2:CJ{})".format(q)
            # rata" SOS ke SOS tambahan dan mapel SOS awal
            ws['CK{}'.format(
                r)] = "=IF($BX${}=0,$AY${},$AY${}-1)".format(r, r, r)
            ws['CK{}'.format(s)] = "=STDEV(CK2:CK{})".format(q)
            ws['CK{}'.format(t)] = "=MAX(CK2:CK{})".format(q)
            ws['CK{}'.format(u)] = "=MIN(CK2:CK{})".format(q)
            # jml MAPEL
            ws['CL{}'.format(r)] = "=ROUND(AVERAGE(CL2:CL{}),2)".format(q)
            ws['CL{}'.format(t)] = "=MAX(CL2:CL{})".format(q)
            ws['CL{}'.format(u)] = "=MIN(CL2:CL{})".format(q)
            # MAX Z SCORE
            ws['CM{}'.format(r)] = "=MAX(CM2:CM{})".format(q)
            ws['CN{}'.format(r)] = "=MAX(CN2:CN{})".format(q)
            ws['CO{}'.format(r)] = "=MAX(CO2:CO{})".format(q)
            ws['CP{}'.format(r)] = "=MAX(CP2:CP{})".format(q)
            ws['CQ{}'.format(r)] = "=MAX(CQ2:CQ{})".format(q)
            ws['CR{}'.format(r)] = "=MAX(CR2:CR{})".format(q)
            ws['CS{}'.format(r)] = "=MAX(CS2:CS{})".format(q)
            # NILAI STANDAR MTK
            ws['CT{}'.format(r)] = "=MAX(CT2:CT{})".format(q)
            ws['CT{}'.format(s)] = "=MIN(CT2:CT{})".format(q)
            ws['CT{}'.format(t)] = "=ROUND(AVERAGE(CT2:CT{}),2)".format(q)
            # NILAI STANDAR IND
            ws['CU{}'.format(r)] = "=MAX(CU2:CU{})".format(q)
            ws['CU{}'.format(s)] = "=MIN(CU2:CU{})".format(q)
            ws['CU{}'.format(t)] = "=ROUND(AVERAGE(CU2:CU{}),2)".format(q)
            # NILAI STANDAR ENG
            ws['CV{}'.format(r)] = "=MAX(CV2:CV{})".format(q)
            ws['CV{}'.format(s)] = "=MIN(CV2:CV{})".format(q)
            ws['CV{}'.format(t)] = "=ROUND(AVERAGE(CV2:CV{}),2)".format(q)
            # NILAI STANDAR SEJ
            ws['CW{}'.format(r)] = "=MAX(CW2:CW{})".format(q)
            ws['CW{}'.format(s)] = "=MIN(CW2:CW{})".format(q)
            ws['CW{}'.format(t)] = "=ROUND(AVERAGE(CW2:CW{}),2)".format(q)
            # NILAI STANDAR GEO
            ws['CX{}'.format(r)] = "=MAX(CX2:CX{})".format(q)
            ws['CX{}'.format(s)] = "=MIN(CX2:CX{})".format(q)
            ws['CX{}'.format(t)] = "=ROUND(AVERAGE(CX2:CX{}),2)".format(q)
            # NILAI STANDAR EKO
            ws['CY{}'.format(r)] = "=MAX(CY2:CY{})".format(q)
            ws['CY{}'.format(s)] = "=MIN(CY2:CY{})".format(q)
            ws['CY{}'.format(t)] = "=ROUND(AVERAGE(CY2:CY{}),2)".format(q)
            # NILAI STANDAR SOS
            ws['CZ{}'.format(r)] = "=MAX(CZ2:CZ{})".format(q)
            ws['CZ{}'.format(s)] = "=MIN(CZ2:CZ{})".format(q)
            ws['CZ{}'.format(t)] = "=ROUND(AVERAGE(CZ2:CZ{}),2)".format(q)
            # NILAI STANDAR S.JML
            ws['DA{}'.format(r)] = "=MAX(DA2:DA{})".format(q)
            ws['DA{}'.format(s)] = "=MIN(DA2:DA{})".format(q)
            ws['DA{}'.format(t)] = "=ROUND(AVERAGE(DA2:DA{}),2)".format(q)

            # TAMBAHAN
            ws['DD{}'.format(r)] = "=SUM(DD2:DD{})".format(q)
            ws['DE{}'.format(r)] = "=SUM(DE2:DE{})".format(q)
            ws['DF{}'.format(r)] = "=SUM(DF2:DF{})".format(q)
            ws['DG{}'.format(r)] = "=SUM(DG2:DG{})".format(q)
            ws['DH{}'.format(r)] = "=SUM(DH2:DH{})".format(q)
            ws['DI{}'.format(r)] = "=SUM(DI2:DI{})".format(q)
            ws['DJ{}'.format(r)] = "=SUM(DJ2:DJ{})".format(q)

            # iterasi 3 rata-rata - 3
            # rata" MTK ke MTK tambahan dan mapel MTK awal
            ws['DQ{}'.format(
                r)] = "=IF($DD${}=0,$CE${},$CE${}-1)".format(r, r, r)
            ws['DQ{}'.format(s)] = "=STDEV(DQ2:DQ{})".format(q)
            ws['DQ{}'.format(t)] = "=MAX(DQ2:DQ{})".format(q)
            ws['DQ{}'.format(u)] = "=MIN(DQ2:DQ{})".format(q)
            # rata" IND ke IND tambahan dan mapel IND awal
            ws['DR{}'.format(
                r)] = "=IF($DE${}=0,$CF${},$CF${}-1)".format(r, r, r)
            ws['DR{}'.format(s)] = "=STDEV(DR2:DR{})".format(q)
            ws['DR{}'.format(t)] = "=MAX(DR2:DR{})".format(q)
            ws['DR{}'.format(u)] = "=MIN(DR2:DR{})".format(q)
            # rata" ENG ke ENG tambahan dan mapel ENG awal
            ws['DS{}'.format(
                r)] = "=IF($DF${}=0,$CG${},$CG{}-1)".format(r, r, r)
            ws['DS{}'.format(s)] = "=STDEV(DS2:DS{})".format(q)
            ws['DS{}'.format(t)] = "=MAX(DS2:DS{})".format(q)
            ws['DS{}'.format(u)] = "=MIN(DS2:DS{})".format(q)
            # rata" SEJ ke SEJ tambahan dan mapel SEJ awal
            ws['DT{}'.format(
                r)] = "=IF($DG${}=0,$CH${},$CH${}-1)".format(r, r, r)
            ws['DT{}'.format(s)] = "=STDEV(DT2:DT{})".format(q)
            ws['DT{}'.format(t)] = "=MAX(DT2:DT{})".format(q)
            ws['DT{}'.format(u)] = "=MIN(DT2:DT{})".format(q)
            # rata" GEO ke GEO tambahan dan mapel GEO awal
            ws['DU{}'.format(
                r)] = "=IF($DH${}=0,$CI${},$CI${}-1)".format(r, r, r)
            ws['DU{}'.format(s)] = "=STDEV(DU2:DU{})".format(q)
            ws['DU{}'.format(t)] = "=MAX(DU2:DU{})".format(q)
            ws['DU{}'.format(u)] = "=MIN(DU2:DU{})".format(q)
            # rata" EKO ke EKO tambahan dan mapel EKO awal
            ws['DV{}'.format(
                r)] = "=IF($DI${}=0,$CJ${},$CJ${}-1)".format(r, r, r)
            ws['DV{}'.format(s)] = "=STDEV(DV2:DV{})".format(q)
            ws['DV{}'.format(t)] = "=MAX(DV2:DV{})".format(q)
            ws['DV{}'.format(u)] = "=MIN(DV2:DV{})".format(q)
            # rata" SOS ke SOS tambahan dan mapel SOS awal
            ws['DW{}'.format(
                r)] = "=IF($DJ${}=0,$CK${},$CK${}-1)".format(r, r, r)
            ws['DW{}'.format(s)] = "=STDEV(DW2:DW{})".format(q)
            ws['DW{}'.format(t)] = "=MAX(DW2:DW{})".format(q)
            ws['DW{}'.format(u)] = "=MIN(DW2:DW{})".format(q)
            # jml MAPEL
            ws['DX{}'.format(r)] = "=ROUND(AVERAGE(DX2:DX{}),2)".format(q)
            ws['DX{}'.format(t)] = "=MAX(DX2:DX{})".format(q)
            ws['DX{}'.format(u)] = "=MIN(DX2:DX{})".format(q)
            # MAX Z SCORE
            ws['DY{}'.format(r)] = "=MAX(DY2:DY{})".format(q)
            ws['DZ{}'.format(r)] = "=MAX(DZ2:DZ{})".format(q)
            ws['EA{}'.format(r)] = "=MAX(EA2:EA{})".format(q)
            ws['EB{}'.format(r)] = "=MAX(EB2:EB{})".format(q)
            ws['EC{}'.format(r)] = "=MAX(EC2:EC{})".format(q)
            ws['ED{}'.format(r)] = "=MAX(ED2:ED{})".format(q)
            ws['EE{}'.format(r)] = "=MAX(EE2:EE{})".format(q)
            # NILAI STANDAR MTK
            ws['EF{}'.format(r)] = "=MAX(EF2:EF{})".format(q)
            ws['EF{}'.format(s)] = "=MIN(EF2:EF{})".format(q)
            ws['EF{}'.format(t)] = "=ROUND(AVERAGE(EF2:EF{}),2)".format(q)
            # NILAI STANDAR IND
            ws['EG{}'.format(r)] = "=MAX(EG2:EG{})".format(q)
            ws['EG{}'.format(s)] = "=MIN(EG2:EG{})".format(q)
            ws['EG{}'.format(t)] = "=ROUND(AVERAGE(EG2:EG{}),2)".format(q)
            # NILAI STANDAR ENG
            ws['EH{}'.format(r)] = "=MAX(EH2:EH{})".format(q)
            ws['EH{}'.format(s)] = "=MIN(EH2:EH{})".format(q)
            ws['EH{}'.format(t)] = "=ROUND(AVERAGE(EH2:EH{}),2)".format(q)
            # NILAI STANDAR SEJ
            ws['EI{}'.format(r)] = "=MAX(EI2:EI{})".format(q)
            ws['EI{}'.format(s)] = "=MIN(EI2:EI{})".format(q)
            ws['EI{}'.format(t)] = "=ROUND(AVERAGE(EI2:EI{}),2)".format(q)
            # NILAI STANDAR GEO
            ws['EJ{}'.format(r)] = "=MAX(EJ2:EJ{})".format(q)
            ws['EJ{}'.format(s)] = "=MIN(EJ2:EJ{})".format(q)
            ws['EJ{}'.format(t)] = "=ROUND(AVERAGE(EJ2:EJ{}),2)".format(q)
            # NILAI STANDAR EKO
            ws['EK{}'.format(r)] = "=MAX(EK2:EK{})".format(q)
            ws['EK{}'.format(s)] = "=MIN(EK2:EK{})".format(q)
            ws['EK{}'.format(t)] = "=ROUND(AVERAGE(EK2:EK{}),2)".format(q)
            # NILAI STANDAR SOS
            ws['EL{}'.format(r)] = "=MAX(EL2:EL{})".format(q)
            ws['EL{}'.format(s)] = "=MIN(EL2:EL{})".format(q)
            ws['EL{}'.format(t)] = "=ROUND(AVERAGE(EL2:EL{}),2)".format(q)
            # NILAI STANDAR S.JML
            ws['EM{}'.format(r)] = "=MAX(EM2:EM{})".format(q)
            ws['EM{}'.format(s)] = "=MIN(EM2:EM{})".format(q)
            ws['EM{}'.format(t)] = "=ROUND(AVERAGE(EM2:EM{}),2)".format(q)

            # TAMBAHAN
            ws['EP{}'.format(r)] = "=SUM(EP2:EP{})".format(q)
            ws['EQ{}'.format(r)] = "=SUM(EQ2:EQ{})".format(q)
            ws['ER{}'.format(r)] = "=SUM(ER2:ER{})".format(q)
            ws['ES{}'.format(r)] = "=SUM(ES2:ES{})".format(q)
            ws['ET{}'.format(r)] = "=SUM(ET2:ET{})".format(q)
            ws['EU{}'.format(r)] = "=SUM(EU2:EU{})".format(q)
            ws['EV{}'.format(r)] = "=SUM(EV2:EV{})".format(q)

            # iterasi 4 rata-rata - 4
            # rata" MTK ke MTK tambahan dan mapel MTK awal
            ws['FC{}'.format(
                r)] = "=IF($EP${}=0,$DQ${},$DQ${}-1)".format(r, r, r)
            ws['FC{}'.format(s)] = "=STDEV(FC2:FC{})".format(q)
            ws['FC{}'.format(t)] = "=MAX(FC2:FC{})".format(q)
            ws['FC{}'.format(u)] = "=MIN(FC2:FC{})".format(q)
            # rata" IND ke IND tambahan dan mapel IND awal
            ws['FD{}'.format(
                r)] = "=IF($EQ${}=0,$DR${},$DR${}-1)".format(r, r, r)
            ws['FD{}'.format(s)] = "=STDEV(FD2:FD{})".format(q)
            ws['FD{}'.format(t)] = "=MAX(FD2:FD{})".format(q)
            ws['FD{}'.format(u)] = "=MIN(FD2:FD{})".format(q)
            # rata" ENG ke ENG tambahan dan mapel ENG awal
            ws['FE{}'.format(
                r)] = "=IF($ER${}=0,$DS${},$DS{}-1)".format(r, r, r)
            ws['FE{}'.format(s)] = "=STDEV(FE2:FE{})".format(q)
            ws['FE{}'.format(t)] = "=MAX(FE2:FE{})".format(q)
            ws['FE{}'.format(u)] = "=MIN(FE2:FE{})".format(q)
            # rata" SEJ ke SEJ tambahan dan mapel SEJ awal
            ws['FF{}'.format(
                r)] = "=IF($ES${}=0,$DT${},$DT${}-1)".format(r, r, r)
            ws['FF{}'.format(s)] = "=STDEV(FF2:FF{})".format(q)
            ws['FF{}'.format(t)] = "=MAX(FF2:FF{})".format(q)
            ws['FF{}'.format(u)] = "=MIN(FF2:FF{})".format(q)
            # rata" GEO ke GEO tambahan dan mapel GEO awal
            ws['FG{}'.format(
                r)] = "=IF($ET${}=0,$DU${},$DU${}-1)".format(r, r, r)
            ws['FG{}'.format(s)] = "=STDEV(FG2:FG{})".format(q)
            ws['FG{}'.format(t)] = "=MAX(FG2:FG{})".format(q)
            ws['FG{}'.format(u)] = "=MIN(FG2:FG{})".format(q)
            # rata" EKO ke EKO tambahan dan mapel EKO awal
            ws['FH{}'.format(
                r)] = "=IF($EU${}=0,$DV${},$DV${}-1)".format(r, r, r)
            ws['FH{}'.format(s)] = "=STDEV(FH2:FH{})".format(q)
            ws['FH{}'.format(t)] = "=MAX(FH2:FH{})".format(q)
            ws['FH{}'.format(u)] = "=MIN(FH2:FH{})".format(q)
            # rata" SOS ke SOS tambahan dan mapel SOS awal
            ws['FI{}'.format(
                r)] = "=IF($EV${}=0,$DW${},$DW${}-1)".format(r, r, r)
            ws['FI{}'.format(s)] = "=STDEV(FI2:FI{})".format(q)
            ws['FI{}'.format(t)] = "=MAX(FI2:FI{})".format(q)
            ws['FI{}'.format(u)] = "=MIN(FI2:FI{})".format(q)
            # jml MAPEL
            ws['FJ{}'.format(r)] = "=ROUND(AVERAGE(FJ2:FJ{}),2)".format(q)
            ws['FJ{}'.format(t)] = "=MAX(FJ2:FJ{})".format(q)
            ws['FJ{}'.format(u)] = "=MIN(FJ2:FJ{})".format(q)
            # MAX Z SCORE
            ws['FK{}'.format(r)] = "=MAX(FK2:FK{})".format(q)
            ws['FL{}'.format(r)] = "=MAX(FL2:FL{})".format(q)
            ws['FM{}'.format(r)] = "=MAX(FM2:FM{})".format(q)
            ws['FN{}'.format(r)] = "=MAX(FN2:FN{})".format(q)
            ws['FO{}'.format(r)] = "=MAX(FO2:FO{})".format(q)
            ws['FP{}'.format(r)] = "=MAX(FP2:FP{})".format(q)
            ws['FQ{}'.format(r)] = "=MAX(FQ2:FQ{})".format(q)
            # NILAI STANDAR MTK
            ws['FR{}'.format(r)] = "=MAX(FR2:FR{})".format(q)
            ws['FR{}'.format(s)] = "=MIN(FR2:FR{})".format(q)
            ws['FR{}'.format(t)] = "=ROUND(AVERAGE(FR2:FR{}),2)".format(q)
            # NILAI STANDAR IND
            ws['FS{}'.format(r)] = "=MAX(FS2:FS{})".format(q)
            ws['FS{}'.format(s)] = "=MIN(FS2:FS{})".format(q)
            ws['FS{}'.format(t)] = "=ROUND(AVERAGE(FS2:FS{}),2)".format(q)
            # NILAI STANDAR ENG
            ws['FT{}'.format(r)] = "=MAX(FT2:FT{})".format(q)
            ws['FT{}'.format(s)] = "=MIN(FT2:FT{})".format(q)
            ws['FT{}'.format(t)] = "=ROUND(AVERAGE(FT2:FT{}),2)".format(q)
            # NILAI STANDAR SEJ
            ws['FU{}'.format(r)] = "=MAX(FU2:FU{})".format(q)
            ws['FU{}'.format(s)] = "=MIN(FU2:FU{})".format(q)
            ws['FU{}'.format(t)] = "=ROUND(AVERAGE(FU2:FU{}),2)".format(q)
            # NILAI STANDAR GEO
            ws['FV{}'.format(r)] = "=MAX(FV2:FV{})".format(q)
            ws['FV{}'.format(s)] = "=MIN(FV2:FV{})".format(q)
            ws['FV{}'.format(t)] = "=ROUND(AVERAGE(FV2:FV{}),2)".format(q)
            # NILAI STANDAR EKO
            ws['FW{}'.format(r)] = "=MAX(FW2:FW{})".format(q)
            ws['FW{}'.format(s)] = "=MIN(FW2:FW{})".format(q)
            ws['FW{}'.format(t)] = "=ROUND(AVERAGE(FW2:FW{}),2)".format(q)
            # NILAI STANDAR SOS
            ws['FX{}'.format(r)] = "=MAX(FX2:FX{})".format(q)
            ws['FX{}'.format(s)] = "=MIN(FX2:FX{})".format(q)
            ws['FX{}'.format(t)] = "=ROUND(AVERAGE(FX2:FX{}),2)".format(q)
            # NILAI STANDAR S.JML
            ws['FY{}'.format(r)] = "=MAX(FY2:FY{})".format(q)
            ws['FY{}'.format(s)] = "=MIN(FY2:FY{})".format(q)
            ws['FY{}'.format(t)] = "=ROUND(AVERAGE(FY2:FY{}),2)".format(q)

            # TAMBAHAN
            ws['GB{}'.format(r)] = "=SUM(GB2:GB{})".format(q)
            ws['GC{}'.format(r)] = "=SUM(GC2:GC{})".format(q)
            ws['GD{}'.format(r)] = "=SUM(GD2:GD{})".format(q)
            ws['GE{}'.format(r)] = "=SUM(GE2:GE{})".format(q)
            ws['GF{}'.format(r)] = "=SUM(GF2:GF{})".format(q)
            ws['GG{}'.format(r)] = "=SUM(GG2:GG{})".format(q)
            ws['GH{}'.format(r)] = "=SUM(GH2:GH{})".format(q)

            # iterasi 5 rata-rata - 5
            # rata" MTK ke MTK tambahan dan mapel MTK awal
            ws['GO{}'.format(
                r)] = "=IF($GB${}=0,$FC${},$FC${}-1)".format(r, r, r)
            ws['GO{}'.format(s)] = "=STDEV(GO2:GO{})".format(q)
            ws['GO{}'.format(t)] = "=MAX(GO2:GO{})".format(q)
            ws['GO{}'.format(u)] = "=MIN(GO2:GO{})".format(q)
            # rata" IND ke IND tambahan dan mapel IND awal
            ws['GP{}'.format(
                r)] = "=IF($GC${}=0,$FD${},$FD${}-1)".format(r, r, r)
            ws['GP{}'.format(s)] = "=STDEV(GP2:GP{})".format(q)
            ws['GP{}'.format(t)] = "=MAX(GP2:GP{})".format(q)
            ws['GP{}'.format(u)] = "=MIN(GP2:GP{})".format(q)
            # rata" ENG ke ENG tambahan dan mapel ENG awal
            ws['GQ{}'.format(
                r)] = "=IF($GD${}=0,$FE${},$FE{}-1)".format(r, r, r)
            ws['GQ{}'.format(s)] = "=STDEV(GQ2:GQ{})".format(q)
            ws['GQ{}'.format(t)] = "=MAX(GQ2:GQ{})".format(q)
            ws['GQ{}'.format(u)] = "=MIN(GQ2:GQ{})".format(q)
            # rata" SEJ ke SEJ tambahan dan mapel SEJ awal
            ws['GR{}'.format(
                r)] = "=IF($GE${}=0,$FF${},$FF${}-1)".format(r, r, r)
            ws['GR{}'.format(s)] = "=STDEV(GR2:GR{})".format(q)
            ws['GR{}'.format(t)] = "=MAX(GR2:GR{})".format(q)
            ws['GR{}'.format(u)] = "=MIN(GR2:GR{})".format(q)
            # rata" GEO ke GEO tambahan dan mapel GEO awal
            ws['GS{}'.format(
                r)] = "=IF($GF${}=0,$FG${},$FG${}-1)".format(r, r, r)
            ws['GS{}'.format(s)] = "=STDEV(GS2:GS{})".format(q)
            ws['GS{}'.format(t)] = "=MAX(GS2:GS{})".format(q)
            ws['GS{}'.format(u)] = "=MIN(GS2:GS{})".format(q)
            # rata" EKO ke EKO tambahan dan mapel EKO awal
            ws['GT{}'.format(
                r)] = "=IF($GG${}=0,$FH${},$FH${}-1)".format(r, r, r)
            ws['GT{}'.format(s)] = "=STDEV(GT2:GT{})".format(q)
            ws['GT{}'.format(t)] = "=MAX(GT2:GT{})".format(q)
            ws['GT{}'.format(u)] = "=MIN(GT2:GT{})".format(q)
            # rata" SOS ke SOS tambahan dan mapel SOS awal
            ws['GU{}'.format(
                r)] = "=IF($GH${}=0,$FI${},$FI${}-1)".format(r, r, r)
            ws['GU{}'.format(s)] = "=STDEV(GU2:GU{})".format(q)
            ws['GU{}'.format(t)] = "=MAX(GU2:GU{})".format(q)
            ws['GU{}'.format(u)] = "=MIN(GU2:GU{})".format(q)
            # jml MAPEL
            ws['GV{}'.format(r)] = "=ROUND(AVERAGE(GV2:GV{}),2)".format(q)
            ws['GV{}'.format(t)] = "=MAX(GV2:GV{})".format(q)
            ws['GV{}'.format(u)] = "=MIN(GV2:GV{})".format(q)
            # MAX Z SCORE
            ws['GW{}'.format(r)] = "=MAX(GW2:GW{})".format(q)
            ws['GX{}'.format(r)] = "=MAX(GX2:GX{})".format(q)
            ws['GY{}'.format(r)] = "=MAX(GY2:GY{})".format(q)
            ws['GZ{}'.format(r)] = "=MAX(GZ2:GZ{})".format(q)
            ws['HA{}'.format(r)] = "=MAX(HA2:HA{})".format(q)
            ws['HB{}'.format(r)] = "=MAX(HB2:HB{})".format(q)
            ws['HC{}'.format(r)] = "=MAX(HC2:HC{})".format(q)
            # NILAI STANDAR MTK
            ws['HD{}'.format(r)] = "=MAX(HD2:HD{})".format(q)
            ws['HD{}'.format(s)] = "=MIN(HD2:HD{})".format(q)
            ws['HD{}'.format(t)] = "=ROUND(AVERAGE(HD2:HD{}),2)".format(q)
            # NILAI STANDAR IND
            ws['HE{}'.format(r)] = "=MAX(HE2:HE{})".format(q)
            ws['HE{}'.format(s)] = "=MIN(HE2:HE{})".format(q)
            ws['HE{}'.format(t)] = "=ROUND(AVERAGE(HE2:HE{}),2)".format(q)
            # NILAI STANDAR ENG
            ws['HF{}'.format(r)] = "=MAX(HF2:HF{})".format(q)
            ws['HF{}'.format(s)] = "=MIN(HF2:HF{})".format(q)
            ws['HF{}'.format(t)] = "=ROUND(AVERAGE(HF2:HF{}),2)".format(q)
            # NILAI STANDAR SEJ
            ws['HG{}'.format(r)] = "=MAX(HG2:HG{})".format(q)
            ws['HG{}'.format(s)] = "=MIN(HG2:HG{})".format(q)
            ws['HG{}'.format(t)] = "=ROUND(AVERAGE(HG2:HG{}),2)".format(q)
            # NILAI STANDAR GEO
            ws['HH{}'.format(r)] = "=MAX(HH2:HH{})".format(q)
            ws['HH{}'.format(s)] = "=MIN(HH2:HH{})".format(q)
            ws['HH{}'.format(t)] = "=ROUND(AVERAGE(HH2:HH{}),2)".format(q)
            # NILAI STANDAR EKO
            ws['HI{}'.format(r)] = "=MAX(HI2:HI{})".format(q)
            ws['HI{}'.format(s)] = "=MIN(HI2:HI{})".format(q)
            ws['HI{}'.format(t)] = "=ROUND(AVERAGE(HI2:HI{}),2)".format(q)
            # NILAI STANDAR SOS
            ws['HJ{}'.format(r)] = "=MAX(HJ2:HJ{})".format(q)
            ws['HJ{}'.format(s)] = "=MIN(HJ2:HJ{})".format(q)
            ws['HJ{}'.format(t)] = "=ROUND(AVERAGE(HJ2:HJ{}),2)".format(q)
            # NILAI STANDAR S.JML
            ws['HK{}'.format(r)] = "=MAX(HK2:HK{})".format(q)
            ws['HK{}'.format(s)] = "=MIN(HK2:HK{})".format(q)
            ws['HK{}'.format(t)] = "=ROUND(AVERAGE(HK2:HK{}),2)".format(q)

            # TAMBAHAN
            ws['HN{}'.format(r)] = "=SUM(HN2:HN{})".format(q)
            ws['HO{}'.format(r)] = "=SUM(HO2:HO{})".format(q)
            ws['HP{}'.format(r)] = "=SUM(HP2:HP{})".format(q)
            ws['HQ{}'.format(r)] = "=SUM(HQ2:HQ{})".format(q)
            ws['HR{}'.format(r)] = "=SUM(HR2:HR{})".format(q)
            ws['HS{}'.format(r)] = "=SUM(HS2:HS{})".format(q)
            ws['HT{}'.format(r)] = "=SUM(HT2:HT{})".format(q)

            # Z Score [1]
            ws['B1'] = 'NAMA SISWA_A'
            ws['C1'] = 'NOMOR NF_A'
            ws['D1'] = 'KELAS_A'
            ws['E1'] = 'NAMA SEKOLAH_A'
            ws['F1'] = 'LOKASI_A'

            ws['G1'] = 'MAT_A'
            ws['H1'] = 'IND_A'
            ws['I1'] = 'ENG_A'
            ws['J1'] = 'SEJ_A'
            ws['K1'] = 'GEO_A'
            ws['L1'] = 'EKO_A'
            ws['M1'] = 'SOS_A'
            ws['N1'] = 'JML_A'

            ws['O1'] = 'Z_MAT_A'
            ws['P1'] = 'Z_IND_A'
            ws['Q1'] = 'Z_ENG_A'
            ws['R1'] = 'Z_SEJ_A'
            ws['S1'] = 'Z_GEO_A'
            ws['T1'] = 'Z_EKO_A'
            ws['U1'] = 'Z_SOS_A'

            ws['V1'] = 'S_MAT_A'
            ws['W1'] = 'S_IND_A'
            ws['X1'] = 'S_ENG_A'
            ws['Y1'] = 'S_SEJ_A'
            ws['Z1'] = 'S_GEO_A'
            ws['AA1'] = 'S_EKO_A'
            ws['AB1'] = 'S_SOS_A'
            ws['AC1'] = 'S_JML_A'

            ws['AD1'] = 'RANK NAS._A'
            ws['AE1'] = 'RANK LOK._A'

            ws['O1'].font = Font(bold=False, name='Calibri', size=11)
            ws['P1'].font = Font(bold=False, name='Calibri', size=11)
            ws['Q1'].font = Font(bold=False, name='Calibri', size=11)
            ws['R1'].font = Font(bold=False, name='Calibri', size=11)
            ws['S1'].font = Font(bold=False, name='Calibri', size=11)
            ws['T1'].font = Font(bold=False, name='Calibri', size=11)
            ws['U1'].font = Font(bold=False, name='Calibri', size=11)
            ws['V1'].font = Font(bold=False, name='Calibri', size=11)
            ws['W1'].font = Font(bold=False, name='Calibri', size=11)
            ws['X1'].font = Font(bold=False, name='Calibri', size=11)
            ws['Y1'].font = Font(bold=False, name='Calibri', size=11)
            ws['Z1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AA1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AB1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AC1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AD1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AE1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['B1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['C1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['D1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['E1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['F1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['G1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['H1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['I1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['J1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['K1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['L1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['M1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['N1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['O1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['P1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['Q1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['R1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['S1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['T1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['U1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['V1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['W1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['X1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['Y1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['Z1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AA1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AB1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AC1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AD1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AE1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')

            # tambahan
            ws['AF1'] = 'MAT_20_A'
            ws['AG1'] = 'IND_20_A'
            ws['AH1'] = 'ENG_20_A'
            ws['AI1'] = 'SEJ_20_A'
            ws['AJ1'] = 'GEO_20_A'
            ws['AK1'] = 'EKO_20_A'
            ws['AL1'] = 'SOS_20_A'

            ws['AF1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AG1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AH1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AI1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AJ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AK1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AL1'].font = Font(bold=False, name='Calibri', size=11)

            ws['AF1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AG1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AH1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AI1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AJ1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AK1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AL1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')

            for row in range(2, q+1):
                ws['N{}'.format(
                    row)] = '=SUM(G{}:M{})'.format(row, row, row)
                ws['O{}'.format(
                    row)] = '=IFERROR(ROUND(IF(G{}="","",(G{}-G${})/G${}),2),"")'.format(row, row, r, s)
                ws['P{}'.format(
                    row)] = '=IFERROR(ROUND(IF(H{}="","",(H{}-H${})/H${}),2),"")'.format(row, row, r, s)
                ws['Q{}'.format(
                    row)] = '=IFERROR(ROUND(IF(I{}="","",(I{}-I${})/I${}),2),"")'.format(row, row, r, s)
                ws['R{}'.format(
                    row)] = '=IFERROR(ROUND(IF(J{}="","",(J{}-J${})/J${}),2),"")'.format(row, row, r, s)
                ws['S{}'.format(
                    row)] = '=IFERROR(ROUND(IF(K{}="","",(K{}-K${})/K${}),2),"")'.format(row, row, r, s)
                ws['T{}'.format(
                    row)] = '=IFERROR(ROUND(IF(L{}="","",(L{}-L${})/L${}),2),"")'.format(row, row, r, s)
                ws['U{}'.format(
                    row)] = '=IFERROR(ROUND(IF(M{}="","",(M{}-M${})/M${}),2),"")'.format(row, row, r, s)

                ws['V{}'.format(
                    row)] = '=IFERROR(ROUND(IF(G{}="","",IF(70+30*O{}/$O${}<20,20,70+30*O{}/$O${})),2),"")'.format(row, row, r, row, r)
                ws['W{}'.format(
                    row)] = '=IFERROR(ROUND(IF(H{}="","",IF(70+30*P{}/$P${}<20,20,70+30*P{}/$P${})),2),"")'.format(row, row, r, row, r)
                ws['X{}'.format(
                    row)] = '=IFERROR(ROUND(IF(I{}="","",IF(70+30*Q{}/$Q${}<20,20,70+30*Q{}/$Q${})),2),"")'.format(row, row, r, row, r)
                ws['Y{}'.format(
                    row)] = '=IFERROR(ROUND(IF(J{}="","",IF(70+30*R{}/$R${}<20,20,70+30*R{}/$R${})),2),"")'.format(row, row, r, row, r)
                ws['Z{}'.format(
                    row)] = '=IFERROR(ROUND(IF(K{}="","",IF(70+30*S{}/$S${}<20,20,70+30*S{}/$S${})),2),"")'.format(row, row, r, row, r)
                ws['AA{}'.format(
                    row)] = '=IFERROR(ROUND(IF(L{}="","",IF(70+30*T{}/$T${}<20,20,70+30*T{}/$T${})),2),"")'.format(row, row, r, row, r)
                ws['AB{}'.format(
                    row)] = '=IFERROR(ROUND(IF(M{}="","",IF(70+30*U{}/$U${}<20,20,70+30*U{}/$U${})),2),"")'.format(row, row, r, row, r)

                ws['AC{}'.format(row)] = '=IF(SUM(V{}:AB{})=0,"",SUM(V{}:AB{}))'.format(
                    row, row, row, row)
                ws['AD{}'.format(row)] = '=IF(AC{}="","",RANK(AC{},$AC$2:$AC${}))'.format(
                    row, row, q)
                ws['AE{}'.format(
                    row)] = '=IF(AD{}="","",COUNTIFS($F$2:$F${},F{},$AD$2:$AD${},"<"&AD{})+1)'.format(row, q, row, q, row)
            # TAMBAHAN
                ws['AF{}'.format(row)] = '=IF($G${}=20,IF(AND(G{}>3,V{}=20),1,""),IF($G${}=25,IF(AND(G{}>4,V{}=20),1,""),IF($G${}=30,IF(AND(G{}>5,V{}=20),1,""),IF($G${}=35,IF(AND(G{}>6,V{}=20),1,""),IF($G${}=40,IF(AND(G{}>7,V{}=20),1,""),IF($G${}=45,IF(AND(G{}>8,V{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['AG{}'.format(row)] = '=IF($H${}=20,IF(AND(H{}>3,W{}=20),1,""),IF($H${}=25,IF(AND(H{}>4,W{}=20),1,""),IF($H${}=30,IF(AND(H{}>5,W{}=20),1,""),IF($H${}=35,IF(AND(H{}>6,W{}=20),1,""),IF($H${}=40,IF(AND(H{}>7,W{}=20),1,""),IF($H${}=45,IF(AND(H{}>8,W{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['AH{}'.format(row)] = '=IF($I${}=20,IF(AND(I{}>3,X{}=20),1,""),IF($I${}=25,IF(AND(I{}>4,X{}=20),1,""),IF($I${}=30,IF(AND(I{}>5,X{}=20),1,""),IF($I${}=35,IF(AND(I{}>6,X{}=20),1,""),IF($I${}=40,IF(AND(I{}>7,X{}=20),1,""),IF($I${}=45,IF(AND(I{}>8,X{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['AI{}'.format(row)] = '=IF($J${}=20,IF(AND(J{}>3,Y{}=20),1,""),IF($J${}=25,IF(AND(J{}>4,Y{}=20),1,""),IF($J${}=30,IF(AND(J{}>5,Y{}=20),1,""),IF($J${}=35,IF(AND(J{}>6,Y{}=20),1,""),IF($J${}=40,IF(AND(J{}>7,Y{}=20),1,""),IF($J${}=45,IF(AND(J{}>8,Y{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['AJ{}'.format(row)] = '=IF($K${}=20,IF(AND(K{}>3,Z{}=20),1,""),IF($K${}=25,IF(AND(K{}>4,Z{}=20),1,""),IF($K${}=30,IF(AND(K{}>5,Z{}=20),1,""),IF($K${}=35,IF(AND(K{}>6,Z{}=20),1,""),IF($K${}=40,IF(AND(K{}>7,Z{}=20),1,""),IF($K${}=45,IF(AND(K{}>8,Z{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['AK{}'.format(row)] = '=IF($L${}=20,IF(AND(L{}>3,AA{}=20),1,""),IF($L${}=25,IF(AND(L{}>4,AA{}=20),1,""),IF($L${}=30,IF(AND(L{}>5,AA{}=20),1,""),IF($L${}=35,IF(AND(L{}>6,AA{}=20),1,""),IF($L${}=40,IF(AND(L{}>7,AA{}=20),1,""),IF($L${}=45,IF(AND(L{}>8,AA{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['AL{}'.format(row)] = '=IF($M${}=20,IF(AND(M{}>3,AB{}=20),1,""),IF($M${}=25,IF(AND(M{}>4,AB{}=20),1,""),IF($M${}=30,IF(AND(M{}>5,AB{}=20),1,""),IF($M${}=35,IF(AND(M{}>6,AB{}=20),1,""),IF($M${}=40,IF(AND(M{}>7,AB{}=20),1,""),IF($M${}=45,IF(AND(M{}>8,AB{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

            # Z Score [2]
            ws['AN1'] = 'NAMA SISWA_B'
            ws['AO1'] = 'NOMOR NF_B'
            ws['AP1'] = 'KELAS_B'
            ws['AQ1'] = 'NAMA SEKOLAH_B'
            ws['AR1'] = 'LOKASI_B'

            ws['AS1'] = 'MAT_B'
            ws['AT1'] = 'IND_B'
            ws['AU1'] = 'ENG_B'
            ws['AV1'] = 'SEJ_B'
            ws['AW1'] = 'GEO_B'
            ws['AX1'] = 'EKO_B'
            ws['AY1'] = 'SOS_B'
            ws['AZ1'] = 'JML_B'

            ws['BA1'] = 'Z_MAT_B'
            ws['BB1'] = 'Z_IND_B'
            ws['BC1'] = 'Z_ENG_B'
            ws['BD1'] = 'Z_SEJ_B'
            ws['BE1'] = 'Z_GEO_B'
            ws['BF1'] = 'Z_EKO_B'
            ws['BG1'] = 'Z_SOS_B'

            ws['BH1'] = 'S_MAT_B'
            ws['BI1'] = 'S_IND_B'
            ws['BJ1'] = 'S_ENG_B'
            ws['BK1'] = 'S_SEJ_B'
            ws['BL1'] = 'S_GEO_B'
            ws['BM1'] = 'S_EKO_B'
            ws['BN1'] = 'S_SOS_B'
            ws['BO1'] = 'S_JML_B'

            ws['BP1'] = 'RANK NAS._B'
            ws['BQ1'] = 'RANK LOK._B'

            ws['BA1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BB1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BC1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BD1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BE1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BF1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BG1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BH1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BI1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BJ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BK1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BL1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BM1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BN1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BO1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BP1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BQ1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['AN1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AO1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AP1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AQ1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AR1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AS1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AT1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AU1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AV1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AW1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AX1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AY1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AZ1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BA1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BB1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BC1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BD1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BE1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BF1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BG1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BH1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BI1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BJ1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BK1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BL1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BM1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BN1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BO1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BP1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BQ1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')

            # tambahan
            ws['BR1'] = 'MAT_20_B'
            ws['BS1'] = 'IND_20_B'
            ws['BT1'] = 'ENG_20_B'
            ws['BU1'] = 'SEJ_20_B'
            ws['BV1'] = 'GEO_20_B'
            ws['BW1'] = 'EKO_20_B'
            ws['BX1'] = 'SOS_20_B'

            ws['BR1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BS1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BT1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BU1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BV1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BW1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BX1'].font = Font(bold=False, name='Calibri', size=11)

            ws['BR1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BS1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BT1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BU1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BV1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BW1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BX1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')

            for row in range(2, q+1):
                ws['AN{}'.format(row)] = '=B{}'.format(row)
                ws['AO{}'.format(row)] = '=C{}'.format(row, row)
                ws['AP{}'.format(row)] = '=D{}'.format(row, row)
                ws['AQ{}'.format(row)] = '=E{}'.format(row, row)
                ws['AR{}'.format(row)] = '=F{}'.format(row, row)
                ws['AS{}'.format(row)] = '=IF(G{}="","",G{})'.format(row, row)
                ws['AT{}'.format(row)] = '=IF(H{}="","",H{})'.format(row, row)
                ws['AU{}'.format(row)] = '=IF(I{}="","",I{})'.format(row, row)
                ws['AV{}'.format(row)] = '=IF(J{}="","",J{})'.format(row, row)
                ws['AW{}'.format(row)] = '=IF(K{}="","",K{})'.format(row, row)
                ws['AX{}'.format(row)] = '=IF(L{}="","",L{})'.format(row, row)
                ws['AY{}'.format(row)] = '=IF(M{}="","",M{})'.format(row, row)
                ws['AZ{}'.format(row)] = '=IF(N{}="","",N{})'.format(row, row)
            # Z Ke mapel
                ws['BA{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AS{}="","",(AS{}-AS${})/AS${}),2),"")'.format(row, row, r, s)
                ws['BB{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AT{}="","",(AT{}-AT${})/AT${}),2),"")'.format(row, row, r, s)
                ws['BC{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AU{}="","",(AU{}-AU${})/AU${}),2),"")'.format(row, row, r, s)
                ws['BD{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AV{}="","",(AV{}-AV${})/AV${}),2),"")'.format(row, row, r, s)
                ws['BE{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AW{}="","",(AW{}-AW${})/AW${}),2),"")'.format(row, row, r, s)
                ws['BF{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AX{}="","",(AX{}-AX${})/AX${}),2),"")'.format(row, row, r, s)
                ws['BG{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AY{}="","",(AY{}-AY${})/AY${}),2),"")'.format(row, row, r, s)
            # NILAI STANDAR ke mapel dan Z score
                ws['BH{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AS{}="","",IF(70+30*BA{}/$BA${}<20,20,70+30*BA{}/$BA${})),2),"")'.format(row, row, r, row, r)
                ws['BI{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AT{}="","",IF(70+30*BB{}/$BB${}<20,20,70+30*BB{}/$BB${})),2),"")'.format(row, row, r, row, r)
                ws['BJ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AU{}="","",IF(70+30*BC{}/$BC${}<20,20,70+30*BC{}/$BC${})),2),"")'.format(row, row, r, row, r)
                ws['BK{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AV{}="","",IF(70+30*BD{}/$BD${}<20,20,70+30*BD{}/$BD${})),2),"")'.format(row, row, r, row, r)
                ws['BL{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AW{}="","",IF(70+30*BE{}/$BE${}<20,20,70+30*BE{}/$BE${})),2),"")'.format(row, row, r, row, r)
                ws['BM{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AX{}="","",IF(70+30*BF{}/$BF${}<20,20,70+30*BF{}/$BF${})),2),"")'.format(row, row, r, row, r)
                ws['BN{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AY{}="","",IF(70+30*BG{}/$BG${}<20,20,70+30*BG{}/$BG${})),2),"")'.format(row, row, r, row, r)
            # JUMLAH SELURUH NILAI STANDAR
                ws['BO{}'.format(row)] = '=IF(SUM(BH{}:BN{})=0,"",SUM(BH{}:BN{}))'.format(
                    row, row, row, row)
                ws['BP{}'.format(row)] = '=IF(BO{}="","",RANK(BO{},$BO$2:$BO${}))'.format(
                    row, row, q)
                ws['BQ{}'.format(
                    row)] = '=IF(BP{}="","",COUNTIFS($AR$2:$AR${},AR{},$BP$2:$BP${},"<"&BP{})+1)'.format(row, q, row, q, row)
            # TAMBAHAN, MAPEL DAN NILAI STANDAR
                ws['BR{}'.format(row)] = '=IF($G${}=20,IF(AND(AS{}>3,BH{}=20),1,""),IF($G${}=25,IF(AND(AS{}>4,BH{}=20),1,""),IF($G${}=30,IF(AND(AS{}>5,BH{}=20),1,""),IF($G${}=35,IF(AND(AS{}>6,BH{}=20),1,""),IF($G${}=40,IF(AND(AS{}>7,BH{}=20),1,""),IF($G${}=45,IF(AND(AS{}>8,BH{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['BS{}'.format(row)] = '=IF($H${}=20,IF(AND(AT{}>3,BI{}=20),1,""),IF($H${}=25,IF(AND(AT{}>4,BI{}=20),1,""),IF($H${}=30,IF(AND(AT{}>5,BI{}=20),1,""),IF($H${}=35,IF(AND(AT{}>6,BI{}=20),1,""),IF($H${}=40,IF(AND(AT{}>7,BI{}=20),1,""),IF($H${}=45,IF(AND(AT{}>8,BI{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['BT{}'.format(row)] = '=IF($I${}=20,IF(AND(AU{}>3,BJ{}=20),1,""),IF($I${}=25,IF(AND(AU{}>4,BJ{}=20),1,""),IF($I${}=30,IF(AND(AU{}>5,BJ{}=20),1,""),IF($I${}=35,IF(AND(AU{}>6,BJ{}=20),1,""),IF($I${}=40,IF(AND(AU{}>7,BJ{}=20),1,""),IF($I${}=45,IF(AND(AU{}>8,BJ{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['BU{}'.format(row)] = '=IF($J${}=20,IF(AND(AV{}>3,BK{}=20),1,""),IF($J${}=25,IF(AND(AV{}>4,BK{}=20),1,""),IF($J${}=30,IF(AND(AV{}>5,BK{}=20),1,""),IF($J${}=35,IF(AND(AV{}>6,BK{}=20),1,""),IF($J${}=40,IF(AND(AV{}>7,BK{}=20),1,""),IF($J${}=45,IF(AND(AV{}>8,BK{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['BV{}'.format(row)] = '=IF($K${}=20,IF(AND(AW{}>3,BL{}=20),1,""),IF($K${}=25,IF(AND(AW{}>4,BL{}=20),1,""),IF($K${}=30,IF(AND(AW{}>5,BL{}=20),1,""),IF($K${}=35,IF(AND(AW{}>6,BL{}=20),1,""),IF($K${}=40,IF(AND(AW{}>7,BL{}=20),1,""),IF($K${}=45,IF(AND(AW{}>8,BL{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['BW{}'.format(row)] = '=IF($L${}=20,IF(AND(AX{}>3,BM{}=20),1,""),IF($L${}=25,IF(AND(AX{}>4,BM{}=20),1,""),IF($L${}=30,IF(AND(AX{}>5,BM{}=20),1,""),IF($L${}=35,IF(AND(AX{}>6,BM{}=20),1,""),IF($L${}=40,IF(AND(AX{}>7,BM{}=20),1,""),IF($L${}=45,IF(AND(AX{}>8,BM{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['BX{}'.format(row)] = '=IF($M${}=20,IF(AND(AY{}>3,BN{}=20),1,""),IF($M${}=25,IF(AND(AY{}>4,BN{}=20),1,""),IF($M${}=30,IF(AND(AY{}>5,BN{}=20),1,""),IF($M${}=35,IF(AND(AY{}>6,BN{}=20),1,""),IF($M${}=40,IF(AND(AY{}>7,BN{}=20),1,""),IF($M${}=45,IF(AND(AY{}>8,BN{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

            # Z Score [3]
            ws['BZ1'] = 'NAMA SISWA_C'
            ws['CA1'] = 'NOMOR NF_C'
            ws['CB1'] = 'KELAS_C'
            ws['CC1'] = 'NAMA SEKOLAH_C'
            ws['CD1'] = 'LOKASI_C'

            ws['CE1'] = 'MAT_C'
            ws['CF1'] = 'IND_C'
            ws['CG1'] = 'ENG_C'
            ws['CH1'] = 'SEJ_C'
            ws['CI1'] = 'GEO_C'
            ws['CJ1'] = 'EKO_C'
            ws['CK1'] = 'SOS_C'
            ws['CL1'] = 'JML_C'

            ws['CM1'] = 'Z_MAT_C'
            ws['CN1'] = 'Z_IND_C'
            ws['CO1'] = 'Z_ENG_C'
            ws['CP1'] = 'Z_SEJ_C'
            ws['CQ1'] = 'Z_GEO_C'
            ws['CR1'] = 'Z_EKO_C'
            ws['CS1'] = 'Z_SOS_C'

            ws['CT1'] = 'S_MAT_C'
            ws['CU1'] = 'S_IND_C'
            ws['CV1'] = 'S_ENG_C'
            ws['CW1'] = 'S_SEJ_C'
            ws['CX1'] = 'S_GEO_C'
            ws['CY1'] = 'S_EKO_C'
            ws['CZ1'] = 'S_SOS_C'
            ws['DA1'] = 'S_JML_C'

            ws['DB1'] = 'RANK NAS._C'
            ws['DC1'] = 'RANK LOK._C'

            ws['CM1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CN1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CO1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CP1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CQ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CR1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CS1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CT1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CU1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CV1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CW1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CX1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CY1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CZ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DA1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DB1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DC1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['BZ1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CA1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CB1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CC1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CD1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CE1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CF1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CG1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CH1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CI1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CJ1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CK1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CL1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CM1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CN1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CO1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CP1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CQ1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CR1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CS1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CT1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CU1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CV1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CW1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CX1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CY1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CZ1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['DA1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['DB1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['DC1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')

            # tambahan
            ws['DD1'] = 'MAT_20_C'
            ws['DE1'] = 'IND_20_C'
            ws['DF1'] = 'ENG_20_C'
            ws['DG1'] = 'SEJ_20_C'
            ws['DH1'] = 'GEO_20_C'
            ws['DI1'] = 'EKO_20_C'
            ws['DJ1'] = 'SOS_20_C'

            ws['DD1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DE1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DF1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DG1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DH1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DI1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DJ1'].font = Font(bold=False, name='Calibri', size=11)

            ws['DD1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['DE1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['DF1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['DG1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['DH1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['DI1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['DJ1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')

            for row in range(2, q+1):
                ws['BZ{}'.format(row)] = '=AN{}'.format(row)
                ws['CA{}'.format(row)] = '=AO{}'.format(row, row)
                ws['CB{}'.format(row)] = '=AP{}'.format(row, row)
                ws['CC{}'.format(row)] = '=AQ{}'.format(row, row)
                ws['CD{}'.format(row)] = '=AR{}'.format(row, row)
                ws['CE{}'.format(row)] = '=IF(AS{}="","",AS{})'.format(
                    row, row)
                ws['CF{}'.format(row)] = '=IF(AT{}="","",AT{})'.format(
                    row, row)
                ws['CG{}'.format(row)] = '=IF(AU{}="","",AU{})'.format(
                    row, row)
                ws['CH{}'.format(row)] = '=IF(AV{}="","",AV{})'.format(
                    row, row)
                ws['CI{}'.format(row)] = '=IF(AW{}="","",AW{})'.format(
                    row, row)
                ws['CJ{}'.format(row)] = '=IF(AX{}="","",AX{})'.format(
                    row, row)
                ws['CK{}'.format(row)] = '=IF(AY{}="","",AY{})'.format(
                    row, row)
                ws['CL{}'.format(row)] = '=IF(AZ{}="","",AZ{})'.format(
                    row, row)
            # Z Ke mapel
                ws['CM{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CE{}="","",(CE{}-CE${})/CE${}),2),"")'.format(row, row, r, s)
                ws['CN{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CF{}="","",(CF{}-CF${})/CF${}),2),"")'.format(row, row, r, s)
                ws['CO{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CG{}="","",(CG{}-CG${})/CG${}),2),"")'.format(row, row, r, s)
                ws['CP{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CH{}="","",(CH{}-CH${})/CH${}),2),"")'.format(row, row, r, s)
                ws['CQ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CI{}="","",(CI{}-CI${})/CI${}),2),"")'.format(row, row, r, s)
                ws['CR{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CJ{}="","",(CJ{}-CJ${})/CJ${}),2),"")'.format(row, row, r, s)
                ws['CS{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CK{}="","",(CK{}-CK${})/CK${}),2),"")'.format(row, row, r, s)
            # NILAI STANDAR ke mapel dan Z score
                ws['CT{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CE{}="","",IF(70+30*CM{}/$CM${}<20,20,70+30*CM{}/$CM${})),2),"")'.format(row, row, r, row, r)
                ws['CU{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CF{}="","",IF(70+30*CN{}/$CN${}<20,20,70+30*CN{}/$CN${})),2),"")'.format(row, row, r, row, r)
                ws['CV{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CG{}="","",IF(70+30*CO{}/$CO${}<20,20,70+30*CO{}/$CO${})),2),"")'.format(row, row, r, row, r)
                ws['CW{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CH{}="","",IF(70+30*CP{}/$CP${}<20,20,70+30*CP{}/$CP${})),2),"")'.format(row, row, r, row, r)
                ws['CX{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CI{}="","",IF(70+30*CQ{}/$CQ${}<20,20,70+30*CQ{}/$CQ${})),2),"")'.format(row, row, r, row, r)
                ws['CY{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CJ{}="","",IF(70+30*CR{}/$CR${}<20,20,70+30*CR{}/$CR${})),2),"")'.format(row, row, r, row, r)
                ws['CZ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CK{}="","",IF(70+30*CS{}/$CS${}<20,20,70+30*CS{}/$CS${})),2),"")'.format(row, row, r, row, r)
            # JUMLAH SELURUH NILAI STANDAR
                ws['DA{}'.format(row)] = '=IF(SUM(CT{}:CZ{})=0,"",SUM(CT{}:CZ{}))'.format(
                    row, row, row, row)
                ws['DB{}'.format(row)] = '=IF(DA{}="","",RANK(DA{},$DA$2:$DA${}))'.format(
                    row, row, q)
                ws['DC{}'.format(
                    row)] = '=IF(DB{}="","",COUNTIFS($CD$2:$CD${},CD{},$DB$2:$DB${},"<"&DB{})+1)'.format(row, q, row, q, row)
            # TAMBAHAN, MAPEL DAN NILAI STANDAR
                ws['DD{}'.format(row)] = '=IF($G${}=20,IF(AND(CE{}>3,CT{}=20),1,""),IF($G${}=25,IF(AND(CE{}>4,CT{}=20),1,""),IF($G${}=30,IF(AND(CE{}>5,CT{}=20),1,""),IF($G${}=35,IF(AND(CE{}>6,CT{}=20),1,""),IF($G${}=40,IF(AND(CE{}>7,CT{}=20),1,""),IF($G${}=45,IF(AND(CE{}>8,CT{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['DE{}'.format(row)] = '=IF($H${}=20,IF(AND(CF{}>3,CU{}=20),1,""),IF($H${}=25,IF(AND(CF{}>4,CU{}=20),1,""),IF($H${}=30,IF(AND(CF{}>5,CU{}=20),1,""),IF($H${}=35,IF(AND(CF{}>6,CU{}=20),1,""),IF($H${}=40,IF(AND(CF{}>7,CU{}=20),1,""),IF($H${}=45,IF(AND(CF{}>8,CU{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['DF{}'.format(row)] = '=IF($I${}=20,IF(AND(CG{}>3,CV{}=20),1,""),IF($I${}=25,IF(AND(CG{}>4,CV{}=20),1,""),IF($I${}=30,IF(AND(CG{}>5,CV{}=20),1,""),IF($I${}=35,IF(AND(CG{}>6,CV{}=20),1,""),IF($I${}=40,IF(AND(CG{}>7,CV{}=20),1,""),IF($I${}=45,IF(AND(CG{}>8,CV{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['DG{}'.format(row)] = '=IF($J${}=20,IF(AND(CH{}>3,CW{}=20),1,""),IF($J${}=25,IF(AND(CH{}>4,CW{}=20),1,""),IF($J${}=30,IF(AND(CH{}>5,CW{}=20),1,""),IF($J${}=35,IF(AND(CH{}>6,CW{}=20),1,""),IF($J${}=40,IF(AND(CH{}>7,CW{}=20),1,""),IF($J${}=45,IF(AND(CH{}>8,CW{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['DH{}'.format(row)] = '=IF($K${}=20,IF(AND(CI{}>3,CX{}=20),1,""),IF($K${}=25,IF(AND(CI{}>4,CX{}=20),1,""),IF($K${}=30,IF(AND(CI{}>5,CX{}=20),1,""),IF($K${}=35,IF(AND(CI{}>6,CX{}=20),1,""),IF($K${}=40,IF(AND(CI{}>7,CX{}=20),1,""),IF($K${}=45,IF(AND(CI{}>8,CX{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['DI{}'.format(row)] = '=IF($L${}=20,IF(AND(CJ{}>3,CY{}=20),1,""),IF($L${}=25,IF(AND(CJ{}>4,CY{}=20),1,""),IF($L${}=30,IF(AND(CJ{}>5,CY{}=20),1,""),IF($L${}=35,IF(AND(CJ{}>6,CY{}=20),1,""),IF($L${}=40,IF(AND(CJ{}>7,CY{}=20),1,""),IF($L${}=45,IF(AND(CJ{}>8,CY{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['DJ{}'.format(row)] = '=IF($M${}=20,IF(AND(CK{}>3,CZ{}=20),1,""),IF($M${}=25,IF(AND(CK{}>4,CZ{}=20),1,""),IF($M${}=30,IF(AND(CK{}>5,CZ{}=20),1,""),IF($M${}=35,IF(AND(CK{}>6,CZ{}=20),1,""),IF($M${}=40,IF(AND(CK{}>7,CZ{}=20),1,""),IF($M${}=45,IF(AND(CK{}>8,CZ{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

            # Z Score [4]
            ws['DL1'] = 'NAMA SISWA_D'
            ws['DM1'] = 'NOMOR NF_D'
            ws['DN1'] = 'KELAS_D'
            ws['DO1'] = 'NAMA SEKOLAH_D'
            ws['DP1'] = 'LOKASI_D'

            ws['DQ1'] = 'MAT_D'
            ws['DR1'] = 'IND_D'
            ws['DS1'] = 'ENG_D'
            ws['DT1'] = 'SEJ_D'
            ws['DU1'] = 'GEO_D'
            ws['DV1'] = 'EKO_D'
            ws['DW1'] = 'SOS_D'
            ws['DX1'] = 'JML_D'

            ws['DY1'] = 'Z_MAT_D'
            ws['DZ1'] = 'Z_IND_D'
            ws['EA1'] = 'Z_ENG_D'
            ws['EB1'] = 'Z_SEJ_D'
            ws['EC1'] = 'Z_GEO_D'
            ws['ED1'] = 'Z_EKO_D'
            ws['EE1'] = 'Z_SOS_D'

            ws['EF1'] = 'S_MAT_D'
            ws['EG1'] = 'S_IND_D'
            ws['EH1'] = 'S_ENG_D'
            ws['EI1'] = 'S_SEJ_D'
            ws['EJ1'] = 'S_GEO_D'
            ws['EK1'] = 'S_EKO_D'
            ws['EL1'] = 'S_SOS_D'
            ws['EM1'] = 'S_JML_D'

            ws['EN1'] = 'RANK NAS._D'
            ws['EO1'] = 'RANK LOK._D'

            ws['DY1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DZ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EA1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EB1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EC1'].font = Font(bold=False, name='Calibri', size=11)
            ws['ED1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EE1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EF1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EG1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EH1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EI1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EJ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EK1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EL1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EM1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EN1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EO1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['DL1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DM1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DN1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DO1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DP1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DQ1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DR1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DS1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DT1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DU1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DV1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DW1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DX1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DY1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DZ1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['EA1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['EB1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['EC1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['ED1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['EE1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['EF1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['EG1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['EH1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['EI1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['EJ1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['EK1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['EL1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['EM1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['EN1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['EO1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')

            # tambahan
            ws['EP1'] = 'MAT_20_D'
            ws['EQ1'] = 'IND_20_D'
            ws['ER1'] = 'ENG_20_D'
            ws['ES1'] = 'SEJ_20_D'
            ws['ET1'] = 'GEO_20_D'
            ws['EU1'] = 'EKO_20_D'
            ws['EV1'] = 'SOS_20_D'

            ws['EP1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EQ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['ER1'].font = Font(bold=False, name='Calibri', size=11)
            ws['ES1'].font = Font(bold=False, name='Calibri', size=11)
            ws['ET1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EU1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EV1'].font = Font(bold=False, name='Calibri', size=11)

            ws['EP1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['EQ1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['ER1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['ES1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['ET1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['EU1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['EV1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')

            for row in range(2, q+1):
                ws['DL{}'.format(row)] = '=BZ{}'.format(row)
                ws['DM{}'.format(row)] = '=CA{}'.format(row, row)
                ws['DN{}'.format(row)] = '=CB{}'.format(row, row)
                ws['DO{}'.format(row)] = '=CC{}'.format(row, row)
                ws['DP{}'.format(row)] = '=CD{}'.format(row, row)
                ws['DQ{}'.format(row)] = '=IF(CE{}="","",CE{})'.format(
                    row, row)
                ws['DR{}'.format(row)] = '=IF(CF{}="","",CF{})'.format(
                    row, row)
                ws['DS{}'.format(row)] = '=IF(CG{}="","",CG{})'.format(
                    row, row)
                ws['DT{}'.format(row)] = '=IF(CH{}="","",CH{})'.format(
                    row, row)
                ws['DU{}'.format(row)] = '=IF(CI{}="","",CI{})'.format(
                    row, row)
                ws['DV{}'.format(row)] = '=IF(CJ{}="","",CJ{})'.format(
                    row, row)
                ws['DW{}'.format(row)] = '=IF(CK{}="","",CK{})'.format(
                    row, row)
                ws['DX{}'.format(row)] = '=IF(CL{}="","",CL{})'.format(
                    row, row)
            # Z Ke mapel
                ws['DY{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DQ{}="","",(DQ{}-DQ${})/DQ${}),2),"")'.format(row, row, r, s)
                ws['DZ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DR{}="","",(DR{}-DR${})/DR${}),2),"")'.format(row, row, r, s)
                ws['EA{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DS{}="","",(DS{}-DS${})/DS${}),2),"")'.format(row, row, r, s)
                ws['EB{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DT{}="","",(DT{}-DT${})/DT${}),2),"")'.format(row, row, r, s)
                ws['EC{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DU{}="","",(DU{}-DU${})/DU${}),2),"")'.format(row, row, r, s)
                ws['ED{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DV{}="","",(DV{}-DV${})/DV${}),2),"")'.format(row, row, r, s)
                ws['EE{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DW{}="","",(DW{}-DW${})/DW${}),2),"")'.format(row, row, r, s)
            # NILAI STANDAR ke mapel dan Z score
                ws['EF{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DQ{}="","",IF(70+30*DY{}/$DY${}<20,20,70+30*DY{}/$DY${})),2),"")'.format(row, row, r, row, r)
                ws['EG{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DR{}="","",IF(70+30*DZ{}/$DZ${}<20,20,70+30*DZ{}/$DZ${})),2),"")'.format(row, row, r, row, r)
                ws['EH{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DS{}="","",IF(70+30*EA{}/$EA${}<20,20,70+30*EA{}/$EA${})),2),"")'.format(row, row, r, row, r)
                ws['EI{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DT{}="","",IF(70+30*EB{}/$EB${}<20,20,70+30*EB{}/$EB${})),2),"")'.format(row, row, r, row, r)
                ws['EJ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DU{}="","",IF(70+30*EC{}/$EC${}<20,20,70+30*EC{}/$EC${})),2),"")'.format(row, row, r, row, r)
                ws['EK{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DV{}="","",IF(70+30*ED{}/$ED${}<20,20,70+30*ED{}/$ED${})),2),"")'.format(row, row, r, row, r)
                ws['EL{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DW{}="","",IF(70+30*EE{}/$EE${}<20,20,70+30*EE{}/$EE${})),2),"")'.format(row, row, r, row, r)
            # JUMLAH SELURUH NILAI STANDAR
                ws['EM{}'.format(row)] = '=IF(SUM(EF{}:EL{})=0,"",SUM(EF{}:EL{}))'.format(
                    row, row, row, row)
                ws['EN{}'.format(row)] = '=IF(EM{}="","",RANK(EM{},$EM$2:$EM${}))'.format(
                    row, row, q)
                ws['EO{}'.format(
                    row)] = '=IF(EN{}="","",COUNTIFS($DP$2:$DP${},DP{},$EN$2:$EN${},"<"&EN{})+1)'.format(row, q, row, q, row)
            # TAMBAHAN, MAPEL DAN NILAI STANDAR
                ws['EP{}'.format(row)] = '=IF($G${}=20,IF(AND(DQ{}>3,EF{}=20),1,""),IF($G${}=25,IF(AND(DQ{}>4,EF{}=20),1,""),IF($G${}=30,IF(AND(DQ{}>5,EF{}=20),1,""),IF($G${}=35,IF(AND(DQ{}>6,EF{}=20),1,""),IF($G${}=40,IF(AND(DQ{}>7,EF{}=20),1,""),IF($G${}=45,IF(AND(DQ{}>8,EF{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['EQ{}'.format(row)] = '=IF($H${}=20,IF(AND(DR{}>3,EG{}=20),1,""),IF($H${}=25,IF(AND(DR{}>4,EG{}=20),1,""),IF($H${}=30,IF(AND(DR{}>5,EG{}=20),1,""),IF($H${}=35,IF(AND(DR{}>6,EG{}=20),1,""),IF($H${}=40,IF(AND(DR{}>7,EG{}=20),1,""),IF($H${}=45,IF(AND(DR{}>8,EG{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['ER{}'.format(row)] = '=IF($I${}=20,IF(AND(DS{}>3,EH{}=20),1,""),IF($I${}=25,IF(AND(DS{}>4,EH{}=20),1,""),IF($I${}=30,IF(AND(DS{}>5,EH{}=20),1,""),IF($I${}=35,IF(AND(DS{}>6,EH{}=20),1,""),IF($I${}=40,IF(AND(DS{}>7,EH{}=20),1,""),IF($I${}=45,IF(AND(DS{}>8,EH{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['ES{}'.format(row)] = '=IF($J${}=20,IF(AND(DT{}>3,EI{}=20),1,""),IF($J${}=25,IF(AND(DT{}>4,EI{}=20),1,""),IF($J${}=30,IF(AND(DT{}>5,EI{}=20),1,""),IF($J${}=35,IF(AND(DT{}>6,EI{}=20),1,""),IF($J${}=40,IF(AND(DT{}>7,EI{}=20),1,""),IF($J${}=45,IF(AND(DT{}>8,EI{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['ET{}'.format(row)] = '=IF($K${}=20,IF(AND(DU{}>3,EJ{}=20),1,""),IF($K${}=25,IF(AND(DU{}>4,EJ{}=20),1,""),IF($K${}=30,IF(AND(DU{}>5,EJ{}=20),1,""),IF($K${}=35,IF(AND(DU{}>6,EJ{}=20),1,""),IF($K${}=40,IF(AND(DU{}>7,EJ{}=20),1,""),IF($K${}=45,IF(AND(DU{}>8,EJ{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['EU{}'.format(row)] = '=IF($L${}=20,IF(AND(DV{}>3,EK{}=20),1,""),IF($L${}=25,IF(AND(DV{}>4,EK{}=20),1,""),IF($L${}=30,IF(AND(DV{}>5,EK{}=20),1,""),IF($L${}=35,IF(AND(DV{}>6,EK{}=20),1,""),IF($L${}=40,IF(AND(DV{}>7,EK{}=20),1,""),IF($L${}=45,IF(AND(DV{}>8,EK{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['EV{}'.format(row)] = '=IF($M${}=20,IF(AND(DW{}>3,EL{}=20),1,""),IF($M${}=25,IF(AND(DW{}>4,EL{}=20),1,""),IF($M${}=30,IF(AND(DW{}>5,EL{}=20),1,""),IF($M${}=35,IF(AND(DW{}>6,EL{}=20),1,""),IF($M${}=40,IF(AND(DW{}>7,EL{}=20),1,""),IF($M${}=45,IF(AND(DW{}>8,EL{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

            # Z Score [5]
            ws['EX1'] = 'NAMA SISWA_E'
            ws['EY1'] = 'NOMOR NF_E'
            ws['EZ1'] = 'KELAS_E'
            ws['FA1'] = 'NAMA SEKOLAH_E'
            ws['FB1'] = 'LOKASI_E'

            ws['FC1'] = 'MAT_E'
            ws['FD1'] = 'IND_E'
            ws['FE1'] = 'ENG_E'
            ws['FF1'] = 'SEJ_E'
            ws['FG1'] = 'GEO_E'
            ws['FH1'] = 'EKO_E'
            ws['FI1'] = 'SOS_E'
            ws['FJ1'] = 'JML_E'

            ws['FK1'] = 'Z_MAT_E'
            ws['FL1'] = 'Z_IND_E'
            ws['FM1'] = 'Z_ENG_E'
            ws['FN1'] = 'Z_SEJ_E'
            ws['FO1'] = 'Z_GEO_E'
            ws['FP1'] = 'Z_EKO_E'
            ws['FQ1'] = 'Z_SOS_E'

            ws['FR1'] = 'S_MAT_E'
            ws['FS1'] = 'S_IND_E'
            ws['FT1'] = 'S_ENG_E'
            ws['FU1'] = 'S_SEJ_E'
            ws['FV1'] = 'S_GEO_E'
            ws['FW1'] = 'S_EKO_E'
            ws['FX1'] = 'S_SOS_E'
            ws['FY1'] = 'S_JML_E'

            ws['FZ1'] = 'RANK NAS._E'
            ws['GA1'] = 'RANK LOK._E'

            ws['FK1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FL1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FM1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FN1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FO1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FP1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FQ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FR1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FS1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FT1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FU1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FV1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FW1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FX1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FY1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FZ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['GA1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['EX1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EY1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EZ1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['FA1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['FB1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['FC1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['FD1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['FE1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['FF1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['FG1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['FH1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['FI1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['FJ1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['FK1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['FL1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['FM1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['FN1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['FO1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['FP1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['FQ1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['FR1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['FS1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['FT1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['FU1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['FV1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['FW1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['FX1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['FY1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['FZ1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['GA1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')

            # tambahan
            ws['GB1'] = 'MAT_20_E'
            ws['GC1'] = 'IND_20_E'
            ws['GD1'] = 'ENG_20_E'
            ws['GE1'] = 'SEJ_20_E'
            ws['GF1'] = 'GEO_20_E'
            ws['GG1'] = 'EKO_20_E'
            ws['GH1'] = 'SOS_20_E'

            ws['GB1'].font = Font(bold=False, name='Calibri', size=11)
            ws['GC1'].font = Font(bold=False, name='Calibri', size=11)
            ws['GD1'].font = Font(bold=False, name='Calibri', size=11)
            ws['GE1'].font = Font(bold=False, name='Calibri', size=11)
            ws['GF1'].font = Font(bold=False, name='Calibri', size=11)
            ws['GG1'].font = Font(bold=False, name='Calibri', size=11)
            ws['GH1'].font = Font(bold=False, name='Calibri', size=11)

            ws['GB1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['GC1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['GD1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['GE1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['GF1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['GG1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['GH1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')

            for row in range(2, q+1):
                ws['EX{}'.format(row)] = '=DL{}'.format(row)
                ws['EY{}'.format(row)] = '=DM{}'.format(row, row)
                ws['EZ{}'.format(row)] = '=DN{}'.format(row, row)
                ws['FA{}'.format(row)] = '=DO{}'.format(row, row)
                ws['FB{}'.format(row)] = '=DP{}'.format(row, row)
                ws['FC{}'.format(row)] = '=IF(DQ{}="","",DQ{})'.format(
                    row, row)
                ws['FD{}'.format(row)] = '=IF(DR{}="","",DR{})'.format(
                    row, row)
                ws['FE{}'.format(row)] = '=IF(DS{}="","",DS{})'.format(
                    row, row)
                ws['FF{}'.format(row)] = '=IF(DT{}="","",DT{})'.format(
                    row, row)
                ws['FG{}'.format(row)] = '=IF(DU{}="","",DU{})'.format(
                    row, row)
                ws['FH{}'.format(row)] = '=IF(DV{}="","",DV{})'.format(
                    row, row)
                ws['FI{}'.format(row)] = '=IF(DW{}="","",DW{})'.format(
                    row, row)
                ws['FJ{}'.format(row)] = '=IF(DX{}="","",DX{})'.format(
                    row, row)
            # Z Ke mapel
                ws['FK{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FC{}="","",(FC{}-FC${})/FC${}),2),"")'.format(row, row, r, s)
                ws['FL{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FD{}="","",(FD{}-FD${})/FD${}),2),"")'.format(row, row, r, s)
                ws['FM{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FE{}="","",(FE{}-FE${})/FE${}),2),"")'.format(row, row, r, s)
                ws['FN{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FF{}="","",(FF{}-FF${})/FF${}),2),"")'.format(row, row, r, s)
                ws['FO{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FG{}="","",(FG{}-FG${})/FG${}),2),"")'.format(row, row, r, s)
                ws['FP{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FH{}="","",(FH{}-FH${})/FH${}),2),"")'.format(row, row, r, s)
                ws['FQ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FI{}="","",(FI{}-FI${})/FI${}),2),"")'.format(row, row, r, s)
            # NILAI STANDAR ke mapel dan Z score
                ws['FR{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FC{}="","",IF(70+30*FK{}/$FK${}<20,20,70+30*FK{}/$FK${})),2),"")'.format(row, row, r, row, r)
                ws['FS{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FD{}="","",IF(70+30*FL{}/$FL${}<20,20,70+30*FL{}/$FL${})),2),"")'.format(row, row, r, row, r)
                ws['FT{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FE{}="","",IF(70+30*FM{}/$FM${}<20,20,70+30*FM{}/$FM${})),2),"")'.format(row, row, r, row, r)
                ws['FU{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FF{}="","",IF(70+30*FN{}/$FN${}<20,20,70+30*FN{}/$FN${})),2),"")'.format(row, row, r, row, r)
                ws['FV{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FG{}="","",IF(70+30*FO{}/$FO${}<20,20,70+30*FO{}/$FO${})),2),"")'.format(row, row, r, row, r)
                ws['FW{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FH{}="","",IF(70+30*FP{}/$FP${}<20,20,70+30*FP{}/$FP${})),2),"")'.format(row, row, r, row, r)
                ws['FX{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FI{}="","",IF(70+30*FQ{}/$FQ${}<20,20,70+30*FQ{}/$FQ${})),2),"")'.format(row, row, r, row, r)
            # JUMLAH SELURUH NILAI STANDAR
                ws['FY{}'.format(row)] = '=IF(SUM(FR{}:FX{})=0,"",SUM(FR{}:FX{}))'.format(
                    row, row, row, row)
                ws['FZ{}'.format(row)] = '=IF(FY{}="","",RANK(FY{},$FY$2:$FY${}))'.format(
                    row, row, q)
                ws['GA{}'.format(
                    row)] = '=IF(FZ{}="","",COUNTIFS($FB$2:$FB${},FB{},$FZ$2:$FZ${},"<"&FZ{})+1)'.format(row, q, row, q, row)
            # TAMBAHAN, MAPEL DAN NILAI STANDAR
                ws['GB{}'.format(row)] = '=IF($G${}=20,IF(AND(FC{}>3,FR{}=20),1,""),IF($G${}=25,IF(AND(FC{}>4,FR{}=20),1,""),IF($G${}=30,IF(AND(FC{}>5,FR{}=20),1,""),IF($G${}=35,IF(AND(FC{}>6,FR{}=20),1,""),IF($G${}=40,IF(AND(FC{}>7,FR{}=20),1,""),IF($G${}=45,IF(AND(FC{}>8,FR{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['GC{}'.format(row)] = '=IF($H${}=20,IF(AND(FD{}>3,FS{}=20),1,""),IF($H${}=25,IF(AND(FD{}>4,FS{}=20),1,""),IF($H${}=30,IF(AND(FD{}>5,FS{}=20),1,""),IF($H${}=35,IF(AND(FD{}>6,FS{}=20),1,""),IF($H${}=40,IF(AND(FD{}>7,FS{}=20),1,""),IF($H${}=45,IF(AND(FD{}>8,FS{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['GD{}'.format(row)] = '=IF($I${}=20,IF(AND(FE{}>3,FT{}=20),1,""),IF($I${}=25,IF(AND(FE{}>4,FT{}=20),1,""),IF($I${}=30,IF(AND(FE{}>5,FT{}=20),1,""),IF($I${}=35,IF(AND(FE{}>6,FT{}=20),1,""),IF($I${}=40,IF(AND(FE{}>7,FT{}=20),1,""),IF($I${}=45,IF(AND(FE{}>8,FT{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['GE{}'.format(row)] = '=IF($J${}=20,IF(AND(FF{}>3,FU{}=20),1,""),IF($J${}=25,IF(AND(FF{}>4,FU{}=20),1,""),IF($J${}=30,IF(AND(FF{}>5,FU{}=20),1,""),IF($J${}=35,IF(AND(FF{}>6,FU{}=20),1,""),IF($J${}=40,IF(AND(FF{}>7,FU{}=20),1,""),IF($J${}=45,IF(AND(FF{}>8,FU{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['GF{}'.format(row)] = '=IF($K${}=20,IF(AND(FG{}>3,FV{}=20),1,""),IF($K${}=25,IF(AND(FG{}>4,FV{}=20),1,""),IF($K${}=30,IF(AND(FG{}>5,FV{}=20),1,""),IF($K${}=35,IF(AND(FG{}>6,FV{}=20),1,""),IF($K${}=40,IF(AND(FG{}>7,FV{}=20),1,""),IF($K${}=45,IF(AND(FG{}>8,FV{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['GG{}'.format(row)] = '=IF($L${}=20,IF(AND(FH{}>3,FW{}=20),1,""),IF($L${}=25,IF(AND(FH{}>4,FW{}=20),1,""),IF($L${}=30,IF(AND(FH{}>5,FW{}=20),1,""),IF($L${}=35,IF(AND(FH{}>6,FW{}=20),1,""),IF($L${}=40,IF(AND(FH{}>7,FW{}=20),1,""),IF($L${}=45,IF(AND(FH{}>8,FW{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['GH{}'.format(row)] = '=IF($M${}=20,IF(AND(FI{}>3,FX{}=20),1,""),IF($M${}=25,IF(AND(FI{}>4,FX{}=20),1,""),IF($M${}=30,IF(AND(FI{}>5,FX{}=20),1,""),IF($M${}=35,IF(AND(FI{}>6,FX{}=20),1,""),IF($M${}=40,IF(AND(FI{}>7,FX{}=20),1,""),IF($M${}=45,IF(AND(FI{}>8,FX{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

            # Z Score
            ws['GJ1'] = 'NAMA SISWA'
            ws['GK1'] = 'NOMOR NF'
            ws['GL1'] = 'KELAS'
            ws['GM1'] = 'NAMA SEKOLAH'
            ws['GN1'] = 'LOKASI'

            ws['GO1'] = 'MAT'
            ws['GP1'] = 'IND'
            ws['GQ1'] = 'ENG'
            ws['GR1'] = 'SEJ'
            ws['GS1'] = 'GEO'
            ws['GT1'] = 'EKO'
            ws['GU1'] = 'SOS'
            ws['GV1'] = 'JML'

            ws['GW1'] = 'Z_MAT'
            ws['GX1'] = 'Z_IND'
            ws['GY1'] = 'Z_ENG'
            ws['GZ1'] = 'Z_SEJ'
            ws['HA1'] = 'Z_GEO'
            ws['HB1'] = 'Z_EKO'
            ws['HC1'] = 'Z_SOS'

            ws['HD1'] = 'S_MAT'
            ws['HE1'] = 'S_IND'
            ws['HF1'] = 'S_ENG'
            ws['HG1'] = 'S_SEJ'
            ws['HH1'] = 'S_GEO'
            ws['HI1'] = 'S_EKO'
            ws['HJ1'] = 'S_SOS'
            ws['HK1'] = 'S_JML'

            ws['HL1'] = 'RANK NAS.'
            ws['HM1'] = 'RANK LOK.'

            ws['GW1'].font = Font(bold=False, name='Calibri', size=11)
            ws['GX1'].font = Font(bold=False, name='Calibri', size=11)
            ws['GY1'].font = Font(bold=False, name='Calibri', size=11)
            ws['GZ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HA1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HB1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HC1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HD1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HE1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HF1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HG1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HH1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HI1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HJ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HK1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HL1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HM1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['GJ1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['GK1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['GL1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['GM1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['GN1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['GO1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['GP1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['GQ1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['GR1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['GS1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['GT1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['GU1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['GV1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['GW1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['GX1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['GY1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['GZ1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['HA1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['HB1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['HC1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['HD1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['HE1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['HF1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['HG1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['HH1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['HI1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['HJ1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['HK1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['HL1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['HM1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')

            # tambahan
            ws['HN1'] = 'MAT_20'
            ws['HO1'] = 'IND_20'
            ws['HP1'] = 'ENG_20'
            ws['HQ1'] = 'SEJ_20'
            ws['HR1'] = 'GEO_20'
            ws['HS1'] = 'EKO_20'
            ws['HT1'] = 'SOS_20'

            ws['HN1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HO1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HP1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HQ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HR1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HS1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HT1'].font = Font(bold=False, name='Calibri', size=11)

            ws['HN1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['HO1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['HP1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['HQ1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['HR1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['HS1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['HT1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')

            for row in range(2, q+1):
                ws['GJ{}'.format(row)] = '=EX{}'.format(row)
                ws['GK{}'.format(row)] = '=EY{}'.format(row, row)
                ws['GL{}'.format(row)] = '=EZ{}'.format(row, row)
                ws['GM{}'.format(row)] = '=FA{}'.format(row, row)
                ws['GN{}'.format(row)] = '=FB{}'.format(row, row)
                ws['GO{}'.format(row)] = '=IF(FC{}="","",FC{})'.format(
                    row, row)
                ws['GP{}'.format(row)] = '=IF(FD{}="","",FD{})'.format(
                    row, row)
                ws['GQ{}'.format(row)] = '=IF(FE{}="","",FE{})'.format(
                    row, row)
                ws['GR{}'.format(row)] = '=IF(FF{}="","",FF{})'.format(
                    row, row)
                ws['GS{}'.format(row)] = '=IF(FG{}="","",FG{})'.format(
                    row, row)
                ws['GT{}'.format(row)] = '=IF(FH{}="","",FH{})'.format(
                    row, row)
                ws['GU{}'.format(row)] = '=IF(FI{}="","",FI{})'.format(
                    row, row)
                ws['GV{}'.format(row)] = '=IF(FJ{}="","",FJ{})'.format(
                    row, row)
            # Z Ke mapel
                ws['GW{}'.format(
                    row)] = '=IFERROR(ROUND(IF(GO{}="","",(GO{}-GO${})/GO${}),2),"")'.format(row, row, r, s)
                ws['GX{}'.format(
                    row)] = '=IFERROR(ROUND(IF(GP{}="","",(GP{}-GP${})/GP${}),2),"")'.format(row, row, r, s)
                ws['GY{}'.format(
                    row)] = '=IFERROR(ROUND(IF(GQ{}="","",(GQ{}-GQ${})/GQ${}),2),"")'.format(row, row, r, s)
                ws['GZ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(GR{}="","",(GR{}-GR${})/GR${}),2),"")'.format(row, row, r, s)
                ws['HA{}'.format(
                    row)] = '=IFERROR(ROUND(IF(GS{}="","",(GS{}-GS${})/GS${}),2),"")'.format(row, row, r, s)
                ws['HB{}'.format(
                    row)] = '=IFERROR(ROUND(IF(GT{}="","",(GT{}-GT${})/GT${}),2),"")'.format(row, row, r, s)
                ws['HC{}'.format(
                    row)] = '=IFERROR(ROUND(IF(GU{}="","",(GU{}-GU${})/GU${}),2),"")'.format(row, row, r, s)
            # NILAI STANDAR ke mapel dan Z score
                ws['HD{}'.format(
                    row)] = '=IFERROR(ROUND(IF(GO{}="","",IF(70+30*GW{}/$GW${}<20,20,70+30*GW{}/$GW${})),2),"")'.format(row, row, r, row, r)
                ws['HE{}'.format(
                    row)] = '=IFERROR(ROUND(IF(GP{}="","",IF(70+30*GX{}/$GX${}<20,20,70+30*GX{}/$GX${})),2),"")'.format(row, row, r, row, r)
                ws['HF{}'.format(
                    row)] = '=IFERROR(ROUND(IF(GQ{}="","",IF(70+30*GY{}/$GY${}<20,20,70+30*GY{}/$GY${})),2),"")'.format(row, row, r, row, r)
                ws['HG{}'.format(
                    row)] = '=IFERROR(ROUND(IF(GR{}="","",IF(70+30*GZ{}/$GZ${}<20,20,70+30*GZ{}/$GZ${})),2),"")'.format(row, row, r, row, r)
                ws['HH{}'.format(
                    row)] = '=IFERROR(ROUND(IF(GS{}="","",IF(70+30*HA{}/$HA${}<20,20,70+30*HA{}/$HA${})),2),"")'.format(row, row, r, row, r)
                ws['HI{}'.format(
                    row)] = '=IFERROR(ROUND(IF(GT{}="","",IF(70+30*HB{}/$HB${}<20,20,70+30*HB{}/$HB${})),2),"")'.format(row, row, r, row, r)
                ws['HJ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(GU{}="","",IF(70+30*HC{}/$HC${}<20,20,70+30*HC{}/$HC${})),2),"")'.format(row, row, r, row, r)
            # JUMLAH SELURUH NILAI STANDAR
                ws['HK{}'.format(row)] = '=IF(SUM(HD{}:HJ{})=0,"",SUM(HD{}:HJ{}))'.format(
                    row, row, row, row)
                ws['HL{}'.format(row)] = '=IF(HK{}="","",RANK(HK{},$HK$2:$HK${}))'.format(
                    row, row, q)
                ws['HM{}'.format(
                    row)] = '=IF(HL{}="","",COUNTIFS($GN$2:$GN${},GN{},$HL$2:$HL${},"<"&HL{})+1)'.format(row, q, row, q, row)
            # TAMBAHAN, MAPEL DAN NILAI STANDAR
                ws['HN{}'.format(row)] = '=IF($G${}=20,IF(AND(GO{}>3,HD{}=20),1,""),IF($G${}=25,IF(AND(GO{}>4,HD{}=20),1,""),IF($G${}=30,IF(AND(GO{}>5,HD{}=20),1,""),IF($G${}=35,IF(AND(GO{}>6,HD{}=20),1,""),IF($G${}=40,IF(AND(GO{}>7,HD{}=20),1,""),IF($G${}=45,IF(AND(GO{}>8,HD{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['HO{}'.format(row)] = '=IF($H${}=20,IF(AND(GP{}>3,HE{}=20),1,""),IF($H${}=25,IF(AND(GP{}>4,HE{}=20),1,""),IF($H${}=30,IF(AND(GP{}>5,HE{}=20),1,""),IF($H${}=35,IF(AND(GP{}>6,HE{}=20),1,""),IF($H${}=40,IF(AND(GP{}>7,HE{}=20),1,""),IF($H${}=45,IF(AND(GP{}>8,HE{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['HP{}'.format(row)] = '=IF($I${}=20,IF(AND(GQ{}>3,HF{}=20),1,""),IF($I${}=25,IF(AND(GQ{}>4,HF{}=20),1,""),IF($I${}=30,IF(AND(GQ{}>5,HF{}=20),1,""),IF($I${}=35,IF(AND(GQ{}>6,HF{}=20),1,""),IF($I${}=40,IF(AND(GQ{}>7,HF{}=20),1,""),IF($I${}=45,IF(AND(GQ{}>8,HF{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['HQ{}'.format(row)] = '=IF($J${}=20,IF(AND(GR{}>3,HG{}=20),1,""),IF($J${}=25,IF(AND(GR{}>4,HG{}=20),1,""),IF($J${}=30,IF(AND(GR{}>5,HG{}=20),1,""),IF($J${}=35,IF(AND(GR{}>6,HG{}=20),1,""),IF($J${}=40,IF(AND(GR{}>7,HG{}=20),1,""),IF($J${}=45,IF(AND(GR{}>8,HG{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['HR{}'.format(row)] = '=IF($K${}=20,IF(AND(GS{}>3,HH{}=20),1,""),IF($K${}=25,IF(AND(GS{}>4,HH{}=20),1,""),IF($K${}=30,IF(AND(GS{}>5,HH{}=20),1,""),IF($K${}=35,IF(AND(GS{}>6,HH{}=20),1,""),IF($K${}=40,IF(AND(GS{}>7,HH{}=20),1,""),IF($K${}=45,IF(AND(GS{}>8,HH{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['HS{}'.format(row)] = '=IF($L${}=20,IF(AND(GT{}>3,HI{}=20),1,""),IF($L${}=25,IF(AND(GT{}>4,HI{}=20),1,""),IF($L${}=30,IF(AND(GT{}>5,HI{}=20),1,""),IF($L${}=35,IF(AND(GT{}>6,HI{}=20),1,""),IF($L${}=40,IF(AND(GT{}>7,HI{}=20),1,""),IF($L${}=45,IF(AND(GT{}>8,HI{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['HT{}'.format(row)] = '=IF($M${}=20,IF(AND(GU{}>3,HJ{}=20),1,""),IF($M${}=25,IF(AND(GU{}>4,HJ{}=20),1,""),IF($M${}=30,IF(AND(GU{}>5,HJ{}=20),1,""),IF($M${}=35,IF(AND(GU{}>6,HJ{}=20),1,""),IF($M${}=40,IF(AND(GU{}>7,HJ{}=20),1,""),IF($M${}=45,IF(AND(GU{}>8,HJ{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

            # Mengubah 'KELAS' sesuai dengan nilai yang dipilih dari selectbox 'KELAS'
            kelas = KELAS.lower().replace(" ", "")
            semester = SEMESTER.lower()
            tahun = TAHUN.replace("-", "")
            penilaian = PENILAIAN.lower()
            kurikulum = KURIKULUM.lower()

            path_file = f"{kelas}_{penilaian}_{semester}_{kurikulum}_{tahun}_nilai_std.xlsx"

            # Simpan file ke direktori temporer
            temp_dir = tempfile.gettempdir()
            file_path = temp_dir + '/' + path_file
            wb.save(file_path)

            st.success(
                "File siap diunduh!")

            # Tombol unduh file
            with open(file_path, "rb") as f:
                bytes_data = f.read()
            st.download_button(label="Unduh File", data=bytes_data,
                               file_name=path_file)

            st.warning(
                "Buka file unduhan, klik 'Enable Editing' dan 'Save'")
            
    if selected_file == "Nilai Std. 11 SMA (KM)":
        # menghilangkan hamburger
        st.markdown("""
        <style>
        .css-1rs6os.edgvbvh3
        {
            visibility:hidden;
        }
        .css-1lsmgbg.egzxvld0
        {
            visibility:hidden;
        }
        </style>
        """, unsafe_allow_html=True)

        image = Image.open('logo resmi nf resize.png')
        st.image(image)

        st.title("Olah Nilai Standar KM")
        st.header("11 SMA")

        col6 = st.container()

        with col6:
            KELAS = st.selectbox(
                "KELAS",
                ("--Pilih Kelas--", "11 SMA"))

        col7 = st.container()

        with col7:
            SEMESTER = st.selectbox(
                "SEMESTER",
                ("--Pilih Semester--", "SEMESTER 1", "SEMESTER 2"))

        col8 = st.container()

        with col8:
            PENILAIAN = st.selectbox(
                "PENILAIAN",
                ("--Pilih Penilaian--", "SUMATIF TENGAH SEMESTER", "SUMATIF AKHIR SEMESTER", "SUMATIF AKHIR TAHUN"))

        col9 = st.container()

        with col9:
            KURIKULUM = st.selectbox(
                "KURIKULUM",
                ("--Pilih Kurikulum--", "KM"))

        TAHUN = st.text_input("Masukkan Tahun Ajaran",
                              placeholder="contoh: 2022-2023")

        col1, col2, col3, col4 = st.columns(4)

        with col1:
            MTK_1 = st.selectbox(
                "JML. SOAL MAT_1.",
                (15, 20, 25, 30, 35, 40, 45, 50))
        
        with col2:
            MTK_2 = st.selectbox(
                "JML. SOAL MAT._2",
                (15, 20, 25, 30, 35, 40, 45, 50))

        with col3:
            IND = st.selectbox(
                "JML. SOAL IND.",
                (15, 20, 25, 30, 35, 40, 45, 50))

        with col4:
            ENG = st.selectbox(
                "JML. SOAL ENG.",
                (15, 20, 25, 30, 35, 40, 45, 50))

        col5, col6, col7, col8, col9 = st.columns(5)
        with col5:
            SEJ = st.selectbox(
                "JML. SOAL SEJ.",
                (15, 20, 25, 30, 35, 40, 45, 50))

        with col6:
            GEO = st.selectbox(
                "JML. SOAL GEO.",
                (15, 20, 25, 30, 35, 40, 45, 50))

        with col7:
            EKO = st.selectbox(
                "JML. SOAL EKO.",
                (15, 20, 25, 30, 35, 40, 45, 50))

        with col8:
            SOS = st.selectbox(
                "JML. SOAL SOS.",
                (15, 20, 25, 30, 35, 40, 45, 50))
        
        with col9:
            ANT = st.selectbox(
                "JML. SOAL ANT.",
                (15, 20, 25, 30, 35, 40, 45, 50))

        col10, col11, col12, col13 = st.columns(4)
        with col10:
            BIO = st.selectbox(
                "JML. SOAL BIO.",
                (15, 20, 25, 30, 35, 40, 45, 50))

        with col11:
            FIS = st.selectbox(
                "JML. SOAL FIS.",
                (15, 20, 25, 30, 35, 40, 45, 50))

        with col12:
            KIM_1 = st.selectbox(
                "JML. SOAL KIM_1.",
                (15, 20, 25, 30, 35, 40, 45, 50))
        
        with col13:
            KIM_2 = st.selectbox(
                "JML. SOAL KIM_2.",
                (15, 20, 25, 30, 35, 40, 45, 50))

        JML_SOAL_MAT_1 = MTK_1
        JML_SOAL_MAT_2 = MTK_2
        JML_SOAL_IND = IND
        JML_SOAL_ENG = ENG
        JML_SOAL_SEJ = SEJ
        JML_SOAL_GEO = GEO
        JML_SOAL_EKO = EKO
        JML_SOAL_SOS = SOS
        JML_SOAL_ANT = ANT
        JML_SOAL_BIO = BIO
        JML_SOAL_FIS = FIS
        JML_SOAL_KIM_1 = KIM_1
        JML_SOAL_KIM_2 = KIM_2

        uploaded_file = st.file_uploader(
            'Letakkan file excel Pivot', type='xlsx')

        if uploaded_file is not None:
            wb = openpyxl.load_workbook(uploaded_file)
            ws = wb['Sheet1']

            q = len(ws['K'])
            r = len(ws['K'])+2
            s = len(ws['K'])+3
            t = len(ws['K'])+4
            u = len(ws['K'])+5
            v = len(ws['K'])+6
            w = len(ws['K'])+7
            x = len(ws['K'])+8

            # JML BENAR MAT 1, MAT 2, IND, ENG
            ws['G{}'.format(r)] = "=ROUND(AVERAGE(G2:G{}),2)".format(q)
            ws['H{}'.format(r)] = "=ROUND(AVERAGE(H2:H{}),2)".format(q)
            ws['I{}'.format(r)] = "=ROUND(AVERAGE(I2:I{}),2)".format(q)
            ws['J{}'.format(r)] = "=ROUND(AVERAGE(J2:J{}),2)".format(q)
            # JML BENAR SEJ, EKO, SOS, GEO, ANT
            ws['K{}'.format(r)] = "=ROUND(AVERAGE(K2:K{}),2)".format(q)
            ws['L{}'.format(r)] = "=ROUND(AVERAGE(L2:L{}),2)".format(q)
            ws['M{}'.format(r)] = "=ROUND(AVERAGE(M2:M{}),2)".format(q)
            ws['N{}'.format(r)] = "=ROUND(AVERAGE(N2:N{}),2)".format(q)
            ws['O{}'.format(r)] = "=ROUND(AVERAGE(O2:O{}),2)".format(q)
            # JML BENAR BIO, FIS, KIM 1, KIM 2
            ws['P{}'.format(r)] = "=ROUND(AVERAGE(P2:P{}),2)".format(q)
            ws['Q{}'.format(r)] = "=ROUND(AVERAGE(Q2:Q{}),2)".format(q)
            ws['R{}'.format(r)] = "=ROUND(AVERAGE(R2:R{}),2)".format(q)
            ws['S{}'.format(r)] = "=ROUND(AVERAGE(S2:S{}),2)".format(q)

            # STDEV MAT 1, MAT 2, IND, ENG
            ws['G{}'.format(s)] = "=STDEV(G2:G{})".format(q)
            ws['H{}'.format(s)] = "=STDEV(H2:H{})".format(q)
            ws['I{}'.format(s)] = "=STDEV(I2:I{})".format(q)
            ws['J{}'.format(s)] = "=STDEV(J2:J{})".format(q)
            # STDEV SEJ, EKO, SOS, GEO, ANT
            ws['K{}'.format(s)] = "=STDEV(K2:K{})".format(q)
            ws['L{}'.format(s)] = "=STDEV(L2:L{})".format(q)
            ws['M{}'.format(s)] = "=STDEV(M2:M{})".format(q)
            ws['N{}'.format(s)] = "=STDEV(N2:N{})".format(q)
            ws['O{}'.format(s)] = "=STDEV(O2:O{})".format(q)
            # STDEV BIO, FIS, KIM 1, KIM 2
            ws['P{}'.format(s)] = "=STDEV(P2:P{})".format(q)
            ws['Q{}'.format(s)] = "=STDEV(Q2:Q{})".format(q)
            ws['R{}'.format(s)] = "=STDEV(R2:R{})".format(q)
            ws['S{}'.format(s)] = "=STDEV(S2:S{})".format(q)

            # MAX MAT 1, MAT 2, IND, ENG
            ws['G{}'.format(t)] = "=MAX(G2:G{})".format(q)
            ws['H{}'.format(t)] = "=MAX(H2:H{})".format(q)
            ws['I{}'.format(t)] = "=MAX(I2:I{})".format(q)
            ws['J{}'.format(t)] = "=MAX(J2:J{})".format(q)
            # MAX SEJ, EKO, SOS, GEO, ANT
            ws['K{}'.format(t)] = "=MAX(K2:K{})".format(q)
            ws['L{}'.format(t)] = "=MAX(L2:L{})".format(q)
            ws['M{}'.format(t)] = "=MAX(M2:M{})".format(q)
            ws['N{}'.format(t)] = "=MAX(N2:N{})".format(q)
            ws['O{}'.format(t)] = "=MAX(O2:O{})".format(q)
            # MAX BIO, FIS, KIM 1, KIM 2
            ws['P{}'.format(t)] = "=MAX(P2:P{})".format(q)
            ws['Q{}'.format(t)] = "=MAX(Q2:Q{})".format(q)
            ws['R{}'.format(t)] = "=MAX(R2:R{})".format(q)
            ws['S{}'.format(t)] = "=MAX(S2:S{})".format(q)

            # JML
            ws['T{}'.format(r)] = "=MAX(T2:T{})".format(q)
            # Z MAT 1, MAT 2, IND, ENG
            ws['U{}'.format(r)] = "=MAX(U2:U{})".format(q)
            ws['V{}'.format(r)] = "=MAX(V2:V{})".format(q)
            ws['W{}'.format(r)] = "=MAX(W2:W{})".format(q)
            ws['X{}'.format(r)] = "=MAX(X2:X{})".format(q)
            # Z SEJ, EKO, SOS, GEO, ANT
            ws['Y{}'.format(r)] = "=MAX(Y2:Y{})".format(q)
            ws['Z{}'.format(r)] = "=MAX(Z2:Z{})".format(q)
            ws['AA{}'.format(r)] = "=MAX(AA2:AA{})".format(q)
            ws['AB{}'.format(r)] = "=MAX(AB2:AB{})".format(q)
            ws['AC{}'.format(r)] = "=MAX(AC2:AC{})".format(q)
            # Z BIO, FIS, KIM 1, KIM 2
            ws['AD{}'.format(r)] = "=MAX(AD2:AD{})".format(q)
            ws['AE{}'.format(r)] = "=MAX(AE2:AE{})".format(q)
            ws['AF{}'.format(r)] = "=MAX(AF2:AF{})".format(q)
            ws['AG{}'.format(r)] = "=MAX(AG2:AG{})".format(q)

            # S MAT 1, MAT 2, IND, ENG
            ws['AH{}'.format(r)] = "=MAX(AH2:AH{})".format(q)
            ws['AI{}'.format(r)] = "=MAX(AI2:AI{})".format(q)
            ws['AJ{}'.format(r)] = "=MAX(AJ2:AJ{})".format(q)
            ws['AK{}'.format(r)] = "=MAX(AK2:AK{})".format(q)
            # S SEJ, EKO, SOS, GEO, ANT
            ws['AL{}'.format(r)] = "=MAX(AL2:AL{})".format(q)
            ws['AM{}'.format(r)] = "=MAX(AM2:AM{})".format(q)
            ws['AN{}'.format(r)] = "=MAX(AN2:AN{})".format(q)
            ws['AO{}'.format(r)] = "=MAX(AO2:AO{})".format(q)
            ws['AP{}'.format(r)] = "=MAX(AP2:AP{})".format(q)
            # S BIO, FIS, KIM 1, KIM 2
            ws['AQ{}'.format(r)] = "=MAX(AQ2:AQ{})".format(q)
            ws['AR{}'.format(r)] = "=MAX(AR2:AR{})".format(q)
            ws['AS{}'.format(r)] = "=MAX(AS2:AS{})".format(q)
            ws['AT{}'.format(r)] = "=MAX(AT2:AT{})".format(q)
            # S JML, RANK NAS, RANK LOK
            ws['AU{}'.format(r)] = "=MAX(AU2:AU{})".format(q)
            ws['AV{}'.format(r)] = "=MAX(AV2:AV{})".format(q)
            ws['AW{}'.format(r)] = "=MAX(AW2:AW{})".format(q)

            # MIN MAT 1, MAT 2, IND, ENG
            ws['G{}'.format(u)] = "=MIN(G2:G{})".format(q)
            ws['H{}'.format(u)] = "=MIN(H2:H{})".format(q)
            ws['I{}'.format(u)] = "=MIN(I2:I{})".format(q)
            ws['J{}'.format(u)] = "=MIN(J2:J{})".format(q)
            # MIN SEJ, EKO, SOS, GEO, ANT
            ws['K{}'.format(u)] = "=MIN(K2:K{})".format(q)
            ws['L{}'.format(u)] = "=MIN(L2:L{})".format(q)
            ws['M{}'.format(u)] = "=MIN(M2:M{})".format(q)
            ws['N{}'.format(u)] = "=MIN(N2:N{})".format(q)
            ws['O{}'.format(u)] = "=MIN(O2:O{})".format(q)
            # MIN BIO, FIS, KIM 1, KIM 2
            ws['P{}'.format(u)] = "=MIN(P2:P{})".format(q)
            ws['Q{}'.format(u)] = "=MIN(Q2:Q{})".format(q)
            ws['R{}'.format(u)] = "=MIN(R2:R{})".format(q)
            ws['S{}'.format(u)] = "=MIN(S2:S{})".format(q)

            # S MAT 1, MAT 2, IND, ENG
            ws['AH{}'.format(s)] = "=MIN(AH2:AH{})".format(q)
            ws['AI{}'.format(s)] = "=MIN(AI2:AI{})".format(q)
            ws['AJ{}'.format(s)] = "=MIN(AJ2:AJ{})".format(q)
            ws['AK{}'.format(s)] = "=MIN(AK2:AK{})".format(q)
            # S SEJ, EKO, SOS, GEO, ANT
            ws['AL{}'.format(s)] = "=MIN(AL2:AL{})".format(q)
            ws['AM{}'.format(s)] = "=MIN(AM2:AM{})".format(q)
            ws['AN{}'.format(s)] = "=MIN(AN2:AN{})".format(q)
            ws['AO{}'.format(s)] = "=MIN(AO2:AO{})".format(q)
            ws['AP{}'.format(s)] = "=MIN(AP2:AP{})".format(q)
            # S BIO, FIS, KIM 1, KIM 2
            ws['AQ{}'.format(s)] = "=MIN(AQ2:AQ{})".format(q)
            ws['AR{}'.format(s)] = "=MIN(AR2:AR{})".format(q)
            ws['AS{}'.format(s)] = "=MIN(AS2:AS{})".format(q)
            ws['AT{}'.format(s)] = "=MIN(AT2:AT{})".format(q)
            # S JML
            ws['AU{}'.format(s)] = "=MIN(AU2:AU{})".format(q)
            
            # S MAT 1, MAT 2, IND, ENG
            ws['AH{}'.format(t)] = "=ROUND(AVERAGE(AH2:AH{}),2)".format(q)
            ws['AI{}'.format(t)] = "=ROUND(AVERAGE(AI2:AI{}),2)".format(q)
            ws['AJ{}'.format(t)] = "=ROUND(AVERAGE(AJ2:AJ{}),2)".format(q)
            ws['AK{}'.format(t)] = "=ROUND(AVERAGE(AK2:AK{}),2)".format(q)
            # S SEJ, EKO, SOS, GEO, ANT
            ws['AL{}'.format(t)] = "=ROUND(AVERAGE(AL2:AL{}),2)".format(q)
            ws['AM{}'.format(t)] = "=ROUND(AVERAGE(AM2:AM{}),2)".format(q)
            ws['AN{}'.format(t)] = "=ROUND(AVERAGE(AN2:AN{}),2)".format(q)
            ws['AO{}'.format(t)] = "=ROUND(AVERAGE(AO2:AO{}),2)".format(q)
            ws['AP{}'.format(t)] = "=ROUND(AVERAGE(AP2:AP{}),2)".format(q)
            # S BIO, FIS, KIM 1, KIM 2
            ws['AQ{}'.format(t)] = "=ROUND(AVERAGE(AQ2:AQ{}),2)".format(q)
            ws['AR{}'.format(t)] = "=ROUND(AVERAGE(AR2:AR{}),2)".format(q)
            ws['AS{}'.format(t)] = "=ROUND(AVERAGE(AS2:AS{}),2)".format(q)
            ws['AT{}'.format(t)] = "=ROUND(AVERAGE(AT2:AT{}),2)".format(q)
            # S JML
            ws['AU{}'.format(t)] = "=ROUND(AVERAGE(AU2:AU{}),2)".format(q)

            # MAT 1 20, MAT 2 20, IND 20, ENG 20 
            ws['AX{}'.format(r)] = "=SUM(AX2:AX{})".format(q)
            ws['AY{}'.format(r)] = "=SUM(AY2:AY{})".format(q)
            ws['AZ{}'.format(r)] = "=SUM(AZ2:AZ{})".format(q)
            ws['BA{}'.format(r)] = "=SUM(BA2:BA{})".format(q)
            # SEJ 20, EKO 2 20, SOS 20, GEO 20, ANT 20
            ws['BB{}'.format(r)] = "=SUM(BB2:BB{})".format(q)
            ws['BC{}'.format(r)] = "=SUM(BC2:BC{})".format(q)
            ws['BD{}'.format(r)] = "=SUM(BD2:BD{})".format(q)
            ws['BE{}'.format(r)] = "=SUM(BE2:BE{})".format(q)
            ws['BF{}'.format(r)] = "=SUM(BF2:BF{})".format(q)
            # BIO 20, FIS 20, KIM 1 20, KIM 2 20
            ws['BG{}'.format(r)] = "=SUM(BG2:BG{})".format(q)
            ws['BH{}'.format(r)] = "=SUM(BH2:BH{})".format(q)
            ws['BI{}'.format(r)] = "=SUM(BI2:BI{})".format(q)
            ws['BJ{}'.format(r)] = "=SUM(BJ2:BJ{})".format(q)

            # Jumlah Soal
            ws['F{}'.format(v)] = 'JUMLAH SOAL'
            ws['G{}'.format(v)] = JML_SOAL_MAT_1
            ws['H{}'.format(v)] = JML_SOAL_MAT_2
            ws['I{}'.format(v)] = JML_SOAL_IND
            ws['J{}'.format(v)] = JML_SOAL_ENG
            ws['K{}'.format(v)] = JML_SOAL_SEJ
            ws['L{}'.format(v)] = JML_SOAL_GEO
            ws['M{}'.format(v)] = JML_SOAL_EKO
            ws['N{}'.format(v)] = JML_SOAL_SOS
            ws['O{}'.format(v)] = JML_SOAL_ANT
            ws['P{}'.format(v)] = JML_SOAL_BIO
            ws['Q{}'.format(v)] = JML_SOAL_FIS
            ws['R{}'.format(v)] = JML_SOAL_KIM_1
            ws['S{}'.format(v)] = JML_SOAL_KIM_2

            # new
            # iterasi 1 rata-rata - 1
            # rata" MTK 1 ke MTK 1 tambahan dan mapel MTK 1 awal
            ws['BQ{}'.format(r)] = "=IF($AX${}=0,$G${},$G${}-1)".format(r, r, r)
            ws['BQ{}'.format(s)] = "=STDEV(BQ2:BQ{})".format(q)
            ws['BQ{}'.format(t)] = "=MAX(BQ2:BQ{})".format(q)
            ws['BQ{}'.format(u)] = "=MIN(BQ2:BQ{})".format(q)
            # rata" MTK 2 ke MTK 2 tambahan dan mapel MTK 2 awal
            ws['BR{}'.format(r)] = "=IF($AY${}=0,$H${},$H${}-1)".format(r, r, r)
            ws['BR{}'.format(s)] = "=STDEV(BR2:BR{})".format(q)
            ws['BR{}'.format(t)] = "=MAX(BR2:BR{})".format(q)
            ws['BR{}'.format(u)] = "=MIN(BR2:BR{})".format(q)
            # rata" IND ke IND tambahan dan mapel IND awal
            ws['BS{}'.format(r)] = "=IF($AZ${}=0,$I${},$I${}-1)".format(r, r, r)
            ws['BS{}'.format(s)] = "=STDEV(BS2:BS{})".format(q)
            ws['BS{}'.format(t)] = "=MAX(BS2:BS{})".format(q)
            ws['BS{}'.format(u)] = "=MIN(BS2:BS{})".format(q)
            # rata" ENG ke ENG tambahan dan mapel ENG awal
            ws['BT{}'.format(r)] = "=IF($BA${}=0,$J${},$J${}-1)".format(r, r, r)
            ws['BT{}'.format(s)] = "=STDEV(BT2:BT{})".format(q)
            ws['BT{}'.format(t)] = "=MAX(BT2:BT{})".format(q)
            ws['BT{}'.format(u)] = "=MIN(BT2:BT{})".format(q)
            # rata" SEJ ke SEJ tambahan dan mapel SEJ awal
            ws['BU{}'.format(r)] = "=IF($BB${}=0,$K${},$K${}-1)".format(r, r, r)
            ws['BU{}'.format(s)] = "=STDEV(BU2:BU{})".format(q)
            ws['BU{}'.format(t)] = "=MAX(BU2:BU{})".format(q)
            ws['BU{}'.format(u)] = "=MIN(BU2:BU{})".format(q)
            # rata" EKO ke EKO tambahan dan mapel EKO awal
            ws['BV{}'.format(r)] = "=IF($BC${}=0,$L${},$L${}-1)".format(r, r, r)
            ws['BV{}'.format(s)] = "=STDEV(BV2:BV{})".format(q)
            ws['BV{}'.format(t)] = "=MAX(BV2:BV{})".format(q)
            ws['BV{}'.format(u)] = "=MIN(BV2:BV{})".format(q)
            # rata" SOS ke SOS tambahan dan mapel SOS awal
            ws['BW{}'.format(r)] = "=IF($BD${}=0,$M${},$M${}-1)".format(r, r, r)
            ws['BW{}'.format(s)] = "=STDEV(BW2:BW{})".format(q)
            ws['BW{}'.format(t)] = "=MAX(BW2:BW{})".format(q)
            ws['BW{}'.format(u)] = "=MIN(BW2:BW{})".format(q)
            # rata" GEO ke GEO tambahan dan mapel GEO awal
            ws['BX{}'.format(r)] = "=IF($BE${}=0,$N${},$N${}-1)".format(r, r, r)
            ws['BX{}'.format(s)] = "=STDEV(BX2:BX{})".format(q)
            ws['BX{}'.format(t)] = "=MAX(BX2:BX{})".format(q)
            ws['BX{}'.format(u)] = "=MIN(BX2:BX{})".format(q)
            # rata" ANT ke ANT tambahan dan mapel ANT awal
            ws['BY{}'.format(r)] = "=IF($BF${}=0,$O${},$O${}-1)".format(r, r, r)
            ws['BY{}'.format(s)] = "=STDEV(BY2:BY{})".format(q)
            ws['BY{}'.format(t)] = "=MAX(BY2:BY{})".format(q)
            ws['BY{}'.format(u)] = "=MIN(BY2:BY{})".format(q)
            # rata" BIO ke BIO tambahan dan mapel BIO awal
            ws['BZ{}'.format(r)] = "=IF($BG${}=0,$P${},$P${}-1)".format(r, r, r)
            ws['BZ{}'.format(s)] = "=STDEV(BZ2:BZ{})".format(q)
            ws['BZ{}'.format(t)] = "=MAX(BZ2:BZ{})".format(q)
            ws['BZ{}'.format(u)] = "=MIN(BZ2:BZ{})".format(q)
            # rata" FIS ke FIS tambahan dan mapel FIS awal
            ws['CA{}'.format(r)] = "=IF($BH${}=0,$Q${},$Q${}-1)".format(r, r, r)
            ws['CA{}'.format(s)] = "=STDEV(CA2:CA{})".format(q)
            ws['CA{}'.format(t)] = "=MAX(CA2:CA{})".format(q)
            ws['CA{}'.format(u)] = "=MIN(CA2:CA{})".format(q)
            # rata" KIM 1 ke KIM 1 tambahan dan mapel KIM 1 awal
            ws['CB{}'.format(r)] = "=IF($BI${}=0,$R${},$R${}-1)".format(r, r, r)
            ws['CB{}'.format(s)] = "=STDEV(CB2:CB{})".format(q)
            ws['CB{}'.format(t)] = "=MAX(CB2:CB{})".format(q)
            ws['CB{}'.format(u)] = "=MIN(CB2:CB{})".format(q)
            # rata" KIM 2 ke KIM 2 tambahan dan mapel KIM 2 awal
            ws['CC{}'.format(r)] = "=IF($BJ${}=0,$S${},$S${}-1)".format(r, r, r)
            ws['CC{}'.format(s)] = "=STDEV(CC2:CC{})".format(q)
            ws['CC{}'.format(t)] = "=MAX(CC2:CC{})".format(q)
            ws['CC{}'.format(u)] = "=MIN(CC2:CC{})".format(q)
            # JML BENAR
            ws['CD{}'.format(r)] = "=ROUND(AVERAGE(CD2:CD{}),2)".format(q)
            ws['CD{}'.format(t)] = "=MAX(CD2:CD{})".format(q)
            ws['CD{}'.format(u)] = "=MIN(CD2:CD{})".format(q)
            # MAX Z SCORE MAT 1, MAT 2, IND, ENG
            ws['CE{}'.format(r)] = "=MAX(CE2:CE{})".format(q)
            ws['CF{}'.format(r)] = "=MAX(CF2:CF{})".format(q)
            ws['CG{}'.format(r)] = "=MAX(CG2:CG{})".format(q)
            ws['CH{}'.format(r)] = "=MAX(CH2:CH{})".format(q)
            # MAX Z SCORE SEJ, EKO, SOS, GEO, ANT
            ws['CI{}'.format(r)] = "=MAX(CI2:CI{})".format(q)
            ws['CJ{}'.format(r)] = "=MAX(CJ2:CJ{})".format(q)
            ws['CK{}'.format(r)] = "=MAX(CK2:CK{})".format(q)
            ws['CL{}'.format(r)] = "=MAX(CL2:CL{})".format(q)
            ws['CM{}'.format(r)] = "=MAX(CM2:CM{})".format(q)
            # MAX Z SCORE BIO, FIS, KIM 1, KIM 2
            ws['CN{}'.format(r)] = "=MAX(CN2:CN{})".format(q)
            ws['CO{}'.format(r)] = "=MAX(CO2:CO{})".format(q)
            ws['CP{}'.format(r)] = "=MAX(CP2:CP{})".format(q)
            ws['CQ{}'.format(r)] = "=MAX(CQ2:CQ{})".format(q)

            # NILAI STANDAR MTK 1
            ws['CR{}'.format(r)] = "=MAX(CR2:CR{})".format(q)
            ws['CR{}'.format(s)] = "=MIN(CR2:CR{})".format(q)
            ws['CR{}'.format(t)] = "=ROUND(AVERAGE(CR2:CR{}),2)".format(q)
            # NILAI STANDAR MTK 1
            ws['CS{}'.format(r)] = "=MAX(CS2:CS{})".format(q)
            ws['CS{}'.format(s)] = "=MIN(CS2:CS{})".format(q)
            ws['CS{}'.format(t)] = "=ROUND(AVERAGE(CS2:CS{}),2)".format(q)
            # NILAI STANDAR IND
            ws['CT{}'.format(r)] = "=MAX(CT2:CT{})".format(q)
            ws['CT{}'.format(s)] = "=MIN(CT2:CT{})".format(q)
            ws['CT{}'.format(t)] = "=ROUND(AVERAGE(CT2:CT{}),2)".format(q)
            # NILAI STANDAR ENG
            ws['CU{}'.format(r)] = "=MAX(CU2:CU{})".format(q)
            ws['CU{}'.format(s)] = "=MIN(CU2:CU{})".format(q)
            ws['CU{}'.format(t)] = "=ROUND(AVERAGE(CU2:CU{}),2)".format(q)
            # NILAI STANDAR SEJ
            ws['CV{}'.format(r)] = "=MAX(CV2:CV{})".format(q)
            ws['CV{}'.format(s)] = "=MIN(CV2:CV{})".format(q)
            ws['CV{}'.format(t)] = "=ROUND(AVERAGE(CV2:CV{}),2)".format(q)
            # NILAI STANDAR EKO
            ws['CW{}'.format(r)] = "=MAX(CW2:CW{})".format(q)
            ws['CW{}'.format(s)] = "=MIN(CW2:CW{})".format(q)
            ws['CW{}'.format(t)] = "=ROUND(AVERAGE(CW2:CW{}),2)".format(q)
            # NILAI STANDAR SOS
            ws['CX{}'.format(r)] = "=MAX(CX2:CX{})".format(q)
            ws['CX{}'.format(s)] = "=MIN(CX2:CX{})".format(q)
            ws['CX{}'.format(t)] = "=ROUND(AVERAGE(CX2:CX{}),2)".format(q)
            # NILAI STANDAR GEO
            ws['CY{}'.format(r)] = "=MAX(CY2:CY{})".format(q)
            ws['CY{}'.format(s)] = "=MIN(CY2:CY{})".format(q)
            ws['CY{}'.format(t)] = "=ROUND(AVERAGE(CY2:CY{}),2)".format(q)
            # NILAI STANDAR ANT
            ws['CZ{}'.format(r)] = "=MAX(CZ2:CZ{})".format(q)
            ws['CZ{}'.format(s)] = "=MIN(CZ2:CZ{})".format(q)
            ws['CZ{}'.format(t)] = "=ROUND(AVERAGE(CZ2:CZ{}),2)".format(q)
            # NILAI STANDAR BIO
            ws['DA{}'.format(r)] = "=MAX(DA2:DA{})".format(q)
            ws['DA{}'.format(s)] = "=MIN(DA2:DA{})".format(q)
            ws['DA{}'.format(t)] = "=ROUND(AVERAGE(DA2:DA{}),2)".format(q)
            # NILAI STANDAR FIS
            ws['DB{}'.format(r)] = "=MAX(DB2:DB{})".format(q)
            ws['DB{}'.format(s)] = "=MIN(DB2:DB{})".format(q)
            ws['DB{}'.format(t)] = "=ROUND(AVERAGE(DB2:DB{}),2)".format(q)
            # NILAI STANDAR KIM 1
            ws['DC{}'.format(r)] = "=MAX(DC2:DC{})".format(q)
            ws['DC{}'.format(s)] = "=MIN(DC2:DC{})".format(q)
            ws['DC{}'.format(t)] = "=ROUND(AVERAGE(DC2:DC{}),2)".format(q)
            # NILAI STANDAR KIM 2
            ws['DD{}'.format(r)] = "=MAX(DD2:DD{})".format(q)
            ws['DD{}'.format(s)] = "=MIN(DD2:DD{})".format(q)
            ws['DD{}'.format(t)] = "=ROUND(AVERAGE(DD2:DD{}),2)".format(q)
            # NILAI STANDAR JML
            ws['DE{}'.format(r)] = "=MAX(DE2:DE{})".format(q)
            ws['DE{}'.format(s)] = "=MIN(DE2:DE{})".format(q)
            ws['DE{}'.format(t)] = "=ROUND(AVERAGE(DE2:DE{}),2)".format(q)

            # TAMBAHAN
            # MTK 1, MTK 2, IND, ENG
            ws['DH{}'.format(r)] = "=SUM(DH2:DH{})".format(q)
            ws['DI{}'.format(r)] = "=SUM(DI2:DI{})".format(q)
            ws['DJ{}'.format(r)] = "=SUM(DJ2:DJ{})".format(q)
            ws['DK{}'.format(r)] = "=SUM(DK2:DK{})".format(q)
            # SEJ, EKO, SOS, GEO, ANT
            ws['DL{}'.format(r)] = "=SUM(DL2:DL{})".format(q)
            ws['DM{}'.format(r)] = "=SUM(DM2:DM{})".format(q)
            ws['DN{}'.format(r)] = "=SUM(DN2:DN{})".format(q)
            ws['DO{}'.format(r)] = "=SUM(DO2:DO{})".format(q)
            ws['DP{}'.format(r)] = "=SUM(DP2:DP{})".format(q)
            # BIO, FIS, KIM 1, KIM 2
            ws['DQ{}'.format(r)] = "=SUM(DQ2:DQ{})".format(q)
            ws['DR{}'.format(r)] = "=SUM(DR2:DR{})".format(q)
            ws['DS{}'.format(r)] = "=SUM(DS2:DS{})".format(q)
            ws['DT{}'.format(r)] = "=SUM(DT2:DT{})".format(q)

            # iterasi 2 rata-rata - 2
            # rata" MTK 1 ke MTK 1 tambahan dan mapel MTK 1 awal
            ws['EA{}'.format(r)] = "=IF($DH${}=0,$G${},$G${}-1)".format(r, r, r)
            ws['EA{}'.format(s)] = "=STDEV(EA2:EA{})".format(q)
            ws['EA{}'.format(t)] = "=MAX(EA2:EA{})".format(q)
            ws['EA{}'.format(u)] = "=MIN(EA2:EA{})".format(q)
            # rata" MTK 2 ke MTK 2 tambahan dan mapel MTK 2 awal
            ws['EB{}'.format(r)] = "=IF($DI${}=0,$H${},$H${}-1)".format(r, r, r)
            ws['EB{}'.format(s)] = "=STDEV(EB2:EB{})".format(q)
            ws['EB{}'.format(t)] = "=MAX(EB2:EB{})".format(q)
            ws['EB{}'.format(u)] = "=MIN(EB2:EB{})".format(q)
            # rata" IND ke IND tambahan dan mapel IND awal
            ws['EC{}'.format(r)] = "=IF($DJ${}=0,$I${},$I${}-1)".format(r, r, r)
            ws['EC{}'.format(s)] = "=STDEV(EC2:EC{})".format(q)
            ws['EC{}'.format(t)] = "=MAX(EC2:EC{})".format(q)
            ws['EC{}'.format(u)] = "=MIN(EC2:EC{})".format(q)
            # rata" ENG ke ENG tambahan dan mapel ENG awal
            ws['ED{}'.format(r)] = "=IF($DK${}=0,$J${},$J${}-1)".format(r, r, r)
            ws['ED{}'.format(s)] = "=STDEV(ED2:ED{})".format(q)
            ws['ED{}'.format(t)] = "=MAX(ED2:ED{})".format(q)
            ws['ED{}'.format(u)] = "=MIN(ED2:ED{})".format(q)
            # rata" SEJ ke SEJ tambahan dan mapel SEJ awal
            ws['EE{}'.format(r)] = "=IF($DL${}=0,$K${},$K${}-1)".format(r, r, r)
            ws['EE{}'.format(s)] = "=STDEV(EE2:EE{})".format(q)
            ws['EE{}'.format(t)] = "=MAX(EE2:EE{})".format(q)
            ws['EE{}'.format(u)] = "=MIN(EE2:EE{})".format(q)
            # rata" EKO ke EKO tambahan dan mapel EKO awal
            ws['EF{}'.format(r)] = "=IF($DM${}=0,$L${},$L${}-1)".format(r, r, r)
            ws['EF{}'.format(s)] = "=STDEV(EF2:EF{})".format(q)
            ws['EF{}'.format(t)] = "=MAX(EF2:EF{})".format(q)
            ws['EF{}'.format(u)] = "=MIN(EF2:EF{})".format(q)
            # rata" SOS ke SOS tambahan dan mapel SOS awal
            ws['EG{}'.format(r)] = "=IF($DN${}=0,$M${},$M${}-1)".format(r, r, r)
            ws['EG{}'.format(s)] = "=STDEV(EG2:EG{})".format(q)
            ws['EG{}'.format(t)] = "=MAX(EG2:EG{})".format(q)
            ws['EG{}'.format(u)] = "=MIN(EG2:EG{})".format(q)
            # rata" GEO ke GEO tambahan dan mapel GEO awal
            ws['EH{}'.format(r)] = "=IF($DO${}=0,$N${},$N${}-1)".format(r, r, r)
            ws['EH{}'.format(s)] = "=STDEV(EH2:EH{})".format(q)
            ws['EH{}'.format(t)] = "=MAX(EH2:EH{})".format(q)
            ws['EH{}'.format(u)] = "=MIN(EH2:EH{})".format(q)
            # rata" ANT ke ANT tambahan dan mapel ANT awal
            ws['EI{}'.format(r)] = "=IF($DP${}=0,$O${},$O${}-1)".format(r, r, r)
            ws['EI{}'.format(s)] = "=STDEV(EI2:EI{})".format(q)
            ws['EI{}'.format(t)] = "=MAX(EI2:EI{})".format(q)
            ws['EI{}'.format(u)] = "=MIN(EI2:EI{})".format(q)
            # rata" BIO ke BIO tambahan dan mapel BIO awal
            ws['EJ{}'.format(r)] = "=IF($DQ${}=0,$P${},$P${}-1)".format(r, r, r)
            ws['EJ{}'.format(s)] = "=STDEV(EJ2:EJ{})".format(q)
            ws['EJ{}'.format(t)] = "=MAX(EJ2:EJ{})".format(q)
            ws['EJ{}'.format(u)] = "=MIN(EJ2:EJ{})".format(q)
            # rata" FIS ke FIS tambahan dan mapel FIS awal
            ws['EK{}'.format(r)] = "=IF($DR${}=0,$Q${},$Q${}-1)".format(r, r, r)
            ws['EK{}'.format(s)] = "=STDEV(EK2:EK{})".format(q)
            ws['EK{}'.format(t)] = "=MAX(EK2:EK{})".format(q)
            ws['EK{}'.format(u)] = "=MIN(EK2:EK{})".format(q)
            # rata" KIM 1 ke KIM 1 tambahan dan mapel KIM 1 awal
            ws['EL{}'.format(r)] = "=IF($DS${}=0,$R${},$R${}-1)".format(r, r, r)
            ws['EL{}'.format(s)] = "=STDEV(EL2:EL{})".format(q)
            ws['EL{}'.format(t)] = "=MAX(EL2:EL{})".format(q)
            ws['EL{}'.format(u)] = "=MIN(EL2:EL{})".format(q)
            # rata" KIM 2 ke KIM 2 tambahan dan mapel KIM 2 awal
            ws['EM{}'.format(r)] = "=IF($DT${}=0,$S${},$S${}-1)".format(r, r, r)
            ws['EM{}'.format(s)] = "=STDEV(EM2:EM{})".format(q)
            ws['EM{}'.format(t)] = "=MAX(EM2:EM{})".format(q)
            ws['EM{}'.format(u)] = "=MIN(EM2:EM{})".format(q)
            # JML BENAR
            ws['EN{}'.format(r)] = "=ROUND(AVERAGE(EN2:EN{}),2)".format(q)
            ws['EN{}'.format(t)] = "=MAX(EN2:EN{})".format(q)
            ws['EN{}'.format(u)] = "=MIN(EN2:EN{})".format(q)
            # MAX Z SCORE MAT 1, MAT 2, IND, ENG
            ws['EO{}'.format(r)] = "=MAX(EO2:EO{})".format(q)
            ws['EP{}'.format(r)] = "=MAX(EP2:EP{})".format(q)
            ws['EQ{}'.format(r)] = "=MAX(EQ2:EQ{})".format(q)
            ws['ER{}'.format(r)] = "=MAX(ER2:ER{})".format(q)
            # MAX Z SCORE SEJ, EKO, SOS, GEO, ANT
            ws['ES{}'.format(r)] = "=MAX(ES2:ES{})".format(q)
            ws['ET{}'.format(r)] = "=MAX(ET2:ET{})".format(q)
            ws['EU{}'.format(r)] = "=MAX(EU2:EU{})".format(q)
            ws['EV{}'.format(r)] = "=MAX(EV2:EV{})".format(q)
            ws['EW{}'.format(r)] = "=MAX(EW2:EW{})".format(q)
            # MAX Z SCORE BIO, FIS, KIM 1, KIM 2
            ws['EX{}'.format(r)] = "=MAX(EX2:EX{})".format(q)
            ws['EY{}'.format(r)] = "=MAX(EY2:EY{})".format(q)
            ws['EZ{}'.format(r)] = "=MAX(EZ2:EZ{})".format(q)
            ws['FA{}'.format(r)] = "=MAX(FA2:FA{})".format(q)

            # NILAI STANDAR MTK 1
            ws['FB{}'.format(r)] = "=MAX(FB2:FB{})".format(q)
            ws['FB{}'.format(s)] = "=MIN(FB2:FB{})".format(q)
            ws['FB{}'.format(t)] = "=ROUND(AVERAGE(FB2:FB{}),2)".format(q)
            # NILAI STANDAR MTK 1
            ws['FC{}'.format(r)] = "=MAX(FC2:FC{})".format(q)
            ws['FC{}'.format(s)] = "=MIN(FC2:FC{})".format(q)
            ws['FC{}'.format(t)] = "=ROUND(AVERAGE(FC2:FC{}),2)".format(q)
            # NILAI STANDAR IND
            ws['FD{}'.format(r)] = "=MAX(FD2:FD{})".format(q)
            ws['FD{}'.format(s)] = "=MIN(FD2:FD{})".format(q)
            ws['FD{}'.format(t)] = "=ROUND(AVERAGE(FD2:FD{}),2)".format(q)
            # NILAI STANDAR ENG
            ws['FE{}'.format(r)] = "=MAX(FE2:FE{})".format(q)
            ws['FE{}'.format(s)] = "=MIN(FE2:FE{})".format(q)
            ws['FE{}'.format(t)] = "=ROUND(AVERAGE(FE2:FE{}),2)".format(q)
            # NILAI STANDAR SEJ
            ws['FF{}'.format(r)] = "=MAX(FF2:FF{})".format(q)
            ws['FF{}'.format(s)] = "=MIN(FF2:FF{})".format(q)
            ws['FF{}'.format(t)] = "=ROUND(AVERAGE(FF2:FF{}),2)".format(q)
            # NILAI STANDAR EKO
            ws['FG{}'.format(r)] = "=MAX(FG2:FG{})".format(q)
            ws['FG{}'.format(s)] = "=MIN(FG2:FG{})".format(q)
            ws['FG{}'.format(t)] = "=ROUND(AVERAGE(FG2:FG{}),2)".format(q)
            # NILAI STANDAR SOS
            ws['FH{}'.format(r)] = "=MAX(FH2:FH{})".format(q)
            ws['FH{}'.format(s)] = "=MIN(FH2:FH{})".format(q)
            ws['FH{}'.format(t)] = "=ROUND(AVERAGE(FH2:FH{}),2)".format(q)
            # NILAI STANDAR GEO
            ws['FI{}'.format(r)] = "=MAX(FI2:FI{})".format(q)
            ws['FI{}'.format(s)] = "=MIN(FI2:FI{})".format(q)
            ws['FI{}'.format(t)] = "=ROUND(AVERAGE(FI2:FI{}),2)".format(q)
            # NILAI STANDAR ANT
            ws['FJ{}'.format(r)] = "=MAX(FJ2:FJ{})".format(q)
            ws['FJ{}'.format(s)] = "=MIN(FJ2:FJ{})".format(q)
            ws['FJ{}'.format(t)] = "=ROUND(AVERAGE(FJ2:FJ{}),2)".format(q)
            # NILAI STANDAR BIO
            ws['FK{}'.format(r)] = "=MAX(FK2:FK{})".format(q)
            ws['FK{}'.format(s)] = "=MIN(FK2:FK{})".format(q)
            ws['FK{}'.format(t)] = "=ROUND(AVERAGE(FK2:FK{}),2)".format(q)
            # NILAI STANDAR FIS
            ws['FL{}'.format(r)] = "=MAX(FL2:FL{})".format(q)
            ws['FL{}'.format(s)] = "=MIN(FL2:FL{})".format(q)
            ws['FL{}'.format(t)] = "=ROUND(AVERAGE(FL2:FL{}),2)".format(q)
            # NILAI STANDAR KIM 1
            ws['FM{}'.format(r)] = "=MAX(FM2:FM{})".format(q)
            ws['FM{}'.format(s)] = "=MIN(FM2:FM{})".format(q)
            ws['FM{}'.format(t)] = "=ROUND(AVERAGE(FM2:FM{}),2)".format(q)
            # NILAI STANDAR KIM 2
            ws['FN{}'.format(r)] = "=MAX(FN2:FN{})".format(q)
            ws['FN{}'.format(s)] = "=MIN(FN2:FN{})".format(q)
            ws['FN{}'.format(t)] = "=ROUND(AVERAGE(FN2:FN{}),2)".format(q)
            # NILAI STANDAR JML
            ws['FO{}'.format(r)] = "=MAX(FO2:FO{})".format(q)
            ws['FO{}'.format(s)] = "=MIN(FO2:FO{})".format(q)
            ws['FO{}'.format(t)] = "=ROUND(AVERAGE(FO2:FO{}),2)".format(q)

            # TAMBAHAN
            # MTK 1, MTK 2, IND, ENG
            ws['FR{}'.format(r)] = "=SUM(FR2:FR{})".format(q)
            ws['FS{}'.format(r)] = "=SUM(FS2:FS{})".format(q)
            ws['FT{}'.format(r)] = "=SUM(FT2:FT{})".format(q)
            ws['FU{}'.format(r)] = "=SUM(FU2:FU{})".format(q)
            # SEJ, EKO, SOS, GEO, ANT
            ws['FV{}'.format(r)] = "=SUM(FV2:FV{})".format(q)
            ws['FW{}'.format(r)] = "=SUM(FW2:FW{})".format(q)
            ws['FX{}'.format(r)] = "=SUM(FX2:FX{})".format(q)
            ws['FY{}'.format(r)] = "=SUM(FY2:FY{})".format(q)
            ws['FZ{}'.format(r)] = "=SUM(FZ2:FZ{})".format(q)
            # BIO, FIS, KIM 1, KIM 2
            ws['GA{}'.format(r)] = "=SUM(GA2:GA{})".format(q)
            ws['GB{}'.format(r)] = "=SUM(GB2:GB{})".format(q)
            ws['GC{}'.format(r)] = "=SUM(GC2:GC{})".format(q)
            ws['GD{}'.format(r)] = "=SUM(GD2:GD{})".format(q)

            # iterasi 3 rata-rata - 3
            # rata" MTK 1 ke MTK 1 tambahan dan mapel MTK 1 awal
            ws['GK{}'.format(r)] = "=IF($FR${}=0,$G${},$G${}-1)".format(r, r, r)
            ws['GK{}'.format(s)] = "=STDEV(GK2:GK{})".format(q)
            ws['GK{}'.format(t)] = "=MAX(GK2:GK{})".format(q)
            ws['GK{}'.format(u)] = "=MIN(GK2:GK{})".format(q)
            # rata" MTK 2 ke MTK 2 tambahan dan mapel MTK 2 awal
            ws['GL{}'.format(r)] = "=IF($FS${}=0,$H${},$H${}-1)".format(r, r, r)
            ws['GL{}'.format(s)] = "=STDEV(GL2:GL{})".format(q)
            ws['GL{}'.format(t)] = "=MAX(GL2:GL{})".format(q)
            ws['GL{}'.format(u)] = "=MIN(GL2:GL{})".format(q)
            # rata" IND ke IND tambahan dan mapel IND awal
            ws['GM{}'.format(r)] = "=IF($FT${}=0,$I${},$I${}-1)".format(r, r, r)
            ws['GM{}'.format(s)] = "=STDEV(GM2:GM{})".format(q)
            ws['GM{}'.format(t)] = "=MAX(GM2:GM{})".format(q)
            ws['GM{}'.format(u)] = "=MIN(GM2:GM{})".format(q)
            # rata" ENG ke ENG tambahan dan mapel ENG awal
            ws['GN{}'.format(r)] = "=IF($FU${}=0,$J${},$J${}-1)".format(r, r, r)
            ws['GN{}'.format(s)] = "=STDEV(GN2:GN{})".format(q)
            ws['GN{}'.format(t)] = "=MAX(GN2:GN{})".format(q)
            ws['GN{}'.format(u)] = "=MIN(GN2:GN{})".format(q)
            # rata" SEJ ke SEJ tambahan dan mapel SEJ awal
            ws['GO{}'.format(r)] = "=IF($FV${}=0,$K${},$K${}-1)".format(r, r, r)
            ws['GO{}'.format(s)] = "=STDEV(GO2:GO{})".format(q)
            ws['GO{}'.format(t)] = "=MAX(GO2:GO{})".format(q)
            ws['GO{}'.format(u)] = "=MIN(GO2:GO{})".format(q)
            # rata" EKO ke EKO tambahan dan mapel EKO awal
            ws['GP{}'.format(r)] = "=IF($FW${}=0,$L${},$L${}-1)".format(r, r, r)
            ws['GP{}'.format(s)] = "=STDEV(GP2:GP{})".format(q)
            ws['GP{}'.format(t)] = "=MAX(GP2:GP{})".format(q)
            ws['GP{}'.format(u)] = "=MIN(GP2:GP{})".format(q)
            # rata" SOS ke SOS tambahan dan mapel SOS awal
            ws['GQ{}'.format(r)] = "=IF($FX${}=0,$M${},$M${}-1)".format(r, r, r)
            ws['GQ{}'.format(s)] = "=STDEV(GQ2:GQ{})".format(q)
            ws['GQ{}'.format(t)] = "=MAX(GQ2:GQ{})".format(q)
            ws['GQ{}'.format(u)] = "=MIN(GQ2:GQ{})".format(q)
            # rata" GEO ke GEO tambahan dan mapel GEO awal
            ws['GR{}'.format(r)] = "=IF($FY${}=0,$N${},$N${}-1)".format(r, r, r)
            ws['GR{}'.format(s)] = "=STDEV(GR2:GR{})".format(q)
            ws['GR{}'.format(t)] = "=MAX(GR2:GR{})".format(q)
            ws['GR{}'.format(u)] = "=MIN(GR2:GR{})".format(q)
            # rata" ANT ke ANT tambahan dan mapel ANT awal
            ws['GS{}'.format(r)] = "=IF($FZ${}=0,$O${},$O${}-1)".format(r, r, r)
            ws['GS{}'.format(s)] = "=STDEV(GS2:GS{})".format(q)
            ws['GS{}'.format(t)] = "=MAX(GS2:GS{})".format(q)
            ws['GS{}'.format(u)] = "=MIN(GS2:GS{})".format(q)
            # rata" BIO ke BIO tambahan dan mapel BIO awal
            ws['GT{}'.format(r)] = "=IF($GA${}=0,$P${},$P${}-1)".format(r, r, r)
            ws['GT{}'.format(s)] = "=STDEV(GT2:GT{})".format(q)
            ws['GT{}'.format(t)] = "=MAX(GT2:GT{})".format(q)
            ws['GT{}'.format(u)] = "=MIN(GT2:GT{})".format(q)
            # rata" FIS ke FIS tambahan dan mapel FIS awal
            ws['GU{}'.format(r)] = "=IF($GB${}=0,$Q${},$Q${}-1)".format(r, r, r)
            ws['GU{}'.format(s)] = "=STDEV(GU2:GU{})".format(q)
            ws['GU{}'.format(t)] = "=MAX(GU2:GU{})".format(q)
            ws['GU{}'.format(u)] = "=MIN(GU2:GU{})".format(q)
            # rata" KIM 1 ke KIM 1 tambahan dan mapel KIM 1 awal
            ws['GV{}'.format(r)] = "=IF($GC${}=0,$R${},$R${}-1)".format(r, r, r)
            ws['GV{}'.format(s)] = "=STDEV(GV2:GV{})".format(q)
            ws['GV{}'.format(t)] = "=MAX(GV2:GV{})".format(q)
            ws['GV{}'.format(u)] = "=MIN(GV2:GV{})".format(q)
            # rata" KIM 2 ke KIM 2 tambahan dan mapel KIM 2 awal
            ws['GW{}'.format(r)] = "=IF($GD${}=0,$S${},$S${}-1)".format(r, r, r)
            ws['GW{}'.format(s)] = "=STDEV(GW2:GW{})".format(q)
            ws['GW{}'.format(t)] = "=MAX(GW2:GW{})".format(q)
            ws['GW{}'.format(u)] = "=MIN(GW2:GW{})".format(q)
            # JML BENAR
            ws['GX{}'.format(r)] = "=ROUND(AVERAGE(GX2:GX{}),2)".format(q)
            ws['GX{}'.format(t)] = "=MAX(GX2:GX{})".format(q)
            ws['GX{}'.format(u)] = "=MIN(GX2:GX{})".format(q)
            # MAX Z SCORE MAT 1, MAT 2, IND, ENG
            ws['GY{}'.format(r)] = "=MAX(GY2:GY{})".format(q)
            ws['GZ{}'.format(r)] = "=MAX(GZ2:GZ{})".format(q)
            ws['HA{}'.format(r)] = "=MAX(HA2:HA{})".format(q)
            ws['HB{}'.format(r)] = "=MAX(HB2:HB{})".format(q)
            # MAX Z SCORE SEJ, EKO, SOS, GEO, ANT
            ws['HC{}'.format(r)] = "=MAX(HC2:HC{})".format(q)
            ws['HD{}'.format(r)] = "=MAX(HD2:HD{})".format(q)
            ws['HE{}'.format(r)] = "=MAX(HE2:HE{})".format(q)
            ws['HF{}'.format(r)] = "=MAX(HF2:HF{})".format(q)
            ws['HG{}'.format(r)] = "=MAX(HG2:HG{})".format(q)
            # MAX Z SCORE BIO, FIS, KIM 1, KIM 2
            ws['HH{}'.format(r)] = "=MAX(HH2:HH{})".format(q)
            ws['HI{}'.format(r)] = "=MAX(HI2:HI{})".format(q)
            ws['HJ{}'.format(r)] = "=MAX(HJ2:HJ{})".format(q)
            ws['HK{}'.format(r)] = "=MAX(HK2:HK{})".format(q)

            # NILAI STANDAR MTK 1
            ws['HL{}'.format(r)] = "=MAX(HL2:HL{})".format(q)
            ws['HL{}'.format(s)] = "=MIN(HL2:HL{})".format(q)
            ws['HL{}'.format(t)] = "=ROUND(AVERAGE(HL2:HL{}),2)".format(q)
            # NILAI STANDAR MTK 1
            ws['HM{}'.format(r)] = "=MAX(HM2:HM{})".format(q)
            ws['HM{}'.format(s)] = "=MIN(HM2:HM{})".format(q)
            ws['HM{}'.format(t)] = "=ROUND(AVERAGE(HM2:HM{}),2)".format(q)
            # NILAI STANDAR IND
            ws['HN{}'.format(r)] = "=MAX(HN2:HN{})".format(q)
            ws['HN{}'.format(s)] = "=MIN(HN2:HN{})".format(q)
            ws['HN{}'.format(t)] = "=ROUND(AVERAGE(HN2:HN{}),2)".format(q)
            # NILAI STANDAR ENG
            ws['HO{}'.format(r)] = "=MAX(HO2:HO{})".format(q)
            ws['HO{}'.format(s)] = "=MIN(HO2:HO{})".format(q)
            ws['HO{}'.format(t)] = "=ROUND(AVERAGE(HO2:HO{}),2)".format(q)
            # NILAI STANDAR SEJ
            ws['HP{}'.format(r)] = "=MAX(HP2:HP{})".format(q)
            ws['HP{}'.format(s)] = "=MIN(HP2:HP{})".format(q)
            ws['HP{}'.format(t)] = "=ROUND(AVERAGE(HP2:HP{}),2)".format(q)
            # NILAI STANDAR EKO
            ws['HQ{}'.format(r)] = "=MAX(HQ2:HQ{})".format(q)
            ws['HQ{}'.format(s)] = "=MIN(HQ2:HQ{})".format(q)
            ws['HQ{}'.format(t)] = "=ROUND(AVERAGE(HQ2:HQ{}),2)".format(q)
            # NILAI STANDAR SOS
            ws['HR{}'.format(r)] = "=MAX(HR2:HR{})".format(q)
            ws['HR{}'.format(s)] = "=MIN(HR2:HR{})".format(q)
            ws['HR{}'.format(t)] = "=ROUND(AVERAGE(HR2:HR{}),2)".format(q)
            # NILAI STANDAR GEO
            ws['HS{}'.format(r)] = "=MAX(HS2:HS{})".format(q)
            ws['HS{}'.format(s)] = "=MIN(HS2:HS{})".format(q)
            ws['HS{}'.format(t)] = "=ROUND(AVERAGE(HS2:HS{}),2)".format(q)
            # NILAI STANDAR ANT
            ws['HT{}'.format(r)] = "=MAX(HT2:HT{})".format(q)
            ws['HT{}'.format(s)] = "=MIN(HT2:HT{})".format(q)
            ws['HT{}'.format(t)] = "=ROUND(AVERAGE(HT2:HT{}),2)".format(q)
            # NILAI STANDAR BIO
            ws['HU{}'.format(r)] = "=MAX(HU2:HU{})".format(q)
            ws['HU{}'.format(s)] = "=MIN(HU2:HU{})".format(q)
            ws['HU{}'.format(t)] = "=ROUND(AVERAGE(HU2:HU{}),2)".format(q)
            # NILAI STANDAR FIS
            ws['HV{}'.format(r)] = "=MAX(HV2:HV{})".format(q)
            ws['HV{}'.format(s)] = "=MIN(HV2:HV{})".format(q)
            ws['HV{}'.format(t)] = "=ROUND(AVERAGE(HV2:HV{}),2)".format(q)
            # NILAI STANDAR KIM 1
            ws['HW{}'.format(r)] = "=MAX(HW2:HW{})".format(q)
            ws['HW{}'.format(s)] = "=MIN(HW2:HW{})".format(q)
            ws['HW{}'.format(t)] = "=ROUND(AVERAGE(HW2:HW{}),2)".format(q)
            # NILAI STANDAR KIM 2
            ws['HX{}'.format(r)] = "=MAX(HX2:HX{})".format(q)
            ws['HX{}'.format(s)] = "=MIN(HX2:HX{})".format(q)
            ws['HX{}'.format(t)] = "=ROUND(AVERAGE(HX2:HX{}),2)".format(q)
            # NILAI STANDAR JML
            ws['HY{}'.format(r)] = "=MAX(HY2:HY{})".format(q)
            ws['HY{}'.format(s)] = "=MIN(HY2:HY{})".format(q)
            ws['HY{}'.format(t)] = "=ROUND(AVERAGE(HY2:HY{}),2)".format(q)

            # TAMBAHAN
            # MTK 1, MTK 2, IND, ENG
            ws['IB{}'.format(r)] = "=SUM(IB2:IB{})".format(q)
            ws['IC{}'.format(r)] = "=SUM(IC2:IC{})".format(q)
            ws['ID{}'.format(r)] = "=SUM(ID2:ID{})".format(q)
            ws['IE{}'.format(r)] = "=SUM(IE2:IE{})".format(q)
            # SEJ, EKO, SOS, GEO, ANT
            ws['IF{}'.format(r)] = "=SUM(IF2:IF{})".format(q)
            ws['IG{}'.format(r)] = "=SUM(IG2:IG{})".format(q)
            ws['IH{}'.format(r)] = "=SUM(IH2:IH{})".format(q)
            ws['II{}'.format(r)] = "=SUM(II2:II{})".format(q)
            ws['IJ{}'.format(r)] = "=SUM(IJ2:IJ{})".format(q)
            # BIO, FIS, KIM 1, KIM 2
            ws['IK{}'.format(r)] = "=SUM(IK2:IK{})".format(q)
            ws['IL{}'.format(r)] = "=SUM(IL2:IL{})".format(q)
            ws['IM{}'.format(r)] = "=SUM(IM2:IM{})".format(q)
            ws['IN{}'.format(r)] = "=SUM(IN2:IN{})".format(q)

            # iterasi 4 rata-rata - 4
            # rata" MTK 1 ke MTK 1 tambahan dan mapel MTK 1 awal
            ws['IU{}'.format(r)] = "=IF($IB${}=0,$G${},$G${}-1)".format(r, r, r)
            ws['IU{}'.format(s)] = "=STDEV(IU2:IU{})".format(q)
            ws['IU{}'.format(t)] = "=MAX(IU2:IU{})".format(q)
            ws['IU{}'.format(u)] = "=MIN(IU2:IU{})".format(q)
            # rata" MTK 2 ke MTK 2 tambahan dan mapel MTK 2 awal
            ws['IV{}'.format(r)] = "=IF($IC${}=0,$H${},$H${}-1)".format(r, r, r)
            ws['IV{}'.format(s)] = "=STDEV(IV2:IV{})".format(q)
            ws['IV{}'.format(t)] = "=MAX(IV2:IV{})".format(q)
            ws['IV{}'.format(u)] = "=MIN(IV2:IV{})".format(q)
            # rata" IND ke IND tambahan dan mapel IND awal
            ws['IW{}'.format(r)] = "=IF($ID${}=0,$I${},$I${}-1)".format(r, r, r)
            ws['IW{}'.format(s)] = "=STDEV(IW2:IW{})".format(q)
            ws['IW{}'.format(t)] = "=MAX(IW2:IW{})".format(q)
            ws['IW{}'.format(u)] = "=MIN(IW2:IW{})".format(q)
            # rata" ENG ke ENG tambahan dan mapel ENG awal
            ws['IX{}'.format(r)] = "=IF($IE${}=0,$J${},$J${}-1)".format(r, r, r)
            ws['IX{}'.format(s)] = "=STDEV(IX2:IX{})".format(q)
            ws['IX{}'.format(t)] = "=MAX(IX2:IX{})".format(q)
            ws['IX{}'.format(u)] = "=MIN(IX2:IX{})".format(q)
            # rata" SEJ ke SEJ tambahan dan mapel SEJ awal
            ws['IY{}'.format(r)] = "=IF($IF${}=0,$K${},$K${}-1)".format(r, r, r)
            ws['IY{}'.format(s)] = "=STDEV(IY2:IY{})".format(q)
            ws['IY{}'.format(t)] = "=MAX(IY2:IY{})".format(q)
            ws['IY{}'.format(u)] = "=MIN(IY2:IY{})".format(q)
            # rata" EKO ke EKO tambahan dan mapel EKO awal
            ws['IZ{}'.format(r)] = "=IF($IG${}=0,$L${},$L${}-1)".format(r, r, r)
            ws['IZ{}'.format(s)] = "=STDEV(IZ2:IZ{})".format(q)
            ws['IZ{}'.format(t)] = "=MAX(IZ2:IZ{})".format(q)
            ws['IZ{}'.format(u)] = "=MIN(IZ2:IZ{})".format(q)
            # rata" SOS ke SOS tambahan dan mapel SOS awal
            ws['JA{}'.format(r)] = "=IF($IH${}=0,$M${},$M${}-1)".format(r, r, r)
            ws['JA{}'.format(s)] = "=STDEV(JA2:JA{})".format(q)
            ws['JA{}'.format(t)] = "=MAX(JA2:JA{})".format(q)
            ws['JA{}'.format(u)] = "=MIN(JA2:JA{})".format(q)
            # rata" GEO ke GEO tambahan dan mapel GEO awal
            ws['JB{}'.format(r)] = "=IF($II${}=0,$N${},$N${}-1)".format(r, r, r)
            ws['JB{}'.format(s)] = "=STDEV(JB2:JB{})".format(q)
            ws['JB{}'.format(t)] = "=MAX(JB2:JB{})".format(q)
            ws['JB{}'.format(u)] = "=MIN(JB2:JB{})".format(q)
            # rata" ANT ke ANT tambahan dan mapel ANT awal
            ws['JC{}'.format(r)] = "=IF($IJ${}=0,$O${},$O${}-1)".format(r, r, r)
            ws['JC{}'.format(s)] = "=STDEV(JC2:JC{})".format(q)
            ws['JC{}'.format(t)] = "=MAX(JC2:JC{})".format(q)
            ws['JC{}'.format(u)] = "=MIN(JC2:JC{})".format(q)
            # rata" BIO ke BIO tambahan dan mapel BIO awal
            ws['JD{}'.format(r)] = "=IF($IK${}=0,$P${},$P${}-1)".format(r, r, r)
            ws['JD{}'.format(s)] = "=STDEV(JD2:JD{})".format(q)
            ws['JD{}'.format(t)] = "=MAX(JD2:JD{})".format(q)
            ws['JD{}'.format(u)] = "=MIN(JD2:JD{})".format(q)
            # rata" FIS ke FIS tambahan dan mapel FIS awal
            ws['JE{}'.format(r)] = "=IF($IL${}=0,$Q${},$Q${}-1)".format(r, r, r)
            ws['JE{}'.format(s)] = "=STDEV(JE2:JE{})".format(q)
            ws['JE{}'.format(t)] = "=MAX(JE2:JE{})".format(q)
            ws['JE{}'.format(u)] = "=MIN(JE2:JE{})".format(q)
            # rata" KIM 1 ke KIM 1 tambahan dan mapel KIM 1 awal
            ws['JF{}'.format(r)] = "=IF($IM${}=0,$R${},$R${}-1)".format(r, r, r)
            ws['JF{}'.format(s)] = "=STDEV(JF2:JF{})".format(q)
            ws['JF{}'.format(t)] = "=MAX(JF2:JF{})".format(q)
            ws['JF{}'.format(u)] = "=MIN(JF2:JF{})".format(q)
            # rata" KIM 2 ke KIM 2 tambahan dan mapel KIM 2 awal
            ws['JG{}'.format(r)] = "=IF($IN${}=0,$S${},$S${}-1)".format(r, r, r)
            ws['JG{}'.format(s)] = "=STDEV(JG2:JG{})".format(q)
            ws['JG{}'.format(t)] = "=MAX(JG2:JG{})".format(q)
            ws['JG{}'.format(u)] = "=MIN(JG2:JG{})".format(q)
            # JML BENAR
            ws['JH{}'.format(r)] = "=ROUND(AVERAGE(JH2:JH{}),2)".format(q)
            ws['JH{}'.format(t)] = "=MAX(JH2:JH{})".format(q)
            ws['JH{}'.format(u)] = "=MIN(JH2:JH{})".format(q)
            # MAX Z SCORE MAT 1, MAT 2, IND, ENG
            ws['JI{}'.format(r)] = "=MAX(JI2:JI{})".format(q)
            ws['JJ{}'.format(r)] = "=MAX(JJ2:JJ{})".format(q)
            ws['JK{}'.format(r)] = "=MAX(JK2:JK{})".format(q)
            ws['JL{}'.format(r)] = "=MAX(JL2:JL{})".format(q)
            # MAX Z SCORE SEJ, EKO, SOS, GEO, ANT
            ws['JM{}'.format(r)] = "=MAX(JM2:JM{})".format(q)
            ws['JN{}'.format(r)] = "=MAX(JN2:JN{})".format(q)
            ws['JO{}'.format(r)] = "=MAX(JO2:JO{})".format(q)
            ws['JP{}'.format(r)] = "=MAX(JP2:JP{})".format(q)
            ws['JQ{}'.format(r)] = "=MAX(JQ2:JQ{})".format(q)
            # MAX Z SCORE BIO, FIS, KIM 1, KIM 2
            ws['JR{}'.format(r)] = "=MAX(JR2:JR{})".format(q)
            ws['JS{}'.format(r)] = "=MAX(JS2:JS{})".format(q)
            ws['JT{}'.format(r)] = "=MAX(JT2:JT{})".format(q)
            ws['JU{}'.format(r)] = "=MAX(JU2:JU{})".format(q)

            # NILAI STANDAR MTK 1
            ws['JV{}'.format(r)] = "=MAX(JV2:JV{})".format(q)
            ws['JV{}'.format(s)] = "=MIN(JV2:JV{})".format(q)
            ws['JV{}'.format(t)] = "=ROUND(AVERAGE(JV2:JV{}),2)".format(q)
            # NILAI STANDAR MTK 1
            ws['JW{}'.format(r)] = "=MAX(JW2:JW{})".format(q)
            ws['JW{}'.format(s)] = "=MIN(JW2:JW{})".format(q)
            ws['JW{}'.format(t)] = "=ROUND(AVERAGE(JW2:JW{}),2)".format(q)
            # NILAI STANDAR IND
            ws['JX{}'.format(r)] = "=MAX(JX2:JX{})".format(q)
            ws['JX{}'.format(s)] = "=MIN(JX2:JX{})".format(q)
            ws['JX{}'.format(t)] = "=ROUND(AVERAGE(JX2:JX{}),2)".format(q)
            # NILAI STANDAR ENG
            ws['JY{}'.format(r)] = "=MAX(JY2:JY{})".format(q)
            ws['JY{}'.format(s)] = "=MIN(JY2:JY{})".format(q)
            ws['JY{}'.format(t)] = "=ROUND(AVERAGE(JY2:JY{}),2)".format(q)
            # NILAI STANDAR SEJ
            ws['JZ{}'.format(r)] = "=MAX(JZ2:JZ{})".format(q)
            ws['JZ{}'.format(s)] = "=MIN(JZ2:JZ{})".format(q)
            ws['JZ{}'.format(t)] = "=ROUND(AVERAGE(JZ2:JZ{}),2)".format(q)
            # NILAI STANDAR EKO
            ws['KA{}'.format(r)] = "=MAX(KA2:KA{})".format(q)
            ws['KA{}'.format(s)] = "=MIN(KA2:KA{})".format(q)
            ws['KA{}'.format(t)] = "=ROUND(AVERAGE(KA2:KA{}),2)".format(q)
            # NILAI STANDAR SOS
            ws['KB{}'.format(r)] = "=MAX(KB2:KB{})".format(q)
            ws['KB{}'.format(s)] = "=MIN(KB2:KB{})".format(q)
            ws['KB{}'.format(t)] = "=ROUND(AVERAGE(KB2:KB{}),2)".format(q)
            # NILAI STANDAR GEO
            ws['KC{}'.format(r)] = "=MAX(KC2:KC{})".format(q)
            ws['KC{}'.format(s)] = "=MIN(KC2:KC{})".format(q)
            ws['KC{}'.format(t)] = "=ROUND(AVERAGE(KC2:KC{}),2)".format(q)
            # NILAI STANDAR ANT
            ws['KD{}'.format(r)] = "=MAX(KD2:KD{})".format(q)
            ws['KD{}'.format(s)] = "=MIN(KD2:KD{})".format(q)
            ws['KD{}'.format(t)] = "=ROUND(AVERAGE(KD2:KD{}),2)".format(q)
            # NILAI STANDAR BIO
            ws['KE{}'.format(r)] = "=MAX(KE2:KE{})".format(q)
            ws['KE{}'.format(s)] = "=MIN(KE2:KE{})".format(q)
            ws['KE{}'.format(t)] = "=ROUND(AVERAGE(KE2:KE{}),2)".format(q)
            # NILAI STANDAR FIS
            ws['KF{}'.format(r)] = "=MAX(KF2:KF{})".format(q)
            ws['KF{}'.format(s)] = "=MIN(KF2:KF{})".format(q)
            ws['KF{}'.format(t)] = "=ROUND(AVERAGE(KF2:KF{}),2)".format(q)
            # NILAI STANDAR KIM 1
            ws['KG{}'.format(r)] = "=MAX(KG2:KG{})".format(q)
            ws['KG{}'.format(s)] = "=MIN(KG2:KG{})".format(q)
            ws['KG{}'.format(t)] = "=ROUND(AVERAGE(KG2:KG{}),2)".format(q)
            # NILAI STANDAR KIM 2
            ws['KH{}'.format(r)] = "=MAX(KH2:KH{})".format(q)
            ws['KH{}'.format(s)] = "=MIN(KH2:KH{})".format(q)
            ws['KH{}'.format(t)] = "=ROUND(AVERAGE(KH2:KH{}),2)".format(q)
            # NILAI STANDAR JML
            ws['KI{}'.format(r)] = "=MAX(KI2:KI{})".format(q)
            ws['KI{}'.format(s)] = "=MIN(KI2:KI{})".format(q)
            ws['KI{}'.format(t)] = "=ROUND(AVERAGE(KI2:KI{}),2)".format(q)

            # TAMBAHAN
            # MTK 1, MTK 2, IND, ENG
            ws['KL{}'.format(r)] = "=SUM(KL2:KL{})".format(q)
            ws['KM{}'.format(r)] = "=SUM(KM2:KM{})".format(q)
            ws['KN{}'.format(r)] = "=SUM(KN2:KN{})".format(q)
            ws['KO{}'.format(r)] = "=SUM(KO2:KO{})".format(q)
            # SEJ, EKO, SOS, GEO, ANT
            ws['KP{}'.format(r)] = "=SUM(KP2:KP{})".format(q)
            ws['KQ{}'.format(r)] = "=SUM(KQ2:KQ{})".format(q)
            ws['KR{}'.format(r)] = "=SUM(KR2:KR{})".format(q)
            ws['KS{}'.format(r)] = "=SUM(KS2:KS{})".format(q)
            ws['KT{}'.format(r)] = "=SUM(KT2:KT{})".format(q)
            # BIO, FIS, KIM 1, KIM 2
            ws['KU{}'.format(r)] = "=SUM(KU2:KU{})".format(q)
            ws['KV{}'.format(r)] = "=SUM(KV2:KV{})".format(q)
            ws['KW{}'.format(r)] = "=SUM(KW2:KW{})".format(q)
            ws['KX{}'.format(r)] = "=SUM(KX2:KX{})".format(q)

            # iterasi 5 rata-rata - 5
            # rata" MTK 1 ke MTK 1 tambahan dan mapel MTK 1 awal
            ws['LE{}'.format(r)] = "=IF($KL${}=0,$G${},$G${}-1)".format(r, r, r)
            ws['LE{}'.format(s)] = "=STDEV(LE2:LE{})".format(q)
            ws['LE{}'.format(t)] = "=MAX(LE2:LE{})".format(q)
            ws['LE{}'.format(u)] = "=MIN(LE2:LE{})".format(q)
            # rata" MTK 2 ke MTK 2 tambahan dan mapel MTK 2 awal
            ws['LF{}'.format(r)] = "=IF($KM${}=0,$H${},$H${}-1)".format(r, r, r)
            ws['LF{}'.format(s)] = "=STDEV(LF2:LF{})".format(q)
            ws['LF{}'.format(t)] = "=MAX(LF2:LF{})".format(q)
            ws['LF{}'.format(u)] = "=MIN(LF2:LF{})".format(q)
            # rata" IND ke IND tambahan dan mapel IND awal
            ws['LG{}'.format(r)] = "=IF($KN${}=0,$I${},$I${}-1)".format(r, r, r)
            ws['LG{}'.format(s)] = "=STDEV(LG2:LG{})".format(q)
            ws['LG{}'.format(t)] = "=MAX(LG2:LG{})".format(q)
            ws['LG{}'.format(u)] = "=MIN(LG2:LG{})".format(q)
            # rata" ENG ke ENG tambahan dan mapel ENG awal
            ws['LH{}'.format(r)] = "=IF($KO${}=0,$J${},$J${}-1)".format(r, r, r)
            ws['LH{}'.format(s)] = "=STDEV(LH2:LH{})".format(q)
            ws['LH{}'.format(t)] = "=MAX(LH2:LH{})".format(q)
            ws['LH{}'.format(u)] = "=MIN(LH2:LH{})".format(q)
            # rata" SEJ ke SEJ tambahan dan mapel SEJ awal
            ws['LI{}'.format(r)] = "=IF($KP${}=0,$K${},$K${}-1)".format(r, r, r)
            ws['LI{}'.format(s)] = "=STDEV(LI2:LI{})".format(q)
            ws['LI{}'.format(t)] = "=MAX(LI2:LI{})".format(q)
            ws['LI{}'.format(u)] = "=MIN(LI2:LI{})".format(q)
            # rata" EKO ke EKO tambahan dan mapel EKO awal
            ws['LJ{}'.format(r)] = "=IF($KQ${}=0,$L${},$L${}-1)".format(r, r, r)
            ws['LJ{}'.format(s)] = "=STDEV(LJ2:LJ{})".format(q)
            ws['LJ{}'.format(t)] = "=MAX(LJ2:LJ{})".format(q)
            ws['LJ{}'.format(u)] = "=MIN(LJ2:LJ{})".format(q)
            # rata" SOS ke SOS tambahan dan mapel SOS awal
            ws['LK{}'.format(r)] = "=IF($KR${}=0,$M${},$M${}-1)".format(r, r, r)
            ws['LK{}'.format(s)] = "=STDEV(LK2:LK{})".format(q)
            ws['LK{}'.format(t)] = "=MAX(LK2:LK{})".format(q)
            ws['LK{}'.format(u)] = "=MIN(LK2:LK{})".format(q)
            # rata" GEO ke GEO tambahan dan mapel GEO awal
            ws['LL{}'.format(r)] = "=IF($KS${}=0,$N${},$N${}-1)".format(r, r, r)
            ws['LL{}'.format(s)] = "=STDEV(LL2:LL{})".format(q)
            ws['LL{}'.format(t)] = "=MAX(LL2:LL{})".format(q)
            ws['LL{}'.format(u)] = "=MIN(LL2:LL{})".format(q)
            # rata" ANT ke ANT tambahan dan mapel ANT awal
            ws['LM{}'.format(r)] = "=IF($KT${}=0,$O${},$O${}-1)".format(r, r, r)
            ws['LM{}'.format(s)] = "=STDEV(LM2:LM{})".format(q)
            ws['LM{}'.format(t)] = "=MAX(LM2:LM{})".format(q)
            ws['LM{}'.format(u)] = "=MIN(LM2:LM{})".format(q)
            # rata" BIO ke BIO tambahan dan mapel BIO awal
            ws['LN{}'.format(r)] = "=IF($KU${}=0,$P${},$P${}-1)".format(r, r, r)
            ws['LN{}'.format(s)] = "=STDEV(LN2:LN{})".format(q)
            ws['LN{}'.format(t)] = "=MAX(LN2:LN{})".format(q)
            ws['LN{}'.format(u)] = "=MIN(LN2:LN{})".format(q)
            # rata" FIS ke FIS tambahan dan mapel FIS awal
            ws['LO{}'.format(r)] = "=IF($KV${}=0,$Q${},$Q${}-1)".format(r, r, r)
            ws['LO{}'.format(s)] = "=STDEV(LO2:LO{})".format(q)
            ws['LO{}'.format(t)] = "=MAX(LO2:LO{})".format(q)
            ws['LO{}'.format(u)] = "=MIN(LO2:LO{})".format(q)
            # rata" KIM 1 ke KIM 1 tambahan dan mapel KIM 1 awal
            ws['LP{}'.format(r)] = "=IF($KW${}=0,$R${},$R${}-1)".format(r, r, r)
            ws['LP{}'.format(s)] = "=STDEV(LP2:LP{})".format(q)
            ws['LP{}'.format(t)] = "=MAX(LP2:LP{})".format(q)
            ws['LP{}'.format(u)] = "=MIN(LP2:LP{})".format(q)
            # rata" KIM 2 ke KIM 2 tambahan dan mapel KIM 2 awal
            ws['LQ{}'.format(r)] = "=IF($KX${}=0,$S${},$S${}-1)".format(r, r, r)
            ws['LQ{}'.format(s)] = "=STDEV(LQ2:LQ{})".format(q)
            ws['LQ{}'.format(t)] = "=MAX(LQ2:LQ{})".format(q)
            ws['LQ{}'.format(u)] = "=MIN(LQ2:LQ{})".format(q)
            # JML BENAR
            ws['LR{}'.format(r)] = "=ROUND(AVERAGE(LR2:LR{}),2)".format(q)
            ws['LR{}'.format(t)] = "=MAX(LR2:LR{})".format(q)
            ws['LR{}'.format(u)] = "=MIN(LR2:LR{})".format(q)
            # MAX Z SCORE MAT 1, MAT 2, IND, ENG
            ws['LS{}'.format(r)] = "=MAX(LS2:LS{})".format(q)
            ws['LT{}'.format(r)] = "=MAX(LT2:LT{})".format(q)
            ws['LU{}'.format(r)] = "=MAX(LU2:LU{})".format(q)
            ws['LV{}'.format(r)] = "=MAX(LV2:LV{})".format(q)
            # MAX Z SCORE SEJ, EKO, SOS, GEO, ANT
            ws['LW{}'.format(r)] = "=MAX(LW2:LW{})".format(q)
            ws['LX{}'.format(r)] = "=MAX(LX2:LX{})".format(q)
            ws['LY{}'.format(r)] = "=MAX(LY2:LY{})".format(q)
            ws['LZ{}'.format(r)] = "=MAX(LZ2:LZ{})".format(q)
            ws['MA{}'.format(r)] = "=MAX(MA2:MA{})".format(q)
            # MAX Z SCORE BIO, FIS, KIM 1, KIM 2
            ws['MB{}'.format(r)] = "=MAX(MB2:MB{})".format(q)
            ws['MC{}'.format(r)] = "=MAX(MC2:MC{})".format(q)
            ws['MD{}'.format(r)] = "=MAX(MD2:MD{})".format(q)
            ws['ME{}'.format(r)] = "=MAX(ME2:ME{})".format(q)

            # NILAI STANDAR MTK 1
            ws['MF{}'.format(r)] = "=MAX(MF2:MF{})".format(q)
            ws['MF{}'.format(s)] = "=MIN(MF2:MF{})".format(q)
            ws['MF{}'.format(t)] = "=ROUND(AVERAGE(MF2:MF{}),2)".format(q)
            # NILAI STANDAR MTK 1
            ws['MG{}'.format(r)] = "=MAX(MG2:MG{})".format(q)
            ws['MG{}'.format(s)] = "=MIN(MG2:MG{})".format(q)
            ws['MG{}'.format(t)] = "=ROUND(AVERAGE(MG2:MG{}),2)".format(q)
            # NILAI STANDAR IND
            ws['MH{}'.format(r)] = "=MAX(MH2:MH{})".format(q)
            ws['MH{}'.format(s)] = "=MIN(MH2:MH{})".format(q)
            ws['MH{}'.format(t)] = "=ROUND(AVERAGE(MH2:MH{}),2)".format(q)
            # NILAI STANDAR ENG
            ws['MI{}'.format(r)] = "=MAX(MI2:MI{})".format(q)
            ws['MI{}'.format(s)] = "=MIN(MI2:MI{})".format(q)
            ws['MI{}'.format(t)] = "=ROUND(AVERAGE(MI2:MI{}),2)".format(q)
            # NILAI STANDAR SEJ
            ws['MJ{}'.format(r)] = "=MAX(MJ2:MJ{})".format(q)
            ws['MJ{}'.format(s)] = "=MIN(MJ2:MJ{})".format(q)
            ws['MJ{}'.format(t)] = "=ROUND(AVERAGE(MJ2:MJ{}),2)".format(q)
            # NILAI STANDAR EKO
            ws['MK{}'.format(r)] = "=MAX(MK2:MK{})".format(q)
            ws['MK{}'.format(s)] = "=MIN(MK2:MK{})".format(q)
            ws['MK{}'.format(t)] = "=ROUND(AVERAGE(MK2:MK{}),2)".format(q)
            # NILAI STANDAR SOS
            ws['ML{}'.format(r)] = "=MAX(ML2:ML{})".format(q)
            ws['ML{}'.format(s)] = "=MIN(ML2:ML{})".format(q)
            ws['ML{}'.format(t)] = "=ROUND(AVERAGE(ML2:ML{}),2)".format(q)
            # NILAI STANDAR GEO
            ws['MM{}'.format(r)] = "=MAX(MM2:MM{})".format(q)
            ws['MM{}'.format(s)] = "=MIN(MM2:MM{})".format(q)
            ws['MM{}'.format(t)] = "=ROUND(AVERAGE(MM2:MM{}),2)".format(q)
            # NILAI STANDAR ANT
            ws['MN{}'.format(r)] = "=MAX(MN2:MN{})".format(q)
            ws['MN{}'.format(s)] = "=MIN(MN2:MN{})".format(q)
            ws['MN{}'.format(t)] = "=ROUND(AVERAGE(MN2:MN{}),2)".format(q)
            # NILAI STANDAR BIO
            ws['MO{}'.format(r)] = "=MAX(MO2:MO{})".format(q)
            ws['MO{}'.format(s)] = "=MIN(MO2:MO{})".format(q)
            ws['MO{}'.format(t)] = "=ROUND(AVERAGE(MO2:MO{}),2)".format(q)
            # NILAI STANDAR FIS
            ws['MP{}'.format(r)] = "=MAX(MP2:MP{})".format(q)
            ws['MP{}'.format(s)] = "=MIN(MP2:MP{})".format(q)
            ws['MP{}'.format(t)] = "=ROUND(AVERAGE(MP2:MP{}),2)".format(q)
            # NILAI STANDAR KIM 1
            ws['MQ{}'.format(r)] = "=MAX(MQ2:MQ{})".format(q)
            ws['MQ{}'.format(s)] = "=MIN(MQ2:MQ{})".format(q)
            ws['MQ{}'.format(t)] = "=ROUND(AVERAGE(MQ2:MQ{}),2)".format(q)
            # NILAI STANDAR KIM 2
            ws['MR{}'.format(r)] = "=MAX(MR2:MR{})".format(q)
            ws['MR{}'.format(s)] = "=MIN(MR2:MR{})".format(q)
            ws['MR{}'.format(t)] = "=ROUND(AVERAGE(MR2:MR{}),2)".format(q)
            # NILAI STANDAR JML
            ws['MS{}'.format(r)] = "=MAX(MS2:MS{})".format(q)
            ws['MS{}'.format(s)] = "=MIN(MS2:MS{})".format(q)
            ws['MS{}'.format(t)] = "=ROUND(AVERAGE(MS2:MS{}),2)".format(q)

            # TAMBAHAN
            # MTK 1, MTK 2, IND, ENG
            ws['MV{}'.format(r)] = "=SUM(MV2:MV{})".format(q)
            ws['MW{}'.format(r)] = "=SUM(MW2:MW{})".format(q)
            ws['MX{}'.format(r)] = "=SUM(MX2:MX{})".format(q)
            ws['MY{}'.format(r)] = "=SUM(MY2:MY{})".format(q)
            # SEJ, EKO, SOS, GEO, ANT
            ws['MZ{}'.format(r)] = "=SUM(MZ2:MZ{})".format(q)
            ws['NA{}'.format(r)] = "=SUM(NA2:NA{})".format(q)
            ws['NB{}'.format(r)] = "=SUM(NB2:NB{})".format(q)
            ws['NC{}'.format(r)] = "=SUM(NC2:NC{})".format(q)
            ws['ND{}'.format(r)] = "=SUM(ND2:ND{})".format(q)
            # BIO, FIS, KIM 1, KIM 2
            ws['NE{}'.format(r)] = "=SUM(NE2:NE{})".format(q)
            ws['NF{}'.format(r)] = "=SUM(NF2:NF{})".format(q)
            ws['NG{}'.format(r)] = "=SUM(NG2:NG{})".format(q)
            ws['NH{}'.format(r)] = "=SUM(NH2:NH{})".format(q)

            # -----------------------------------------------------------------------------

            # Z Score [1]
            ws['B1'] = 'NAMA SISWA_A'
            ws['C1'] = 'NOMOR NF_A'
            ws['D1'] = 'KELAS_A'
            ws['E1'] = 'NAMA SEKOLAH_A'
            ws['F1'] = 'LOKASI_A'

            ws['G1'] = 'MAT_1_A'
            ws['H1'] = 'MAT_2_A'
            ws['I1'] = 'IND_A'
            ws['J1'] = 'ENG_A'
            ws['K1'] = 'SEJ_A'
            ws['L1'] = 'EKO_A'
            ws['M1'] = 'SOS_A'
            ws['N1'] = 'GEO_A'
            ws['O1'] = 'ANT_A'
            ws['P1'] = 'BIO_A'
            ws['Q1'] = 'FIS_A'
            ws['R1'] = 'KIM_1_A'
            ws['S1'] = 'KIM_2_A'
            ws['T1'] = 'JML_A'

            ws['U1'] = 'Z_MAT_1_A'
            ws['V1'] = 'Z_MAT_2_A'
            ws['W1'] = 'Z_IND_A'
            ws['X1'] = 'Z_ENG_A'
            ws['Y1'] = 'Z_SEJ_A'
            ws['Z1'] = 'Z_EKO_A'
            ws['AA1'] = 'Z_SOS_A'
            ws['AB1'] = 'Z_GEO_A'
            ws['AC1'] = 'Z_ANT_A'
            ws['AD1'] = 'Z_BIO_A'
            ws['AE1'] = 'Z_FIS_A'
            ws['AF1'] = 'Z_KIM_1_A'
            ws['AG1'] = 'Z_KIM_2_A'

            ws['AH1'] = 'S_MAT_1_A'
            ws['AI1'] = 'S_MAT_2_A'
            ws['AJ1'] = 'S_IND_A'
            ws['AK1'] = 'S_ENG_A'
            ws['AL1'] = 'S_SEJ_A'
            ws['AM1'] = 'S_EKO_A'
            ws['AN1'] = 'S_SOS_A'
            ws['AO1'] = 'S_GEO_A'
            ws['AP1'] = 'S_ANT_A'
            ws['AQ1'] = 'S_BIO_A'
            ws['AR1'] = 'S_FIS_A'
            ws['AS1'] = 'S_KIM_1_A'
            ws['AT1'] = 'S_KIM_2_A'
            ws['AU1'] = 'S_JML_A'

            ws['AV1'] = 'RANK NAS._A'
            ws['AW1'] = 'RANK LOK._A'

            ws['U1'].font = Font(bold=False, name='Calibri', size=11)
            ws['V1'].font = Font(bold=False, name='Calibri', size=11)
            ws['W1'].font = Font(bold=False, name='Calibri', size=11)
            ws['X1'].font = Font(bold=False, name='Calibri', size=11)
            ws['Y1'].font = Font(bold=False, name='Calibri', size=11)
            ws['Z1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AA1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AB1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AC1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AD1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AE1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AF1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AG1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AH1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AI1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AJ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AK1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AL1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AM1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AN1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AO1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AP1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AQ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AR1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AS1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AT1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AU1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AV1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AW1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['B1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['C1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['D1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['E1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['F1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['G1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['H1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['I1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['J1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['K1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['L1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['M1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['N1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['O1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['P1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['Q1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['R1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['S1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['T1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['U1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['V1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['W1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['X1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['Y1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['Z1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AA1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AB1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AC1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AD1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AE1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AF1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AG1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AH1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AI1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AJ1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AK1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AL1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AM1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AN1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AO1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AP1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AQ1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AR1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AS1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AT1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AU1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AV1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AW1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')

            # tambahan
            ws['AX1'] = 'MAT_1_20_A'
            ws['AY1'] = 'MAT_2_20_A'
            ws['AZ1'] = 'IND_20_A'
            ws['BA1'] = 'ENG_20_A'
            ws['BB1'] = 'SEJ_20_A'
            ws['BC1'] = 'EKO_20_A'
            ws['BD1'] = 'SOS_20_A'
            ws['BE1'] = 'GEO_20_A'
            ws['BF1'] = 'ANT_20_A'
            ws['BG1'] = 'BIO_20_A'
            ws['BH1'] = 'FIS_20_A'
            ws['BI1'] = 'KIM_1_20_A'
            ws['BJ1'] = 'KIM_2_20_A'

            ws['AX1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AY1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AZ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BA1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BB1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BC1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BD1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BE1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BF1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BG1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BH1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BI1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BJ1'].font = Font(bold=False, name='Calibri', size=11)

            ws['AX1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AY1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AZ1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['BA1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['BB1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['BC1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['BD1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['BE1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['BF1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['BG1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['BH1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['BI1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['BJ1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')

            for row in range(2, q+1):
                ws['T{}'.format(
                    row)] = '=SUM(G{}:M{})'.format(row, row, row)
                ws['U{}'.format(
                    row)] = '=IFERROR(ROUND(IF(G{}="","",(G{}-G${})/G${}),2),"")'.format(row, row, r, s)
                ws['V{}'.format(
                    row)] = '=IFERROR(ROUND(IF(H{}="","",(H{}-H${})/H${}),2),"")'.format(row, row, r, s)
                ws['W{}'.format(
                    row)] = '=IFERROR(ROUND(IF(I{}="","",(I{}-I${})/I${}),2),"")'.format(row, row, r, s)
                ws['X{}'.format(
                    row)] = '=IFERROR(ROUND(IF(J{}="","",(J{}-J${})/J${}),2),"")'.format(row, row, r, s)
                ws['Y{}'.format(
                    row)] = '=IFERROR(ROUND(IF(K{}="","",(K{}-K${})/K${}),2),"")'.format(row, row, r, s)
                ws['Z{}'.format(
                    row)] = '=IFERROR(ROUND(IF(L{}="","",(L{}-L${})/L${}),2),"")'.format(row, row, r, s)
                ws['AA{}'.format(
                    row)] = '=IFERROR(ROUND(IF(M{}="","",(M{}-M${})/M${}),2),"")'.format(row, row, r, s)
                ws['AB{}'.format(
                    row)] = '=IFERROR(ROUND(IF(N{}="","",(N{}-N${})/N${}),2),"")'.format(row, row, r, s)
                ws['AC{}'.format(
                    row)] = '=IFERROR(ROUND(IF(O{}="","",(O{}-O${})/O${}),2),"")'.format(row, row, r, s)
                ws['AD{}'.format(
                    row)] = '=IFERROR(ROUND(IF(P{}="","",(P{}-P${})/P${}),2),"")'.format(row, row, r, s)
                ws['AE{}'.format(
                    row)] = '=IFERROR(ROUND(IF(Q{}="","",(Q{}-Q${})/Q${}),2),"")'.format(row, row, r, s)
                ws['AF{}'.format(
                    row)] = '=IFERROR(ROUND(IF(R{}="","",(R{}-R${})/R${}),2),"")'.format(row, row, r, s)
                ws['AG{}'.format(
                    row)] = '=IFERROR(ROUND(IF(S{}="","",(S{}-S${})/S${}),2),"")'.format(row, row, r, s)

                ws['AH{}'.format(
                    row)] = '=IFERROR(ROUND(IF(G{}="","",IF(70+30*U{}/$U${}<20,20,70+30*U{}/$U${})),2),"")'.format(row, row, r, row, r)
                ws['AI{}'.format(
                    row)] = '=IFERROR(ROUND(IF(H{}="","",IF(70+30*V{}/$V${}<20,20,70+30*V{}/$V${})),2),"")'.format(row, row, r, row, r)
                ws['AJ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(I{}="","",IF(70+30*W{}/$W${}<20,20,70+30*W{}/$W${})),2),"")'.format(row, row, r, row, r)
                ws['AK{}'.format(
                    row)] = '=IFERROR(ROUND(IF(J{}="","",IF(70+30*X{}/$X${}<20,20,70+30*X{}/$X${})),2),"")'.format(row, row, r, row, r)
                ws['AL{}'.format(
                    row)] = '=IFERROR(ROUND(IF(K{}="","",IF(70+30*Y{}/$Y${}<20,20,70+30*Y{}/$Y${})),2),"")'.format(row, row, r, row, r)
                ws['AM{}'.format(
                    row)] = '=IFERROR(ROUND(IF(L{}="","",IF(70+30*Z{}/$Z${}<20,20,70+30*Z{}/$Z${})),2),"")'.format(row, row, r, row, r)
                ws['AN{}'.format(
                    row)] = '=IFERROR(ROUND(IF(M{}="","",IF(70+30*AA{}/$AA${}<20,20,70+30*AA{}/$AA${})),2),"")'.format(row, row, r, row, r)
                ws['AO{}'.format(
                    row)] = '=IFERROR(ROUND(IF(N{}="","",IF(70+30*AB{}/$AB${}<20,20,70+30*AB{}/$AB${})),2),"")'.format(row, row, r, row, r)
                ws['AP{}'.format(
                    row)] = '=IFERROR(ROUND(IF(O{}="","",IF(70+30*AC{}/$AC${}<20,20,70+30*AC{}/$AC${})),2),"")'.format(row, row, r, row, r)
                ws['AQ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(P{}="","",IF(70+30*AD{}/$AD${}<20,20,70+30*AD{}/$AD${})),2),"")'.format(row, row, r, row, r)
                ws['AR{}'.format(
                    row)] = '=IFERROR(ROUND(IF(Q{}="","",IF(70+30*AE{}/$AE${}<20,20,70+30*AE{}/$AE${})),2),"")'.format(row, row, r, row, r)
                ws['AS{}'.format(
                    row)] = '=IFERROR(ROUND(IF(R{}="","",IF(70+30*AF{}/$AF${}<20,20,70+30*AF{}/$AF${})),2),"")'.format(row, row, r, row, r)
                ws['AT{}'.format(
                    row)] = '=IFERROR(ROUND(IF(S{}="","",IF(70+30*AG{}/$AG${}<20,20,70+30*AG{}/$AG${})),2),"")'.format(row, row, r, row, r)

                ws['AU{}'.format(row)] = '=IF(SUM(AH{}:AT{})=0,"",SUM(AH{}:AT{}))'.format(
                    row, row, row, row)
                ws['AV{}'.format(row)] = '=IF(AU{}="","",RANK(AU{},$AU$2:$AU${}))'.format(
                    row, row, q)
                ws['AW{}'.format(
                    row)] = '=IF(AV{}="","",COUNTIFS($F$2:$F${},F{},$AV$2:$AV${},"<"&AV{})+1)'.format(row, q, row, q, row)
            # TAMBAHAN
                ws['AX{}'.format(row)] = '=IF($G${}=20,IF(AND(G{}>3,V{}=20),1,""),IF($G${}=25,IF(AND(G{}>4,V{}=20),1,""),IF($G${}=30,IF(AND(G{}>5,V{}=20),1,""),IF($G${}=35,IF(AND(G{}>6,V{}=20),1,""),IF($G${}=40,IF(AND(G{}>7,V{}=20),1,""),IF($G${}=45,IF(AND(G{}>8,V{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['AY{}'.format(row)] = '=IF($H${}=20,IF(AND(H{}>3,W{}=20),1,""),IF($H${}=25,IF(AND(H{}>4,W{}=20),1,""),IF($H${}=30,IF(AND(H{}>5,W{}=20),1,""),IF($H${}=35,IF(AND(H{}>6,W{}=20),1,""),IF($H${}=40,IF(AND(H{}>7,W{}=20),1,""),IF($H${}=45,IF(AND(H{}>8,W{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['AZ{}'.format(row)] = '=IF($I${}=20,IF(AND(I{}>3,X{}=20),1,""),IF($I${}=25,IF(AND(I{}>4,X{}=20),1,""),IF($I${}=30,IF(AND(I{}>5,X{}=20),1,""),IF($I${}=35,IF(AND(I{}>6,X{}=20),1,""),IF($I${}=40,IF(AND(I{}>7,X{}=20),1,""),IF($I${}=45,IF(AND(I{}>8,X{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['BA{}'.format(row)] = '=IF($J${}=20,IF(AND(J{}>3,Y{}=20),1,""),IF($J${}=25,IF(AND(J{}>4,Y{}=20),1,""),IF($J${}=30,IF(AND(J{}>5,Y{}=20),1,""),IF($J${}=35,IF(AND(J{}>6,Y{}=20),1,""),IF($J${}=40,IF(AND(J{}>7,Y{}=20),1,""),IF($J${}=45,IF(AND(J{}>8,Y{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['BB{}'.format(row)] = '=IF($K${}=20,IF(AND(K{}>3,Z{}=20),1,""),IF($K${}=25,IF(AND(K{}>4,Z{}=20),1,""),IF($K${}=30,IF(AND(K{}>5,Z{}=20),1,""),IF($K${}=35,IF(AND(K{}>6,Z{}=20),1,""),IF($K${}=40,IF(AND(K{}>7,Z{}=20),1,""),IF($K${}=45,IF(AND(K{}>8,Z{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['BC{}'.format(row)] = '=IF($L${}=20,IF(AND(L{}>3,AA{}=20),1,""),IF($L${}=25,IF(AND(L{}>4,AA{}=20),1,""),IF($L${}=30,IF(AND(L{}>5,AA{}=20),1,""),IF($L${}=35,IF(AND(L{}>6,AA{}=20),1,""),IF($L${}=40,IF(AND(L{}>7,AA{}=20),1,""),IF($L${}=45,IF(AND(L{}>8,AA{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['BD{}'.format(row)] = '=IF($M${}=20,IF(AND(M{}>3,AB{}=20),1,""),IF($M${}=25,IF(AND(M{}>4,AB{}=20),1,""),IF($M${}=30,IF(AND(M{}>5,AB{}=20),1,""),IF($M${}=35,IF(AND(M{}>6,AB{}=20),1,""),IF($M${}=40,IF(AND(M{}>7,AB{}=20),1,""),IF($M${}=45,IF(AND(M{}>8,AB{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['BE{}'.format(row)] = '=IF($N${}=20,IF(AND(N{}>3,AC{}=20),1,""),IF($N${}=25,IF(AND(N{}>4,AC{}=20),1,""),IF($N${}=30,IF(AND(N{}>5,AC{}=20),1,""),IF($N${}=35,IF(AND(N{}>6,AC{}=20),1,""),IF($N${}=40,IF(AND(N{}>7,AC{}=20),1,""),IF($N${}=45,IF(AND(N{}>8,AC{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['BF{}'.format(row)] = '=IF($O${}=20,IF(AND(O{}>3,AD{}=20),1,""),IF($O${}=25,IF(AND(O{}>4,AD{}=20),1,""),IF($O${}=30,IF(AND(O{}>5,AD{}=20),1,""),IF($O${}=35,IF(AND(O{}>6,AD{}=20),1,""),IF($O${}=40,IF(AND(O{}>7,AD{}=20),1,""),IF($O${}=45,IF(AND(O{}>8,AD{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['BG{}'.format(row)] = '=IF($P${}=20,IF(AND(P{}>3,AE{}=20),1,""),IF($P${}=25,IF(AND(P{}>4,AE{}=20),1,""),IF($P${}=30,IF(AND(P{}>5,AE{}=20),1,""),IF($P${}=35,IF(AND(P{}>6,AE{}=20),1,""),IF($P${}=40,IF(AND(P{}>7,AE{}=20),1,""),IF($P${}=45,IF(AND(P{}>8,AE{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['BH{}'.format(row)] = '=IF($Q${}=20,IF(AND(Q{}>3,AF{}=20),1,""),IF($Q${}=25,IF(AND(Q{}>4,AF{}=20),1,""),IF($Q${}=30,IF(AND(Q{}>5,AF{}=20),1,""),IF($Q${}=35,IF(AND(Q{}>6,AF{}=20),1,""),IF($Q${}=40,IF(AND(Q{}>7,AF{}=20),1,""),IF($Q${}=45,IF(AND(Q{}>8,AF{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['BI{}'.format(row)] = '=IF($R${}=20,IF(AND(R{}>3,AG{}=20),1,""),IF($R${}=25,IF(AND(R{}>4,AG{}=20),1,""),IF($R${}=30,IF(AND(R{}>5,AG{}=20),1,""),IF($R${}=35,IF(AND(R{}>6,AG{}=20),1,""),IF($R${}=40,IF(AND(R{}>7,AG{}=20),1,""),IF($R${}=45,IF(AND(R{}>8,AG{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['BJ{}'.format(row)] = '=IF($S${}=20,IF(AND(S{}>3,AH{}=20),1,""),IF($S${}=25,IF(AND(S{}>4,AH{}=20),1,""),IF($S${}=30,IF(AND(S{}>5,AH{}=20),1,""),IF($S${}=35,IF(AND(S{}>6,AH{}=20),1,""),IF($S${}=40,IF(AND(S{}>7,AH{}=20),1,""),IF($S${}=45,IF(AND(S{}>8,AH{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

            # Z Score [2]
            ws['BL1'] = 'NAMA SISWA_B'
            ws['BM1'] = 'NOMOR NF_B'
            ws['BN1'] = 'KELAS_B'
            ws['BO1'] = 'NAMA SEKOLAH_B'
            ws['BP1'] = 'LOKASI_B'

            ws['BQ1'] = 'MAT_1_B'
            ws['BR1'] = 'MAT_2_B'
            ws['BS1'] = 'IND_B'
            ws['BT1'] = 'ENG_B'
            ws['BU1'] = 'SEJ_B'
            ws['BV1'] = 'EKO_B'
            ws['BW1'] = 'SOS_B'
            ws['BX1'] = 'GEO_B'
            ws['BY1'] = 'ANT_B'
            ws['BZ1'] = 'BIO_B'
            ws['CA1'] = 'FIS_B'
            ws['CB1'] = 'KIM_1_B'
            ws['CC1'] = 'KIM_2_B'
            ws['CD1'] = 'JML_B'

            ws['CE1'] = 'Z_MAT_1_B'
            ws['CF1'] = 'Z_MAT_2_B'
            ws['CG1'] = 'Z_IND_B'
            ws['CH1'] = 'Z_ENG_B'
            ws['CI1'] = 'Z_SEJ_B'
            ws['CJ1'] = 'Z_EKO_B'
            ws['CK1'] = 'Z_SOS_B'
            ws['CL1'] = 'Z_GEO_B'
            ws['CM1'] = 'Z_ANT_B'
            ws['CN1'] = 'Z_BIO_B'
            ws['CO1'] = 'Z_FIS_B'
            ws['CP1'] = 'Z_KIM_1_B'
            ws['CQ1'] = 'Z_KIM_2_B'

            ws['CR1'] = 'S_MAT_1_B'
            ws['CS1'] = 'S_MAT_2_B'
            ws['CT1'] = 'S_IND_B'
            ws['CU1'] = 'S_ENG_B'
            ws['CV1'] = 'S_SEJ_B'
            ws['CW1'] = 'S_EKO_B'
            ws['CX1'] = 'S_SOS_B'
            ws['CY1'] = 'S_GEO_B'
            ws['CZ1'] = 'S_ANT_B'
            ws['DA1'] = 'S_BIO_B'
            ws['DB1'] = 'S_FIS_B'
            ws['DC1'] = 'S_KIM_1_B'
            ws['DD1'] = 'S_KIM_2_B'
            ws['DE1'] = 'S_JML_B'

            ws['DF1'] = 'RANK NAS._B'
            ws['DG1'] = 'RANK LOK._B'

            # Z MAT 1
            ws['CE1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CF1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CG1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CH1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CI1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CJ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CK1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CL1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CM1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CN1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CO1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CP1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CQ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CR1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CS1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CT1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CU1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CV1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CW1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CX1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CY1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CZ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DA1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DB1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DC1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DD1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DE1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DF1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DG1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL DARI NAMA
            ws['BL1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BM1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BN1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BO1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BP1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BQ1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BR1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BS1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BT1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BU1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BV1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BW1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BX1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BY1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BZ1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['CA1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['CB1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['CC1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['CD1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['CE1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['CF1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['CG1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['CH1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['CI1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['CJ1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['CK1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['CL1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['CM1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['CN1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['CO1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['CP1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['CQ1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['CR1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['CS1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['CT1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['CU1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['CV1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['CW1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['CX1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['CY1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['CZ1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['DA1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['DB1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['DC1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['DD1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['DE1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['DF1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['DG1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')

            # tambahan
            ws['DH1'] = 'MAT_1_20_B'
            ws['DI1'] = 'MAT_2_20_B'
            ws['DJ1'] = 'IND_20_B'
            ws['DK1'] = 'ENG_20_B'
            ws['DL1'] = 'SEJ_20_B'
            ws['DM1'] = 'EKO_20_B'
            ws['DN1'] = 'SOS_20_B'
            ws['DO1'] = 'GEO_20_B'
            ws['DP1'] = 'ANT_20_B'
            ws['DQ1'] = 'BIO_20_B'
            ws['DR1'] = 'FIS_20_B'
            ws['DS1'] = 'KIM_1_20_B'
            ws['DT1'] = 'KIM_2_20_B'

            ws['DH1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DI1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DJ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DK1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DL1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DM1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DN1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DO1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DP1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DQ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DR1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DS1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DT1'].font = Font(bold=False, name='Calibri', size=11)

            ws['DH1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['DI1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['DJ1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['DK1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['DL1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['DM1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['DN1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['DO1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['DP1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['DQ1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['DR1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['DS1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['DT1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')

            # DARI JML
            for row in range(2, q+1):
                ws['BL{}'.format(row)] = '=B{}'.format(row)
                ws['BM{}'.format(row)] = '=C{}'.format(row, row)
                ws['BN{}'.format(row)] = '=D{}'.format(row, row)
                ws['BO{}'.format(row)] = '=E{}'.format(row, row)
                ws['BP{}'.format(row)] = '=F{}'.format(row, row)
                ws['BQ{}'.format(row)] = '=IF(G{}="","",G{})'.format(row, row)
                ws['BR{}'.format(row)] = '=IF(H{}="","",H{})'.format(row, row)
                ws['BS{}'.format(row)] = '=IF(I{}="","",I{})'.format(row, row)
                ws['BT{}'.format(row)] = '=IF(J{}="","",J{})'.format(row, row)
                ws['BU{}'.format(row)] = '=IF(K{}="","",K{})'.format(row, row)
                ws['BV{}'.format(row)] = '=IF(L{}="","",L{})'.format(row, row)
                ws['BW{}'.format(row)] = '=IF(M{}="","",M{})'.format(row, row)
                ws['BX{}'.format(row)] = '=IF(N{}="","",N{})'.format(row, row)
                ws['BY{}'.format(row)] = '=IF(O{}="","",O{})'.format(row, row)
                ws['BZ{}'.format(row)] = '=IF(P{}="","",P{})'.format(row, row)
                ws['CA{}'.format(row)] = '=IF(Q{}="","",Q{})'.format(row, row)
                ws['CB{}'.format(row)] = '=IF(R{}="","",R{})'.format(row, row)
                ws['CC{}'.format(row)] = '=IF(S{}="","",S{})'.format(row, row)
                ws['CD{}'.format(row)] = '=IF(T{}="","",T{})'.format(row, row)
                
                ws['CE{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BQ{}="","",(BQ{}-BQ${})/BQ${}),2),"")'.format(row, row, r, s)
                ws['CF{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BR{}="","",(BR{}-BR${})/BR${}),2),"")'.format(row, row, r, s)
                ws['CG{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BS{}="","",(BS{}-BS${})/BS${}),2),"")'.format(row, row, r, s)
                ws['CH{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BT{}="","",(BT{}-BT${})/BT${}),2),"")'.format(row, row, r, s)
                ws['CI{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BU{}="","",(BU{}-BU${})/BU${}),2),"")'.format(row, row, r, s)
                ws['CJ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BV{}="","",(BV{}-BV${})/BV${}),2),"")'.format(row, row, r, s)
                ws['CK{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BW{}="","",(BW{}-BW${})/BW${}),2),"")'.format(row, row, r, s)
                ws['CL{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BX{}="","",(BX{}-BX${})/BX${}),2),"")'.format(row, row, r, s)
                ws['CM{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BY{}="","",(BY{}-BY${})/BY${}),2),"")'.format(row, row, r, s)
                ws['CN{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BZ{}="","",(BZ{}-BZ${})/BZ${}),2),"")'.format(row, row, r, s)
                ws['CO{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CA{}="","",(CA{}-CA${})/CA${}),2),"")'.format(row, row, r, s)
                ws['CP{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CB{}="","",(CB{}-CB${})/CB${}),2),"")'.format(row, row, r, s)
                ws['CQ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CC{}="","",(CC{}-CC${})/CC${}),2),"")'.format(row, row, r, s)

                ws['CR{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BQ{}="","",IF(70+30*CE{}/$CE${}<20,20,70+30*CE{}/$CE${})),2),"")'.format(row, row, r, row, r)
                ws['CS{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BR{}="","",IF(70+30*CF{}/$CF${}<20,20,70+30*CF{}/$CF${})),2),"")'.format(row, row, r, row, r)
                ws['CT{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BS{}="","",IF(70+30*CG{}/$CG${}<20,20,70+30*CG{}/$CG${})),2),"")'.format(row, row, r, row, r)
                ws['CU{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BT{}="","",IF(70+30*CH{}/$CH${}<20,20,70+30*CH{}/$CH${})),2),"")'.format(row, row, r, row, r)
                ws['CV{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BU{}="","",IF(70+30*CI{}/$CI${}<20,20,70+30*CI{}/$CI${})),2),"")'.format(row, row, r, row, r)
                ws['CW{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BV{}="","",IF(70+30*CJ{}/$CJ${}<20,20,70+30*CJ{}/$CJ${})),2),"")'.format(row, row, r, row, r)
                ws['CX{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BW{}="","",IF(70+30*CK{}/$CK${}<20,20,70+30*CK{}/$CK${})),2),"")'.format(row, row, r, row, r)
                ws['CY{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BX{}="","",IF(70+30*CL{}/$CL${}<20,20,70+30*CL{}/$CL${})),2),"")'.format(row, row, r, row, r)
                ws['CZ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BY{}="","",IF(70+30*CM{}/$CM${}<20,20,70+30*CM{}/$CM${})),2),"")'.format(row, row, r, row, r)
                ws['DA{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BZ{}="","",IF(70+30*CN{}/$CN${}<20,20,70+30*CN{}/$CN${})),2),"")'.format(row, row, r, row, r)
                ws['DB{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CA{}="","",IF(70+30*CO{}/$CO${}<20,20,70+30*CO{}/$CO${})),2),"")'.format(row, row, r, row, r)
                ws['DC{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CB{}="","",IF(70+30*CP{}/$CP${}<20,20,70+30*CP{}/$CP${})),2),"")'.format(row, row, r, row, r)
                ws['DD{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CC{}="","",IF(70+30*CQ{}/$CQ${}<20,20,70+30*CQ{}/$CQ${})),2),"")'.format(row, row, r, row, r)

                ws['DE{}'.format(row)] = '=IF(SUM(CR{}:DD{})=0,"",SUM(CR{}:DD{}))'.format(
                    row, row, row, row)
                ws['DF{}'.format(row)] = '=IF(DE{}="","",RANK(DE{},$DE$2:$DE${}))'.format(
                    row, row, q)
                ws['DG{}'.format(
                    row)] = '=IF(DF{}="","",COUNTIFS($BP$2:$BP${},BP{},$DF$2:$DF${},"<"&DF{})+1)'.format(row, q, row, q, row)
            # TAMBAHAN
                ws['DH{}'.format(row)] = '=IF($G${}=20,IF(AND(G{}>3,CE{}=20),1,""),IF($G${}=25,IF(AND(G{}>4,CE{}=20),1,""),IF($G${}=30,IF(AND(G{}>5,CE{}=20),1,""),IF($G${}=35,IF(AND(G{}>6,CE{}=20),1,""),IF($G${}=40,IF(AND(G{}>7,CE{}=20),1,""),IF($G${}=45,IF(AND(G{}>8,CE{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['DI{}'.format(row)] = '=IF($H${}=20,IF(AND(H{}>3,CF{}=20),1,""),IF($H${}=25,IF(AND(H{}>4,CF{}=20),1,""),IF($H${}=30,IF(AND(H{}>5,CF{}=20),1,""),IF($H${}=35,IF(AND(H{}>6,CF{}=20),1,""),IF($H${}=40,IF(AND(H{}>7,CF{}=20),1,""),IF($H${}=45,IF(AND(H{}>8,CF{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['DJ{}'.format(row)] = '=IF($I${}=20,IF(AND(I{}>3,CG{}=20),1,""),IF($I${}=25,IF(AND(I{}>4,CG{}=20),1,""),IF($I${}=30,IF(AND(I{}>5,CG{}=20),1,""),IF($I${}=35,IF(AND(I{}>6,CG{}=20),1,""),IF($I${}=40,IF(AND(I{}>7,CG{}=20),1,""),IF($I${}=45,IF(AND(I{}>8,CG{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['DK{}'.format(row)] = '=IF($J${}=20,IF(AND(J{}>3,CH{}=20),1,""),IF($J${}=25,IF(AND(J{}>4,CH{}=20),1,""),IF($J${}=30,IF(AND(J{}>5,CH{}=20),1,""),IF($J${}=35,IF(AND(J{}>6,CH{}=20),1,""),IF($J${}=40,IF(AND(J{}>7,CH{}=20),1,""),IF($J${}=45,IF(AND(J{}>8,CH{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['DL{}'.format(row)] = '=IF($K${}=20,IF(AND(K{}>3,CI{}=20),1,""),IF($K${}=25,IF(AND(K{}>4,CI{}=20),1,""),IF($K${}=30,IF(AND(K{}>5,CI{}=20),1,""),IF($K${}=35,IF(AND(K{}>6,CI{}=20),1,""),IF($K${}=40,IF(AND(K{}>7,CI{}=20),1,""),IF($K${}=45,IF(AND(K{}>8,CI{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['DM{}'.format(row)] = '=IF($L${}=20,IF(AND(L{}>3,CJ{}=20),1,""),IF($L${}=25,IF(AND(L{}>4,CJ{}=20),1,""),IF($L${}=30,IF(AND(L{}>5,CJ{}=20),1,""),IF($L${}=35,IF(AND(L{}>6,CJ{}=20),1,""),IF($L${}=40,IF(AND(L{}>7,CJ{}=20),1,""),IF($L${}=45,IF(AND(L{}>8,CJ{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['DN{}'.format(row)] = '=IF($M${}=20,IF(AND(M{}>3,CK{}=20),1,""),IF($M${}=25,IF(AND(M{}>4,CK{}=20),1,""),IF($M${}=30,IF(AND(M{}>5,CK{}=20),1,""),IF($M${}=35,IF(AND(M{}>6,CK{}=20),1,""),IF($M${}=40,IF(AND(M{}>7,CK{}=20),1,""),IF($M${}=45,IF(AND(M{}>8,CK{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['DO{}'.format(row)] = '=IF($N${}=20,IF(AND(N{}>3,CL{}=20),1,""),IF($N${}=25,IF(AND(N{}>4,CL{}=20),1,""),IF($N${}=30,IF(AND(N{}>5,CL{}=20),1,""),IF($N${}=35,IF(AND(N{}>6,CL{}=20),1,""),IF($N${}=40,IF(AND(N{}>7,CL{}=20),1,""),IF($N${}=45,IF(AND(N{}>8,CL{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['DP{}'.format(row)] = '=IF($O${}=20,IF(AND(O{}>3,CM{}=20),1,""),IF($O${}=25,IF(AND(O{}>4,CM{}=20),1,""),IF($O${}=30,IF(AND(O{}>5,CM{}=20),1,""),IF($O${}=35,IF(AND(O{}>6,CM{}=20),1,""),IF($O${}=40,IF(AND(O{}>7,CM{}=20),1,""),IF($O${}=45,IF(AND(O{}>8,CM{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['DQ{}'.format(row)] = '=IF($P${}=20,IF(AND(P{}>3,CN{}=20),1,""),IF($P${}=25,IF(AND(P{}>4,CN{}=20),1,""),IF($P${}=30,IF(AND(P{}>5,CN{}=20),1,""),IF($P${}=35,IF(AND(P{}>6,CN{}=20),1,""),IF($P${}=40,IF(AND(P{}>7,CN{}=20),1,""),IF($P${}=45,IF(AND(P{}>8,CN{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['DR{}'.format(row)] = '=IF($Q${}=20,IF(AND(Q{}>3,CO{}=20),1,""),IF($Q${}=25,IF(AND(Q{}>4,CO{}=20),1,""),IF($Q${}=30,IF(AND(Q{}>5,CO{}=20),1,""),IF($Q${}=35,IF(AND(Q{}>6,CO{}=20),1,""),IF($Q${}=40,IF(AND(Q{}>7,CO{}=20),1,""),IF($Q${}=45,IF(AND(Q{}>8,CO{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['DS{}'.format(row)] = '=IF($R${}=20,IF(AND(R{}>3,CP{}=20),1,""),IF($R${}=25,IF(AND(R{}>4,CP{}=20),1,""),IF($R${}=30,IF(AND(R{}>5,CP{}=20),1,""),IF($R${}=35,IF(AND(R{}>6,CP{}=20),1,""),IF($R${}=40,IF(AND(R{}>7,CP{}=20),1,""),IF($R${}=45,IF(AND(R{}>8,CP{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['DT{}'.format(row)] = '=IF($S${}=20,IF(AND(S{}>3,CQ{}=20),1,""),IF($S${}=25,IF(AND(S{}>4,CQ{}=20),1,""),IF($S${}=30,IF(AND(S{}>5,CQ{}=20),1,""),IF($S${}=35,IF(AND(S{}>6,CQ{}=20),1,""),IF($S${}=40,IF(AND(S{}>7,CQ{}=20),1,""),IF($S${}=45,IF(AND(S{}>8,CQ{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

            # Z Score [3]
            ws['DV1'] = 'NAMA SISWA_C'
            ws['DW1'] = 'NOMOR NF_C'
            ws['DX1'] = 'KELAS_C'
            ws['DY1'] = 'NAMA SEKOLAH_C'
            ws['DZ1'] = 'LOKASI_C'

            ws['EA1'] = 'MAT_1_C'
            ws['EB1'] = 'MAT_2_C'
            ws['EC1'] = 'IND_C'
            ws['ED1'] = 'ENG_C'
            ws['EE1'] = 'SEJ_C'
            ws['EF1'] = 'EKO_C'
            ws['EG1'] = 'SOS_C'
            ws['EH1'] = 'GEO_C'
            ws['EI1'] = 'ANT_C'
            ws['EJ1'] = 'BIO_C'
            ws['EK1'] = 'FIS_C'
            ws['EL1'] = 'KIM_1_C'
            ws['EM1'] = 'KIM_2_C'
            ws['EN1'] = 'JML_C'

            ws['EO1'] = 'Z_MAT_1_C'
            ws['EP1'] = 'Z_MAT_2_C'
            ws['EQ1'] = 'Z_IND_C'
            ws['ER1'] = 'Z_ENG_C'
            ws['ES1'] = 'Z_SEJ_C'
            ws['ET1'] = 'Z_EKO_C'
            ws['EU1'] = 'Z_SOS_C'
            ws['EV1'] = 'Z_GEO_C'
            ws['EW1'] = 'Z_ANT_C'
            ws['EX1'] = 'Z_BIO_C'
            ws['EY1'] = 'Z_FIS_C'
            ws['EZ1'] = 'Z_KIM_1_C'
            ws['FA1'] = 'Z_KIM_2_C'

            ws['FB1'] = 'S_MAT_1_C'
            ws['FC1'] = 'S_MAT_2_C'
            ws['FD1'] = 'S_IND_C'
            ws['FE1'] = 'S_ENG_C'
            ws['FF1'] = 'S_SEJ_C'
            ws['FG1'] = 'S_EKO_C'
            ws['FH1'] = 'S_SOS_C'
            ws['FI1'] = 'S_GEO_C'
            ws['FJ1'] = 'S_ANT_C'
            ws['FK1'] = 'S_BIO_C'
            ws['FL1'] = 'S_FIS_C'
            ws['FM1'] = 'S_KIM_1_C'
            ws['FN1'] = 'S_KIM_2_C'
            ws['FO1'] = 'S_JML_C'

            ws['FP1'] = 'RANK NAS._C'
            ws['FQ1'] = 'RANK LOK._C'

            # Z MAT 1
            ws['EO1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EP1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EQ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['ER1'].font = Font(bold=False, name='Calibri', size=11)
            ws['ES1'].font = Font(bold=False, name='Calibri', size=11)
            ws['ET1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EU1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EV1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EW1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EX1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EY1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EZ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FA1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FB1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FC1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FD1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FE1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FF1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FG1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FH1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FI1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FJ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FK1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FL1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FM1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FN1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FO1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FP1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FQ1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL DARI NAMA
            ws['DV1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['DW1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['DX1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['DY1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['DZ1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['EA1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['EB1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['EC1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['ED1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['EE1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['EF1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['EG1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['EH1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['EI1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['EJ1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['EK1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['EL1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['EM1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['EN1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['EO1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['EP1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['EQ1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['ER1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['ES1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['ET1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['EU1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['EV1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['EW1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['EX1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['EY1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['EZ1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['FA1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['FB1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['FC1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['FD1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['FE1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['FF1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['FG1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['FH1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['FI1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['FJ1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['FK1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['FL1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['FM1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['FN1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['FO1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['FP1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['FQ1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            
            # tambahan
            ws['FR1'] = 'MAT_1_20_C'
            ws['FS1'] = 'MAT_2_20_C'
            ws['FT1'] = 'IND_20_C'
            ws['FU1'] = 'ENG_20_C'
            ws['FV1'] = 'SEJ_20_C'
            ws['FW1'] = 'EKO_20_C'
            ws['FX1'] = 'SOS_20_C'
            ws['FY1'] = 'GEO_20_C'
            ws['FZ1'] = 'ANT_20_C'
            ws['GA1'] = 'BIO_20_C'
            ws['GB1'] = 'FIS_20_C'
            ws['GC1'] = 'KIM_1_20_C'
            ws['GD1'] = 'KIM_2_20_C'

            ws['FR1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FS1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FT1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FU1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FV1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FW1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FX1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FY1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FZ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['GA1'].font = Font(bold=False, name='Calibri', size=11)
            ws['GB1'].font = Font(bold=False, name='Calibri', size=11)
            ws['GC1'].font = Font(bold=False, name='Calibri', size=11)
            ws['GD1'].font = Font(bold=False, name='Calibri', size=11)

            ws['FR1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['FS1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['FT1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['FU1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['FV1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['FW1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['FX1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['FY1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['FZ1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['GA1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['GB1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['GC1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['GD1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')

            # DARI JML
            for row in range(2, q+1):
                ws['DV{}'.format(row)] = '=B{}'.format(row)
                ws['DW{}'.format(row)] = '=C{}'.format(row, row)
                ws['DX{}'.format(row)] = '=D{}'.format(row, row)
                ws['DY{}'.format(row)] = '=E{}'.format(row, row)
                ws['DZ{}'.format(row)] = '=F{}'.format(row, row)
                ws['EA{}'.format(row)] = '=IF(G{}="","",G{})'.format(row, row)
                ws['EB{}'.format(row)] = '=IF(H{}="","",H{})'.format(row, row)
                ws['EC{}'.format(row)] = '=IF(I{}="","",I{})'.format(row, row)
                ws['ED{}'.format(row)] = '=IF(J{}="","",J{})'.format(row, row)
                ws['EE{}'.format(row)] = '=IF(K{}="","",K{})'.format(row, row)
                ws['EF{}'.format(row)] = '=IF(L{}="","",L{})'.format(row, row)
                ws['EG{}'.format(row)] = '=IF(M{}="","",M{})'.format(row, row)
                ws['EH{}'.format(row)] = '=IF(N{}="","",N{})'.format(row, row)
                ws['EI{}'.format(row)] = '=IF(O{}="","",O{})'.format(row, row)
                ws['EJ{}'.format(row)] = '=IF(P{}="","",P{})'.format(row, row)
                ws['EK{}'.format(row)] = '=IF(Q{}="","",Q{})'.format(row, row)
                ws['EL{}'.format(row)] = '=IF(R{}="","",R{})'.format(row, row)
                ws['EM{}'.format(row)] = '=IF(S{}="","",S{})'.format(row, row)
                ws['EN{}'.format(row)] = '=IF(T{}="","",T{})'.format(row, row)
                
                ws['EO{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EA{}="","",(EA{}-EA${})/EA${}),2),"")'.format(row, row, r, s)
                ws['EP{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EB{}="","",(EB{}-EB${})/EB${}),2),"")'.format(row, row, r, s)
                ws['EQ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EC{}="","",(EC{}-EC${})/EC${}),2),"")'.format(row, row, r, s)
                ws['ER{}'.format(
                    row)] = '=IFERROR(ROUND(IF(ED{}="","",(ED{}-ED${})/ED${}),2),"")'.format(row, row, r, s)
                ws['ES{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EE{}="","",(EE{}-EE${})/EE${}),2),"")'.format(row, row, r, s)
                ws['ET{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EF{}="","",(EF{}-EF${})/EF${}),2),"")'.format(row, row, r, s)
                ws['EU{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EG{}="","",(EG{}-EG${})/EG${}),2),"")'.format(row, row, r, s)
                ws['EV{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EH{}="","",(EH{}-EH${})/EH${}),2),"")'.format(row, row, r, s)
                ws['EW{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EI{}="","",(EI{}-EI${})/EI${}),2),"")'.format(row, row, r, s)
                ws['EX{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EJ{}="","",(EJ{}-EJ${})/EJ${}),2),"")'.format(row, row, r, s)
                ws['EY{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EK{}="","",(EK{}-EK${})/EK${}),2),"")'.format(row, row, r, s)
                ws['EZ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EL{}="","",(EL{}-EL${})/EL${}),2),"")'.format(row, row, r, s)
                ws['FA{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EM{}="","",(EM{}-EM${})/EM${}),2),"")'.format(row, row, r, s)

                ws['FB{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EA{}="","",IF(70+30*EO{}/$EO${}<20,20,70+30*EO{}/$EO${})),2),"")'.format(row, row, r, row, r)
                ws['FC{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EB{}="","",IF(70+30*EP{}/$EP${}<20,20,70+30*EP{}/$EP${})),2),"")'.format(row, row, r, row, r)
                ws['FD{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EC{}="","",IF(70+30*EQ{}/$EQ${}<20,20,70+30*EQ{}/$EQ${})),2),"")'.format(row, row, r, row, r)
                ws['FE{}'.format(
                    row)] = '=IFERROR(ROUND(IF(ED{}="","",IF(70+30*ER{}/$ER${}<20,20,70+30*ER{}/$ER${})),2),"")'.format(row, row, r, row, r)
                ws['FF{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EE{}="","",IF(70+30*ES{}/$ES${}<20,20,70+30*ES{}/$ES${})),2),"")'.format(row, row, r, row, r)
                ws['FG{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EF{}="","",IF(70+30*ET{}/$ET${}<20,20,70+30*ET{}/$ET${})),2),"")'.format(row, row, r, row, r)
                ws['FH{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EG{}="","",IF(70+30*EU{}/$EU${}<20,20,70+30*EU{}/$EU${})),2),"")'.format(row, row, r, row, r)
                ws['FI{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EH{}="","",IF(70+30*EV{}/$EV${}<20,20,70+30*EV{}/$EV${})),2),"")'.format(row, row, r, row, r)
                ws['FJ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EI{}="","",IF(70+30*EW{}/$EW${}<20,20,70+30*EW{}/$EW${})),2),"")'.format(row, row, r, row, r)
                ws['FK{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EJ{}="","",IF(70+30*EX{}/$EX${}<20,20,70+30*EX{}/$EX${})),2),"")'.format(row, row, r, row, r)
                ws['FL{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EK{}="","",IF(70+30*EY{}/$EY${}<20,20,70+30*EY{}/$EY${})),2),"")'.format(row, row, r, row, r)
                ws['FM{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EL{}="","",IF(70+30*EZ{}/$EZ${}<20,20,70+30*EZ{}/$EZ${})),2),"")'.format(row, row, r, row, r)
                ws['FN{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EM{}="","",IF(70+30*FA{}/$FA${}<20,20,70+30*FA{}/$FA${})),2),"")'.format(row, row, r, row, r)

                ws['FO{}'.format(row)] = '=IF(SUM(FB{}:FN{})=0,"",SUM(FB{}:FN{}))'.format(
                    row, row, row, row)
                ws['FP{}'.format(row)] = '=IF(FO{}="","",RANK(FO{},$FO$2:$FO${}))'.format(
                    row, row, q)
                ws['FQ{}'.format(
                    row)] = '=IF(FP{}="","",COUNTIFS($DZ$2:$DZ${},DZ{},$FP$2:$FP${},"<"&FP{})+1)'.format(row, q, row, q, row)
            # TAMBAHAN
                ws['FR{}'.format(row)] = '=IF($G${}=20,IF(AND(G{}>3,EO{}=20),1,""),IF($G${}=25,IF(AND(G{}>4,EO{}=20),1,""),IF($G${}=30,IF(AND(G{}>5,EO{}=20),1,""),IF($G${}=35,IF(AND(G{}>6,EO{}=20),1,""),IF($G${}=40,IF(AND(G{}>7,EO{}=20),1,""),IF($G${}=45,IF(AND(G{}>8,EO{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['FS{}'.format(row)] = '=IF($H${}=20,IF(AND(H{}>3,EP{}=20),1,""),IF($H${}=25,IF(AND(H{}>4,EP{}=20),1,""),IF($H${}=30,IF(AND(H{}>5,EP{}=20),1,""),IF($H${}=35,IF(AND(H{}>6,EP{}=20),1,""),IF($H${}=40,IF(AND(H{}>7,EP{}=20),1,""),IF($H${}=45,IF(AND(H{}>8,EP{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['FT{}'.format(row)] = '=IF($I${}=20,IF(AND(I{}>3,EQ{}=20),1,""),IF($I${}=25,IF(AND(I{}>4,EQ{}=20),1,""),IF($I${}=30,IF(AND(I{}>5,EQ{}=20),1,""),IF($I${}=35,IF(AND(I{}>6,EQ{}=20),1,""),IF($I${}=40,IF(AND(I{}>7,EQ{}=20),1,""),IF($I${}=45,IF(AND(I{}>8,EQ{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['FU{}'.format(row)] = '=IF($J${}=20,IF(AND(J{}>3,ER{}=20),1,""),IF($J${}=25,IF(AND(J{}>4,ER{}=20),1,""),IF($J${}=30,IF(AND(J{}>5,ER{}=20),1,""),IF($J${}=35,IF(AND(J{}>6,ER{}=20),1,""),IF($J${}=40,IF(AND(J{}>7,ER{}=20),1,""),IF($J${}=45,IF(AND(J{}>8,ER{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['FV{}'.format(row)] = '=IF($K${}=20,IF(AND(K{}>3,ES{}=20),1,""),IF($K${}=25,IF(AND(K{}>4,ES{}=20),1,""),IF($K${}=30,IF(AND(K{}>5,ES{}=20),1,""),IF($K${}=35,IF(AND(K{}>6,ES{}=20),1,""),IF($K${}=40,IF(AND(K{}>7,ES{}=20),1,""),IF($K${}=45,IF(AND(K{}>8,ES{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['FW{}'.format(row)] = '=IF($L${}=20,IF(AND(L{}>3,ET{}=20),1,""),IF($L${}=25,IF(AND(L{}>4,ET{}=20),1,""),IF($L${}=30,IF(AND(L{}>5,ET{}=20),1,""),IF($L${}=35,IF(AND(L{}>6,ET{}=20),1,""),IF($L${}=40,IF(AND(L{}>7,ET{}=20),1,""),IF($L${}=45,IF(AND(L{}>8,ET{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['FX{}'.format(row)] = '=IF($M${}=20,IF(AND(M{}>3,EU{}=20),1,""),IF($M${}=25,IF(AND(M{}>4,EU{}=20),1,""),IF($M${}=30,IF(AND(M{}>5,EU{}=20),1,""),IF($M${}=35,IF(AND(M{}>6,EU{}=20),1,""),IF($M${}=40,IF(AND(M{}>7,EU{}=20),1,""),IF($M${}=45,IF(AND(M{}>8,EU{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['FY{}'.format(row)] = '=IF($N${}=20,IF(AND(N{}>3,EV{}=20),1,""),IF($N${}=25,IF(AND(N{}>4,EV{}=20),1,""),IF($N${}=30,IF(AND(N{}>5,EV{}=20),1,""),IF($N${}=35,IF(AND(N{}>6,EV{}=20),1,""),IF($N${}=40,IF(AND(N{}>7,EV{}=20),1,""),IF($N${}=45,IF(AND(N{}>8,EV{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['FZ{}'.format(row)] = '=IF($O${}=20,IF(AND(O{}>3,EW{}=20),1,""),IF($O${}=25,IF(AND(O{}>4,EW{}=20),1,""),IF($O${}=30,IF(AND(O{}>5,EW{}=20),1,""),IF($O${}=35,IF(AND(O{}>6,EW{}=20),1,""),IF($O${}=40,IF(AND(O{}>7,EW{}=20),1,""),IF($O${}=45,IF(AND(O{}>8,EW{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['GA{}'.format(row)] = '=IF($P${}=20,IF(AND(P{}>3,EX{}=20),1,""),IF($P${}=25,IF(AND(P{}>4,EX{}=20),1,""),IF($P${}=30,IF(AND(P{}>5,EX{}=20),1,""),IF($P${}=35,IF(AND(P{}>6,EX{}=20),1,""),IF($P${}=40,IF(AND(P{}>7,EX{}=20),1,""),IF($P${}=45,IF(AND(P{}>8,EX{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['GB{}'.format(row)] = '=IF($Q${}=20,IF(AND(Q{}>3,EY{}=20),1,""),IF($Q${}=25,IF(AND(Q{}>4,EY{}=20),1,""),IF($Q${}=30,IF(AND(Q{}>5,EY{}=20),1,""),IF($Q${}=35,IF(AND(Q{}>6,EY{}=20),1,""),IF($Q${}=40,IF(AND(Q{}>7,EY{}=20),1,""),IF($Q${}=45,IF(AND(Q{}>8,EY{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['GC{}'.format(row)] = '=IF($R${}=20,IF(AND(R{}>3,EZ{}=20),1,""),IF($R${}=25,IF(AND(R{}>4,EZ{}=20),1,""),IF($R${}=30,IF(AND(R{}>5,EZ{}=20),1,""),IF($R${}=35,IF(AND(R{}>6,EZ{}=20),1,""),IF($R${}=40,IF(AND(R{}>7,EZ{}=20),1,""),IF($R${}=45,IF(AND(R{}>8,EZ{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['GD{}'.format(row)] = '=IF($S${}=20,IF(AND(S{}>3,FA{}=20),1,""),IF($S${}=25,IF(AND(S{}>4,FA{}=20),1,""),IF($S${}=30,IF(AND(S{}>5,FA{}=20),1,""),IF($S${}=35,IF(AND(S{}>6,FA{}=20),1,""),IF($S${}=40,IF(AND(S{}>7,FA{}=20),1,""),IF($S${}=45,IF(AND(S{}>8,FA{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

            # Z Score [4]
            ws['GF1'] = 'NAMA SISWA_D'
            ws['GG1'] = 'NOMOR NF_D'
            ws['GH1'] = 'KELAS_D'
            ws['GI1'] = 'NAMA SEKOLAH_D'
            ws['GJ1'] = 'LOKASI_D'

            ws['GK1'] = 'MAT_1_D'
            ws['GL1'] = 'MAT_2_D'
            ws['GM1'] = 'IND_D'
            ws['GN1'] = 'ENG_D'
            ws['GO1'] = 'SEJ_D'
            ws['GP1'] = 'EKO_D'
            ws['GQ1'] = 'SOS_D'
            ws['GR1'] = 'GEO_D'
            ws['GS1'] = 'ANT_D'
            ws['GT1'] = 'BIO_D'
            ws['GU1'] = 'FIS_D'
            ws['GV1'] = 'KIM_1_D'
            ws['GW1'] = 'KIM_2_D'
            ws['GX1'] = 'JML_D'

            ws['GY1'] = 'Z_MAT_1_D'
            ws['GZ1'] = 'Z_MAT_2_D'
            ws['HA1'] = 'Z_IND_D'
            ws['HB1'] = 'Z_ENG_D'
            ws['HC1'] = 'Z_SEJ_D'
            ws['HD1'] = 'Z_EKO_D'
            ws['HE1'] = 'Z_SOS_D'
            ws['HF1'] = 'Z_GEO_D'
            ws['HG1'] = 'Z_ANT_D'
            ws['HH1'] = 'Z_BIO_D'
            ws['HI1'] = 'Z_FIS_D'
            ws['HJ1'] = 'Z_KIM_1_D'
            ws['HK1'] = 'Z_KIM_2_D'

            ws['HL1'] = 'S_MAT_1_D'
            ws['HM1'] = 'S_MAT_2_D'
            ws['HN1'] = 'S_IND_D'
            ws['HO1'] = 'S_ENG_D'
            ws['HP1'] = 'S_SEJ_D'
            ws['HQ1'] = 'S_EKO_D'
            ws['HR1'] = 'S_SOS_D'
            ws['HS1'] = 'S_GEO_D'
            ws['HT1'] = 'S_ANT_D'
            ws['HU1'] = 'S_BIO_D'
            ws['HV1'] = 'S_FIS_D'
            ws['HW1'] = 'S_KIM_1_D'
            ws['HX1'] = 'S_KIM_2_D'
            ws['HY1'] = 'S_JML_D'

            ws['HZ1'] = 'RANK NAS._D'
            ws['IA1'] = 'RANK LOK._D'

            # Z MAT 1
            ws['GY1'].font = Font(bold=False, name='Calibri', size=11)
            ws['GZ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HA1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HB1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HC1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HD1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HE1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HF1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HG1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HH1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HI1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HJ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HK1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HL1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HM1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HN1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HO1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HP1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HQ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HR1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HS1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HT1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HU1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HV1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HW1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HX1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HY1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HZ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['IA1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL DARI NAMA
            ws['GF1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['GG1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['GH1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['GI1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['GJ1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['GK1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['GL1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['GM1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['GN1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['GO1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['GP1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['GQ1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['GR1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['GS1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['GT1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['GU1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['GV1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['GW1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['GX1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['GY1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['GZ1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['HA1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['HB1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['HC1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['HD1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['HE1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['HF1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['HG1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['HH1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['HI1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['HJ1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['HK1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['HL1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['HM1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['HN1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['HO1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['HP1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['HQ1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['HR1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['HS1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['HT1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['HU1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['HV1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['HW1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['HX1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['HY1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['HZ1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['IA1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            
            # tambahan
            ws['IB1'] = 'MAT_1_20_D'
            ws['IC1'] = 'MAT_2_20_D'
            ws['ID1'] = 'IND_20_D'
            ws['IE1'] = 'ENG_20_D'
            ws['IF1'] = 'SEJ_20_D'
            ws['IG1'] = 'EKO_20_D'
            ws['IH1'] = 'SOS_20_D'
            ws['II1'] = 'GEO_20_D'
            ws['IJ1'] = 'ANT_20_D'
            ws['IK1'] = 'BIO_20_D'
            ws['IL1'] = 'FIS_20_D'
            ws['IM1'] = 'KIM_1_20_D'
            ws['IN1'] = 'KIM_2_20_D'

            ws['IB1'].font = Font(bold=False, name='Calibri', size=11)
            ws['IC1'].font = Font(bold=False, name='Calibri', size=11)
            ws['ID1'].font = Font(bold=False, name='Calibri', size=11)
            ws['IE1'].font = Font(bold=False, name='Calibri', size=11)
            ws['IF1'].font = Font(bold=False, name='Calibri', size=11)
            ws['IG1'].font = Font(bold=False, name='Calibri', size=11)
            ws['IH1'].font = Font(bold=False, name='Calibri', size=11)
            ws['II1'].font = Font(bold=False, name='Calibri', size=11)
            ws['IJ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['IK1'].font = Font(bold=False, name='Calibri', size=11)
            ws['IL1'].font = Font(bold=False, name='Calibri', size=11)
            ws['IM1'].font = Font(bold=False, name='Calibri', size=11)
            ws['IN1'].font = Font(bold=False, name='Calibri', size=11)

            ws['IB1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['IC1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['ID1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['IE1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['IF1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['IG1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['IH1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['II1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['IJ1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['IK1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['IL1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['IM1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['IN1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')

            # DARI NAMA
            for row in range(2, q+1):
                ws['GF{}'.format(row)] = '=B{}'.format(row)
                ws['GG{}'.format(row)] = '=C{}'.format(row, row)
                ws['GH{}'.format(row)] = '=D{}'.format(row, row)
                ws['GI{}'.format(row)] = '=E{}'.format(row, row)
                ws['GJ{}'.format(row)] = '=F{}'.format(row, row)
                ws['GK{}'.format(row)] = '=IF(G{}="","",G{})'.format(row, row)
                ws['GL{}'.format(row)] = '=IF(H{}="","",H{})'.format(row, row)
                ws['GM{}'.format(row)] = '=IF(I{}="","",I{})'.format(row, row)
                ws['GN{}'.format(row)] = '=IF(J{}="","",J{})'.format(row, row)
                ws['GO{}'.format(row)] = '=IF(K{}="","",K{})'.format(row, row)
                ws['GP{}'.format(row)] = '=IF(L{}="","",L{})'.format(row, row)
                ws['GQ{}'.format(row)] = '=IF(M{}="","",M{})'.format(row, row)
                ws['GR{}'.format(row)] = '=IF(N{}="","",N{})'.format(row, row)
                ws['GS{}'.format(row)] = '=IF(O{}="","",O{})'.format(row, row)
                ws['GT{}'.format(row)] = '=IF(P{}="","",P{})'.format(row, row)
                ws['GU{}'.format(row)] = '=IF(Q{}="","",Q{})'.format(row, row)
                ws['GV{}'.format(row)] = '=IF(R{}="","",R{})'.format(row, row)
                ws['GW{}'.format(row)] = '=IF(S{}="","",S{})'.format(row, row)
                ws['GX{}'.format(row)] = '=IF(T{}="","",T{})'.format(row, row)
                
                ws['GY{}'.format(
                    row)] = '=IFERROR(ROUND(IF(GK{}="","",(GK{}-GK${})/GK${}),2),"")'.format(row, row, r, s)
                ws['GZ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(GL{}="","",(GL{}-GL${})/GL${}),2),"")'.format(row, row, r, s)
                ws['HA{}'.format(
                    row)] = '=IFERROR(ROUND(IF(GM{}="","",(GM{}-GM${})/GM${}),2),"")'.format(row, row, r, s)
                ws['HB{}'.format(
                    row)] = '=IFERROR(ROUND(IF(GN{}="","",(GN{}-GN${})/GN${}),2),"")'.format(row, row, r, s)
                ws['HC{}'.format(
                    row)] = '=IFERROR(ROUND(IF(GO{}="","",(GO{}-GO${})/GO${}),2),"")'.format(row, row, r, s)
                ws['HD{}'.format(
                    row)] = '=IFERROR(ROUND(IF(GP{}="","",(GP{}-GP${})/GP${}),2),"")'.format(row, row, r, s)
                ws['HE{}'.format(
                    row)] = '=IFERROR(ROUND(IF(GQ{}="","",(GQ{}-GQ${})/GQ${}),2),"")'.format(row, row, r, s)
                ws['HF{}'.format(
                    row)] = '=IFERROR(ROUND(IF(GR{}="","",(GR{}-GR${})/GR${}),2),"")'.format(row, row, r, s)
                ws['HG{}'.format(
                    row)] = '=IFERROR(ROUND(IF(GS{}="","",(GS{}-GS${})/GS${}),2),"")'.format(row, row, r, s)
                ws['HH{}'.format(
                    row)] = '=IFERROR(ROUND(IF(GT{}="","",(GT{}-GT${})/GT${}),2),"")'.format(row, row, r, s)
                ws['HI{}'.format(
                    row)] = '=IFERROR(ROUND(IF(GU{}="","",(GU{}-GU${})/GU${}),2),"")'.format(row, row, r, s)
                ws['HJ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(GV{}="","",(GV{}-GV${})/GV${}),2),"")'.format(row, row, r, s)
                ws['HK{}'.format(
                    row)] = '=IFERROR(ROUND(IF(GW{}="","",(GW{}-GW${})/GW${}),2),"")'.format(row, row, r, s)

                ws['HL{}'.format(
                    row)] = '=IFERROR(ROUND(IF(GK{}="","",IF(70+30*GY{}/$GY${}<20,20,70+30*GY{}/$GY${})),2),"")'.format(row, row, r, row, r)
                ws['HM{}'.format(
                    row)] = '=IFERROR(ROUND(IF(GL{}="","",IF(70+30*GZ{}/$GZ${}<20,20,70+30*GZ{}/$GZ${})),2),"")'.format(row, row, r, row, r)
                ws['HN{}'.format(
                    row)] = '=IFERROR(ROUND(IF(GM{}="","",IF(70+30*HA{}/$HA${}<20,20,70+30*HA{}/$HA${})),2),"")'.format(row, row, r, row, r)
                ws['HO{}'.format(
                    row)] = '=IFERROR(ROUND(IF(GN{}="","",IF(70+30*HB{}/$HB${}<20,20,70+30*HB{}/$HB${})),2),"")'.format(row, row, r, row, r)
                ws['HP{}'.format(
                    row)] = '=IFERROR(ROUND(IF(GO{}="","",IF(70+30*HC{}/$HC${}<20,20,70+30*HC{}/$HC${})),2),"")'.format(row, row, r, row, r)
                ws['HQ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(GP{}="","",IF(70+30*HD{}/$HD${}<20,20,70+30*HD{}/$HD${})),2),"")'.format(row, row, r, row, r)
                ws['HR{}'.format(
                    row)] = '=IFERROR(ROUND(IF(GQ{}="","",IF(70+30*HE{}/$HE${}<20,20,70+30*HE{}/$HE${})),2),"")'.format(row, row, r, row, r)
                ws['HS{}'.format(
                    row)] = '=IFERROR(ROUND(IF(GR{}="","",IF(70+30*HF{}/$HF${}<20,20,70+30*HF{}/$HF${})),2),"")'.format(row, row, r, row, r)
                ws['HT{}'.format(
                    row)] = '=IFERROR(ROUND(IF(GS{}="","",IF(70+30*HG{}/$HG${}<20,20,70+30*HG{}/$HG${})),2),"")'.format(row, row, r, row, r)
                ws['HU{}'.format(
                    row)] = '=IFERROR(ROUND(IF(GT{}="","",IF(70+30*HH{}/$HH${}<20,20,70+30*HH{}/$HH${})),2),"")'.format(row, row, r, row, r)
                ws['HV{}'.format(
                    row)] = '=IFERROR(ROUND(IF(GU{}="","",IF(70+30*HI{}/$HI${}<20,20,70+30*HI{}/$HI${})),2),"")'.format(row, row, r, row, r)
                ws['HW{}'.format(
                    row)] = '=IFERROR(ROUND(IF(GV{}="","",IF(70+30*HJ{}/$HJ${}<20,20,70+30*HJ{}/$HJ${})),2),"")'.format(row, row, r, row, r)
                ws['HX{}'.format(
                    row)] = '=IFERROR(ROUND(IF(GW{}="","",IF(70+30*HK{}/$HK${}<20,20,70+30*HK{}/$HK${})),2),"")'.format(row, row, r, row, r)

                ws['HY{}'.format(row)] = '=IF(SUM(HL{}:HX{})=0,"",SUM(HL{}:HX{}))'.format(
                    row, row, row, row)
                ws['HZ{}'.format(row)] = '=IF(HY{}="","",RANK(HY{},$HY$2:$HY${}))'.format(
                    row, row, q)
                ws['IA{}'.format(
                    row)] = '=IF(HZ{}="","",COUNTIFS($GJ$2:$GJ${},GJ{},$HZ$2:$HZ${},"<"&HZ{})+1)'.format(row, q, row, q, row)
            # TAMBAHAN
                ws['IB{}'.format(row)] = '=IF($G${}=20,IF(AND(G{}>3,GY{}=20),1,""),IF($G${}=25,IF(AND(G{}>4,GY{}=20),1,""),IF($G${}=30,IF(AND(G{}>5,GY{}=20),1,""),IF($G${}=35,IF(AND(G{}>6,GY{}=20),1,""),IF($G${}=40,IF(AND(G{}>7,GY{}=20),1,""),IF($G${}=45,IF(AND(G{}>8,GY{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['IC{}'.format(row)] = '=IF($H${}=20,IF(AND(H{}>3,GZ{}=20),1,""),IF($H${}=25,IF(AND(H{}>4,GZ{}=20),1,""),IF($H${}=30,IF(AND(H{}>5,GZ{}=20),1,""),IF($H${}=35,IF(AND(H{}>6,GZ{}=20),1,""),IF($H${}=40,IF(AND(H{}>7,GZ{}=20),1,""),IF($H${}=45,IF(AND(H{}>8,GZ{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['ID{}'.format(row)] = '=IF($I${}=20,IF(AND(I{}>3,HA{}=20),1,""),IF($I${}=25,IF(AND(I{}>4,HA{}=20),1,""),IF($I${}=30,IF(AND(I{}>5,HA{}=20),1,""),IF($I${}=35,IF(AND(I{}>6,HA{}=20),1,""),IF($I${}=40,IF(AND(I{}>7,HA{}=20),1,""),IF($I${}=45,IF(AND(I{}>8,HA{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['IE{}'.format(row)] = '=IF($J${}=20,IF(AND(J{}>3,HB{}=20),1,""),IF($J${}=25,IF(AND(J{}>4,HB{}=20),1,""),IF($J${}=30,IF(AND(J{}>5,HB{}=20),1,""),IF($J${}=35,IF(AND(J{}>6,HB{}=20),1,""),IF($J${}=40,IF(AND(J{}>7,HB{}=20),1,""),IF($J${}=45,IF(AND(J{}>8,HB{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['IF{}'.format(row)] = '=IF($K${}=20,IF(AND(K{}>3,HC{}=20),1,""),IF($K${}=25,IF(AND(K{}>4,HC{}=20),1,""),IF($K${}=30,IF(AND(K{}>5,HC{}=20),1,""),IF($K${}=35,IF(AND(K{}>6,HC{}=20),1,""),IF($K${}=40,IF(AND(K{}>7,HC{}=20),1,""),IF($K${}=45,IF(AND(K{}>8,HC{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['IG{}'.format(row)] = '=IF($L${}=20,IF(AND(L{}>3,HD{}=20),1,""),IF($L${}=25,IF(AND(L{}>4,HD{}=20),1,""),IF($L${}=30,IF(AND(L{}>5,HD{}=20),1,""),IF($L${}=35,IF(AND(L{}>6,HD{}=20),1,""),IF($L${}=40,IF(AND(L{}>7,HD{}=20),1,""),IF($L${}=45,IF(AND(L{}>8,HD{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['IH{}'.format(row)] = '=IF($M${}=20,IF(AND(M{}>3,HE{}=20),1,""),IF($M${}=25,IF(AND(M{}>4,HE{}=20),1,""),IF($M${}=30,IF(AND(M{}>5,HE{}=20),1,""),IF($M${}=35,IF(AND(M{}>6,HE{}=20),1,""),IF($M${}=40,IF(AND(M{}>7,HE{}=20),1,""),IF($M${}=45,IF(AND(M{}>8,HE{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['II{}'.format(row)] = '=IF($N${}=20,IF(AND(N{}>3,HF{}=20),1,""),IF($N${}=25,IF(AND(N{}>4,HF{}=20),1,""),IF($N${}=30,IF(AND(N{}>5,HF{}=20),1,""),IF($N${}=35,IF(AND(N{}>6,HF{}=20),1,""),IF($N${}=40,IF(AND(N{}>7,HF{}=20),1,""),IF($N${}=45,IF(AND(N{}>8,HF{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['IJ{}'.format(row)] = '=IF($O${}=20,IF(AND(O{}>3,HG{}=20),1,""),IF($O${}=25,IF(AND(O{}>4,HG{}=20),1,""),IF($O${}=30,IF(AND(O{}>5,HG{}=20),1,""),IF($O${}=35,IF(AND(O{}>6,HG{}=20),1,""),IF($O${}=40,IF(AND(O{}>7,HG{}=20),1,""),IF($O${}=45,IF(AND(O{}>8,HG{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['IK{}'.format(row)] = '=IF($P${}=20,IF(AND(P{}>3,HH{}=20),1,""),IF($P${}=25,IF(AND(P{}>4,HH{}=20),1,""),IF($P${}=30,IF(AND(P{}>5,HH{}=20),1,""),IF($P${}=35,IF(AND(P{}>6,HH{}=20),1,""),IF($P${}=40,IF(AND(P{}>7,HH{}=20),1,""),IF($P${}=45,IF(AND(P{}>8,HH{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['IL{}'.format(row)] = '=IF($Q${}=20,IF(AND(Q{}>3,HI{}=20),1,""),IF($Q${}=25,IF(AND(Q{}>4,HI{}=20),1,""),IF($Q${}=30,IF(AND(Q{}>5,HI{}=20),1,""),IF($Q${}=35,IF(AND(Q{}>6,HI{}=20),1,""),IF($Q${}=40,IF(AND(Q{}>7,HI{}=20),1,""),IF($Q${}=45,IF(AND(Q{}>8,HI{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['IM{}'.format(row)] = '=IF($R${}=20,IF(AND(R{}>3,HJ{}=20),1,""),IF($R${}=25,IF(AND(R{}>4,HJ{}=20),1,""),IF($R${}=30,IF(AND(R{}>5,HJ{}=20),1,""),IF($R${}=35,IF(AND(R{}>6,HJ{}=20),1,""),IF($R${}=40,IF(AND(R{}>7,HJ{}=20),1,""),IF($R${}=45,IF(AND(R{}>8,HJ{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['IN{}'.format(row)] = '=IF($S${}=20,IF(AND(S{}>3,HK{}=20),1,""),IF($S${}=25,IF(AND(S{}>4,HK{}=20),1,""),IF($S${}=30,IF(AND(S{}>5,HK{}=20),1,""),IF($S${}=35,IF(AND(S{}>6,HK{}=20),1,""),IF($S${}=40,IF(AND(S{}>7,HK{}=20),1,""),IF($S${}=45,IF(AND(S{}>8,HK{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
            
            # Z Score [5]
            ws['IP1'] = 'NAMA SISWA_E'
            ws['IQ1'] = 'NOMOR NF_E'
            ws['IR1'] = 'KELAS_E'
            ws['IS1'] = 'NAMA SEKOLAH_E'
            ws['IT1'] = 'LOKASI_E'

            ws['IU1'] = 'MAT_1_E'
            ws['IV1'] = 'MAT_2_E'
            ws['IW1'] = 'IND_E'
            ws['IX1'] = 'ENG_E'
            ws['IY1'] = 'SEJ_E'
            ws['IZ1'] = 'EKO_E'
            ws['JA1'] = 'SOS_E'
            ws['JB1'] = 'GEO_E'
            ws['JC1'] = 'ANT_E'
            ws['JD1'] = 'BIO_E'
            ws['JE1'] = 'FIS_E'
            ws['JF1'] = 'KIM_1_E'
            ws['JG1'] = 'KIM_2_E'
            ws['JH1'] = 'JML_E'

            ws['JI1'] = 'Z_MAT_1_E'
            ws['JJ1'] = 'Z_MAT_2_E'
            ws['JK1'] = 'Z_IND_E'
            ws['JL1'] = 'Z_ENG_E'
            ws['JM1'] = 'Z_SEJ_E'
            ws['JN1'] = 'Z_EKO_E'
            ws['JO1'] = 'Z_SOS_E'
            ws['JP1'] = 'Z_GEO_E'
            ws['JQ1'] = 'Z_ANT_E'
            ws['JR1'] = 'Z_BIO_E'
            ws['JS1'] = 'Z_FIS_E'
            ws['JT1'] = 'Z_KIM_1_E'
            ws['JU1'] = 'Z_KIM_2_E'

            ws['JV1'] = 'S_MAT_1_E'
            ws['JW1'] = 'S_MAT_2_E'
            ws['JX1'] = 'S_IND_E'
            ws['JY1'] = 'S_ENG_E'
            ws['JZ1'] = 'S_SEJ_E'
            ws['KA1'] = 'S_EKO_E'
            ws['KB1'] = 'S_SOS_E'
            ws['KC1'] = 'S_GEO_E'
            ws['KD1'] = 'S_ANT_E'
            ws['KE1'] = 'S_BIO_E'
            ws['KF1'] = 'S_FIS_E'
            ws['KG1'] = 'S_KIM_1_E'
            ws['KH1'] = 'S_KIM_2_E'
            ws['KI1'] = 'S_JML_E'

            ws['KJ1'] = 'RANK NAS._E'
            ws['KK1'] = 'RANK LOK._E'

            # Z MAT 1
            ws['JI1'].font = Font(bold=False, name='Calibri', size=11)
            ws['JJ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['JK1'].font = Font(bold=False, name='Calibri', size=11)
            ws['JL1'].font = Font(bold=False, name='Calibri', size=11)
            ws['JM1'].font = Font(bold=False, name='Calibri', size=11)
            ws['JN1'].font = Font(bold=False, name='Calibri', size=11)
            ws['JO1'].font = Font(bold=False, name='Calibri', size=11)
            ws['JP1'].font = Font(bold=False, name='Calibri', size=11)
            ws['JQ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['JR1'].font = Font(bold=False, name='Calibri', size=11)
            ws['JS1'].font = Font(bold=False, name='Calibri', size=11)
            ws['JT1'].font = Font(bold=False, name='Calibri', size=11)
            ws['JU1'].font = Font(bold=False, name='Calibri', size=11)
            ws['JV1'].font = Font(bold=False, name='Calibri', size=11)
            ws['JW1'].font = Font(bold=False, name='Calibri', size=11)
            ws['JX1'].font = Font(bold=False, name='Calibri', size=11)
            ws['JY1'].font = Font(bold=False, name='Calibri', size=11)
            ws['JZ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['KA1'].font = Font(bold=False, name='Calibri', size=11)
            ws['KB1'].font = Font(bold=False, name='Calibri', size=11)
            ws['KC1'].font = Font(bold=False, name='Calibri', size=11)
            ws['KD1'].font = Font(bold=False, name='Calibri', size=11)
            ws['KE1'].font = Font(bold=False, name='Calibri', size=11)
            ws['KF1'].font = Font(bold=False, name='Calibri', size=11)
            ws['KG1'].font = Font(bold=False, name='Calibri', size=11)
            ws['KH1'].font = Font(bold=False, name='Calibri', size=11)
            ws['KI1'].font = Font(bold=False, name='Calibri', size=11)
            ws['KJ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['KK1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL DARI NAMA
            ws['IP1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['IQ1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['IR1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['IS1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['IT1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['IU1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['IV1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['IW1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['IX1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['IY1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['IZ1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['JA1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['JB1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['JC1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['JD1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['JE1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['JF1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['JG1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['JH1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['JI1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['JJ1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['JK1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['JL1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['JM1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['JN1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['JO1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['JP1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['JQ1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['JR1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['JS1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['JT1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['JU1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['JV1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['JW1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['JX1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['JY1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['JZ1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['KA1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['KB1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['KC1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['KD1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['KE1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['KF1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['KG1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['KH1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['KI1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['KJ1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['KK1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            
            # tambahan
            ws['KL1'] = 'MAT_1_20_E'
            ws['KM1'] = 'MAT_2_20_E'
            ws['KN1'] = 'IND_20_E'
            ws['KO1'] = 'ENG_20_E'
            ws['KP1'] = 'SEJ_20_E'
            ws['KQ1'] = 'EKO_20_E'
            ws['KR1'] = 'SOS_20_E'
            ws['KS1'] = 'GEO_20_E'
            ws['KT1'] = 'ANT_20_E'
            ws['KU1'] = 'BIO_20_E'
            ws['KV1'] = 'FIS_20_E'
            ws['KW1'] = 'KIM_1_20_E'
            ws['KX1'] = 'KIM_2_20_E'

            ws['KL1'].font = Font(bold=False, name='Calibri', size=11)
            ws['KM1'].font = Font(bold=False, name='Calibri', size=11)
            ws['KN1'].font = Font(bold=False, name='Calibri', size=11)
            ws['KO1'].font = Font(bold=False, name='Calibri', size=11)
            ws['KP1'].font = Font(bold=False, name='Calibri', size=11)
            ws['KQ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['KR1'].font = Font(bold=False, name='Calibri', size=11)
            ws['KS1'].font = Font(bold=False, name='Calibri', size=11)
            ws['KT1'].font = Font(bold=False, name='Calibri', size=11)
            ws['KU1'].font = Font(bold=False, name='Calibri', size=11)
            ws['KV1'].font = Font(bold=False, name='Calibri', size=11)
            ws['KW1'].font = Font(bold=False, name='Calibri', size=11)
            ws['KX1'].font = Font(bold=False, name='Calibri', size=11)

            ws['KL1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['KM1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['KN1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['KO1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['KP1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['KQ1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['KR1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['KS1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['KT1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['KU1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['KV1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['KW1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['KX1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')

            # DARI NAMA
            for row in range(2, q+1):
                ws['IP{}'.format(row)] = '=B{}'.format(row)
                ws['IQ{}'.format(row)] = '=C{}'.format(row, row)
                ws['IR{}'.format(row)] = '=D{}'.format(row, row)
                ws['IS{}'.format(row)] = '=E{}'.format(row, row)
                ws['IT{}'.format(row)] = '=F{}'.format(row, row)
                ws['IU{}'.format(row)] = '=IF(G{}="","",G{})'.format(row, row)
                ws['IV{}'.format(row)] = '=IF(H{}="","",H{})'.format(row, row)
                ws['IW{}'.format(row)] = '=IF(I{}="","",I{})'.format(row, row)
                ws['IX{}'.format(row)] = '=IF(J{}="","",J{})'.format(row, row)
                ws['IY{}'.format(row)] = '=IF(K{}="","",K{})'.format(row, row)
                ws['IZ{}'.format(row)] = '=IF(L{}="","",L{})'.format(row, row)
                ws['JA{}'.format(row)] = '=IF(M{}="","",M{})'.format(row, row)
                ws['JB{}'.format(row)] = '=IF(N{}="","",N{})'.format(row, row)
                ws['JC{}'.format(row)] = '=IF(O{}="","",O{})'.format(row, row)
                ws['JD{}'.format(row)] = '=IF(P{}="","",P{})'.format(row, row)
                ws['JE{}'.format(row)] = '=IF(Q{}="","",Q{})'.format(row, row)
                ws['JF{}'.format(row)] = '=IF(R{}="","",R{})'.format(row, row)
                ws['JG{}'.format(row)] = '=IF(S{}="","",S{})'.format(row, row)
                ws['JH{}'.format(row)] = '=IF(T{}="","",T{})'.format(row, row)
                
                ws['JI{}'.format(
                    row)] = '=IFERROR(ROUND(IF(IU{}="","",(IU{}-IU${})/IU${}),2),"")'.format(row, row, r, s)
                ws['JJ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(IV{}="","",(IV{}-IV${})/IV${}),2),"")'.format(row, row, r, s)
                ws['JK{}'.format(
                    row)] = '=IFERROR(ROUND(IF(IW{}="","",(IW{}-IW${})/IW${}),2),"")'.format(row, row, r, s)
                ws['JL{}'.format(
                    row)] = '=IFERROR(ROUND(IF(IX{}="","",(IX{}-IX${})/IX${}),2),"")'.format(row, row, r, s)
                ws['JM{}'.format(
                    row)] = '=IFERROR(ROUND(IF(IY{}="","",(IY{}-IY${})/IY${}),2),"")'.format(row, row, r, s)
                ws['JN{}'.format(
                    row)] = '=IFERROR(ROUND(IF(IZ{}="","",(IZ{}-IZ${})/IZ${}),2),"")'.format(row, row, r, s)
                ws['JO{}'.format(
                    row)] = '=IFERROR(ROUND(IF(JA{}="","",(JA{}-JA${})/JA${}),2),"")'.format(row, row, r, s)
                ws['JP{}'.format(
                    row)] = '=IFERROR(ROUND(IF(JB{}="","",(JB{}-JB${})/JB${}),2),"")'.format(row, row, r, s)
                ws['JQ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(JC{}="","",(JC{}-JC${})/JC${}),2),"")'.format(row, row, r, s)
                ws['JR{}'.format(
                    row)] = '=IFERROR(ROUND(IF(JD{}="","",(JD{}-JD${})/JD${}),2),"")'.format(row, row, r, s)
                ws['JS{}'.format(
                    row)] = '=IFERROR(ROUND(IF(JE{}="","",(JE{}-JE${})/JE${}),2),"")'.format(row, row, r, s)
                ws['JT{}'.format(
                    row)] = '=IFERROR(ROUND(IF(JF{}="","",(JF{}-JF${})/JF${}),2),"")'.format(row, row, r, s)
                ws['JU{}'.format(
                    row)] = '=IFERROR(ROUND(IF(JG{}="","",(JG{}-JG${})/JG${}),2),"")'.format(row, row, r, s)

                ws['JV{}'.format(
                    row)] = '=IFERROR(ROUND(IF(IU{}="","",IF(70+30*JI{}/$JI${}<20,20,70+30*JI{}/$JI${})),2),"")'.format(row, row, r, row, r)
                ws['JW{}'.format(
                    row)] = '=IFERROR(ROUND(IF(IV{}="","",IF(70+30*JJ{}/$JJ${}<20,20,70+30*JJ{}/$JJ${})),2),"")'.format(row, row, r, row, r)
                ws['JX{}'.format(
                    row)] = '=IFERROR(ROUND(IF(IW{}="","",IF(70+30*JK{}/$JK${}<20,20,70+30*JK{}/$JK${})),2),"")'.format(row, row, r, row, r)
                ws['JY{}'.format(
                    row)] = '=IFERROR(ROUND(IF(IX{}="","",IF(70+30*JL{}/$JL${}<20,20,70+30*JL{}/$JL${})),2),"")'.format(row, row, r, row, r)
                ws['JZ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(IY{}="","",IF(70+30*JM{}/$JM${}<20,20,70+30*JM{}/$JM${})),2),"")'.format(row, row, r, row, r)
                ws['KA{}'.format(
                    row)] = '=IFERROR(ROUND(IF(IZ{}="","",IF(70+30*JN{}/$JN${}<20,20,70+30*JN{}/$JN${})),2),"")'.format(row, row, r, row, r)
                ws['KB{}'.format(
                    row)] = '=IFERROR(ROUND(IF(JA{}="","",IF(70+30*JO{}/$JO${}<20,20,70+30*JO{}/$JO${})),2),"")'.format(row, row, r, row, r)
                ws['KC{}'.format(
                    row)] = '=IFERROR(ROUND(IF(JB{}="","",IF(70+30*JP{}/$JP${}<20,20,70+30*JP{}/$JP${})),2),"")'.format(row, row, r, row, r)
                ws['KD{}'.format(
                    row)] = '=IFERROR(ROUND(IF(JC{}="","",IF(70+30*JQ{}/$JQ${}<20,20,70+30*JQ{}/$JQ${})),2),"")'.format(row, row, r, row, r)
                ws['KE{}'.format(
                    row)] = '=IFERROR(ROUND(IF(JD{}="","",IF(70+30*JR{}/$JR${}<20,20,70+30*JR{}/$JR${})),2),"")'.format(row, row, r, row, r)
                ws['KF{}'.format(
                    row)] = '=IFERROR(ROUND(IF(JE{}="","",IF(70+30*JS{}/$JS${}<20,20,70+30*JS{}/$JS${})),2),"")'.format(row, row, r, row, r)
                ws['KG{}'.format(
                    row)] = '=IFERROR(ROUND(IF(JF{}="","",IF(70+30*JT{}/$JT${}<20,20,70+30*JT{}/$JT${})),2),"")'.format(row, row, r, row, r)
                ws['KH{}'.format(
                    row)] = '=IFERROR(ROUND(IF(JG{}="","",IF(70+30*JU{}/$JU${}<20,20,70+30*JU{}/$JU${})),2),"")'.format(row, row, r, row, r)

                ws['KI{}'.format(row)] = '=IF(SUM(JV{}:KH{})=0,"",SUM(JV{}:KH{}))'.format(
                    row, row, row, row)
                ws['KJ{}'.format(row)] = '=IF(KI{}="","",RANK(KI{},$KI$2:$KI${}))'.format(
                    row, row, q)
                ws['KK{}'.format(
                    row)] = '=IF(KJ{}="","",COUNTIFS($IT$2:$IT${},IT{},$KJ$2:$KJ${},"<"&KJ{})+1)'.format(row, q, row, q, row)
            # TAMBAHAN
                ws['KL{}'.format(row)] = '=IF($G${}=20,IF(AND(G{}>3,JI{}=20),1,""),IF($G${}=25,IF(AND(G{}>4,JI{}=20),1,""),IF($G${}=30,IF(AND(G{}>5,JI{}=20),1,""),IF($G${}=35,IF(AND(G{}>6,JI{}=20),1,""),IF($G${}=40,IF(AND(G{}>7,JI{}=20),1,""),IF($G${}=45,IF(AND(G{}>8,JI{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['KM{}'.format(row)] = '=IF($H${}=20,IF(AND(H{}>3,JJ{}=20),1,""),IF($H${}=25,IF(AND(H{}>4,JJ{}=20),1,""),IF($H${}=30,IF(AND(H{}>5,JJ{}=20),1,""),IF($H${}=35,IF(AND(H{}>6,JJ{}=20),1,""),IF($H${}=40,IF(AND(H{}>7,JJ{}=20),1,""),IF($H${}=45,IF(AND(H{}>8,JJ{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['KN{}'.format(row)] = '=IF($I${}=20,IF(AND(I{}>3,JK{}=20),1,""),IF($I${}=25,IF(AND(I{}>4,JK{}=20),1,""),IF($I${}=30,IF(AND(I{}>5,JK{}=20),1,""),IF($I${}=35,IF(AND(I{}>6,JK{}=20),1,""),IF($I${}=40,IF(AND(I{}>7,JK{}=20),1,""),IF($I${}=45,IF(AND(I{}>8,JK{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['KO{}'.format(row)] = '=IF($J${}=20,IF(AND(J{}>3,JL{}=20),1,""),IF($J${}=25,IF(AND(J{}>4,JL{}=20),1,""),IF($J${}=30,IF(AND(J{}>5,JL{}=20),1,""),IF($J${}=35,IF(AND(J{}>6,JL{}=20),1,""),IF($J${}=40,IF(AND(J{}>7,JL{}=20),1,""),IF($J${}=45,IF(AND(J{}>8,JL{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['KP{}'.format(row)] = '=IF($K${}=20,IF(AND(K{}>3,JM{}=20),1,""),IF($K${}=25,IF(AND(K{}>4,JM{}=20),1,""),IF($K${}=30,IF(AND(K{}>5,JM{}=20),1,""),IF($K${}=35,IF(AND(K{}>6,JM{}=20),1,""),IF($K${}=40,IF(AND(K{}>7,JM{}=20),1,""),IF($K${}=45,IF(AND(K{}>8,JM{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['KQ{}'.format(row)] = '=IF($L${}=20,IF(AND(L{}>3,JN{}=20),1,""),IF($L${}=25,IF(AND(L{}>4,JN{}=20),1,""),IF($L${}=30,IF(AND(L{}>5,JN{}=20),1,""),IF($L${}=35,IF(AND(L{}>6,JN{}=20),1,""),IF($L${}=40,IF(AND(L{}>7,JN{}=20),1,""),IF($L${}=45,IF(AND(L{}>8,JN{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['KR{}'.format(row)] = '=IF($M${}=20,IF(AND(M{}>3,JO{}=20),1,""),IF($M${}=25,IF(AND(M{}>4,JO{}=20),1,""),IF($M${}=30,IF(AND(M{}>5,JO{}=20),1,""),IF($M${}=35,IF(AND(M{}>6,JO{}=20),1,""),IF($M${}=40,IF(AND(M{}>7,JO{}=20),1,""),IF($M${}=45,IF(AND(M{}>8,JO{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['KS{}'.format(row)] = '=IF($N${}=20,IF(AND(N{}>3,JP{}=20),1,""),IF($N${}=25,IF(AND(N{}>4,JP{}=20),1,""),IF($N${}=30,IF(AND(N{}>5,JP{}=20),1,""),IF($N${}=35,IF(AND(N{}>6,JP{}=20),1,""),IF($N${}=40,IF(AND(N{}>7,JP{}=20),1,""),IF($N${}=45,IF(AND(N{}>8,JP{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['KT{}'.format(row)] = '=IF($O${}=20,IF(AND(O{}>3,JQ{}=20),1,""),IF($O${}=25,IF(AND(O{}>4,JQ{}=20),1,""),IF($O${}=30,IF(AND(O{}>5,JQ{}=20),1,""),IF($O${}=35,IF(AND(O{}>6,JQ{}=20),1,""),IF($O${}=40,IF(AND(O{}>7,JQ{}=20),1,""),IF($O${}=45,IF(AND(O{}>8,JQ{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['KU{}'.format(row)] = '=IF($P${}=20,IF(AND(P{}>3,JR{}=20),1,""),IF($P${}=25,IF(AND(P{}>4,JR{}=20),1,""),IF($P${}=30,IF(AND(P{}>5,JR{}=20),1,""),IF($P${}=35,IF(AND(P{}>6,JR{}=20),1,""),IF($P${}=40,IF(AND(P{}>7,JR{}=20),1,""),IF($P${}=45,IF(AND(P{}>8,JR{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['KV{}'.format(row)] = '=IF($Q${}=20,IF(AND(Q{}>3,JS{}=20),1,""),IF($Q${}=25,IF(AND(Q{}>4,JS{}=20),1,""),IF($Q${}=30,IF(AND(Q{}>5,JS{}=20),1,""),IF($Q${}=35,IF(AND(Q{}>6,JS{}=20),1,""),IF($Q${}=40,IF(AND(Q{}>7,JS{}=20),1,""),IF($Q${}=45,IF(AND(Q{}>8,JS{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['KW{}'.format(row)] = '=IF($R${}=20,IF(AND(R{}>3,JT{}=20),1,""),IF($R${}=25,IF(AND(R{}>4,JT{}=20),1,""),IF($R${}=30,IF(AND(R{}>5,JT{}=20),1,""),IF($R${}=35,IF(AND(R{}>6,JT{}=20),1,""),IF($R${}=40,IF(AND(R{}>7,JT{}=20),1,""),IF($R${}=45,IF(AND(R{}>8,JT{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['KX{}'.format(row)] = '=IF($S${}=20,IF(AND(S{}>3,JU{}=20),1,""),IF($S${}=25,IF(AND(S{}>4,JU{}=20),1,""),IF($S${}=30,IF(AND(S{}>5,JU{}=20),1,""),IF($S${}=35,IF(AND(S{}>6,JU{}=20),1,""),IF($S${}=40,IF(AND(S{}>7,JU{}=20),1,""),IF($S${}=45,IF(AND(S{}>8,JU{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

            # Z Score
            ws['KZ1'] = 'NAMA SISWA'
            ws['LA1'] = 'NOMOR NF'
            ws['LB1'] = 'KELAS'
            ws['LC1'] = 'NAMA SEKOLAH'
            ws['LD1'] = 'LOKASI'

            ws['LE1'] = 'MAT_1'
            ws['LF1'] = 'MAT_2'
            ws['LG1'] = 'IND'
            ws['LH1'] = 'ENG'
            ws['LI1'] = 'SEJ'
            ws['LJ1'] = 'EKO'
            ws['LK1'] = 'SOS'
            ws['LL1'] = 'GEO'
            ws['LM1'] = 'ANT'
            ws['LN1'] = 'BIO'
            ws['LO1'] = 'FIS'
            ws['LP1'] = 'KIM_1'
            ws['LQ1'] = 'KIM_2'
            ws['LR1'] = 'JML'

            ws['LS1'] = 'Z_MAT_1'
            ws['LT1'] = 'Z_MAT_2'
            ws['LU1'] = 'Z_IND'
            ws['LV1'] = 'Z_ENG'
            ws['LW1'] = 'Z_SEJ'
            ws['LX1'] = 'Z_EKO'
            ws['LY1'] = 'Z_SOS'
            ws['LZ1'] = 'Z_GEO'
            ws['MA1'] = 'Z_ANT'
            ws['MB1'] = 'Z_BIO'
            ws['MC1'] = 'Z_FIS'
            ws['MD1'] = 'Z_KIM_1'
            ws['ME1'] = 'Z_KIM_2'

            ws['MF1'] = 'S_MAT_1'
            ws['MG1'] = 'S_MAT_2'
            ws['MH1'] = 'S_IND'
            ws['MI1'] = 'S_ENG'
            ws['MJ1'] = 'S_SEJ'
            ws['MK1'] = 'S_EKO'
            ws['ML1'] = 'S_SOS'
            ws['MM1'] = 'S_GEO'
            ws['MN1'] = 'S_ANT'
            ws['MO1'] = 'S_BIO'
            ws['MP1'] = 'S_FIS'
            ws['MQ1'] = 'S_KIM_1'
            ws['MR1'] = 'S_KIM_2'
            ws['MS1'] = 'S_JML'

            ws['MT1'] = 'RANK NAS.'
            ws['MU1'] = 'RANK LOK.'

            # Z MAT 1
            ws['LS1'].font = Font(bold=False, name='Calibri', size=11)
            ws['LT1'].font = Font(bold=False, name='Calibri', size=11)
            ws['LU1'].font = Font(bold=False, name='Calibri', size=11)
            ws['LV1'].font = Font(bold=False, name='Calibri', size=11)
            ws['LW1'].font = Font(bold=False, name='Calibri', size=11)
            ws['LX1'].font = Font(bold=False, name='Calibri', size=11)
            ws['LY1'].font = Font(bold=False, name='Calibri', size=11)
            ws['LZ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['MA1'].font = Font(bold=False, name='Calibri', size=11)
            ws['MB1'].font = Font(bold=False, name='Calibri', size=11)
            ws['MC1'].font = Font(bold=False, name='Calibri', size=11)
            ws['MD1'].font = Font(bold=False, name='Calibri', size=11)
            ws['ME1'].font = Font(bold=False, name='Calibri', size=11)
            ws['MF1'].font = Font(bold=False, name='Calibri', size=11)
            ws['MG1'].font = Font(bold=False, name='Calibri', size=11)
            ws['MH1'].font = Font(bold=False, name='Calibri', size=11)
            ws['MI1'].font = Font(bold=False, name='Calibri', size=11)
            ws['MJ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['MK1'].font = Font(bold=False, name='Calibri', size=11)
            ws['ML1'].font = Font(bold=False, name='Calibri', size=11)
            ws['MM1'].font = Font(bold=False, name='Calibri', size=11)
            ws['MN1'].font = Font(bold=False, name='Calibri', size=11)
            ws['MO1'].font = Font(bold=False, name='Calibri', size=11)
            ws['MP1'].font = Font(bold=False, name='Calibri', size=11)
            ws['MQ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['MR1'].font = Font(bold=False, name='Calibri', size=11)
            ws['MS1'].font = Font(bold=False, name='Calibri', size=11)
            ws['MT1'].font = Font(bold=False, name='Calibri', size=11)
            ws['KK1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL DARI NAMA
            ws['KZ1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['LA1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['LB1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['LC1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['LD1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['LE1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['LF1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['LG1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['LH1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['LI1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['LJ1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['LK1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['LL1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['LM1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['LN1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['LO1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['LP1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['LQ1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['LR1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['LS1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['LT1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['LU1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['LV1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['LW1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['LX1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['LY1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['LZ1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['MA1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['MB1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['MC1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['MD1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['ME1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['MF1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['MG1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['MH1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['MI1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['MJ1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['MK1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['ML1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['MM1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['MN1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['MO1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['MP1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['MQ1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['MR1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['MS1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['MT1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['MU1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            
            # tambahan
            ws['MV1'] = 'MAT_1_20'
            ws['MW1'] = 'MAT_2_20'
            ws['MX1'] = 'IND_20'
            ws['MY1'] = 'ENG_20'
            ws['MZ1'] = 'SEJ_20'
            ws['NA1'] = 'EKO_20'
            ws['NB1'] = 'SOS_20'
            ws['NC1'] = 'GEO_20'
            ws['ND1'] = 'ANT_20'
            ws['NE1'] = 'BIO_20'
            ws['NF1'] = 'FIS_20'
            ws['NG1'] = 'KIM_1_20'
            ws['NH1'] = 'KIM_2_20'

            ws['MV1'].font = Font(bold=False, name='Calibri', size=11)
            ws['MW1'].font = Font(bold=False, name='Calibri', size=11)
            ws['MX1'].font = Font(bold=False, name='Calibri', size=11)
            ws['MY1'].font = Font(bold=False, name='Calibri', size=11)
            ws['MZ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['NA1'].font = Font(bold=False, name='Calibri', size=11)
            ws['NB1'].font = Font(bold=False, name='Calibri', size=11)
            ws['NC1'].font = Font(bold=False, name='Calibri', size=11)
            ws['ND1'].font = Font(bold=False, name='Calibri', size=11)
            ws['NE1'].font = Font(bold=False, name='Calibri', size=11)
            ws['NF1'].font = Font(bold=False, name='Calibri', size=11)
            ws['NG1'].font = Font(bold=False, name='Calibri', size=11)
            ws['NH1'].font = Font(bold=False, name='Calibri', size=11)

            ws['MV1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['MW1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['MX1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['MY1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['MZ1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['NA1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['NB1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['NC1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['ND1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['NE1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['NF1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['NG1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['NH1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')

            # DARI NAMA
            for row in range(2, q+1):
                ws['KZ{}'.format(row)] = '=B{}'.format(row)
                ws['LA{}'.format(row)] = '=C{}'.format(row, row)
                ws['LB{}'.format(row)] = '=D{}'.format(row, row)
                ws['LC{}'.format(row)] = '=E{}'.format(row, row)
                ws['LD{}'.format(row)] = '=F{}'.format(row, row)
                ws['LE{}'.format(row)] = '=IF(G{}="","",G{})'.format(row, row)
                ws['LF{}'.format(row)] = '=IF(H{}="","",H{})'.format(row, row)
                ws['LG{}'.format(row)] = '=IF(I{}="","",I{})'.format(row, row)
                ws['LH{}'.format(row)] = '=IF(J{}="","",J{})'.format(row, row)
                ws['LI{}'.format(row)] = '=IF(K{}="","",K{})'.format(row, row)
                ws['LJ{}'.format(row)] = '=IF(L{}="","",L{})'.format(row, row)
                ws['LK{}'.format(row)] = '=IF(M{}="","",M{})'.format(row, row)
                ws['LL{}'.format(row)] = '=IF(N{}="","",N{})'.format(row, row)
                ws['LM{}'.format(row)] = '=IF(O{}="","",O{})'.format(row, row)
                ws['LN{}'.format(row)] = '=IF(P{}="","",P{})'.format(row, row)
                ws['LO{}'.format(row)] = '=IF(Q{}="","",Q{})'.format(row, row)
                ws['LP{}'.format(row)] = '=IF(R{}="","",R{})'.format(row, row)
                ws['LQ{}'.format(row)] = '=IF(S{}="","",S{})'.format(row, row)
                ws['LR{}'.format(row)] = '=IF(T{}="","",T{})'.format(row, row)
                
                ws['LS{}'.format(
                    row)] = '=IFERROR(ROUND(IF(IU{}="","",(IU{}-IU${})/IU${}),2),"")'.format(row, row, r, s)
                ws['LT{}'.format(
                    row)] = '=IFERROR(ROUND(IF(IV{}="","",(IV{}-IV${})/IV${}),2),"")'.format(row, row, r, s)
                ws['LU{}'.format(
                    row)] = '=IFERROR(ROUND(IF(IW{}="","",(IW{}-IW${})/IW${}),2),"")'.format(row, row, r, s)
                ws['LV{}'.format(
                    row)] = '=IFERROR(ROUND(IF(IX{}="","",(IX{}-IX${})/IX${}),2),"")'.format(row, row, r, s)
                ws['LW{}'.format(
                    row)] = '=IFERROR(ROUND(IF(IY{}="","",(IY{}-IY${})/IY${}),2),"")'.format(row, row, r, s)
                ws['LX{}'.format(
                    row)] = '=IFERROR(ROUND(IF(IZ{}="","",(IZ{}-IZ${})/IZ${}),2),"")'.format(row, row, r, s)
                ws['LY{}'.format(
                    row)] = '=IFERROR(ROUND(IF(JA{}="","",(JA{}-JA${})/JA${}),2),"")'.format(row, row, r, s)
                ws['LZ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(JB{}="","",(JB{}-JB${})/JB${}),2),"")'.format(row, row, r, s)
                ws['MA{}'.format(
                    row)] = '=IFERROR(ROUND(IF(JC{}="","",(JC{}-JC${})/JC${}),2),"")'.format(row, row, r, s)
                ws['MB{}'.format(
                    row)] = '=IFERROR(ROUND(IF(JD{}="","",(JD{}-JD${})/JD${}),2),"")'.format(row, row, r, s)
                ws['MC{}'.format(
                    row)] = '=IFERROR(ROUND(IF(JE{}="","",(JE{}-JE${})/JE${}),2),"")'.format(row, row, r, s)
                ws['MD{}'.format(
                    row)] = '=IFERROR(ROUND(IF(JF{}="","",(JF{}-JF${})/JF${}),2),"")'.format(row, row, r, s)
                ws['ME{}'.format(
                    row)] = '=IFERROR(ROUND(IF(JG{}="","",(JG{}-JG${})/JG${}),2),"")'.format(row, row, r, s)

                ws['MF{}'.format(
                    row)] = '=IFERROR(ROUND(IF(IU{}="","",IF(70+30*LS{}/$LS${}<20,20,70+30*LS{}/$LS${})),2),"")'.format(row, row, r, row, r)
                ws['MG{}'.format(
                    row)] = '=IFERROR(ROUND(IF(IV{}="","",IF(70+30*LT{}/$LT${}<20,20,70+30*LT{}/$LT${})),2),"")'.format(row, row, r, row, r)
                ws['MH{}'.format(
                    row)] = '=IFERROR(ROUND(IF(IW{}="","",IF(70+30*LU{}/$LU${}<20,20,70+30*LU{}/$LU${})),2),"")'.format(row, row, r, row, r)
                ws['MI{}'.format(
                    row)] = '=IFERROR(ROUND(IF(IX{}="","",IF(70+30*LV{}/$LV${}<20,20,70+30*LV{}/$LV${})),2),"")'.format(row, row, r, row, r)
                ws['MJ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(IY{}="","",IF(70+30*LW{}/$LW${}<20,20,70+30*LW{}/$LW${})),2),"")'.format(row, row, r, row, r)
                ws['MK{}'.format(
                    row)] = '=IFERROR(ROUND(IF(IZ{}="","",IF(70+30*LX{}/$LX${}<20,20,70+30*LX{}/$LX${})),2),"")'.format(row, row, r, row, r)
                ws['ML{}'.format(
                    row)] = '=IFERROR(ROUND(IF(JA{}="","",IF(70+30*LY{}/$LY${}<20,20,70+30*LY{}/$LY${})),2),"")'.format(row, row, r, row, r)
                ws['MM{}'.format(
                    row)] = '=IFERROR(ROUND(IF(JB{}="","",IF(70+30*LZ{}/$LZ${}<20,20,70+30*LZ{}/$LZ${})),2),"")'.format(row, row, r, row, r)
                ws['MN{}'.format(
                    row)] = '=IFERROR(ROUND(IF(JC{}="","",IF(70+30*MA{}/$MA${}<20,20,70+30*MA{}/$MA${})),2),"")'.format(row, row, r, row, r)
                ws['MO{}'.format(
                    row)] = '=IFERROR(ROUND(IF(JD{}="","",IF(70+30*MB{}/$MB${}<20,20,70+30*MB{}/$MB${})),2),"")'.format(row, row, r, row, r)
                ws['MP{}'.format(
                    row)] = '=IFERROR(ROUND(IF(JE{}="","",IF(70+30*MC{}/$MC${}<20,20,70+30*MC{}/$MC${})),2),"")'.format(row, row, r, row, r)
                ws['MQ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(JF{}="","",IF(70+30*MD{}/$MD${}<20,20,70+30*MD{}/$MD${})),2),"")'.format(row, row, r, row, r)
                ws['MR{}'.format(
                    row)] = '=IFERROR(ROUND(IF(JG{}="","",IF(70+30*ME{}/$ME${}<20,20,70+30*ME{}/$ME${})),2),"")'.format(row, row, r, row, r)

                ws['MS{}'.format(row)] = '=IF(SUM(MF{}:MR{})=0,"",SUM(MF{}:MR{}))'.format(
                    row, row, row, row)
                ws['MT{}'.format(row)] = '=IF(MS{}="","",RANK(MS{},$MS$2:$MS${}))'.format(
                    row, row, q)
                ws['MU{}'.format(
                    row)] = '=IF(KJ{}="","",COUNTIFS($IT$2:$IT${},IT{},$KJ$2:$KJ${},"<"&KJ{})+1)'.format(row, q, row, q, row)
            # TAMBAHAN
                ws['MV{}'.format(row)] = '=IF($G${}=20,IF(AND(G{}>3,LS{}=20),1,""),IF($G${}=25,IF(AND(G{}>4,LS{}=20),1,""),IF($G${}=30,IF(AND(G{}>5,LS{}=20),1,""),IF($G${}=35,IF(AND(G{}>6,LS{}=20),1,""),IF($G${}=40,IF(AND(G{}>7,LS{}=20),1,""),IF($G${}=45,IF(AND(G{}>8,LS{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['MW{}'.format(row)] = '=IF($H${}=20,IF(AND(H{}>3,LT{}=20),1,""),IF($H${}=25,IF(AND(H{}>4,LT{}=20),1,""),IF($H${}=30,IF(AND(H{}>5,LT{}=20),1,""),IF($H${}=35,IF(AND(H{}>6,LT{}=20),1,""),IF($H${}=40,IF(AND(H{}>7,LT{}=20),1,""),IF($H${}=45,IF(AND(H{}>8,LT{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['MX{}'.format(row)] = '=IF($I${}=20,IF(AND(I{}>3,LU{}=20),1,""),IF($I${}=25,IF(AND(I{}>4,LU{}=20),1,""),IF($I${}=30,IF(AND(I{}>5,LU{}=20),1,""),IF($I${}=35,IF(AND(I{}>6,LU{}=20),1,""),IF($I${}=40,IF(AND(I{}>7,LU{}=20),1,""),IF($I${}=45,IF(AND(I{}>8,LU{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['MY{}'.format(row)] = '=IF($J${}=20,IF(AND(J{}>3,LV{}=20),1,""),IF($J${}=25,IF(AND(J{}>4,LV{}=20),1,""),IF($J${}=30,IF(AND(J{}>5,LV{}=20),1,""),IF($J${}=35,IF(AND(J{}>6,LV{}=20),1,""),IF($J${}=40,IF(AND(J{}>7,LV{}=20),1,""),IF($J${}=45,IF(AND(J{}>8,LV{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['MZ{}'.format(row)] = '=IF($K${}=20,IF(AND(K{}>3,LW{}=20),1,""),IF($K${}=25,IF(AND(K{}>4,LW{}=20),1,""),IF($K${}=30,IF(AND(K{}>5,LW{}=20),1,""),IF($K${}=35,IF(AND(K{}>6,LW{}=20),1,""),IF($K${}=40,IF(AND(K{}>7,LW{}=20),1,""),IF($K${}=45,IF(AND(K{}>8,LW{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['NA{}'.format(row)] = '=IF($L${}=20,IF(AND(L{}>3,LX{}=20),1,""),IF($L${}=25,IF(AND(L{}>4,LX{}=20),1,""),IF($L${}=30,IF(AND(L{}>5,LX{}=20),1,""),IF($L${}=35,IF(AND(L{}>6,LX{}=20),1,""),IF($L${}=40,IF(AND(L{}>7,LX{}=20),1,""),IF($L${}=45,IF(AND(L{}>8,LX{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['NB{}'.format(row)] = '=IF($M${}=20,IF(AND(M{}>3,LY{}=20),1,""),IF($M${}=25,IF(AND(M{}>4,LY{}=20),1,""),IF($M${}=30,IF(AND(M{}>5,LY{}=20),1,""),IF($M${}=35,IF(AND(M{}>6,LY{}=20),1,""),IF($M${}=40,IF(AND(M{}>7,LY{}=20),1,""),IF($M${}=45,IF(AND(M{}>8,LY{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['NC{}'.format(row)] = '=IF($N${}=20,IF(AND(N{}>3,LZ{}=20),1,""),IF($N${}=25,IF(AND(N{}>4,LZ{}=20),1,""),IF($N${}=30,IF(AND(N{}>5,LZ{}=20),1,""),IF($N${}=35,IF(AND(N{}>6,LZ{}=20),1,""),IF($N${}=40,IF(AND(N{}>7,LZ{}=20),1,""),IF($N${}=45,IF(AND(N{}>8,LZ{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['ND{}'.format(row)] = '=IF($O${}=20,IF(AND(O{}>3,MA{}=20),1,""),IF($O${}=25,IF(AND(O{}>4,MA{}=20),1,""),IF($O${}=30,IF(AND(O{}>5,MA{}=20),1,""),IF($O${}=35,IF(AND(O{}>6,MA{}=20),1,""),IF($O${}=40,IF(AND(O{}>7,MA{}=20),1,""),IF($O${}=45,IF(AND(O{}>8,MA{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['NE{}'.format(row)] = '=IF($P${}=20,IF(AND(P{}>3,MB{}=20),1,""),IF($P${}=25,IF(AND(P{}>4,MB{}=20),1,""),IF($P${}=30,IF(AND(P{}>5,MB{}=20),1,""),IF($P${}=35,IF(AND(P{}>6,MB{}=20),1,""),IF($P${}=40,IF(AND(P{}>7,MB{}=20),1,""),IF($P${}=45,IF(AND(P{}>8,MB{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['NF{}'.format(row)] = '=IF($Q${}=20,IF(AND(Q{}>3,MC{}=20),1,""),IF($Q${}=25,IF(AND(Q{}>4,MC{}=20),1,""),IF($Q${}=30,IF(AND(Q{}>5,MC{}=20),1,""),IF($Q${}=35,IF(AND(Q{}>6,MC{}=20),1,""),IF($Q${}=40,IF(AND(Q{}>7,MC{}=20),1,""),IF($Q${}=45,IF(AND(Q{}>8,MC{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['NG{}'.format(row)] = '=IF($R${}=20,IF(AND(R{}>3,MD{}=20),1,""),IF($R${}=25,IF(AND(R{}>4,MD{}=20),1,""),IF($R${}=30,IF(AND(R{}>5,MD{}=20),1,""),IF($R${}=35,IF(AND(R{}>6,MD{}=20),1,""),IF($R${}=40,IF(AND(R{}>7,MD{}=20),1,""),IF($R${}=45,IF(AND(R{}>8,MD{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['NH{}'.format(row)] = '=IF($S${}=20,IF(AND(S{}>3,ME{}=20),1,""),IF($S${}=25,IF(AND(S{}>4,ME{}=20),1,""),IF($S${}=30,IF(AND(S{}>5,ME{}=20),1,""),IF($S${}=35,IF(AND(S{}>6,ME{}=20),1,""),IF($S${}=40,IF(AND(S{}>7,ME{}=20),1,""),IF($S${}=45,IF(AND(S{}>8,ME{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

            # Mengubah 'KELAS' sesuai dengan nilai yang dipilih dari selectbox 'KELAS'
            kelas = KELAS.lower().replace(" ", "")
            semester = SEMESTER.lower()
            tahun = TAHUN.replace("-", "")
            penilaian = PENILAIAN.lower()
            kurikulum = KURIKULUM.lower()

            path_file = f"{kelas}_{penilaian}_{semester}_{kurikulum}_{tahun}_nilai_std.xlsx"

            # Simpan file ke direktori temporer
            temp_dir = tempfile.gettempdir()
            file_path = temp_dir + '/' + path_file
            wb.save(file_path)

            st.success(
                "File siap diunduh!")

            # Tombol unduh file
            with open(file_path, "rb") as f:
                bytes_data = f.read()
            st.download_button(label="Unduh File", data=bytes_data,
                               file_name=path_file)

            st.warning(
                "Buka file unduhan, klik 'Enable Editing' dan 'Save'")

    if selected_file == "Nilai Std. 10, 11, PPLS IPA":
        # menghilangkan hamburger
        st.markdown("""
        <style>
        .css-1rs6os.edgvbvh3
        {
            visibility:hidden;
        }
        .css-1lsmgbg.egzxvld0
        {
            visibility:hidden;
        }
        </style>
        """, unsafe_allow_html=True)

        image = Image.open('logo resmi nf resize.png')
        st.image(image)

        st.title("Olah Nilai Standar PPLS")
        st.header("PPLS")

        col6 = st.container()

        with col6:
            KELAS = st.selectbox(
                "KELAS",
                ("--Pilih Kelas--", "10 IPA", "11 IPA", "PPLS IPA"))

        col7 = st.container()

        with col7:
            SEMESTER = st.selectbox(
                "SEMESTER",
                ("--Pilih Semester--", "SEMESTER 1", "SEMESTER 2"))

        col8 = st.container()

        with col8:
            PENILAIAN = st.selectbox(
                "PENILAIAN",
                ("--Pilih Penilaian--", "PENILAIAN TENGAH SEMESTER", "PENILAIAN AKHIR TAHUN"))

        col9 = st.container()

        with col9:
            KURIKULUM = st.selectbox(
                "KURIKULUM",
                ("--Pilih Kurikulum--", "K13", "PPLS"))

        TAHUN = st.text_input("Masukkan Tahun Ajaran",
                              placeholder="contoh: 2022-2023")

        col1, col2, col3, col4 = st.columns(4)

        with col1:
            MTK = st.selectbox(
                "JML. SOAL MAT.",
                (15, 20, 25, 30, 35, 40, 45, 50))

        with col2:
            FIS = st.selectbox(
                "JML. SOAL FIS.",
                (15, 20, 25, 30, 35, 40, 45, 50))

        with col3:
            KIM = st.selectbox(
                "JML. SOAL KIM.",
                (15, 20, 25, 30, 35, 40, 45, 50))

        with col4:
            BIO = st.selectbox(
                "JML. SOAL BIO.",
                (15, 20, 25, 30, 35, 40, 45, 50))

        JML_SOAL_MAT = MTK
        JML_SOAL_FIS = FIS
        JML_SOAL_KIM = KIM
        JML_SOAL_BIO = BIO

        uploaded_file = st.file_uploader(
            'Letakkan file excel', type='xlsx')

        if uploaded_file is not None:
            wb = openpyxl.load_workbook(uploaded_file)
            ws = wb['Sheet1']

            q = len(ws['K'])
            r = len(ws['K'])+2
            s = len(ws['K'])+3
            t = len(ws['K'])+4
            u = len(ws['K'])+5
            v = len(ws['K'])+6
            w = len(ws['K'])+7
            x = len(ws['K'])+8

            ws['G{}'.format(r)] = "=ROUND(AVERAGE(G2:G{}),2)".format(q)
            ws['H{}'.format(r)] = "=ROUND(AVERAGE(H2:H{}),2)".format(q)
            ws['I{}'.format(r)] = "=ROUND(AVERAGE(I2:I{}),2)".format(q)
            ws['J{}'.format(r)] = "=ROUND(AVERAGE(J2:J{}),2)".format(q)
            ws['K{}'.format(r)] = "=ROUND(AVERAGE(K2:K{}),2)".format(q)
            ws['G{}'.format(s)] = "=STDEV(G2:G{})".format(q)
            ws['H{}'.format(s)] = "=STDEV(H2:H{})".format(q)
            ws['I{}'.format(s)] = "=STDEV(I2:I{})".format(q)
            ws['J{}'.format(s)] = "=STDEV(J2:J{})".format(q)
            ws['G{}'.format(t)] = "=MAX(G2:G{})".format(q)
            ws['H{}'.format(t)] = "=MAX(H2:H{})".format(q)
            ws['I{}'.format(t)] = "=MAX(I2:I{})".format(q)
            ws['J{}'.format(t)] = "=MAX(J2:J{})".format(q)
            ws['K{}'.format(t)] = "=MAX(K2:K{})".format(q)
            ws['L{}'.format(r)] = "=MAX(L2:L{})".format(q)
            ws['M{}'.format(r)] = "=MAX(M2:M{})".format(q)
            ws['N{}'.format(r)] = "=MAX(N2:N{})".format(q)
            ws['O{}'.format(r)] = "=MAX(O2:O{})".format(q)
            ws['P{}'.format(r)] = "=MAX(P2:P{})".format(q)
            ws['Q{}'.format(r)] = "=MAX(Q2:Q{})".format(q)
            ws['R{}'.format(r)] = "=MAX(R2:R{})".format(q)
            ws['S{}'.format(r)] = "=MAX(S2:S{})".format(q)
            ws['T{}'.format(r)] = "=ROUND(MAX(T2:T{}),2)".format(q)
            ws['U{}'.format(r)] = "=MAX(U2:U{})".format(q)
            ws['G{}'.format(u)] = "=MIN(G2:G{})".format(q)
            ws['H{}'.format(u)] = "=MIN(H2:H{})".format(q)
            ws['I{}'.format(u)] = "=MIN(I2:I{})".format(q)
            ws['J{}'.format(u)] = "=MIN(J2:J{})".format(q)
            ws['K{}'.format(u)] = "=MIN(K2:K{})".format(q)
            ws['P{}'.format(s)] = "=MIN(P2:P{})".format(q)
            ws['Q{}'.format(s)] = "=MIN(Q2:R{})".format(q)
            ws['R{}'.format(s)] = "=MIN(R2:S{})".format(q)
            ws['S{}'.format(s)] = "=MIN(S2:T{})".format(q)
            ws['T{}'.format(s)] = "=MIN(T2:T{})".format(q)
            ws['P{}'.format(t)] = "=ROUND(AVERAGE(P2:P{}),2)".format(q)
            ws['Q{}'.format(t)] = "=ROUND(AVERAGE(Q2:Q{}),2)".format(q)
            ws['R{}'.format(t)] = "=ROUND(AVERAGE(R2:R{}),2)".format(q)
            ws['S{}'.format(t)] = "=ROUND(AVERAGE(S2:S{}),2)".format(q)
            ws['T{}'.format(t)] = "=ROUND(AVERAGE(T2:T{}),2)".format(q)
            ws['W{}'.format(r)] = "=SUM(W2:W{})".format(q)
            ws['X{}'.format(r)] = "=SUM(X2:X{})".format(q)
            ws['Y{}'.format(r)] = "=SUM(Y2:Y{})".format(q)
            ws['Z{}'.format(r)] = "=SUM(Z2:Z{})".format(q)

            # new
            # iterasi 1 rata-rata - 1

            # MAPEL NORMAL
            ws['AG{}'.format(r)] = "=IF($W${}=0,$G${},$G${}-1)".format(r, r, r)
            ws['AG{}'.format(s)] = "=STDEV(AG2:AG{})".format(q)
            ws['AG{}'.format(t)] = "=MAX(AG2:AG{})".format(q)
            ws['AG{}'.format(u)] = "=MIN(AG2:AG{})".format(q)
            ws['AH{}'.format(r)] = "=IF($X${}=0,$H${},$H${}-1)".format(r, r, r)
            ws['AH{}'.format(s)] = "=STDEV(AH2:AH{})".format(q)
            ws['AH{}'.format(t)] = "=MAX(AH2:AH{})".format(q)
            ws['AH{}'.format(u)] = "=MIN(AH2:AH{})".format(q)
            ws['AI{}'.format(r)] = "=IF($Y${}=0,$I${},$I${}-1)".format(r, r, r)
            ws['AI{}'.format(s)] = "=STDEV(AI2:AI{})".format(q)
            ws['AI{}'.format(t)] = "=MAX(AI2:AI{})".format(q)
            ws['AI{}'.format(u)] = "=MIN(AI2:AI{})".format(q)
            ws['AJ{}'.format(r)] = "=IF($Z${}=0,$J${},$J${}-1)".format(r, r, r)
            ws['AJ{}'.format(s)] = "=STDEV(AJ2:AJ{})".format(q)
            ws['AJ{}'.format(t)] = "=MAX(AJ2:AJ{})".format(q)
            ws['AJ{}'.format(u)] = "=MIN(AJ2:AJ{})".format(q)

            # JUMLAH MAPEL NORMAL
            ws['AK{}'.format(r)] = "=ROUND(AVERAGE(AK2:AK{}),2)".format(q)
            ws['AK{}'.format(t)] = "=MAX(AK2:AK{})".format(q)
            ws['AK{}'.format(u)] = "=MIN(AK2:AK{})".format(q)

            # Z SCORE
            ws['AL{}'.format(r)] = "=MAX(AL2:AL{})".format(q)
            ws['AM{}'.format(r)] = "=MAX(AM2:AM{})".format(q)
            ws['AN{}'.format(r)] = "=MAX(AN2:AN{})".format(q)
            ws['AO{}'.format(r)] = "=MAX(AO2:AO{})".format(q)

            # NILAI STANDAR
            ws['AP{}'.format(r)] = "=MAX(AP2:AP{})".format(q)
            ws['AP{}'.format(s)] = "=MIN(AP2:AP{})".format(q)
            ws['AP{}'.format(t)] = "=ROUND(AVERAGE(AP2:AP{}),2)".format(q)
            ws['AQ{}'.format(r)] = "=MAX(AQ2:AQ{})".format(q)
            ws['AQ{}'.format(s)] = "=MIN(AQ2:AQ{})".format(q)
            ws['AQ{}'.format(t)] = "=ROUND(AVERAGE(AQ2:AQ{}),2)".format(q)
            ws['AR{}'.format(r)] = "=MAX(AR2:AR{})".format(q)
            ws['AR{}'.format(s)] = "=MIN(AR2:AR{})".format(q)
            ws['AR{}'.format(t)] = "=ROUND(AVERAGE(AR2:AR{}),2)".format(q)
            ws['AS{}'.format(r)] = "=MAX(AS2:AS{})".format(q)
            ws['AS{}'.format(s)] = "=MIN(AS2:AS{})".format(q)
            ws['AS{}'.format(t)] = "=ROUND(AVERAGE(AS2:AS{}),2)".format(q)
            ws['AT{}'.format(r)] = "=MAX(AT2:AT{})".format(q)
            ws['AT{}'.format(s)] = "=MIN(AT2:AT{})".format(q)
            ws['AT{}'.format(t)] = "=ROUND(AVERAGE(AT2:AT{}),2)".format(q)

            # INISIASI MAPEL
            ws['AW{}'.format(r)] = "=SUM(AW2:AW{})".format(q)
            ws['AX{}'.format(r)] = "=SUM(AX2:AX{})".format(q)
            ws['AY{}'.format(r)] = "=SUM(AY2:AY{})".format(q)
            ws['AZ{}'.format(r)] = "=SUM(AZ2:AZ{})".format(q)

            # iterasi 2 rata-rata - 1
            # MAPEL NORMAL
            ws['BG{}'.format(
                r)] = "=IF($AW${}=0,$AG${},$AG${}-1)".format(r, r, r)
            ws['BG{}'.format(s)] = "=STDEV(BG2:BG{})".format(q)
            ws['BG{}'.format(t)] = "=MAX(BG2:BG{})".format(q)
            ws['BG{}'.format(u)] = "=MIN(BG2:BG{})".format(q)
            ws['BH{}'.format(
                r)] = "=IF($AX${}=0,$AH${},$AH${}-1)".format(r, r, r)
            ws['BH{}'.format(s)] = "=STDEV(BH2:BH{})".format(q)
            ws['BH{}'.format(t)] = "=MAX(BH2:BH{})".format(q)
            ws['BH{}'.format(u)] = "=MIN(BH2:BH{})".format(q)
            ws['BI{}'.format(
                r)] = "=IF($AY${}=0,$AI${},$AI${}-1)".format(r, r, r)
            ws['BI{}'.format(s)] = "=STDEV(BI2:BI{})".format(q)
            ws['BI{}'.format(t)] = "=MAX(BI2:BI{})".format(q)
            ws['BI{}'.format(u)] = "=MIN(BI2:BI{})".format(q)
            ws['BJ{}'.format(
                r)] = "=IF($AZ${}=0,$AJ${},$AJ${}-1)".format(r, r, r)
            ws['BJ{}'.format(s)] = "=STDEV(BJ2:BJ{})".format(q)
            ws['BJ{}'.format(t)] = "=MAX(BJ2:BJ{})".format(q)
            ws['BJ{}'.format(u)] = "=MIN(BJ2:BJ{})".format(q)

            # JUMLAH MAPEL NORMAL
            ws['BK{}'.format(r)] = "=ROUND(AVERAGE(BK2:BK{}),2)".format(q)
            ws['BK{}'.format(t)] = "=MAX(BK2:BK{})".format(q)
            ws['BK{}'.format(u)] = "=MIN(BK2:BK{})".format(q)

            # Z SCORE
            ws['BL{}'.format(r)] = "=MAX(BL2:BL{})".format(q)
            ws['BM{}'.format(r)] = "=MAX(BM2:BM{})".format(q)
            ws['BN{}'.format(r)] = "=MAX(BN2:BN{})".format(q)
            ws['BO{}'.format(r)] = "=MAX(BO2:BO{})".format(q)

            # NILAI STANDAR
            ws['BP{}'.format(r)] = "=MAX(BP2:BP{})".format(q)
            ws['BP{}'.format(s)] = "=MIN(BP2:BP{})".format(q)
            ws['BP{}'.format(t)] = "=ROUND(AVERAGE(BP2:BP{}),2)".format(q)
            ws['BQ{}'.format(r)] = "=MAX(BQ2:BQ{})".format(q)
            ws['BQ{}'.format(s)] = "=MIN(BQ2:BQ{})".format(q)
            ws['BQ{}'.format(t)] = "=ROUND(AVERAGE(BQ2:BQ{}),2)".format(q)
            ws['BR{}'.format(r)] = "=MAX(BR2:BR{})".format(q)
            ws['BR{}'.format(s)] = "=MIN(BR2:BR{})".format(q)
            ws['BR{}'.format(t)] = "=ROUND(AVERAGE(BR2:BR{}),2)".format(q)
            ws['BS{}'.format(r)] = "=MAX(BS2:BS{})".format(q)
            ws['BS{}'.format(s)] = "=MIN(BS2:BS{})".format(q)
            ws['BS{}'.format(t)] = "=ROUND(AVERAGE(BS2:BS{}),2)".format(q)
            ws['BT{}'.format(r)] = "=MAX(BT2:BT{})".format(q)
            ws['BT{}'.format(s)] = "=MIN(BT2:BT{})".format(q)
            ws['BT{}'.format(t)] = "=ROUND(AVERAGE(BT2:BT{}),2)".format(q)

            # INISIASI MAPEL
            ws['BW{}'.format(r)] = "=SUM(BW2:BW{})".format(q)
            ws['BX{}'.format(r)] = "=SUM(BX2:BX{})".format(q)
            ws['BY{}'.format(r)] = "=SUM(BY2:BY{})".format(q)
            ws['BZ{}'.format(r)] = "=SUM(BZ2:BZ{})".format(q)

            # iterasi 3 rata-rata - 1
            # MAPEL NORMAL
            ws['CG{}'.format(
                r)] = "=IF($BW${}=0,$BG${},$BG${}-1)".format(r, r, r)
            ws['CG{}'.format(s)] = "=STDEV(CG2:CG{})".format(q)
            ws['CG{}'.format(t)] = "=MAX(CG2:CG{})".format(q)
            ws['CG{}'.format(u)] = "=MIN(CG2:CG{})".format(q)
            ws['CH{}'.format(
                r)] = "=IF($BX${}=0,$BH${},$BH${}-1)".format(r, r, r)
            ws['CH{}'.format(s)] = "=STDEV(CH2:CH{})".format(q)
            ws['CH{}'.format(t)] = "=MAX(CH2:CH{})".format(q)
            ws['CH{}'.format(u)] = "=MIN(CH2:CH{})".format(q)
            ws['CI{}'.format(
                r)] = "=IF($BY${}=0,$BI${},$BI${}-1)".format(r, r, r)
            ws['CI{}'.format(s)] = "=STDEV(CI2:CI{})".format(q)
            ws['CI{}'.format(t)] = "=MAX(CI2:CI{})".format(q)
            ws['CI{}'.format(u)] = "=MIN(CI2:CI{})".format(q)
            ws['CJ{}'.format(
                r)] = "=IF($BZ${}=0,$BJ${},$BJ${}-1)".format(r, r, r)
            ws['CJ{}'.format(s)] = "=STDEV(CJ2:CJ{})".format(q)
            ws['CJ{}'.format(t)] = "=MAX(CJ2:CJ{})".format(q)
            ws['CJ{}'.format(u)] = "=MIN(CJ2:CJ{})".format(q)

            # JUMLAH MAPEL NORMAL
            ws['CK{}'.format(r)] = "=ROUND(AVERAGE(CK2:CK{}),2)".format(q)
            ws['CK{}'.format(t)] = "=MAX(CK2:CK{})".format(q)
            ws['CK{}'.format(u)] = "=MIN(CK2:CK{})".format(q)

            # Z SCORE
            ws['CL{}'.format(r)] = "=MAX(CL2:CL{})".format(q)
            ws['CM{}'.format(r)] = "=MAX(CM2:CM{})".format(q)
            ws['CN{}'.format(r)] = "=MAX(CN2:CN{})".format(q)
            ws['CO{}'.format(r)] = "=MAX(CO2:CO{})".format(q)

            # NILAI STANDAR
            ws['CP{}'.format(r)] = "=MAX(CP2:CP{})".format(q)
            ws['CP{}'.format(s)] = "=MIN(CP2:CP{})".format(q)
            ws['CP{}'.format(t)] = "=ROUND(AVERAGE(CP2:CP{}),2)".format(q)
            ws['CQ{}'.format(r)] = "=MAX(CQ2:CQ{})".format(q)
            ws['CQ{}'.format(s)] = "=MIN(CQ2:CQ{})".format(q)
            ws['CQ{}'.format(t)] = "=ROUND(AVERAGE(CQ2:CQ{}),2)".format(q)
            ws['CR{}'.format(r)] = "=MAX(CR2:CR{})".format(q)
            ws['CR{}'.format(s)] = "=MIN(CR2:CR{})".format(q)
            ws['CR{}'.format(t)] = "=ROUND(AVERAGE(CR2:CR{}),2)".format(q)
            ws['CS{}'.format(r)] = "=MAX(CS2:CS{})".format(q)
            ws['CS{}'.format(s)] = "=MIN(CS2:CS{})".format(q)
            ws['CS{}'.format(t)] = "=ROUND(AVERAGE(CS2:CS{}),2)".format(q)
            ws['CT{}'.format(r)] = "=MAX(CT2:CT{})".format(q)
            ws['CT{}'.format(s)] = "=MIN(CT2:CT{})".format(q)
            ws['CT{}'.format(t)] = "=ROUND(AVERAGE(CT2:CT{}),2)".format(q)

            # INISIASI MAPEL
            ws['CW{}'.format(r)] = "=SUM(CW2:CW{})".format(q)
            ws['CX{}'.format(r)] = "=SUM(CX2:CX{})".format(q)
            ws['CY{}'.format(r)] = "=SUM(CY2:CY{})".format(q)
            ws['CZ{}'.format(r)] = "=SUM(CZ2:CZ{})".format(q)

            # iterasi 4 rata-rata - 1
            # MAPEL NORMAL
            ws['DG{}'.format(
                r)] = "=IF($CW${}=0,$CG${},$CG${}-1)".format(r, r, r)
            ws['DG{}'.format(s)] = "=STDEV(DG2:DG{})".format(q)
            ws['DG{}'.format(t)] = "=MAX(DG2:DG{})".format(q)
            ws['DG{}'.format(u)] = "=MIN(DG2:DG{})".format(q)
            ws['DH{}'.format(
                r)] = "=IF($CX${}=0,$CH${},$CH${}-1)".format(r, r, r)
            ws['DH{}'.format(s)] = "=STDEV(DH2:DH{})".format(q)
            ws['DH{}'.format(t)] = "=MAX(DH2:DH{})".format(q)
            ws['DH{}'.format(u)] = "=MIN(DH2:DH{})".format(q)
            ws['DI{}'.format(
                r)] = "=IF($CY${}=0,$CI${},$CI${}-1)".format(r, r, r)
            ws['DI{}'.format(s)] = "=STDEV(DI2:DI{})".format(q)
            ws['DI{}'.format(t)] = "=MAX(DI2:DI{})".format(q)
            ws['DI{}'.format(u)] = "=MIN(DI2:DI{})".format(q)
            ws['DJ{}'.format(
                r)] = "=IF($CZ${}=0,$CJ${},$CJ${}-1)".format(r, r, r)
            ws['DJ{}'.format(s)] = "=STDEV(DJ2:DJ{})".format(q)
            ws['DJ{}'.format(t)] = "=MAX(DJ2:DJ{})".format(q)
            ws['DJ{}'.format(u)] = "=MIN(DJ2:DJ{})".format(q)

            # JUMLAH MAPEL NORMAL
            ws['DK{}'.format(r)] = "=ROUND(AVERAGE(DK2:DK{}),2)".format(q)
            ws['DK{}'.format(t)] = "=MAX(DK2:DK{})".format(q)
            ws['DK{}'.format(u)] = "=MIN(DK2:DK{})".format(q)

            # Z SCORE
            ws['DL{}'.format(r)] = "=MAX(DL2:DL{})".format(q)
            ws['DM{}'.format(r)] = "=MAX(DM2:DM{})".format(q)
            ws['DN{}'.format(r)] = "=MAX(DN2:DN{})".format(q)
            ws['DO{}'.format(r)] = "=MAX(DO2:DO{})".format(q)

            # NILAI STANDAR
            ws['DP{}'.format(r)] = "=MAX(DP2:DP{})".format(q)
            ws['DP{}'.format(s)] = "=MIN(DP2:DP{})".format(q)
            ws['DP{}'.format(t)] = "=ROUND(AVERAGE(DP2:DP{}),2)".format(q)
            ws['DQ{}'.format(r)] = "=MAX(DQ2:DQ{})".format(q)
            ws['DQ{}'.format(s)] = "=MIN(DQ2:DQ{})".format(q)
            ws['DQ{}'.format(t)] = "=ROUND(AVERAGE(DQ2:DQ{}),2)".format(q)
            ws['DR{}'.format(r)] = "=MAX(DR2:DR{})".format(q)
            ws['DR{}'.format(s)] = "=MIN(DR2:DR{})".format(q)
            ws['DR{}'.format(t)] = "=ROUND(AVERAGE(DR2:DR{}),2)".format(q)
            ws['DS{}'.format(r)] = "=MAX(DS2:DS{})".format(q)
            ws['DS{}'.format(s)] = "=MIN(DS2:DS{})".format(q)
            ws['DS{}'.format(t)] = "=ROUND(AVERAGE(DS2:DS{}),2)".format(q)
            ws['DT{}'.format(r)] = "=MAX(DT2:DT{})".format(q)
            ws['DT{}'.format(s)] = "=MIN(DT2:DT{})".format(q)
            ws['DT{}'.format(t)] = "=ROUND(AVERAGE(DT2:DT{}),2)".format(q)

            # INISIASI MAPEL
            ws['DW{}'.format(r)] = "=SUM(DW2:DW{})".format(q)
            ws['DX{}'.format(r)] = "=SUM(DX2:DX{})".format(q)
            ws['DY{}'.format(r)] = "=SUM(DY2:DY{})".format(q)
            ws['DZ{}'.format(r)] = "=SUM(DZ2:DZ{})".format(q)

            # iterasi 5 rata-rata - 1
            # MAPEL NORMAL
            ws['EG{}'.format(
                r)] = "=IF($DW${}=0,$DG${},$DG${}-1)".format(r, r, r)
            ws['EG{}'.format(s)] = "=STDEV(EG2:EG{})".format(q)
            ws['EG{}'.format(t)] = "=MAX(EG2:EG{})".format(q)
            ws['EG{}'.format(u)] = "=MIN(EG2:EG{})".format(q)
            ws['EH{}'.format(
                r)] = "=IF($DX${}=0,$DH${},$DH${}-1)".format(r, r, r)
            ws['EH{}'.format(s)] = "=STDEV(EH2:EH{})".format(q)
            ws['EH{}'.format(t)] = "=MAX(EH2:EH{})".format(q)
            ws['EH{}'.format(u)] = "=MIN(EH2:EH{})".format(q)
            ws['EI{}'.format(
                r)] = "=IF($DY${}=0,$DI${},$DI${}-1)".format(r, r, r)
            ws['EI{}'.format(s)] = "=STDEV(EI2:EI{})".format(q)
            ws['EI{}'.format(t)] = "=MAX(EI2:EI{})".format(q)
            ws['EI{}'.format(u)] = "=MIN(EI2:EI{})".format(q)
            ws['EJ{}'.format(
                r)] = "=IF($DZ${}=0,$DJ${},$DJ${}-1)".format(r, r, r)
            ws['EJ{}'.format(s)] = "=STDEV(EJ2:EJ{})".format(q)
            ws['EJ{}'.format(t)] = "=MAX(EJ2:EJ{})".format(q)
            ws['EJ{}'.format(u)] = "=MIN(EJ2:EJ{})".format(q)

            # JUMLAH MAPEL NORMAL
            ws['EK{}'.format(r)] = "=ROUND(AVERAGE(EK2:EK{}),2)".format(q)
            ws['EK{}'.format(t)] = "=MAX(EK2:EK{})".format(q)
            ws['EK{}'.format(u)] = "=MIN(EK2:EK{})".format(q)

            # Z SCORE
            ws['EL{}'.format(r)] = "=MAX(EL2:EL{})".format(q)
            ws['EM{}'.format(r)] = "=MAX(EM2:EM{})".format(q)
            ws['EN{}'.format(r)] = "=MAX(EN2:EN{})".format(q)
            ws['EO{}'.format(r)] = "=MAX(EO2:EO{})".format(q)

            # NILAI STANDAR
            ws['EP{}'.format(r)] = "=MAX(EP2:EP{})".format(q)
            ws['EP{}'.format(s)] = "=MIN(EP2:EP{})".format(q)
            ws['EP{}'.format(t)] = "=ROUND(AVERAGE(EP2:EP{}),2)".format(q)
            ws['EQ{}'.format(r)] = "=MAX(EQ2:EQ{})".format(q)
            ws['EQ{}'.format(s)] = "=MIN(EQ2:EQ{})".format(q)
            ws['EQ{}'.format(t)] = "=ROUND(AVERAGE(EQ2:EQ{}),2)".format(q)
            ws['ER{}'.format(r)] = "=MAX(ER2:ER{})".format(q)
            ws['ER{}'.format(s)] = "=MIN(ER2:ER{})".format(q)
            ws['ER{}'.format(t)] = "=ROUND(AVERAGE(ER2:ER{}),2)".format(q)
            ws['ES{}'.format(r)] = "=MAX(ES2:ES{})".format(q)
            ws['ES{}'.format(s)] = "=MIN(ES2:ES{})".format(q)
            ws['ES{}'.format(t)] = "=ROUND(AVERAGE(ES2:ES{}),2)".format(q)
            ws['ET{}'.format(r)] = "=MAX(ET2:ET{})".format(q)
            ws['ET{}'.format(s)] = "=MIN(ET2:ET{})".format(q)
            ws['ET{}'.format(t)] = "=ROUND(AVERAGE(ET2:ET{}),2)".format(q)

            # INISIASI MAPEL
            ws['EW{}'.format(r)] = "=SUM(EW2:EW{})".format(q)
            ws['EX{}'.format(r)] = "=SUM(EX2:EX{})".format(q)
            ws['EY{}'.format(r)] = "=SUM(EY2:EY{})".format(q)
            ws['EZ{}'.format(r)] = "=SUM(EZ2:EZ{})".format(q)

            # Jumlah Soal
            ws['F{}'.format(v)] = 'JUMLAH SOAL'
            ws['G{}'.format(v)] = JML_SOAL_MAT
            ws['H{}'.format(v)] = JML_SOAL_FIS
            ws['I{}'.format(v)] = JML_SOAL_KIM
            ws['J{}'.format(v)] = JML_SOAL_BIO

            # Z Score
            ws['B1'] = 'NAMA SISWA_A'
            ws['C1'] = 'NOMOR NF_A'
            ws['D1'] = 'KELAS_A'
            ws['E1'] = 'NAMA SEKOLAH_A'
            ws['F1'] = 'LOKASI_A'
            ws['G1'] = 'MAT_A'
            ws['H1'] = 'FIS_A'
            ws['I1'] = 'KIM_A'
            ws['J1'] = 'BIO_A'
            ws['K1'] = 'JML_A'
            ws['L1'] = 'Z_MAT_A'
            ws['M1'] = 'Z_FIS_A'
            ws['N1'] = 'Z_KIM_A'
            ws['O1'] = 'Z_BIO_A'
            ws['P1'] = 'S_MAT_A'
            ws['Q1'] = 'S_FIS_A'
            ws['R1'] = 'S_KIM_A'
            ws['S1'] = 'S_BIO_A'
            ws['T1'] = 'S_JML_A'
            ws['U1'] = 'RANK NAS._A'
            ws['V1'] = 'RANK LOK._A'

            ws['L1'].font = Font(bold=False, name='Calibri', size=11)
            ws['M1'].font = Font(bold=False, name='Calibri', size=11)
            ws['N1'].font = Font(bold=False, name='Calibri', size=11)
            ws['O1'].font = Font(bold=False, name='Calibri', size=11)
            ws['P1'].font = Font(bold=False, name='Calibri', size=11)
            ws['Q1'].font = Font(bold=False, name='Calibri', size=11)
            ws['R1'].font = Font(bold=False, name='Calibri', size=11)
            ws['S1'].font = Font(bold=False, name='Calibri', size=11)
            ws['T1'].font = Font(bold=False, name='Calibri', size=11)
            ws['U1'].font = Font(bold=False, name='Calibri', size=11)
            ws['V1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['B1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['C1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['D1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['E1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['F1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['G1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['H1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['I1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['J1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['K1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['L1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['M1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['N1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['O1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['P1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['Q1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['R1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['S1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['T1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['U1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['V1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')

            # tambahan
            ws['W1'] = 'MAT_20_A'
            ws['X1'] = 'FIS_20_A'
            ws['Y1'] = 'KIM_20_A'
            ws['Z1'] = 'BIO_20_A'
            ws['W1'].font = Font(bold=False, name='Calibri', size=11)
            ws['X1'].font = Font(bold=False, name='Calibri', size=11)
            ws['Y1'].font = Font(bold=False, name='Calibri', size=11)
            ws['Z1'].font = Font(bold=False, name='Calibri', size=11)
            ws['W1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['X1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['Y1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['Z1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')

            for row in range(2, q+1):
                ws['K{}'.format(
                    row)] = '=SUM(G{}:J{})'.format(row, row, row)
                ws['L{}'.format(
                    row)] = '=IFERROR(ROUND(IF(G{}="","",(G{}-G${})/G${}),2),"")'.format(row, row, r, s)
                ws['M{}'.format(
                    row)] = '=IFERROR(ROUND(IF(H{}="","",(H{}-H${})/H${}),2),"")'.format(row, row, r, s)
                ws['N{}'.format(
                    row)] = '=IFERROR(ROUND(IF(I{}="","",(I{}-I${})/I${}),2),"")'.format(row, row, r, s)
                ws['O{}'.format(
                    row)] = '=IFERROR(ROUND(IF(J{}="","",(J{}-J${})/J${}),2),"")'.format(row, row, r, s)
                ws['P{}'.format(
                    row)] = '=IFERROR(ROUND(IF(G{}="","",IF(70+30*L{}/$L${}<20,20,70+30*L{}/$L${})),2),"")'.format(row, row, r, row, r)
                ws['Q{}'.format(
                    row)] = '=IFERROR(ROUND(IF(H{}="","",IF(70+30*M{}/$M${}<20,20,70+30*M{}/$M${})),2),"")'.format(row, row, r, row, r)
                ws['R{}'.format(
                    row)] = '=IFERROR(ROUND(IF(I{}="","",IF(70+30*N{}/$N${}<20,20,70+30*N{}/$N${})),2),"")'.format(row, row, r, row, r)
                ws['S{}'.format(
                    row)] = '=IFERROR(ROUND(IF(J{}="","",IF(70+30*O{}/$O${}<20,20,70+30*O{}/$P${})),2),"")'.format(row, row, r, row, r)

                ws['T{}'.format(row)] = '=IF(SUM(P{}:S{})=0,"",SUM(P{}:S{}))'.format(
                    row, row, row, row)
                ws['U{}'.format(row)] = '=IF(T{}="","",RANK(T{},$T$2:$T${}))'.format(
                    row, row, q)
                ws['V{}'.format(
                    row)] = '=IF(U{}="","",COUNTIFS($F$2:$F${},F{},$U$2:$U${},"<"&U{})+1)'.format(row, q, row, q, row)
                # TAMBAHAN
                ws['W{}'.format(row)] = '=IF($G${}=25,IF(AND(G{}>4,P{}=20),1,""),IF($G${}=30,IF(AND(G{}>5,P{}=20),1,""),IF($G${}=35,IF(AND(G{}>6,P{}=20),1,""),IF($G${}=40,IF(AND(G{}>7,P{}=20),1,""),IF($G${}=45,IF(AND(G{}>8,P{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['X{}'.format(row)] = '=IF($H${}=25,IF(AND(H{}>4,Q{}=20),1,""),IF($H${}=30,IF(AND(H{}>5,Q{}=20),1,""),IF($H${}=35,IF(AND(H{}>6,Q{}=20),1,""),IF($H${}=40,IF(AND(H{}>7,Q{}=20),1,""),IF($H${}=45,IF(AND(H{}>8,Q{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['Y{}'.format(row)] = '=IF($I${}=25,IF(AND(I{}>4,R{}=20),1,""),IF($I${}=30,IF(AND(I{}>5,R{}=20),1,""),IF($I${}=35,IF(AND(I{}>6,R{}=20),1,""),IF($I${}=40,IF(AND(I{}>7,R{}=20),1,""),IF($I${}=45,IF(AND(I{}>8,R{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['Z{}'.format(row)] = '=IF($J${}=25,IF(AND(J{}>4,S{}=20),1,""),IF($J${}=30,IF(AND(J{}>5,S{}=20),1,""),IF($J${}=35,IF(AND(J{}>6,S{}=20),1,""),IF($J${}=40,IF(AND(J{}>7,S{}=20),1,""),IF($J${}=45,IF(AND(J{}>8,S{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

            # Z Score iterasi 1
            ws['AB1'] = 'NAMA SISWA_B'
            ws['AC1'] = 'NOMOR NF_B'
            ws['AD1'] = 'KELAS_B'
            ws['AE1'] = 'NAMA SEKOLAH_B'
            ws['AF1'] = 'LOKASI_B'
            ws['AG1'] = 'MAT_B'
            ws['AH1'] = 'FIS_B'
            ws['AI1'] = 'KIM_B'
            ws['AJ1'] = 'BIO_B'
            ws['AK1'] = 'JML_B'
            ws['AL1'] = 'Z_MAT_B'
            ws['AM1'] = 'Z_FIS_B'
            ws['AN1'] = 'Z_KIM_B'
            ws['AO1'] = 'Z_BIO_B'
            ws['AP1'] = 'S_MAT_B'
            ws['AQ1'] = 'S_FIS_B'
            ws['AR1'] = 'S_KIM_B'
            ws['AS1'] = 'S_BIO_B'
            ws['AT1'] = 'S_JML_B'
            ws['AU1'] = 'RANK NAS._B'
            ws['AV1'] = 'RANK LOK._B'

            ws['AL1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AM1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AN1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AO1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AP1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AQ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AR1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AS1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AT1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AU1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AV1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['AB1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AC1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AD1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AE1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AF1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AG1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AH1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AI1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AJ1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AK1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AL1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AM1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AN1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AO1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AP1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AQ1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AR1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AS1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AT1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AU1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AV1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')

            # tambahan
            ws['AW1'] = 'MAT_20'
            ws['AX1'] = 'FIS_20'
            ws['AY1'] = 'KIM_20'
            ws['AZ1'] = 'BIO_20'
            ws['AW1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AX1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AY1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AZ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AW1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AX1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AY1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AZ1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')

            for row in range(2, q+1):
                # Tambahan
                ws['AB{}'.format(row)] = '=B{}'.format(row)
                ws['AC{}'.format(row)] = '=C{}'.format(row, row)
                ws['AD{}'.format(row)] = '=D{}'.format(row, row)
                ws['AE{}'.format(row)] = '=E{}'.format(row, row)
                ws['AF{}'.format(row)] = '=F{}'.format(row, row)
                ws['AG{}'.format(row)] = '=IF(G{}="","",G{})'.format(row, row)
                ws['AH{}'.format(row)] = '=IF(H{}="","",H{})'.format(row, row)
                ws['AI{}'.format(row)] = '=IF(I{}="","",I{})'.format(row, row)
                ws['AJ{}'.format(row)] = '=IF(J{}="","",J{})'.format(row, row)
                ws['AK{}'.format(row)] = '=IF(K{}="","",K{})'.format(row, row)

                ws['AL{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AG{}="","",(AG{}-AG${})/AG${}),2),"")'.format(row, row, r, s)
                ws['AM{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AH{}="","",(AH{}-AH${})/AH${}),2),"")'.format(row, row, r, s)
                ws['AN{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AI{}="","",(AI{}-AI${})/AI${}),2),"")'.format(row, row, r, s)
                ws['AO{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AJ{}="","",(AJ{}-AJ${})/AJ${}),2),"")'.format(row, row, r, s)

                ws['AP{}'.format(
                    row)] = '=IFERROR(ROUND(IF(G{}="","",IF(70+30*AL{}/$AL${}<20,20,70+30*AL{}/$AL${})),2),"")'.format(row, row, r, row, r)
                ws['AQ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(H{}="","",IF(70+30*AM{}/$AM{}<20,20,70+30*AM{}/$AM${})),2),"")'.format(row, row, r, row, r)
                ws['AR{}'.format(
                    row)] = '=IFERROR(ROUND(IF(I{}="","",IF(70+30*AN{}/$AN${}<20,20,70+30*AN{}/$AN${})),2),"")'.format(row, row, r, row, r)
                ws['AS{}'.format(
                    row)] = '=IFERROR(ROUND(IF(J{}="","",IF(70+30*AO{}/$AO${}<20,20,70+30*AO{}/$AO${})),2),"")'.format(row, row, r, row, r)

                ws['AT{}'.format(row)] = '=IF(SUM(AP{}:AS{})=0,"",SUM(AP{}:AS{}))'.format(
                    row, row, row, row)
                ws['AU{}'.format(row)] = '=IF(AT{}="","",RANK(AT{},$AT$2:$AT${}))'.format(
                    row, row, q)
                ws['AV{}'.format(
                    row)] = '=IF(AU{}="","",COUNTIFS($AF$2:$AF${},AF{},$AU$2:$AU${},"<"&AU{})+1)'.format(row, q, row, q, row)
                # TAMBAHAN
                ws['AW{}'.format(row)] = '=IF($G${}=25,IF(AND(AG{}>4,AP{}=20),1,""),IF($G${}=30,IF(AND(AG{}>5,AP{}=20),1,""),IF($G${}=35,IF(AND(AG{}>6,AP{}=20),1,""),IF($G${}=40,IF(AND(AG{}>7,AP{}=20),1,""),IF($G${}=45,IF(AND(AG{}>8,AP{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['AX{}'.format(row)] = '=IF($H${}=25,IF(AND(AH{}>4,AQ{}=20),1,""),IF($H${}=30,IF(AND(AH{}>5,AQ{}=20),1,""),IF($H${}=35,IF(AND(AH{}>6,AQ{}=20),1,""),IF($H${}=40,IF(AND(AH{}>7,AQ{}=20),1,""),IF($H${}=45,IF(AND(AH{}>8,AQ{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['AY{}'.format(row)] = '=IF($I${}=25,IF(AND(AI{}>4,AR{}=20),1,""),IF($I${}=30,IF(AND(AI{}>5,AR{}=20),1,""),IF($I${}=35,IF(AND(AI{}>6,AR{}=20),1,""),IF($I${}=40,IF(AND(AI{}>7,AR{}=20),1,""),IF($I${}=45,IF(AND(AI{}>8,AR{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['AZ{}'.format(row)] = '=IF($J${}=25,IF(AND(AJ{}>4,AS{}=20),1,""),IF($J${}=30,IF(AND(AJ{}>5,AS{}=20),1,""),IF($J${}=35,IF(AND(AJ{}>6,AS{}=20),1,""),IF($J${}=40,IF(AND(AJ{}>7,AS{}=20),1,""),IF($J${}=45,IF(AND(AJ{}>8,AS{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

            # Z Score iterasi 2
            ws['BB1'] = 'NAMA SISWA_C'
            ws['BC1'] = 'NOMOR NF_c'
            ws['BD1'] = 'KELAS_C'
            ws['BE1'] = 'NAMA SEKOLAH_C'
            ws['BF1'] = 'LOKASI_C'
            ws['BG1'] = 'MAT_C'
            ws['BH1'] = 'FIS_C'
            ws['BI1'] = 'KIM_C'
            ws['BJ1'] = 'BIO_C'
            ws['BK1'] = 'JML_C'
            ws['BL1'] = 'Z_MAT_C'
            ws['BM1'] = 'Z_FIS_C'
            ws['BN1'] = 'Z_KIM_C'
            ws['BO1'] = 'Z_BIO_C'
            ws['BP1'] = 'S_MAT_C'
            ws['BQ1'] = 'S_FIS_C'
            ws['BR1'] = 'S_KIM_C'
            ws['BS1'] = 'S_BIO_C'
            ws['BT1'] = 'S_JML_C'
            ws['BU1'] = 'RANK NAS._C'
            ws['BV1'] = 'RANK LOK._C'

            ws['BL1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BM1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BN1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BO1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BP1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BQ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BR1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BS1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BT1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BU1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BV1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['BB1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BC1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BD1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BE1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BF1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BG1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BH1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BI1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BJ1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BK1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BL1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BM1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BN1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BO1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BP1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BQ1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BR1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BS1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BT1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BU1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BV1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')

            # tambahan
            ws['BW1'] = 'MAT_20_C'
            ws['BX1'] = 'FIS_20_C'
            ws['BY1'] = 'KIM_20_C'
            ws['BZ1'] = 'BIO_20_C'
            ws['BW1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BX1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BY1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BZ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BW1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BX1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BY1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BZ1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')

            for row in range(2, q+1):
                # Tambahan
                ws['BB{}'.format(row)] = '=AB{}'.format(row)
                ws['BC{}'.format(row)] = '=AC{}'.format(row, row)
                ws['BD{}'.format(row)] = '=AD{}'.format(row, row)
                ws['BE{}'.format(row)] = '=AE{}'.format(row, row)
                ws['BF{}'.format(row)] = '=AF{}'.format(row, row)
                ws['BG{}'.format(row)] = '=IF(AG{}="","",AG{})'.format(
                    row, row)
                ws['BH{}'.format(row)] = '=IF(AH{}="","",AH{})'.format(
                    row, row)
                ws['BI{}'.format(row)] = '=IF(AI{}="","",AI{})'.format(
                    row, row)
                ws['BJ{}'.format(row)] = '=IF(AJ{}="","",AJ{})'.format(
                    row, row)
                ws['BK{}'.format(row)] = '=IF(AK{}="","",AK{})'.format(
                    row, row)

                ws['BL{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BG{}="","",(BG{}-BG${})/BG${}),2),"")'.format(row, row, r, s)
                ws['BM{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BH{}="","",(BH{}-BH${})/BH${}),2),"")'.format(row, row, r, s)
                ws['BN{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BI{}="","",(BI{}-BI${})/BI${}),2),"")'.format(row, row, r, s)
                ws['BO{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BJ{}="","",(BJ{}-BJ${})/BJ${}),2),"")'.format(row, row, r, s)

                ws['BP{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BG{}="","",IF(70+30*BL{}/$BL${}<20,20,70+30*BL{}/$BL${})),2),"")'.format(row, row, r, row, r)
                ws['BQ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BH{}="","",IF(70+30*BM{}/$BM{}<20,20,70+30*BM{}/$BM${})),2),"")'.format(row, row, r, row, r)
                ws['BR{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BI{}="","",IF(70+30*BN{}/$BN${}<20,20,70+30*BN{}/$BN${})),2),"")'.format(row, row, r, row, r)
                ws['BS{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BJ{}="","",IF(70+30*BO{}/$BO${}<20,20,70+30*BO{}/$BO${})),2),"")'.format(row, row, r, row, r)

                ws['BT{}'.format(row)] = '=IF(SUM(BP{}:BS{})=0,"",SUM(BP{}:BS{}))'.format(
                    row, row, row, row)
                ws['BU{}'.format(row)] = '=IF(BT{}="","",RANK(BT{},$BT$2:$BT${}))'.format(
                    row, row, q)
                ws['BV{}'.format(
                    row)] = '=IF(BU{}="","",COUNTIFS($BF$2:$BF${},BF{},$BU$2:$BU${},"<"&BU{})+1)'.format(row, q, row, q, row)
                # TAMBAHAN
                ws['BW{}'.format(row)] = '=IF($G${}=25,IF(AND(BG{}>4,BP{}=20),1,""),IF($G${}=30,IF(AND(BG{}>5,BP{}=20),1,""),IF($G${}=35,IF(AND(BG{}>6,BP{}=20),1,""),IF($G${}=40,IF(AND(BG{}>7,BP{}=20),1,""),IF($G${}=45,IF(AND(BG{}>8,BP{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['BX{}'.format(row)] = '=IF($H${}=25,IF(AND(BH{}>4,BQ{}=20),1,""),IF($H${}=30,IF(AND(BH{}>5,BQ{}=20),1,""),IF($H${}=35,IF(AND(BH{}>6,BQ{}=20),1,""),IF($H${}=40,IF(AND(BH{}>7,BQ{}=20),1,""),IF($H${}=45,IF(AND(BH{}>8,BQ{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['BY{}'.format(row)] = '=IF($I${}=25,IF(AND(BI{}>4,BR{}=20),1,""),IF($I${}=30,IF(AND(BI{}>5,BR{}=20),1,""),IF($I${}=35,IF(AND(BI{}>6,BR{}=20),1,""),IF($I${}=40,IF(AND(BI{}>7,BR{}=20),1,""),IF($I${}=45,IF(AND(BI{}>8,BR{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['BZ{}'.format(row)] = '=IF($J${}=25,IF(AND(BJ{}>4,BS{}=20),1,""),IF($J${}=30,IF(AND(BJ{}>5,BS{}=20),1,""),IF($J${}=35,IF(AND(BJ{}>6,BS{}=20),1,""),IF($J${}=40,IF(AND(BJ{}>7,BS{}=20),1,""),IF($J${}=45,IF(AND(BJ{}>8,BS{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

            # Z Score iterasi 3
            ws['CB1'] = 'NAMA SISWA_D'
            ws['CC1'] = 'NOMOR NF_D'
            ws['CD1'] = 'KELAS_D'
            ws['CE1'] = 'NAMA SEKOLAH_D'
            ws['CF1'] = 'LOKASI_D'
            ws['CG1'] = 'MAT_D'
            ws['CH1'] = 'FIS_D'
            ws['CI1'] = 'KIM_D'
            ws['CJ1'] = 'BIO_D'
            ws['CK1'] = 'JML_D'
            ws['CL1'] = 'Z_MAT_D'
            ws['CM1'] = 'Z_FIS_D'
            ws['CN1'] = 'Z_KIM_D'
            ws['CO1'] = 'Z_BIO_D'
            ws['CP1'] = 'S_MAT_D'
            ws['CQ1'] = 'S_FIS_D'
            ws['CR1'] = 'S_KIM_D'
            ws['CS1'] = 'S_BIO_D'
            ws['CT1'] = 'S_JML_D'
            ws['CU1'] = 'RANK NAS._D'
            ws['CV1'] = 'RANK LOK._D'

            ws['CL1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CM1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CN1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CO1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CP1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CQ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CR1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CS1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CT1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CU1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CV1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['CB1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CC1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CD1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CE1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CF1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CG1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CH1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CI1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CJ1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CK1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CL1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CM1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CN1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CO1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CP1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CQ1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CR1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CS1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CT1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CU1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CV1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')

            # tambahan
            ws['CW1'] = 'MAT_20_D'
            ws['CX1'] = 'FIS_20_D'
            ws['CY1'] = 'KIM_20_D'
            ws['CZ1'] = 'BIO_20_D'
            ws['CW1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CX1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CY1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CZ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CW1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CX1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CY1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CZ1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')

            for row in range(2, q+1):
                ws['CB{}'.format(row)] = '=BB{}'.format(row)
                ws['CC{}'.format(row)] = '=BC{}'.format(row, row)
                ws['CD{}'.format(row)] = '=BD{}'.format(row, row)
                ws['CE{}'.format(row)] = '=BE{}'.format(row, row)
                ws['CF{}'.format(row)] = '=BF{}'.format(row, row)
                ws['CG{}'.format(row)] = '=IF(BG{}="","",BG{})'.format(
                    row, row)
                ws['CH{}'.format(row)] = '=IF(BH{}="","",BH{})'.format(
                    row, row)
                ws['CI{}'.format(row)] = '=IF(BI{}="","",BI{})'.format(
                    row, row)
                ws['CJ{}'.format(row)] = '=IF(BJ{}="","",BJ{})'.format(
                    row, row)
                ws['CK{}'.format(row)] = '=IF(BK{}="","",BK{})'.format(
                    row, row)

                ws['CL{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CG{}="","",(CG{}-CG${})/CG${}),2),"")'.format(row, row, r, s)
                ws['CM{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CH{}="","",(CH{}-CH${})/CH${}),2),"")'.format(row, row, r, s)
                ws['CN{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CI{}="","",(CI{}-CI${})/CI${}),2),"")'.format(row, row, r, s)
                ws['CO{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CJ{}="","",(CJ{}-CJ${})/CJ${}),2),"")'.format(row, row, r, s)

                ws['CP{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CG{}="","",IF(70+30*CL{}/$CL${}<20,20,70+30*CL{}/$CL${})),2),"")'.format(row, row, r, row, r)
                ws['CQ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CH{}="","",IF(70+30*CM{}/$CM{}<20,20,70+30*CM{}/$CM${})),2),"")'.format(row, row, r, row, r)
                ws['CR{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CI{}="","",IF(70+30*CN{}/$CN${}<20,20,70+30*CN{}/$CN${})),2),"")'.format(row, row, r, row, r)
                ws['CS{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CJ{}="","",IF(70+30*CO{}/$CO${}<20,20,70+30*CO{}/$CO${})),2),"")'.format(row, row, r, row, r)

                ws['CT{}'.format(row)] = '=IF(SUM(CP{}:CS{})=0,"",SUM(CP{}:CS{}))'.format(
                    row, row, row, row)
                ws['CU{}'.format(row)] = '=IF(CT{}="","",RANK(CT{},$CT$2:$CT${}))'.format(
                    row, row, q)
                ws['CV{}'.format(
                    row)] = '=IF(CU{}="","",COUNTIFS($CF$2:$CF${},CF{},$CU$2:$CU${},"<"&CU{})+1)'.format(row, q, row, q, row)
                # TAMBAHAN
                ws['CW{}'.format(row)] = '=IF($G${}=25,IF(AND(CG{}>4,CP{}=20),1,""),IF($G${}=30,IF(AND(CG{}>5,CP{}=20),1,""),IF($G${}=35,IF(AND(CG{}>6,CP{}=20),1,""),IF($G${}=40,IF(AND(CG{}>7,CP{}=20),1,""),IF($G${}=45,IF(AND(CG{}>8,CP{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['CX{}'.format(row)] = '=IF($H${}=25,IF(AND(CH{}>4,CQ{}=20),1,""),IF($H${}=30,IF(AND(CH{}>5,CQ{}=20),1,""),IF($H${}=35,IF(AND(CH{}>6,CQ{}=20),1,""),IF($H${}=40,IF(AND(CH{}>7,CQ{}=20),1,""),IF($H${}=45,IF(AND(CH{}>8,CQ{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['CY{}'.format(row)] = '=IF($I${}=25,IF(AND(CI{}>4,CR{}=20),1,""),IF($I${}=30,IF(AND(CI{}>5,CR{}=20),1,""),IF($I${}=35,IF(AND(CI{}>6,CR{}=20),1,""),IF($I${}=40,IF(AND(CI{}>7,CR{}=20),1,""),IF($I${}=45,IF(AND(CI{}>8,CR{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['CZ{}'.format(row)] = '=IF($J${}=25,IF(AND(CJ{}>4,CS{}=20),1,""),IF($J${}=30,IF(AND(CJ{}>5,CS{}=20),1,""),IF($J${}=35,IF(AND(CJ{}>6,CS{}=20),1,""),IF($J${}=40,IF(AND(CJ{}>7,CS{}=20),1,""),IF($J${}=45,IF(AND(CJ{}>8,CS{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

            # Z Score iterasi 4
            ws['DB1'] = 'NAMA SISWA_E'
            ws['DC1'] = 'NOMOR NF_E'
            ws['DD1'] = 'KELAS_E'
            ws['DE1'] = 'NAMA SEKOLAH_E'
            ws['DF1'] = 'LOKASI_E'
            ws['DG1'] = 'MAT_E'
            ws['DH1'] = 'FIS_E'
            ws['DI1'] = 'KIM_E'
            ws['DJ1'] = 'BIO_E'
            ws['DK1'] = 'JML_E'
            ws['DL1'] = 'Z_MAT_E'
            ws['DM1'] = 'Z_FIS_E'
            ws['DN1'] = 'Z_KIM_E'
            ws['DO1'] = 'Z_BIO_E'
            ws['DP1'] = 'S_MAT_E'
            ws['DQ1'] = 'S_FIS_E'
            ws['DR1'] = 'S_KIM_E'
            ws['DS1'] = 'S_BIO_E'
            ws['DT1'] = 'S_JML_E'
            ws['DU1'] = 'RANK NAS._E'
            ws['DV1'] = 'RANK LOK._E'

            ws['DL1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DM1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DN1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DO1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DP1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DQ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DR1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DS1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DT1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DU1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DV1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['DB1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DC1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DD1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DE1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DF1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DG1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DH1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DI1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DJ1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DK1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DL1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DM1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DN1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DO1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DP1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DQ1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DR1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DS1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DT1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DU1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DV1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')

            # tambahan
            ws['DW1'] = 'MAT_20'
            ws['DX1'] = 'FIS_20'
            ws['DY1'] = 'KIM_20'
            ws['DZ1'] = 'BIO_20'
            ws['DW1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DX1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DY1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DZ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DW1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DX1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DY1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DZ1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')

            for row in range(2, q+1):
                # Tambahan
                ws['DB{}'.format(row)] = '=CB{}'.format(row)
                ws['DC{}'.format(row)] = '=CC{}'.format(row, row)
                ws['DD{}'.format(row)] = '=CD{}'.format(row, row)
                ws['DE{}'.format(row)] = '=CE{}'.format(row, row)
                ws['DF{}'.format(row)] = '=CF{}'.format(row, row)
                ws['DG{}'.format(row)] = '=IF(CG{}="","",CG{})'.format(
                    row, row)
                ws['DH{}'.format(row)] = '=IF(CH{}="","",CH{})'.format(
                    row, row)
                ws['DI{}'.format(row)] = '=IF(CI{}="","",CI{})'.format(
                    row, row)
                ws['DJ{}'.format(row)] = '=IF(CJ{}="","",CJ{})'.format(
                    row, row)
                ws['DK{}'.format(row)] = '=IF(CK{}="","",CK{})'.format(
                    row, row)

                ws['DL{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DG{}="","",(DG{}-DG${})/DG${}),2),"")'.format(row, row, r, s)
                ws['DM{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DH{}="","",(DH{}-DH${})/DH${}),2),"")'.format(row, row, r, s)
                ws['DN{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DI{}="","",(DI{}-DI${})/DI${}),2),"")'.format(row, row, r, s)
                ws['DO{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DJ{}="","",(DJ{}-DJ${})/DJ${}),2),"")'.format(row, row, r, s)

                ws['DP{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DG{}="","",IF(70+30*DL{}/$DL${}<20,20,70+30*DL{}/$DL${})),2),"")'.format(row, row, r, row, r)
                ws['DQ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DH{}="","",IF(70+30*DM{}/$DM{}<20,20,70+30*DM{}/$DM${})),2),"")'.format(row, row, r, row, r)
                ws['DR{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DI{}="","",IF(70+30*DN{}/$DN${}<20,20,70+30*DN{}/$DN${})),2),"")'.format(row, row, r, row, r)
                ws['DS{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DJ{}="","",IF(70+30*DO{}/$DO${}<20,20,70+30*DO{}/$DO${})),2),"")'.format(row, row, r, row, r)

                ws['DT{}'.format(row)] = '=IF(SUM(DP{}:DS{})=0,"",SUM(DP{}:DS{}))'.format(
                    row, row, row, row)
                ws['DU{}'.format(row)] = '=IF(DT{}="","",RANK(DT{},$DT$2:$DT${}))'.format(
                    row, row, q)
                ws['DV{}'.format(
                    row)] = '=IF(DU{}="","",COUNTIFS($DF$2:$DF${},DF{},$DU$2:$DU${},"<"&DU{})+1)'.format(row, q, row, q, row)
                # TAMBAHAN
                ws['DW{}'.format(row)] = '=IF($G${}=25,IF(AND(DG{}>4,DP{}=20),1,""),IF($G${}=30,IF(AND(DG{}>5,DP{}=20),1,""),IF($G${}=35,IF(AND(DG{}>6,DP{}=20),1,""),IF($G${}=40,IF(AND(DG{}>7,DP{}=20),1,""),IF($G${}=45,IF(AND(DG{}>8,DP{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['DX{}'.format(row)] = '=IF($H${}=25,IF(AND(DH{}>4,DQ{}=20),1,""),IF($H${}=30,IF(AND(DH{}>5,DQ{}=20),1,""),IF($H${}=35,IF(AND(DH{}>6,DQ{}=20),1,""),IF($H${}=40,IF(AND(DH{}>7,DQ{}=20),1,""),IF($H${}=45,IF(AND(DH{}>8,DQ{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['DY{}'.format(row)] = '=IF($I${}=25,IF(AND(DI{}>4,DR{}=20),1,""),IF($I${}=30,IF(AND(DI{}>5,DR{}=20),1,""),IF($I${}=35,IF(AND(DI{}>6,DR{}=20),1,""),IF($I${}=40,IF(AND(DI{}>7,DR{}=20),1,""),IF($I${}=45,IF(AND(DI{}>8,DR{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['DZ{}'.format(row)] = '=IF($J${}=25,IF(AND(DJ{}>4,DS{}=20),1,""),IF($J${}=30,IF(AND(DJ{}>5,DS{}=20),1,""),IF($J${}=35,IF(AND(DJ{}>6,DS{}=20),1,""),IF($J${}=40,IF(AND(DJ{}>7,DS{}=20),1,""),IF($J${}=45,IF(AND(DJ{}>8,DS{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

            # Z Score iterasi 5
            ws['EB1'] = 'NAMA SISWA'
            ws['EC1'] = 'NOMOR NF'
            ws['ED1'] = 'KELAS'
            ws['EE1'] = 'NAMA SEKOLAH'
            ws['EF1'] = 'LOKASI'
            ws['EG1'] = 'MAT'
            ws['EH1'] = 'FIS'
            ws['EI1'] = 'KIM'
            ws['EJ1'] = 'BIO'
            ws['EK1'] = 'JML'
            ws['EL1'] = 'Z_MAT'
            ws['EM1'] = 'Z_FIS'
            ws['EN1'] = 'Z_KIM'
            ws['EO1'] = 'Z_BIO'
            ws['EP1'] = 'S_MAT'
            ws['EQ1'] = 'S_FIS'
            ws['ER1'] = 'S_KIM'
            ws['ES1'] = 'S_BIO'
            ws['ET1'] = 'S_JML'
            ws['EU1'] = 'RANK NAS.'
            ws['EV1'] = 'RANK LOK.'

            ws['EL1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EM1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EN1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EO1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EP1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EQ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['ER1'].font = Font(bold=False, name='Calibri', size=11)
            ws['ES1'].font = Font(bold=False, name='Calibri', size=11)
            ws['ET1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EU1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EV1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['EB1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EC1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['ED1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EE1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EF1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EG1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EH1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EI1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EJ1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EK1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EL1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EM1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EN1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EO1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EP1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EQ1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['ER1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['ES1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['ET1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EU1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EV1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')

            # tambahan
            ws['EW1'] = 'MAT_20'
            ws['EX1'] = 'FIS_20'
            ws['EY1'] = 'KIM_20'
            ws['EZ1'] = 'BIO_20'
            ws['EW1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EX1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EY1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EZ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EW1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EX1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EY1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EZ1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')

            for row in range(2, q+1):
                # Tambahan
                ws['EB{}'.format(row)] = '=DB{}'.format(row)
                ws['EC{}'.format(row)] = '=DC{}'.format(row, row)
                ws['ED{}'.format(row)] = '=DD{}'.format(row, row)
                ws['EE{}'.format(row)] = '=DE{}'.format(row, row)
                ws['EF{}'.format(row)] = '=DF{}'.format(row, row)
                ws['EG{}'.format(row)] = '=IF(DG{}="","",DG{})'.format(
                    row, row)
                ws['EH{}'.format(row)] = '=IF(DH{}="","",DH{})'.format(
                    row, row)
                ws['EI{}'.format(row)] = '=IF(DI{}="","",DI{})'.format(
                    row, row)
                ws['EJ{}'.format(row)] = '=IF(DJ{}="","",DJ{})'.format(
                    row, row)
                ws['EK{}'.format(row)] = '=IF(DK{}="","",DK{})'.format(
                    row, row)

                ws['EL{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EG{}="","",(EG{}-EG${})/EG${}),2),"")'.format(row, row, r, s)
                ws['EM{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EH{}="","",(EH{}-EH${})/EH${}),2),"")'.format(row, row, r, s)
                ws['EN{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EI{}="","",(EI{}-EI${})/EI${}),2),"")'.format(row, row, r, s)
                ws['EO{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EJ{}="","",(EJ{}-EJ${})/EJ${}),2),"")'.format(row, row, r, s)

                ws['EP{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EG{}="","",IF(70+30*EL{}/$EL${}<20,20,70+30*EL{}/$EL${})),2),"")'.format(row, row, r, row, r)
                ws['EQ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EH{}="","",IF(70+30*EM{}/$EM{}<20,20,70+30*EM{}/$EM${})),2),"")'.format(row, row, r, row, r)
                ws['ER{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EI{}="","",IF(70+30*EN{}/$EN${}<20,20,70+30*EN{}/$EN${})),2),"")'.format(row, row, r, row, r)
                ws['ES{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EJ{}="","",IF(70+30*EO{}/$EO${}<20,20,70+30*EO{}/$EO${})),2),"")'.format(row, row, r, row, r)

                ws['ET{}'.format(row)] = '=IF(SUM(EP{}:ES{})=0,"",SUM(EP{}:ES{}))'.format(
                    row, row, row, row)
                ws['EU{}'.format(row)] = '=IF(ET{}="","",RANK(ET{},$ET$2:$ET${}))'.format(
                    row, row, q)
                ws['EV{}'.format(
                    row)] = '=IF(EU{}="","",COUNTIFS($EF$2:$EF${},EF{},$EU$2:$EU${},"<"&EU{})+1)'.format(row, q, row, q, row)
                # TAMBAHAN
                ws['EW{}'.format(row)] = '=IF($G${}=25,IF(AND(EG{}>4,EP{}=20),1,""),IF($G${}=30,IF(AND(EG{}>5,EP{}=20),1,""),IF($G${}=35,IF(AND(EG{}>6,EP{}=20),1,""),IF($G${}=40,IF(AND(EG{}>7,EP{}=20),1,""),IF($G${}=45,IF(AND(EG{}>8,EP{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['EX{}'.format(row)] = '=IF($H${}=25,IF(AND(EH{}>4,EQ{}=20),1,""),IF($H${}=30,IF(AND(EH{}>5,EQ{}=20),1,""),IF($H${}=35,IF(AND(EH{}>6,EQ{}=20),1,""),IF($H${}=40,IF(AND(EH{}>7,EQ{}=20),1,""),IF($H${}=45,IF(AND(EH{}>8,EQ{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['EY{}'.format(row)] = '=IF($I${}=25,IF(AND(EI{}>4,ER{}=20),1,""),IF($I${}=30,IF(AND(EI{}>5,ER{}=20),1,""),IF($I${}=35,IF(AND(EI{}>6,ER{}=20),1,""),IF($I${}=40,IF(AND(EI{}>7,ER{}=20),1,""),IF($I${}=45,IF(AND(EI{}>8,ER{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['EZ{}'.format(row)] = '=IF($J${}=25,IF(AND(EJ{}>4,ES{}=20),1,""),IF($J${}=30,IF(AND(EJ{}>5,ES{}=20),1,""),IF($J${}=35,IF(AND(EJ{}>6,ES{}=20),1,""),IF($J${}=40,IF(AND(EJ{}>7,ES{}=20),1,""),IF($J${}=45,IF(AND(EJ{}>8,ES{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

            # Mengubah 'KELAS' sesuai dengan nilai yang dipilih dari selectbox 'KELAS'
            kelas = KELAS.lower().replace(" ", "")
            semester = SEMESTER.lower()
            tahun = TAHUN.replace("-", "")
            penilaian = PENILAIAN.lower()
            kurikulum = KURIKULUM.lower()

            path_file = f"{kelas}_{penilaian}_{semester}_{kurikulum}_{tahun}_nilai_std.xlsx"

            # Simpan file ke direktori temporer
            temp_dir = tempfile.gettempdir()
            file_path = temp_dir + '/' + path_file
            wb.save(file_path)

            st.success(
                "File siap diunduh!")

            # Tombol unduh file
            with open(file_path, "rb") as f:
                bytes_data = f.read()
            st.download_button(label="Unduh File", data=bytes_data,
                               file_name=path_file)

            st.warning(
                "Buka file unduhan, klik 'Enable Editing' dan 'Save'")
    if selected_file == "Nilai Std. PPLS IPS":
        # menghilangkan hamburger
        st.markdown("""
        <style>
        .css-1rs6os.edgvbvh3
        {
            visibility:hidden;
        }
        .css-1lsmgbg.egzxvld0
        {
            visibility:hidden;
        }
        </style>
        """, unsafe_allow_html=True)

        image = Image.open('logo resmi nf resize.png')
        st.image(image)

        st.title("Olah Nilai Standar PPLS")
        st.header("PPLS")

        col6 = st.container()

        with col6:
            KELAS = st.selectbox(
                "KELAS",
                ("--Pilih Kelas--", "PPLS IPS"))

        col7 = st.container()

        with col7:
            SEMESTER = st.selectbox(
                "SEMESTER",
                ("--Pilih Semester--", "SEMESTER 1", "SEMESTER 2"))

        col8 = st.container()

        with col8:
            PENILAIAN = st.selectbox(
                "PENILAIAN",
                ("--Pilih Penilaian--", "PENILAIAN TENGAH SEMESTER", "PENILAIAN AKHIR TAHUN"))

        col9 = st.container()

        with col9:
            KURIKULUM = st.selectbox(
                "KURIKULUM",
                ("--Pilih Kurikulum--", "PPLS"))

        TAHUN = st.text_input("Masukkan Tahun Ajaran",
                              placeholder="contoh: 2022-2023")

        col1, col2, col3, col4 = st.columns(4)

        with col1:
            GEO = st.selectbox(
                "JML. SOAL GEO.",
                (15, 20, 25, 30, 35, 40, 45, 50))

        with col2:
            EKO = st.selectbox(
                "JML. SOAL EKO.",
                (15, 20, 25, 30, 35, 40, 45, 50))

        with col3:
            SEJ = st.selectbox(
                "JML. SOAL SEJ.",
                (15, 20, 25, 30, 35, 40, 45, 50))

        with col4:
            SOS = st.selectbox(
                "JML. SOAL SOS.",
                (15, 20, 25, 30, 35, 40, 45, 50))

        JML_SOAL_GEO = GEO
        JML_SOAL_EKO = EKO
        JML_SOAL_SEJ = SEJ
        JML_SOAL_SOS = SOS

        uploaded_file = st.file_uploader(
            'Letakkan file excel', type='xlsx')

        if uploaded_file is not None:
            wb = openpyxl.load_workbook(uploaded_file)
            ws = wb['Sheet1']

            q = len(ws['K'])
            r = len(ws['K'])+2
            s = len(ws['K'])+3
            t = len(ws['K'])+4
            u = len(ws['K'])+5
            v = len(ws['K'])+6
            w = len(ws['K'])+7
            x = len(ws['K'])+8

            ws['G{}'.format(r)] = "=ROUND(AVERAGE(G2:G{}),2)".format(q)
            ws['H{}'.format(r)] = "=ROUND(AVERAGE(H2:H{}),2)".format(q)
            ws['I{}'.format(r)] = "=ROUND(AVERAGE(I2:I{}),2)".format(q)
            ws['J{}'.format(r)] = "=ROUND(AVERAGE(J2:J{}),2)".format(q)
            ws['K{}'.format(r)] = "=ROUND(AVERAGE(K2:K{}),2)".format(q)
            ws['G{}'.format(s)] = "=STDEV(G2:G{})".format(q)
            ws['H{}'.format(s)] = "=STDEV(H2:H{})".format(q)
            ws['I{}'.format(s)] = "=STDEV(I2:I{})".format(q)
            ws['J{}'.format(s)] = "=STDEV(J2:J{})".format(q)
            ws['G{}'.format(t)] = "=MAX(G2:G{})".format(q)
            ws['H{}'.format(t)] = "=MAX(H2:H{})".format(q)
            ws['I{}'.format(t)] = "=MAX(I2:I{})".format(q)
            ws['J{}'.format(t)] = "=MAX(J2:J{})".format(q)
            ws['K{}'.format(t)] = "=MAX(K2:K{})".format(q)
            ws['L{}'.format(r)] = "=MAX(L2:L{})".format(q)
            ws['M{}'.format(r)] = "=MAX(M2:M{})".format(q)
            ws['N{}'.format(r)] = "=MAX(N2:N{})".format(q)
            ws['O{}'.format(r)] = "=MAX(O2:O{})".format(q)
            ws['P{}'.format(r)] = "=MAX(P2:P{})".format(q)
            ws['Q{}'.format(r)] = "=MAX(Q2:Q{})".format(q)
            ws['R{}'.format(r)] = "=MAX(R2:R{})".format(q)
            ws['S{}'.format(r)] = "=MAX(S2:S{})".format(q)
            ws['T{}'.format(r)] = "=ROUND(MAX(T2:T{}),2)".format(q)
            ws['U{}'.format(r)] = "=MAX(U2:U{})".format(q)
            ws['G{}'.format(u)] = "=MIN(G2:G{})".format(q)
            ws['H{}'.format(u)] = "=MIN(H2:H{})".format(q)
            ws['I{}'.format(u)] = "=MIN(I2:I{})".format(q)
            ws['J{}'.format(u)] = "=MIN(J2:J{})".format(q)
            ws['K{}'.format(u)] = "=MIN(K2:K{})".format(q)
            ws['P{}'.format(s)] = "=MIN(P2:P{})".format(q)
            ws['Q{}'.format(s)] = "=MIN(Q2:R{})".format(q)
            ws['R{}'.format(s)] = "=MIN(R2:S{})".format(q)
            ws['S{}'.format(s)] = "=MIN(S2:T{})".format(q)
            ws['T{}'.format(s)] = "=MIN(T2:T{})".format(q)
            ws['P{}'.format(t)] = "=ROUND(AVERAGE(P2:P{}),2)".format(q)
            ws['Q{}'.format(t)] = "=ROUND(AVERAGE(Q2:Q{}),2)".format(q)
            ws['R{}'.format(t)] = "=ROUND(AVERAGE(R2:R{}),2)".format(q)
            ws['S{}'.format(t)] = "=ROUND(AVERAGE(S2:S{}),2)".format(q)
            ws['T{}'.format(t)] = "=ROUND(AVERAGE(T2:T{}),2)".format(q)
            ws['W{}'.format(r)] = "=SUM(W2:W{})".format(q)
            ws['X{}'.format(r)] = "=SUM(X2:X{})".format(q)
            ws['Y{}'.format(r)] = "=SUM(Y2:Y{})".format(q)
            ws['Z{}'.format(r)] = "=SUM(Z2:Z{})".format(q)

            # new
            # iterasi 1 rata-rata - 1

            # MAPEL NORMAL
            ws['AG{}'.format(r)] = "=IF($W${}=0,$G${},$G${}-1)".format(r, r, r)
            ws['AG{}'.format(s)] = "=STDEV(AG2:AG{})".format(q)
            ws['AG{}'.format(t)] = "=MAX(AG2:AG{})".format(q)
            ws['AG{}'.format(u)] = "=MIN(AG2:AG{})".format(q)
            ws['AH{}'.format(r)] = "=IF($X${}=0,$H${},$H${}-1)".format(r, r, r)
            ws['AH{}'.format(s)] = "=STDEV(AH2:AH{})".format(q)
            ws['AH{}'.format(t)] = "=MAX(AH2:AH{})".format(q)
            ws['AH{}'.format(u)] = "=MIN(AH2:AH{})".format(q)
            ws['AI{}'.format(r)] = "=IF($Y${}=0,$I${},$I${}-1)".format(r, r, r)
            ws['AI{}'.format(s)] = "=STDEV(AI2:AI{})".format(q)
            ws['AI{}'.format(t)] = "=MAX(AI2:AI{})".format(q)
            ws['AI{}'.format(u)] = "=MIN(AI2:AI{})".format(q)
            ws['AJ{}'.format(r)] = "=IF($Z${}=0,$J${},$J${}-1)".format(r, r, r)
            ws['AJ{}'.format(s)] = "=STDEV(AJ2:AJ{})".format(q)
            ws['AJ{}'.format(t)] = "=MAX(AJ2:AJ{})".format(q)
            ws['AJ{}'.format(u)] = "=MIN(AJ2:AJ{})".format(q)

            # JUMLAH MAPEL NORMAL
            ws['AK{}'.format(r)] = "=ROUND(AVERAGE(AK2:AK{}),2)".format(q)
            ws['AK{}'.format(t)] = "=MAX(AK2:AK{})".format(q)
            ws['AK{}'.format(u)] = "=MIN(AK2:AK{})".format(q)

            # Z SCORE
            ws['AL{}'.format(r)] = "=MAX(AL2:AL{})".format(q)
            ws['AM{}'.format(r)] = "=MAX(AM2:AM{})".format(q)
            ws['AN{}'.format(r)] = "=MAX(AN2:AN{})".format(q)
            ws['AO{}'.format(r)] = "=MAX(AO2:AO{})".format(q)

            # NILAI STANDAR
            ws['AP{}'.format(r)] = "=MAX(AP2:AP{})".format(q)
            ws['AP{}'.format(s)] = "=MIN(AP2:AP{})".format(q)
            ws['AP{}'.format(t)] = "=ROUND(AVERAGE(AP2:AP{}),2)".format(q)
            ws['AQ{}'.format(r)] = "=MAX(AQ2:AQ{})".format(q)
            ws['AQ{}'.format(s)] = "=MIN(AQ2:AQ{})".format(q)
            ws['AQ{}'.format(t)] = "=ROUND(AVERAGE(AQ2:AQ{}),2)".format(q)
            ws['AR{}'.format(r)] = "=MAX(AR2:AR{})".format(q)
            ws['AR{}'.format(s)] = "=MIN(AR2:AR{})".format(q)
            ws['AR{}'.format(t)] = "=ROUND(AVERAGE(AR2:AR{}),2)".format(q)
            ws['AS{}'.format(r)] = "=MAX(AS2:AS{})".format(q)
            ws['AS{}'.format(s)] = "=MIN(AS2:AS{})".format(q)
            ws['AS{}'.format(t)] = "=ROUND(AVERAGE(AS2:AS{}),2)".format(q)
            ws['AT{}'.format(r)] = "=MAX(AT2:AT{})".format(q)
            ws['AT{}'.format(s)] = "=MIN(AT2:AT{})".format(q)
            ws['AT{}'.format(t)] = "=ROUND(AVERAGE(AT2:AT{}),2)".format(q)

            # INISIASI MAPEL
            ws['AW{}'.format(r)] = "=SUM(AW2:AW{})".format(q)
            ws['AX{}'.format(r)] = "=SUM(AX2:AX{})".format(q)
            ws['AY{}'.format(r)] = "=SUM(AY2:AY{})".format(q)
            ws['AZ{}'.format(r)] = "=SUM(AZ2:AZ{})".format(q)

            # iterasi 2 rata-rata - 1
            # MAPEL NORMAL
            ws['BG{}'.format(
                r)] = "=IF($AW${}=0,$AG${},$AG${}-1)".format(r, r, r)
            ws['BG{}'.format(s)] = "=STDEV(BG2:BG{})".format(q)
            ws['BG{}'.format(t)] = "=MAX(BG2:BG{})".format(q)
            ws['BG{}'.format(u)] = "=MIN(BG2:BG{})".format(q)
            ws['BH{}'.format(
                r)] = "=IF($AX${}=0,$AH${},$AH${}-1)".format(r, r, r)
            ws['BH{}'.format(s)] = "=STDEV(BH2:BH{})".format(q)
            ws['BH{}'.format(t)] = "=MAX(BH2:BH{})".format(q)
            ws['BH{}'.format(u)] = "=MIN(BH2:BH{})".format(q)
            ws['BI{}'.format(
                r)] = "=IF($AY${}=0,$AI${},$AI${}-1)".format(r, r, r)
            ws['BI{}'.format(s)] = "=STDEV(BI2:BI{})".format(q)
            ws['BI{}'.format(t)] = "=MAX(BI2:BI{})".format(q)
            ws['BI{}'.format(u)] = "=MIN(BI2:BI{})".format(q)
            ws['BJ{}'.format(
                r)] = "=IF($AZ${}=0,$AJ${},$AJ${}-1)".format(r, r, r)
            ws['BJ{}'.format(s)] = "=STDEV(BJ2:BJ{})".format(q)
            ws['BJ{}'.format(t)] = "=MAX(BJ2:BJ{})".format(q)
            ws['BJ{}'.format(u)] = "=MIN(BJ2:BJ{})".format(q)

            # JUMLAH MAPEL NORMAL
            ws['BK{}'.format(r)] = "=ROUND(AVERAGE(BK2:BK{}),2)".format(q)
            ws['BK{}'.format(t)] = "=MAX(BK2:BK{})".format(q)
            ws['BK{}'.format(u)] = "=MIN(BK2:BK{})".format(q)

            # Z SCORE
            ws['BL{}'.format(r)] = "=MAX(BL2:BL{})".format(q)
            ws['BM{}'.format(r)] = "=MAX(BM2:BM{})".format(q)
            ws['BN{}'.format(r)] = "=MAX(BN2:BN{})".format(q)
            ws['BO{}'.format(r)] = "=MAX(BO2:BO{})".format(q)

            # NILAI STANDAR
            ws['BP{}'.format(r)] = "=MAX(BP2:BP{})".format(q)
            ws['BP{}'.format(s)] = "=MIN(BP2:BP{})".format(q)
            ws['BP{}'.format(t)] = "=ROUND(AVERAGE(BP2:BP{}),2)".format(q)
            ws['BQ{}'.format(r)] = "=MAX(BQ2:BQ{})".format(q)
            ws['BQ{}'.format(s)] = "=MIN(BQ2:BQ{})".format(q)
            ws['BQ{}'.format(t)] = "=ROUND(AVERAGE(BQ2:BQ{}),2)".format(q)
            ws['BR{}'.format(r)] = "=MAX(BR2:BR{})".format(q)
            ws['BR{}'.format(s)] = "=MIN(BR2:BR{})".format(q)
            ws['BR{}'.format(t)] = "=ROUND(AVERAGE(BR2:BR{}),2)".format(q)
            ws['BS{}'.format(r)] = "=MAX(BS2:BS{})".format(q)
            ws['BS{}'.format(s)] = "=MIN(BS2:BS{})".format(q)
            ws['BS{}'.format(t)] = "=ROUND(AVERAGE(BS2:BS{}),2)".format(q)
            ws['BT{}'.format(r)] = "=MAX(BT2:BT{})".format(q)
            ws['BT{}'.format(s)] = "=MIN(BT2:BT{})".format(q)
            ws['BT{}'.format(t)] = "=ROUND(AVERAGE(BT2:BT{}),2)".format(q)

            # INISIASI MAPEL
            ws['BW{}'.format(r)] = "=SUM(BW2:BW{})".format(q)
            ws['BX{}'.format(r)] = "=SUM(BX2:BX{})".format(q)
            ws['BY{}'.format(r)] = "=SUM(BY2:BY{})".format(q)
            ws['BZ{}'.format(r)] = "=SUM(BZ2:BZ{})".format(q)

            # iterasi 3 rata-rata - 1
            # MAPEL NORMAL
            ws['CG{}'.format(
                r)] = "=IF($BW${}=0,$BG${},$BG${}-1)".format(r, r, r)
            ws['CG{}'.format(s)] = "=STDEV(CG2:CG{})".format(q)
            ws['CG{}'.format(t)] = "=MAX(CG2:CG{})".format(q)
            ws['CG{}'.format(u)] = "=MIN(CG2:CG{})".format(q)
            ws['CH{}'.format(
                r)] = "=IF($BX${}=0,$BH${},$BH${}-1)".format(r, r, r)
            ws['CH{}'.format(s)] = "=STDEV(CH2:CH{})".format(q)
            ws['CH{}'.format(t)] = "=MAX(CH2:CH{})".format(q)
            ws['CH{}'.format(u)] = "=MIN(CH2:CH{})".format(q)
            ws['CI{}'.format(
                r)] = "=IF($BY${}=0,$BI${},$BI${}-1)".format(r, r, r)
            ws['CI{}'.format(s)] = "=STDEV(CI2:CI{})".format(q)
            ws['CI{}'.format(t)] = "=MAX(CI2:CI{})".format(q)
            ws['CI{}'.format(u)] = "=MIN(CI2:CI{})".format(q)
            ws['CJ{}'.format(
                r)] = "=IF($BZ${}=0,$BJ${},$BJ${}-1)".format(r, r, r)
            ws['CJ{}'.format(s)] = "=STDEV(CJ2:CJ{})".format(q)
            ws['CJ{}'.format(t)] = "=MAX(CJ2:CJ{})".format(q)
            ws['CJ{}'.format(u)] = "=MIN(CJ2:CJ{})".format(q)

            # JUMLAH MAPEL NORMAL
            ws['CK{}'.format(r)] = "=ROUND(AVERAGE(CK2:CK{}),2)".format(q)
            ws['CK{}'.format(t)] = "=MAX(CK2:CK{})".format(q)
            ws['CK{}'.format(u)] = "=MIN(CK2:CK{})".format(q)

            # Z SCORE
            ws['CL{}'.format(r)] = "=MAX(CL2:CL{})".format(q)
            ws['CM{}'.format(r)] = "=MAX(CM2:CM{})".format(q)
            ws['CN{}'.format(r)] = "=MAX(CN2:CN{})".format(q)
            ws['CO{}'.format(r)] = "=MAX(CO2:CO{})".format(q)

            # NILAI STANDAR
            ws['CP{}'.format(r)] = "=MAX(CP2:CP{})".format(q)
            ws['CP{}'.format(s)] = "=MIN(CP2:CP{})".format(q)
            ws['CP{}'.format(t)] = "=ROUND(AVERAGE(CP2:CP{}),2)".format(q)
            ws['CQ{}'.format(r)] = "=MAX(CQ2:CQ{})".format(q)
            ws['CQ{}'.format(s)] = "=MIN(CQ2:CQ{})".format(q)
            ws['CQ{}'.format(t)] = "=ROUND(AVERAGE(CQ2:CQ{}),2)".format(q)
            ws['CR{}'.format(r)] = "=MAX(CR2:CR{})".format(q)
            ws['CR{}'.format(s)] = "=MIN(CR2:CR{})".format(q)
            ws['CR{}'.format(t)] = "=ROUND(AVERAGE(CR2:CR{}),2)".format(q)
            ws['CS{}'.format(r)] = "=MAX(CS2:CS{})".format(q)
            ws['CS{}'.format(s)] = "=MIN(CS2:CS{})".format(q)
            ws['CS{}'.format(t)] = "=ROUND(AVERAGE(CS2:CS{}),2)".format(q)
            ws['CT{}'.format(r)] = "=MAX(CT2:CT{})".format(q)
            ws['CT{}'.format(s)] = "=MIN(CT2:CT{})".format(q)
            ws['CT{}'.format(t)] = "=ROUND(AVERAGE(CT2:CT{}),2)".format(q)

            # INISIASI MAPEL
            ws['CW{}'.format(r)] = "=SUM(CW2:CW{})".format(q)
            ws['CX{}'.format(r)] = "=SUM(CX2:CX{})".format(q)
            ws['CY{}'.format(r)] = "=SUM(CY2:CY{})".format(q)
            ws['CZ{}'.format(r)] = "=SUM(CZ2:CZ{})".format(q)

            # iterasi 4 rata-rata - 1
            # MAPEL NORMAL
            ws['DG{}'.format(
                r)] = "=IF($CW${}=0,$CG${},$CG${}-1)".format(r, r, r)
            ws['DG{}'.format(s)] = "=STDEV(DG2:DG{})".format(q)
            ws['DG{}'.format(t)] = "=MAX(DG2:DG{})".format(q)
            ws['DG{}'.format(u)] = "=MIN(DG2:DG{})".format(q)
            ws['DH{}'.format(
                r)] = "=IF($CX${}=0,$CH${},$CH${}-1)".format(r, r, r)
            ws['DH{}'.format(s)] = "=STDEV(DH2:DH{})".format(q)
            ws['DH{}'.format(t)] = "=MAX(DH2:DH{})".format(q)
            ws['DH{}'.format(u)] = "=MIN(DH2:DH{})".format(q)
            ws['DI{}'.format(
                r)] = "=IF($CY${}=0,$CI${},$CI${}-1)".format(r, r, r)
            ws['DI{}'.format(s)] = "=STDEV(DI2:DI{})".format(q)
            ws['DI{}'.format(t)] = "=MAX(DI2:DI{})".format(q)
            ws['DI{}'.format(u)] = "=MIN(DI2:DI{})".format(q)
            ws['DJ{}'.format(
                r)] = "=IF($CZ${}=0,$CJ${},$CJ${}-1)".format(r, r, r)
            ws['DJ{}'.format(s)] = "=STDEV(DJ2:DJ{})".format(q)
            ws['DJ{}'.format(t)] = "=MAX(DJ2:DJ{})".format(q)
            ws['DJ{}'.format(u)] = "=MIN(DJ2:DJ{})".format(q)

            # JUMLAH MAPEL NORMAL
            ws['DK{}'.format(r)] = "=ROUND(AVERAGE(DK2:DK{}),2)".format(q)
            ws['DK{}'.format(t)] = "=MAX(DK2:DK{})".format(q)
            ws['DK{}'.format(u)] = "=MIN(DK2:DK{})".format(q)

            # Z SCORE
            ws['DL{}'.format(r)] = "=MAX(DL2:DL{})".format(q)
            ws['DM{}'.format(r)] = "=MAX(DM2:DM{})".format(q)
            ws['DN{}'.format(r)] = "=MAX(DN2:DN{})".format(q)
            ws['DO{}'.format(r)] = "=MAX(DO2:DO{})".format(q)

            # NILAI STANDAR
            ws['DP{}'.format(r)] = "=MAX(DP2:DP{})".format(q)
            ws['DP{}'.format(s)] = "=MIN(DP2:DP{})".format(q)
            ws['DP{}'.format(t)] = "=ROUND(AVERAGE(DP2:DP{}),2)".format(q)
            ws['DQ{}'.format(r)] = "=MAX(DQ2:DQ{})".format(q)
            ws['DQ{}'.format(s)] = "=MIN(DQ2:DQ{})".format(q)
            ws['DQ{}'.format(t)] = "=ROUND(AVERAGE(DQ2:DQ{}),2)".format(q)
            ws['DR{}'.format(r)] = "=MAX(DR2:DR{})".format(q)
            ws['DR{}'.format(s)] = "=MIN(DR2:DR{})".format(q)
            ws['DR{}'.format(t)] = "=ROUND(AVERAGE(DR2:DR{}),2)".format(q)
            ws['DS{}'.format(r)] = "=MAX(DS2:DS{})".format(q)
            ws['DS{}'.format(s)] = "=MIN(DS2:DS{})".format(q)
            ws['DS{}'.format(t)] = "=ROUND(AVERAGE(DS2:DS{}),2)".format(q)
            ws['DT{}'.format(r)] = "=MAX(DT2:DT{})".format(q)
            ws['DT{}'.format(s)] = "=MIN(DT2:DT{})".format(q)
            ws['DT{}'.format(t)] = "=ROUND(AVERAGE(DT2:DT{}),2)".format(q)

            # INISIASI MAPEL
            ws['DW{}'.format(r)] = "=SUM(DW2:DW{})".format(q)
            ws['DX{}'.format(r)] = "=SUM(DX2:DX{})".format(q)
            ws['DY{}'.format(r)] = "=SUM(DY2:DY{})".format(q)
            ws['DZ{}'.format(r)] = "=SUM(DZ2:DZ{})".format(q)

            # iterasi 5 rata-rata - 1
            # MAPEL NORMAL
            ws['EG{}'.format(
                r)] = "=IF($DW${}=0,$DG${},$DG${}-1)".format(r, r, r)
            ws['EG{}'.format(s)] = "=STDEV(EG2:EG{})".format(q)
            ws['EG{}'.format(t)] = "=MAX(EG2:EG{})".format(q)
            ws['EG{}'.format(u)] = "=MIN(EG2:EG{})".format(q)
            ws['EH{}'.format(
                r)] = "=IF($DX${}=0,$DH${},$DH${}-1)".format(r, r, r)
            ws['EH{}'.format(s)] = "=STDEV(EH2:EH{})".format(q)
            ws['EH{}'.format(t)] = "=MAX(EH2:EH{})".format(q)
            ws['EH{}'.format(u)] = "=MIN(EH2:EH{})".format(q)
            ws['EI{}'.format(
                r)] = "=IF($DY${}=0,$DI${},$DI${}-1)".format(r, r, r)
            ws['EI{}'.format(s)] = "=STDEV(EI2:EI{})".format(q)
            ws['EI{}'.format(t)] = "=MAX(EI2:EI{})".format(q)
            ws['EI{}'.format(u)] = "=MIN(EI2:EI{})".format(q)
            ws['EJ{}'.format(
                r)] = "=IF($DZ${}=0,$DJ${},$DJ${}-1)".format(r, r, r)
            ws['EJ{}'.format(s)] = "=STDEV(EJ2:EJ{})".format(q)
            ws['EJ{}'.format(t)] = "=MAX(EJ2:EJ{})".format(q)
            ws['EJ{}'.format(u)] = "=MIN(EJ2:EJ{})".format(q)

            # JUMLAH MAPEL NORMAL
            ws['EK{}'.format(r)] = "=ROUND(AVERAGE(EK2:EK{}),2)".format(q)
            ws['EK{}'.format(t)] = "=MAX(EK2:EK{})".format(q)
            ws['EK{}'.format(u)] = "=MIN(EK2:EK{})".format(q)

            # Z SCORE
            ws['EL{}'.format(r)] = "=MAX(EL2:EL{})".format(q)
            ws['EM{}'.format(r)] = "=MAX(EM2:EM{})".format(q)
            ws['EN{}'.format(r)] = "=MAX(EN2:EN{})".format(q)
            ws['EO{}'.format(r)] = "=MAX(EO2:EO{})".format(q)

            # NILAI STANDAR
            ws['EP{}'.format(r)] = "=MAX(EP2:EP{})".format(q)
            ws['EP{}'.format(s)] = "=MIN(EP2:EP{})".format(q)
            ws['EP{}'.format(t)] = "=ROUND(AVERAGE(EP2:EP{}),2)".format(q)
            ws['EQ{}'.format(r)] = "=MAX(EQ2:EQ{})".format(q)
            ws['EQ{}'.format(s)] = "=MIN(EQ2:EQ{})".format(q)
            ws['EQ{}'.format(t)] = "=ROUND(AVERAGE(EQ2:EQ{}),2)".format(q)
            ws['ER{}'.format(r)] = "=MAX(ER2:ER{})".format(q)
            ws['ER{}'.format(s)] = "=MIN(ER2:ER{})".format(q)
            ws['ER{}'.format(t)] = "=ROUND(AVERAGE(ER2:ER{}),2)".format(q)
            ws['ES{}'.format(r)] = "=MAX(ES2:ES{})".format(q)
            ws['ES{}'.format(s)] = "=MIN(ES2:ES{})".format(q)
            ws['ES{}'.format(t)] = "=ROUND(AVERAGE(ES2:ES{}),2)".format(q)
            ws['ET{}'.format(r)] = "=MAX(ET2:ET{})".format(q)
            ws['ET{}'.format(s)] = "=MIN(ET2:ET{})".format(q)
            ws['ET{}'.format(t)] = "=ROUND(AVERAGE(ET2:ET{}),2)".format(q)

            # INISIASI MAPEL
            ws['EW{}'.format(r)] = "=SUM(EW2:EW{})".format(q)
            ws['EX{}'.format(r)] = "=SUM(EX2:EX{})".format(q)
            ws['EY{}'.format(r)] = "=SUM(EY2:EY{})".format(q)
            ws['EZ{}'.format(r)] = "=SUM(EZ2:EZ{})".format(q)

            # Jumlah Soal
            ws['F{}'.format(v)] = 'JUMLAH SOAL'
            ws['G{}'.format(v)] = JML_SOAL_GEO
            ws['H{}'.format(v)] = JML_SOAL_EKO
            ws['I{}'.format(v)] = JML_SOAL_SEJ
            ws['J{}'.format(v)] = JML_SOAL_SOS

            # Z Score
            ws['B1'] = 'NAMA SISWA_A'
            ws['C1'] = 'NOMOR NF_A'
            ws['D1'] = 'KELAS_A'
            ws['E1'] = 'NAMA SEKOLAH_A'
            ws['F1'] = 'LOKASI_A'
            ws['G1'] = 'GEO_A'
            ws['H1'] = 'EKO_A'
            ws['I1'] = 'SEJ_A'
            ws['J1'] = 'SOS_A'
            ws['K1'] = 'JML_A'
            ws['L1'] = 'Z_GEO_A'
            ws['M1'] = 'Z_EKO_A'
            ws['N1'] = 'Z_SEJ_A'
            ws['O1'] = 'Z_SOS_A'
            ws['P1'] = 'S_GEO_A'
            ws['Q1'] = 'S_EKO_A'
            ws['R1'] = 'S_SEJ_A'
            ws['S1'] = 'S_SOS_A'
            ws['T1'] = 'S_JML_A'
            ws['U1'] = 'RANK NAS._A'
            ws['V1'] = 'RANK LOK._A'

            ws['L1'].font = Font(bold=False, name='Calibri', size=11)
            ws['M1'].font = Font(bold=False, name='Calibri', size=11)
            ws['N1'].font = Font(bold=False, name='Calibri', size=11)
            ws['O1'].font = Font(bold=False, name='Calibri', size=11)
            ws['P1'].font = Font(bold=False, name='Calibri', size=11)
            ws['Q1'].font = Font(bold=False, name='Calibri', size=11)
            ws['R1'].font = Font(bold=False, name='Calibri', size=11)
            ws['S1'].font = Font(bold=False, name='Calibri', size=11)
            ws['T1'].font = Font(bold=False, name='Calibri', size=11)
            ws['U1'].font = Font(bold=False, name='Calibri', size=11)
            ws['V1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['B1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['C1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['D1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['E1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['F1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['G1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['H1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['I1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['J1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['K1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['L1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['M1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['N1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['O1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['P1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['Q1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['R1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['S1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['T1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['U1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['V1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')

            # tambahan
            ws['W1'] = 'GEO_20_A'
            ws['X1'] = 'EKO_20_A'
            ws['Y1'] = 'SEJ_20_A'
            ws['Z1'] = 'SOS_20_A'
            ws['W1'].font = Font(bold=False, name='Calibri', size=11)
            ws['X1'].font = Font(bold=False, name='Calibri', size=11)
            ws['Y1'].font = Font(bold=False, name='Calibri', size=11)
            ws['Z1'].font = Font(bold=False, name='Calibri', size=11)
            ws['W1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['X1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['Y1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['Z1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')

            for row in range(2, q+1):
                ws['K{}'.format(
                    row)] = '=SUM(G{}:J{})'.format(row, row, row)
                ws['L{}'.format(
                    row)] = '=IFERROR(ROUND(IF(G{}="","",(G{}-G${})/G${}),2),"")'.format(row, row, r, s)
                ws['M{}'.format(
                    row)] = '=IFERROR(ROUND(IF(H{}="","",(H{}-H${})/H${}),2),"")'.format(row, row, r, s)
                ws['N{}'.format(
                    row)] = '=IFERROR(ROUND(IF(I{}="","",(I{}-I${})/I${}),2),"")'.format(row, row, r, s)
                ws['O{}'.format(
                    row)] = '=IFERROR(ROUND(IF(J{}="","",(J{}-J${})/J${}),2),"")'.format(row, row, r, s)
                ws['P{}'.format(
                    row)] = '=IFERROR(ROUND(IF(G{}="","",IF(70+30*L{}/$L${}<20,20,70+30*L{}/$L${})),2),"")'.format(row, row, r, row, r)
                ws['Q{}'.format(
                    row)] = '=IFERROR(ROUND(IF(H{}="","",IF(70+30*M{}/$M${}<20,20,70+30*M{}/$M${})),2),"")'.format(row, row, r, row, r)
                ws['R{}'.format(
                    row)] = '=IFERROR(ROUND(IF(I{}="","",IF(70+30*N{}/$N${}<20,20,70+30*N{}/$N${})),2),"")'.format(row, row, r, row, r)
                ws['S{}'.format(
                    row)] = '=IFERROR(ROUND(IF(J{}="","",IF(70+30*O{}/$O${}<20,20,70+30*O{}/$P${})),2),"")'.format(row, row, r, row, r)

                ws['T{}'.format(row)] = '=IF(SUM(P{}:S{})=0,"",SUM(P{}:S{}))'.format(
                    row, row, row, row)
                ws['U{}'.format(row)] = '=IF(T{}="","",RANK(T{},$T$2:$T${}))'.format(
                    row, row, q)
                ws['V{}'.format(
                    row)] = '=IF(U{}="","",COUNTIFS($F$2:$F${},F{},$U$2:$U${},"<"&U{})+1)'.format(row, q, row, q, row)
                # TAMBAHAN
                ws['W{}'.format(row)] = '=IF($G${}=25,IF(AND(G{}>4,P{}=20),1,""),IF($G${}=30,IF(AND(G{}>5,P{}=20),1,""),IF($G${}=35,IF(AND(G{}>6,P{}=20),1,""),IF($G${}=40,IF(AND(G{}>7,P{}=20),1,""),IF($G${}=45,IF(AND(G{}>8,P{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['X{}'.format(row)] = '=IF($H${}=25,IF(AND(H{}>4,Q{}=20),1,""),IF($H${}=30,IF(AND(H{}>5,Q{}=20),1,""),IF($H${}=35,IF(AND(H{}>6,Q{}=20),1,""),IF($H${}=40,IF(AND(H{}>7,Q{}=20),1,""),IF($H${}=45,IF(AND(H{}>8,Q{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['Y{}'.format(row)] = '=IF($I${}=25,IF(AND(I{}>4,R{}=20),1,""),IF($I${}=30,IF(AND(I{}>5,R{}=20),1,""),IF($I${}=35,IF(AND(I{}>6,R{}=20),1,""),IF($I${}=40,IF(AND(I{}>7,R{}=20),1,""),IF($I${}=45,IF(AND(I{}>8,R{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['Z{}'.format(row)] = '=IF($J${}=25,IF(AND(J{}>4,S{}=20),1,""),IF($J${}=30,IF(AND(J{}>5,S{}=20),1,""),IF($J${}=35,IF(AND(J{}>6,S{}=20),1,""),IF($J${}=40,IF(AND(J{}>7,S{}=20),1,""),IF($J${}=45,IF(AND(J{}>8,S{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

            # Z Score iterasi 1
            ws['AB1'] = 'NAMA SISWA_B'
            ws['AC1'] = 'NOMOR NF_B'
            ws['AD1'] = 'KELAS_B'
            ws['AE1'] = 'NAMA SEKOLAH_B'
            ws['AF1'] = 'LOKASI_B'
            ws['AG1'] = 'GEO_B'
            ws['AH1'] = 'EKO_B'
            ws['AI1'] = 'SEJ_B'
            ws['AJ1'] = 'SOS_B'
            ws['AK1'] = 'JML_B'
            ws['AL1'] = 'Z_GEO_B'
            ws['AM1'] = 'Z_EKO_B'
            ws['AN1'] = 'Z_SEJ_B'
            ws['AO1'] = 'Z_SOS_B'
            ws['AP1'] = 'S_GEO_B'
            ws['AQ1'] = 'S_EKO_B'
            ws['AR1'] = 'S_SEJ_B'
            ws['AS1'] = 'S_SOS_B'
            ws['AT1'] = 'S_JML_B'
            ws['AU1'] = 'RANK NAS._B'
            ws['AV1'] = 'RANK LOK._B'

            ws['AL1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AM1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AN1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AO1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AP1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AQ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AR1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AS1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AT1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AU1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AV1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['AB1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AC1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AD1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AE1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AF1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AG1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AH1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AI1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AJ1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AK1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AL1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AM1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AN1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AO1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AP1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AQ1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AR1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AS1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AT1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AU1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AV1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')

            # tambahan
            ws['AW1'] = 'GEO_20'
            ws['AX1'] = 'EKO_20'
            ws['AY1'] = 'SEJ_20'
            ws['AZ1'] = 'SOS_20'
            ws['AW1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AX1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AY1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AZ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AW1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AX1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AY1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AZ1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')

            for row in range(2, q+1):
                # Tambahan
                ws['AB{}'.format(row)] = '=B{}'.format(row)
                ws['AC{}'.format(row)] = '=C{}'.format(row, row)
                ws['AD{}'.format(row)] = '=D{}'.format(row, row)
                ws['AE{}'.format(row)] = '=E{}'.format(row, row)
                ws['AF{}'.format(row)] = '=F{}'.format(row, row)
                ws['AG{}'.format(row)] = '=IF(G{}="","",G{})'.format(row, row)
                ws['AH{}'.format(row)] = '=IF(H{}="","",H{})'.format(row, row)
                ws['AI{}'.format(row)] = '=IF(I{}="","",I{})'.format(row, row)
                ws['AJ{}'.format(row)] = '=IF(J{}="","",J{})'.format(row, row)
                ws['AK{}'.format(row)] = '=IF(K{}="","",K{})'.format(row, row)

                ws['AL{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AG{}="","",(AG{}-AG${})/AG${}),2),"")'.format(row, row, r, s)
                ws['AM{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AH{}="","",(AH{}-AH${})/AH${}),2),"")'.format(row, row, r, s)
                ws['AN{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AI{}="","",(AI{}-AI${})/AI${}),2),"")'.format(row, row, r, s)
                ws['AO{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AJ{}="","",(AJ{}-AJ${})/AJ${}),2),"")'.format(row, row, r, s)

                ws['AP{}'.format(
                    row)] = '=IFERROR(ROUND(IF(G{}="","",IF(70+30*AL{}/$AL${}<20,20,70+30*AL{}/$AL${})),2),"")'.format(row, row, r, row, r)
                ws['AQ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(H{}="","",IF(70+30*AM{}/$AM{}<20,20,70+30*AM{}/$AM${})),2),"")'.format(row, row, r, row, r)
                ws['AR{}'.format(
                    row)] = '=IFERROR(ROUND(IF(I{}="","",IF(70+30*AN{}/$AN${}<20,20,70+30*AN{}/$AN${})),2),"")'.format(row, row, r, row, r)
                ws['AS{}'.format(
                    row)] = '=IFERROR(ROUND(IF(J{}="","",IF(70+30*AO{}/$AO${}<20,20,70+30*AO{}/$AO${})),2),"")'.format(row, row, r, row, r)

                ws['AT{}'.format(row)] = '=IF(SUM(AP{}:AS{})=0,"",SUM(AP{}:AS{}))'.format(
                    row, row, row, row)
                ws['AU{}'.format(row)] = '=IF(AT{}="","",RANK(AT{},$AT$2:$AT${}))'.format(
                    row, row, q)
                ws['AV{}'.format(
                    row)] = '=IF(AU{}="","",COUNTIFS($AF$2:$AF${},AF{},$AU$2:$AU${},"<"&AU{})+1)'.format(row, q, row, q, row)
                # TAMBAHAN
                ws['AW{}'.format(row)] = '=IF($G${}=25,IF(AND(AG{}>4,AP{}=20),1,""),IF($G${}=30,IF(AND(AG{}>5,AP{}=20),1,""),IF($G${}=35,IF(AND(AG{}>6,AP{}=20),1,""),IF($G${}=40,IF(AND(AG{}>7,AP{}=20),1,""),IF($G${}=45,IF(AND(AG{}>8,AP{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['AX{}'.format(row)] = '=IF($H${}=25,IF(AND(AH{}>4,AQ{}=20),1,""),IF($H${}=30,IF(AND(AH{}>5,AQ{}=20),1,""),IF($H${}=35,IF(AND(AH{}>6,AQ{}=20),1,""),IF($H${}=40,IF(AND(AH{}>7,AQ{}=20),1,""),IF($H${}=45,IF(AND(AH{}>8,AQ{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['AY{}'.format(row)] = '=IF($I${}=25,IF(AND(AI{}>4,AR{}=20),1,""),IF($I${}=30,IF(AND(AI{}>5,AR{}=20),1,""),IF($I${}=35,IF(AND(AI{}>6,AR{}=20),1,""),IF($I${}=40,IF(AND(AI{}>7,AR{}=20),1,""),IF($I${}=45,IF(AND(AI{}>8,AR{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['AZ{}'.format(row)] = '=IF($J${}=25,IF(AND(AJ{}>4,AS{}=20),1,""),IF($J${}=30,IF(AND(AJ{}>5,AS{}=20),1,""),IF($J${}=35,IF(AND(AJ{}>6,AS{}=20),1,""),IF($J${}=40,IF(AND(AJ{}>7,AS{}=20),1,""),IF($J${}=45,IF(AND(AJ{}>8,AS{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

            # Z Score iterasi 2
            ws['BB1'] = 'NAMA SISWA_C'
            ws['BC1'] = 'NOMOR NF_c'
            ws['BD1'] = 'KELAS_C'
            ws['BE1'] = 'NAMA SEKOLAH_C'
            ws['BF1'] = 'LOKASI_C'
            ws['BG1'] = 'GEO_C'
            ws['BH1'] = 'EKO_C'
            ws['BI1'] = 'SEJ_C'
            ws['BJ1'] = 'SOS_C'
            ws['BK1'] = 'JML_C'
            ws['BL1'] = 'Z_GEO_C'
            ws['BM1'] = 'Z_EKO_C'
            ws['BN1'] = 'Z_SEJ_C'
            ws['BO1'] = 'Z_SOS_C'
            ws['BP1'] = 'S_GEO_C'
            ws['BQ1'] = 'S_EKO_C'
            ws['BR1'] = 'S_SEJ_C'
            ws['BS1'] = 'S_SOS_C'
            ws['BT1'] = 'S_JML_C'
            ws['BU1'] = 'RANK NAS._C'
            ws['BV1'] = 'RANK LOK._C'

            ws['BL1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BM1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BN1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BO1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BP1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BQ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BR1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BS1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BT1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BU1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BV1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['BB1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BC1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BD1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BE1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BF1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BG1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BH1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BI1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BJ1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BK1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BL1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BM1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BN1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BO1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BP1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BQ1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BR1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BS1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BT1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BU1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BV1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')

            # tambahan
            ws['BW1'] = 'GEO_20_C'
            ws['BX1'] = 'EKO_20_C'
            ws['BY1'] = 'SEJ_20_C'
            ws['BZ1'] = 'SOS_20_C'
            ws['BW1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BX1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BY1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BZ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BW1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BX1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BY1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BZ1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')

            for row in range(2, q+1):
                # Tambahan
                ws['BB{}'.format(row)] = '=AB{}'.format(row)
                ws['BC{}'.format(row)] = '=AC{}'.format(row, row)
                ws['BD{}'.format(row)] = '=AD{}'.format(row, row)
                ws['BE{}'.format(row)] = '=AE{}'.format(row, row)
                ws['BF{}'.format(row)] = '=AF{}'.format(row, row)
                ws['BG{}'.format(row)] = '=IF(AG{}="","",AG{})'.format(
                    row, row)
                ws['BH{}'.format(row)] = '=IF(AH{}="","",AH{})'.format(
                    row, row)
                ws['BI{}'.format(row)] = '=IF(AI{}="","",AI{})'.format(
                    row, row)
                ws['BJ{}'.format(row)] = '=IF(AJ{}="","",AJ{})'.format(
                    row, row)
                ws['BK{}'.format(row)] = '=IF(AK{}="","",AK{})'.format(
                    row, row)

                ws['BL{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BG{}="","",(BG{}-BG${})/BG${}),2),"")'.format(row, row, r, s)
                ws['BM{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BH{}="","",(BH{}-BH${})/BH${}),2),"")'.format(row, row, r, s)
                ws['BN{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BI{}="","",(BI{}-BI${})/BI${}),2),"")'.format(row, row, r, s)
                ws['BO{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BJ{}="","",(BJ{}-BJ${})/BJ${}),2),"")'.format(row, row, r, s)

                ws['BP{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BG{}="","",IF(70+30*BL{}/$BL${}<20,20,70+30*BL{}/$BL${})),2),"")'.format(row, row, r, row, r)
                ws['BQ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BH{}="","",IF(70+30*BM{}/$BM{}<20,20,70+30*BM{}/$BM${})),2),"")'.format(row, row, r, row, r)
                ws['BR{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BI{}="","",IF(70+30*BN{}/$BN${}<20,20,70+30*BN{}/$BN${})),2),"")'.format(row, row, r, row, r)
                ws['BS{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BJ{}="","",IF(70+30*BO{}/$BO${}<20,20,70+30*BO{}/$BO${})),2),"")'.format(row, row, r, row, r)

                ws['BT{}'.format(row)] = '=IF(SUM(BP{}:BS{})=0,"",SUM(BP{}:BS{}))'.format(
                    row, row, row, row)
                ws['BU{}'.format(row)] = '=IF(BT{}="","",RANK(BT{},$BT$2:$BT${}))'.format(
                    row, row, q)
                ws['BV{}'.format(
                    row)] = '=IF(BU{}="","",COUNTIFS($BF$2:$BF${},BF{},$BU$2:$BU${},"<"&BU{})+1)'.format(row, q, row, q, row)
                # TAMBAHAN
                ws['BW{}'.format(row)] = '=IF($G${}=25,IF(AND(BG{}>4,BP{}=20),1,""),IF($G${}=30,IF(AND(BG{}>5,BP{}=20),1,""),IF($G${}=35,IF(AND(BG{}>6,BP{}=20),1,""),IF($G${}=40,IF(AND(BG{}>7,BP{}=20),1,""),IF($G${}=45,IF(AND(BG{}>8,BP{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['BX{}'.format(row)] = '=IF($H${}=25,IF(AND(BH{}>4,BQ{}=20),1,""),IF($H${}=30,IF(AND(BH{}>5,BQ{}=20),1,""),IF($H${}=35,IF(AND(BH{}>6,BQ{}=20),1,""),IF($H${}=40,IF(AND(BH{}>7,BQ{}=20),1,""),IF($H${}=45,IF(AND(BH{}>8,BQ{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['BY{}'.format(row)] = '=IF($I${}=25,IF(AND(BI{}>4,BR{}=20),1,""),IF($I${}=30,IF(AND(BI{}>5,BR{}=20),1,""),IF($I${}=35,IF(AND(BI{}>6,BR{}=20),1,""),IF($I${}=40,IF(AND(BI{}>7,BR{}=20),1,""),IF($I${}=45,IF(AND(BI{}>8,BR{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['BZ{}'.format(row)] = '=IF($J${}=25,IF(AND(BJ{}>4,BS{}=20),1,""),IF($J${}=30,IF(AND(BJ{}>5,BS{}=20),1,""),IF($J${}=35,IF(AND(BJ{}>6,BS{}=20),1,""),IF($J${}=40,IF(AND(BJ{}>7,BS{}=20),1,""),IF($J${}=45,IF(AND(BJ{}>8,BS{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

            # Z Score iterasi 3
            ws['CB1'] = 'NAMA SISWA_D'
            ws['CC1'] = 'NOMOR NF_D'
            ws['CD1'] = 'KELAS_D'
            ws['CE1'] = 'NAMA SEKOLAH_D'
            ws['CF1'] = 'LOKASI_D'
            ws['CG1'] = 'GEO_D'
            ws['CH1'] = 'EKO_D'
            ws['CI1'] = 'SEJ_D'
            ws['CJ1'] = 'SOS_D'
            ws['CK1'] = 'JML_D'
            ws['CL1'] = 'Z_GEO_D'
            ws['CM1'] = 'Z_EKO_D'
            ws['CN1'] = 'Z_SEJ_D'
            ws['CO1'] = 'Z_SOS_D'
            ws['CP1'] = 'S_GEO_D'
            ws['CQ1'] = 'S_EKO_D'
            ws['CR1'] = 'S_SEJ_D'
            ws['CS1'] = 'S_SOS_D'
            ws['CT1'] = 'S_JML_D'
            ws['CU1'] = 'RANK NAS._D'
            ws['CV1'] = 'RANK LOK._D'

            ws['CL1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CM1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CN1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CO1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CP1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CQ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CR1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CS1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CT1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CU1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CV1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['CB1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CC1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CD1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CE1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CF1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CG1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CH1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CI1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CJ1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CK1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CL1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CM1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CN1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CO1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CP1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CQ1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CR1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CS1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CT1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CU1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CV1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')

            # tambahan
            ws['CW1'] = 'GEO_20_D'
            ws['CX1'] = 'EKO_20_D'
            ws['CY1'] = 'SEJ_20_D'
            ws['CZ1'] = 'SOS_20_D'
            ws['CW1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CX1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CY1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CZ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CW1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CX1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CY1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CZ1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')

            for row in range(2, q+1):
                ws['CB{}'.format(row)] = '=BB{}'.format(row)
                ws['CC{}'.format(row)] = '=BC{}'.format(row, row)
                ws['CD{}'.format(row)] = '=BD{}'.format(row, row)
                ws['CE{}'.format(row)] = '=BE{}'.format(row, row)
                ws['CF{}'.format(row)] = '=BF{}'.format(row, row)
                ws['CG{}'.format(row)] = '=IF(BG{}="","",BG{})'.format(
                    row, row)
                ws['CH{}'.format(row)] = '=IF(BH{}="","",BH{})'.format(
                    row, row)
                ws['CI{}'.format(row)] = '=IF(BI{}="","",BI{})'.format(
                    row, row)
                ws['CJ{}'.format(row)] = '=IF(BJ{}="","",BJ{})'.format(
                    row, row)
                ws['CK{}'.format(row)] = '=IF(BK{}="","",BK{})'.format(
                    row, row)

                ws['CL{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CG{}="","",(CG{}-CG${})/CG${}),2),"")'.format(row, row, r, s)
                ws['CM{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CH{}="","",(CH{}-CH${})/CH${}),2),"")'.format(row, row, r, s)
                ws['CN{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CI{}="","",(CI{}-CI${})/CI${}),2),"")'.format(row, row, r, s)
                ws['CO{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CJ{}="","",(CJ{}-CJ${})/CJ${}),2),"")'.format(row, row, r, s)

                ws['CP{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CG{}="","",IF(70+30*CL{}/$CL${}<20,20,70+30*CL{}/$CL${})),2),"")'.format(row, row, r, row, r)
                ws['CQ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CH{}="","",IF(70+30*CM{}/$CM{}<20,20,70+30*CM{}/$CM${})),2),"")'.format(row, row, r, row, r)
                ws['CR{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CI{}="","",IF(70+30*CN{}/$CN${}<20,20,70+30*CN{}/$CN${})),2),"")'.format(row, row, r, row, r)
                ws['CS{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CJ{}="","",IF(70+30*CO{}/$CO${}<20,20,70+30*CO{}/$CO${})),2),"")'.format(row, row, r, row, r)

                ws['CT{}'.format(row)] = '=IF(SUM(CP{}:CS{})=0,"",SUM(CP{}:CS{}))'.format(
                    row, row, row, row)
                ws['CU{}'.format(row)] = '=IF(CT{}="","",RANK(CT{},$CT$2:$CT${}))'.format(
                    row, row, q)
                ws['CV{}'.format(
                    row)] = '=IF(CU{}="","",COUNTIFS($CF$2:$CF${},CF{},$CU$2:$CU${},"<"&CU{})+1)'.format(row, q, row, q, row)
                # TAMBAHAN
                ws['CW{}'.format(row)] = '=IF($G${}=25,IF(AND(CG{}>4,CP{}=20),1,""),IF($G${}=30,IF(AND(CG{}>5,CP{}=20),1,""),IF($G${}=35,IF(AND(CG{}>6,CP{}=20),1,""),IF($G${}=40,IF(AND(CG{}>7,CP{}=20),1,""),IF($G${}=45,IF(AND(CG{}>8,CP{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['CX{}'.format(row)] = '=IF($H${}=25,IF(AND(CH{}>4,CQ{}=20),1,""),IF($H${}=30,IF(AND(CH{}>5,CQ{}=20),1,""),IF($H${}=35,IF(AND(CH{}>6,CQ{}=20),1,""),IF($H${}=40,IF(AND(CH{}>7,CQ{}=20),1,""),IF($H${}=45,IF(AND(CH{}>8,CQ{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['CY{}'.format(row)] = '=IF($I${}=25,IF(AND(CI{}>4,CR{}=20),1,""),IF($I${}=30,IF(AND(CI{}>5,CR{}=20),1,""),IF($I${}=35,IF(AND(CI{}>6,CR{}=20),1,""),IF($I${}=40,IF(AND(CI{}>7,CR{}=20),1,""),IF($I${}=45,IF(AND(CI{}>8,CR{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['CZ{}'.format(row)] = '=IF($J${}=25,IF(AND(CJ{}>4,CS{}=20),1,""),IF($J${}=30,IF(AND(CJ{}>5,CS{}=20),1,""),IF($J${}=35,IF(AND(CJ{}>6,CS{}=20),1,""),IF($J${}=40,IF(AND(CJ{}>7,CS{}=20),1,""),IF($J${}=45,IF(AND(CJ{}>8,CS{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

            # Z Score iterasi 4
            ws['DB1'] = 'NAMA SISWA_E'
            ws['DC1'] = 'NOMOR NF_E'
            ws['DD1'] = 'KELAS_E'
            ws['DE1'] = 'NAMA SEKOLAH_E'
            ws['DF1'] = 'LOKASI_E'
            ws['DG1'] = 'GEO_E'
            ws['DH1'] = 'EKO_E'
            ws['DI1'] = 'SEJ_E'
            ws['DJ1'] = 'SOS_E'
            ws['DK1'] = 'JML_E'
            ws['DL1'] = 'Z_GEO_E'
            ws['DM1'] = 'Z_EKO_E'
            ws['DN1'] = 'Z_SEJ_E'
            ws['DO1'] = 'Z_SOS_E'
            ws['DP1'] = 'S_GEO_E'
            ws['DQ1'] = 'S_EKO_E'
            ws['DR1'] = 'S_SEJ_E'
            ws['DS1'] = 'S_SOS_E'
            ws['DT1'] = 'S_JML_E'
            ws['DU1'] = 'RANK NAS._E'
            ws['DV1'] = 'RANK LOK._E'

            ws['DL1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DM1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DN1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DO1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DP1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DQ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DR1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DS1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DT1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DU1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DV1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['DB1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DC1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DD1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DE1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DF1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DG1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DH1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DI1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DJ1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DK1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DL1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DM1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DN1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DO1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DP1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DQ1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DR1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DS1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DT1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DU1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DV1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')

            # tambahan
            ws['DW1'] = 'GEO_20'
            ws['DX1'] = 'EKO_20'
            ws['DY1'] = 'SEJ_20'
            ws['DZ1'] = 'SOS_20'
            ws['DW1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DX1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DY1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DZ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DW1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DX1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DY1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DZ1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')

            for row in range(2, q+1):
                # Tambahan
                ws['DB{}'.format(row)] = '=CB{}'.format(row)
                ws['DC{}'.format(row)] = '=CC{}'.format(row, row)
                ws['DD{}'.format(row)] = '=CD{}'.format(row, row)
                ws['DE{}'.format(row)] = '=CE{}'.format(row, row)
                ws['DF{}'.format(row)] = '=CF{}'.format(row, row)
                ws['DG{}'.format(row)] = '=IF(CG{}="","",CG{})'.format(
                    row, row)
                ws['DH{}'.format(row)] = '=IF(CH{}="","",CH{})'.format(
                    row, row)
                ws['DI{}'.format(row)] = '=IF(CI{}="","",CI{})'.format(
                    row, row)
                ws['DJ{}'.format(row)] = '=IF(CJ{}="","",CJ{})'.format(
                    row, row)
                ws['DK{}'.format(row)] = '=IF(CK{}="","",CK{})'.format(
                    row, row)

                ws['DL{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DG{}="","",(DG{}-DG${})/DG${}),2),"")'.format(row, row, r, s)
                ws['DM{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DH{}="","",(DH{}-DH${})/DH${}),2),"")'.format(row, row, r, s)
                ws['DN{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DI{}="","",(DI{}-DI${})/DI${}),2),"")'.format(row, row, r, s)
                ws['DO{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DJ{}="","",(DJ{}-DJ${})/DJ${}),2),"")'.format(row, row, r, s)

                ws['DP{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DG{}="","",IF(70+30*DL{}/$DL${}<20,20,70+30*DL{}/$DL${})),2),"")'.format(row, row, r, row, r)
                ws['DQ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DH{}="","",IF(70+30*DM{}/$DM{}<20,20,70+30*DM{}/$DM${})),2),"")'.format(row, row, r, row, r)
                ws['DR{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DI{}="","",IF(70+30*DN{}/$DN${}<20,20,70+30*DN{}/$DN${})),2),"")'.format(row, row, r, row, r)
                ws['DS{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DJ{}="","",IF(70+30*DO{}/$DO${}<20,20,70+30*DO{}/$DO${})),2),"")'.format(row, row, r, row, r)

                ws['DT{}'.format(row)] = '=IF(SUM(DP{}:DS{})=0,"",SUM(DP{}:DS{}))'.format(
                    row, row, row, row)
                ws['DU{}'.format(row)] = '=IF(DT{}="","",RANK(DT{},$DT$2:$DT${}))'.format(
                    row, row, q)
                ws['DV{}'.format(
                    row)] = '=IF(DU{}="","",COUNTIFS($DF$2:$DF${},DF{},$DU$2:$DU${},"<"&DU{})+1)'.format(row, q, row, q, row)
                # TAMBAHAN
                ws['DW{}'.format(row)] = '=IF($G${}=25,IF(AND(DG{}>4,DP{}=20),1,""),IF($G${}=30,IF(AND(DG{}>5,DP{}=20),1,""),IF($G${}=35,IF(AND(DG{}>6,DP{}=20),1,""),IF($G${}=40,IF(AND(DG{}>7,DP{}=20),1,""),IF($G${}=45,IF(AND(DG{}>8,DP{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['DX{}'.format(row)] = '=IF($H${}=25,IF(AND(DH{}>4,DQ{}=20),1,""),IF($H${}=30,IF(AND(DH{}>5,DQ{}=20),1,""),IF($H${}=35,IF(AND(DH{}>6,DQ{}=20),1,""),IF($H${}=40,IF(AND(DH{}>7,DQ{}=20),1,""),IF($H${}=45,IF(AND(DH{}>8,DQ{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['DY{}'.format(row)] = '=IF($I${}=25,IF(AND(DI{}>4,DR{}=20),1,""),IF($I${}=30,IF(AND(DI{}>5,DR{}=20),1,""),IF($I${}=35,IF(AND(DI{}>6,DR{}=20),1,""),IF($I${}=40,IF(AND(DI{}>7,DR{}=20),1,""),IF($I${}=45,IF(AND(DI{}>8,DR{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['DZ{}'.format(row)] = '=IF($J${}=25,IF(AND(DJ{}>4,DS{}=20),1,""),IF($J${}=30,IF(AND(DJ{}>5,DS{}=20),1,""),IF($J${}=35,IF(AND(DJ{}>6,DS{}=20),1,""),IF($J${}=40,IF(AND(DJ{}>7,DS{}=20),1,""),IF($J${}=45,IF(AND(DJ{}>8,DS{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

            # Z Score iterasi 5
            ws['EB1'] = 'NAMA SISWA'
            ws['EC1'] = 'NOMOR NF'
            ws['ED1'] = 'KELAS'
            ws['EE1'] = 'NAMA SEKOLAH'
            ws['EF1'] = 'LOKASI'
            ws['EG1'] = 'GEO'
            ws['EH1'] = 'EKO'
            ws['EI1'] = 'SEJ'
            ws['EJ1'] = 'SOS'
            ws['EK1'] = 'JML'
            ws['EL1'] = 'Z_GEO'
            ws['EM1'] = 'Z_EKO'
            ws['EN1'] = 'Z_SEJ'
            ws['EO1'] = 'Z_SOS'
            ws['EP1'] = 'S_GEO'
            ws['EQ1'] = 'S_EKO'
            ws['ER1'] = 'S_SEJ'
            ws['ES1'] = 'S_SOS'
            ws['ET1'] = 'S_JML'
            ws['EU1'] = 'RANK NAS.'
            ws['EV1'] = 'RANK LOK.'

            ws['EL1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EM1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EN1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EO1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EP1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EQ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['ER1'].font = Font(bold=False, name='Calibri', size=11)
            ws['ES1'].font = Font(bold=False, name='Calibri', size=11)
            ws['ET1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EU1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EV1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['EB1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EC1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['ED1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EE1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EF1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EG1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EH1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EI1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EJ1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EK1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EL1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EM1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EN1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EO1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EP1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EQ1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['ER1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['ES1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['ET1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EU1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EV1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')

            # tambahan
            ws['EW1'] = 'GEO_20'
            ws['EX1'] = 'EKO_20'
            ws['EY1'] = 'SEJ_20'
            ws['EZ1'] = 'SOS_20'
            ws['EW1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EX1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EY1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EZ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EW1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EX1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EY1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EZ1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')

            for row in range(2, q+1):
                # Tambahan
                ws['EB{}'.format(row)] = '=DB{}'.format(row)
                ws['EC{}'.format(row)] = '=DC{}'.format(row, row)
                ws['ED{}'.format(row)] = '=DD{}'.format(row, row)
                ws['EE{}'.format(row)] = '=DE{}'.format(row, row)
                ws['EF{}'.format(row)] = '=DF{}'.format(row, row)
                ws['EG{}'.format(row)] = '=IF(DG{}="","",DG{})'.format(
                    row, row)
                ws['EH{}'.format(row)] = '=IF(DH{}="","",DH{})'.format(
                    row, row)
                ws['EI{}'.format(row)] = '=IF(DI{}="","",DI{})'.format(
                    row, row)
                ws['EJ{}'.format(row)] = '=IF(DJ{}="","",DJ{})'.format(
                    row, row)
                ws['EK{}'.format(row)] = '=IF(DK{}="","",DK{})'.format(
                    row, row)

                ws['EL{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EG{}="","",(EG{}-EG${})/EG${}),2),"")'.format(row, row, r, s)
                ws['EM{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EH{}="","",(EH{}-EH${})/EH${}),2),"")'.format(row, row, r, s)
                ws['EN{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EI{}="","",(EI{}-EI${})/EI${}),2),"")'.format(row, row, r, s)
                ws['EO{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EJ{}="","",(EJ{}-EJ${})/EJ${}),2),"")'.format(row, row, r, s)

                ws['EP{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EG{}="","",IF(70+30*EL{}/$EL${}<20,20,70+30*EL{}/$EL${})),2),"")'.format(row, row, r, row, r)
                ws['EQ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EH{}="","",IF(70+30*EM{}/$EM{}<20,20,70+30*EM{}/$EM${})),2),"")'.format(row, row, r, row, r)
                ws['ER{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EI{}="","",IF(70+30*EN{}/$EN${}<20,20,70+30*EN{}/$EN${})),2),"")'.format(row, row, r, row, r)
                ws['ES{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EJ{}="","",IF(70+30*EO{}/$EO${}<20,20,70+30*EO{}/$EO${})),2),"")'.format(row, row, r, row, r)

                ws['ET{}'.format(row)] = '=IF(SUM(EP{}:ES{})=0,"",SUM(EP{}:ES{}))'.format(
                    row, row, row, row)
                ws['EU{}'.format(row)] = '=IF(ET{}="","",RANK(ET{},$ET$2:$ET${}))'.format(
                    row, row, q)
                ws['EV{}'.format(
                    row)] = '=IF(EU{}="","",COUNTIFS($EF$2:$EF${},EF{},$EU$2:$EU${},"<"&EU{})+1)'.format(row, q, row, q, row)
                # TAMBAHAN
                ws['EW{}'.format(row)] = '=IF($G${}=25,IF(AND(EG{}>4,EP{}=20),1,""),IF($G${}=30,IF(AND(EG{}>5,EP{}=20),1,""),IF($G${}=35,IF(AND(EG{}>6,EP{}=20),1,""),IF($G${}=40,IF(AND(EG{}>7,EP{}=20),1,""),IF($G${}=45,IF(AND(EG{}>8,EP{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['EX{}'.format(row)] = '=IF($H${}=25,IF(AND(EH{}>4,EQ{}=20),1,""),IF($H${}=30,IF(AND(EH{}>5,EQ{}=20),1,""),IF($H${}=35,IF(AND(EH{}>6,EQ{}=20),1,""),IF($H${}=40,IF(AND(EH{}>7,EQ{}=20),1,""),IF($H${}=45,IF(AND(EH{}>8,EQ{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['EY{}'.format(row)] = '=IF($I${}=25,IF(AND(EI{}>4,ER{}=20),1,""),IF($I${}=30,IF(AND(EI{}>5,ER{}=20),1,""),IF($I${}=35,IF(AND(EI{}>6,ER{}=20),1,""),IF($I${}=40,IF(AND(EI{}>7,ER{}=20),1,""),IF($I${}=45,IF(AND(EI{}>8,ER{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['EZ{}'.format(row)] = '=IF($J${}=25,IF(AND(EJ{}>4,ES{}=20),1,""),IF($J${}=30,IF(AND(EJ{}>5,ES{}=20),1,""),IF($J${}=35,IF(AND(EJ{}>6,ES{}=20),1,""),IF($J${}=40,IF(AND(EJ{}>7,ES{}=20),1,""),IF($J${}=45,IF(AND(EJ{}>8,ES{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

            # Mengubah 'KELAS' sesuai dengan nilai yang dipilih dari selectbox 'KELAS'
            kelas = KELAS.lower().replace(" ", "")
            semester = SEMESTER.lower()
            tahun = TAHUN.replace("-", "")
            penilaian = PENILAIAN.lower()
            kurikulum = KURIKULUM.lower()

            path_file = f"{kelas}_{penilaian}_{semester}_{kurikulum}_{tahun}_nilai_std.xlsx"

            # Simpan file ke direktori temporer
            temp_dir = tempfile.gettempdir()
            file_path = temp_dir + '/' + path_file
            wb.save(file_path)

            st.success(
                "File siap diunduh!")

            # Tombol unduh file
            with open(file_path, "rb") as f:
                bytes_data = f.read()
            st.download_button(label="Unduh File", data=bytes_data,
                               file_name=path_file)

            st.warning(
                "Buka file unduhan, klik 'Enable Editing' dan 'Save'")
